import streamlit as st
import pandas as pd
import datetime
import io
import os
import base64
import math
from PIL import Image as PILImage
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
import sqlite3
import toml

# ==========================================
# GLOBAL STYLES & BACKGROUND
# ==========================================
st.markdown("""
<style>

/* Ensure all buttons center their text perfectly */
.stButton > button {
    display: flex !important;
    justify-content: center !important;
    align-items: center !important;
    text-align: center !important;
}
.stButton > button p {
    text-align: center !important;
    width: 100% !important;
    margin: 0 !important;
}

/* Global background */
[data-testid="stAppViewContainer"] {
    background: linear-gradient(135deg, #F0F9FF 0%, #E0F2FE 100%);
    overflow-x: hidden;
}

/* Cherry blossoms */
.sakura {
    position: fixed;
    opacity: 0.25;
    z-index: 0;
    pointer-events: none;
    user-select: none;
    animation: sway 6s ease-in-out infinite alternate;
    filter: drop-shadow(0 4px 6px rgba(244, 114, 182, 0.2));
}
@keyframes sway {
    0% { transform: rotate(-10deg) translateY(0); }
    100% { transform: rotate(10deg) translateY(-30px); }
}
.s1 { top: 10%; left: 5%; font-size: 140px; animation-delay: 0s; }
.s2 { top: 20%; right: 10%; font-size: 110px; animation-delay: 1s; }
.s3 { bottom: 15%; left: 15%; font-size: 120px; animation-delay: 2s; }
.s4 { bottom: 25%; right: 5%; font-size: 160px; animation-delay: 3s; }
</style>

<!-- Cherry blossom elements -->
<div class="sakura s1"><svg viewBox="0 0 100 100" xmlns="http://www.w3.org/2000/svg" style="width: 1em; height: 1em;"><defs><linearGradient id="petalGrad" x1="0%" y1="0%" x2="0%" y2="100%"><stop offset="0%" stop-color="#fdf2f8"/><stop offset="100%" stop-color="#f472b6"/></linearGradient></defs><path d="M 50 90 C 20 60, 15 30, 25 15 C 35 5, 45 10, 50 20 C 55 10, 65 5, 75 15 C 85 30, 80 60, 50 90 Z" fill="url(#petalGrad)"/></svg></div>
<div class="sakura s2"><svg viewBox="0 0 100 100" xmlns="http://www.w3.org/2000/svg" style="width: 1em; height: 1em;"><path d="M 50 90 C 20 60, 15 30, 25 15 C 35 5, 45 10, 50 20 C 55 10, 65 5, 75 15 C 85 30, 80 60, 50 90 Z" fill="url(#petalGrad)"/></svg></div>
<div class="sakura s3"><svg viewBox="0 0 100 100" xmlns="http://www.w3.org/2000/svg" style="width: 1em; height: 1em;"><path d="M 50 90 C 20 60, 15 30, 25 15 C 35 5, 45 10, 50 20 C 55 10, 65 5, 75 15 C 85 30, 80 60, 50 90 Z" fill="url(#petalGrad)"/></svg></div>
<div class="sakura s4"><svg viewBox="0 0 100 100" xmlns="http://www.w3.org/2000/svg" style="width: 1em; height: 1em;"><path d="M 50 90 C 20 60, 15 30, 25 15 C 35 5, 45 10, 50 20 C 55 10, 65 5, 75 15 C 85 30, 80 60, 50 90 Z" fill="url(#petalGrad)"/></svg></div>
""", unsafe_allow_html=True)


# ==========================================
# ROUTING: TRANG HƯỚNG DẪN
# ==========================================
try:
    if hasattr(st, "query_params"):
        page = st.query_params.get("page", "")
    else:
        page = st.experimental_get_query_params().get("page", [""])[0]
except Exception:
    page = ""

if page == "huong_dan":
    st.set_page_config(page_title="Hướng dẫn sử dụng", page_icon="📖", layout="wide")
    st.markdown("<style>#MainMenu {visibility: hidden;} footer {visibility: hidden;} header {visibility: hidden;}</style>", unsafe_allow_html=True)
    
    if st.button("⬅ Trở về giao diện tính bảng chấm công", type="primary"):
        st.session_state.app_page = "chamcong"
        if hasattr(st, "query_params"):
            st.query_params.clear()
        else:
            st.experimental_set_query_params()
        st.rerun()
        
    st.title("📖 Hướng dẫn sử dụng Ứng dụng Chấm công")
    st.markdown("""
    Chào mừng bạn đến với trang hướng dẫn sử dụng! 
    Dưới đây là các bước cơ bản để tính giờ làm việc từ file Excel máy chấm công:
    
    ### 1. Tải lên file Excel
    - Trở về **[Trang chủ](/)** và nhìn sang vùng điều khiển bên trái màn hình.
    - Nhấp vào nút **Tải file lên** (hoặc kéo thả file xuất từ máy chấm công của bạn).
    
    ### 2. Định dạng cột dữ liệu
    - Hệ thống sẽ tự động quét và nhận diện các cột (Mã NV, Tên, Ngày, Giờ vào, Giờ ra).
    - Nếu hệ thống chọn sai, bạn có thể nhấp vào menu thả xuống để tự chỉnh lại cột tương ứng.
    - Nhấn Xác nhận cấu hình để tiếp tục.
    
    ### 3. Cài đặt thời gian chuẩn
    - **Giờ làm việc chuẩn:** Nhập giờ vào ca và giờ tan ca chuẩn của công ty (Ví dụ: 08:00 - 17:00).
    - **Giờ nghỉ trưa:** Điền thời gian bắt đầu và kết thúc nghỉ trưa (hệ thống sẽ tự động trừ giờ này ra khỏi tổng giờ làm).
    
    ### 4. Xử lý và Xuất báo cáo
    - Nhấn nút **Xử lý dữ liệu**.
    - Bảng Dashboard sẽ hiện ra chi tiết: tổng giờ làm, giờ tăng ca (OT), số phút đi trễ, về sớm.
    - Cuối cùng, nhấn **Tải file Excel báo cáo** để tải kết quả dạng lưới chuyên nghiệp về máy tính.
    
    ---
    *Nếu bạn cần hỗ trợ thêm, vui lòng nhấn vào mục Liên hệ ở trang chủ.*
    """)
    st.stop()

# ==========================================
# KHỞI TẠO LƯU TRỮ API KEY
# ==========================================

SECRETS_DIR = ".streamlit"
SECRETS_FILE = os.path.join(SECRETS_DIR, "secrets.toml")

def load_saved_api_key():
    import streamlit as st
    if "GEMINI_API_KEY" in st.secrets:
        return st.secrets["GEMINI_API_KEY"]
    if os.path.exists(SECRETS_FILE):
        try:
            with open(SECRETS_FILE, "r", encoding="utf-8") as f:
                secrets = toml.load(f)
                return secrets.get("GEMINI_API_KEY", "")
        except:
            return ""
    return ""

def save_api_key(key):
    if not os.path.exists(SECRETS_DIR):
        os.makedirs(SECRETS_DIR)
    try:
        with open(SECRETS_FILE, "w", encoding="utf-8") as f:
            toml.dump({"GEMINI_API_KEY": key}, f)
    except Exception:
        pass

# ==========================================
# CƠ SỞ DỮ LIỆU SQLITE
# ==========================================
DB_FILE = 'attendance.db'

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS records (
            ma_nv TEXT, ten_nv TEXT, ngay TEXT,
            gio_vao TEXT, gio_ra TEXT,
            di_tre INTEGER, ve_som INTEGER, ot REAL, tong_gio REAL, ghi_chu TEXT,
            UNIQUE(ma_nv, ngay)
        )
    ''')
    conn.commit()
    conn.close()

init_db()

def save_to_db(df_filtered, mapping):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    records = []
    for _, row in df_filtered.iterrows():
        records.append((
            str(row[mapping['ma_nv']]), str(row[mapping['ten_nv']]),
            str(row["Ngày"]), str(row[mapping['gio_vao']]), str(row[mapping['gio_ra']]),
            int(row.get("Phút đi trễ", 0)), int(row.get("Phút về sớm", 0)),
            float(row.get("Giờ OT", 0)), float(row.get("Số giờ làm thực tế", 0)),
            str(row.get("Ghi chú", ""))
        ))
    c.executemany('''
        INSERT OR REPLACE INTO records
        (ma_nv, ten_nv, ngay, gio_vao, gio_ra, di_tre, ve_som, ot, tong_gio, ghi_chu)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', records)
    conn.commit()
    conn.close()

# ==========================================
# 1. CẤU HÌNH & ASSETS
# ==========================================
APP_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(APP_DIR, "assets")
FAVICON_PATH = os.path.join(ASSETS_DIR, "favicon.png")
LOGO_HEADER_PATH = os.path.join(ASSETS_DIR, "logo_header.png")

def load_favicon():
    try:
        return PILImage.open(FAVICON_PATH)
    except Exception:
        return "📊"

def load_logo_base64():
    try:
        with open(LOGO_HEADER_PATH, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception:
        return None

LOGO_HEADER_B64 = load_logo_base64()

def load_bg_base64():
    bg_path = os.path.join(APP_DIR, "torii.jpg")
    try:
        with open(bg_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    except Exception:
        return ""

BG_B64 = load_bg_base64()

st.set_page_config(page_title="Bảng Chấm Công", page_icon=load_favicon(), layout="wide")

if BG_B64:
    st.markdown(f"""
    <style>
    [data-testid="stAppViewContainer"], .stApp {{
        background: linear-gradient(rgba(240, 244, 248, 0.9), rgba(219, 234, 254, 0.85)), url("data:image/jpeg;base64,{BG_B64}") !important;
        background-size: cover !important;
        background-position: center !important;
        background-attachment: fixed !important;
    }}
    </style>
    """, unsafe_allow_html=True)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Be+Vietnam+Pro:wght@400;500;600;700;800&family=Inter:wght@400;500;600;700&display=swap');

:root {
    --brand-50:  #EFF6FF;
    --brand-100: #DBEAFE;
    --brand-400: #60A5FA;
    --brand-500: #3B82F6;
    --brand-600: #2563EB;
    --brand-700: #1D4ED8;
    --blue-500:  #2563EB;
    --ink-900:   #0F172A;
    --ink-700:   #334155;
    --ink-500:   #64748B;
    --line:      #E2E8F0;
    --bg-color:  #F8FAFC;
    --card-shadow: 0 4px 20px rgba(0, 0, 0, 0.05);
}

html, body, .stApp {
    font-family: 'Inter', 'Be Vietnam Pro', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background-color: var(--bg-color);
}

#MainMenu {visibility: hidden;}
footer {visibility: hidden;}

.stApp { background: var(--bg-color); }
.block-container {
    padding-top: 1.5rem !important;
    padding-left: 2rem !important;
    padding-right: 2rem !important;
    max-width: 98% !important;
}

/* ===== App Header Banner ===== */
.app-header {
    position: relative; overflow: hidden;
    background: linear-gradient(135deg, #1E3A8A 0%, #2563EB 50%, #3B82F6 100%);
    border-radius: 24px; padding: 36px 48px;
    display: flex; align-items: center; gap: 36px;
    margin-bottom: 32px;
    box-shadow: 0 16px 40px rgba(37, 99, 235, 0.2);
    color: white;
}
.app-header::before {
    content: ''; position: absolute; top: -50%; right: -5%;
    width: 350px; height: 350px;
    background: radial-gradient(circle, rgba(255,255,255,0.15) 0%, rgba(255,255,255,0) 70%);
    border-radius: 50%; pointer-events: none;
}
.app-header::after {
    content: ''; position: absolute; bottom: -40%; right: 15%;
    width: 250px; height: 250px;
    background: radial-gradient(circle, rgba(96, 165, 250, 0.2) 0%, rgba(255,255,255,0) 70%);
    border-radius: 50%; pointer-events: none;
}
.app-header-icon {
    font-size: 32px; display: flex; align-items: center; justify-content: center;
    background: #FFFFFF; border-radius: 100px; padding: 12px 32px;
    box-shadow: 0 10px 30px rgba(0, 0, 0, 0.15); z-index: 1;
    transition: transform 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275);
}
.app-header-icon:hover { transform: scale(1.05) translateY(-2px); }
.app-header-icon img { height: 64px; width: auto; display: block; mix-blend-mode: multiply; }
.app-header-title { font-size: 30px; font-weight: 800; margin: 0; letter-spacing: -0.02em; z-index: 1; position: relative; }
.app-header-sub { font-size: 16px; color: rgba(255,255,255,0.9); margin: 8px 0 0 0; z-index: 1; position: relative; font-weight: 500; }
.app-header-badge {
    margin-left: auto;
    background: rgba(255,255,255,0.1); backdrop-filter: blur(12px);
    border: 1px solid rgba(255,255,255,0.25); color: white;
    font-size: 14.5px; font-weight: 600; padding: 12px 24px;
    border-radius: 100px; box-shadow: 0 8px 24px rgba(0,0,0,0.1);
    z-index: 1; transition: all 0.3s ease;
}
.app-header-badge:hover { background: rgba(255,255,255,0.2); transform: translateY(-2px); }

/* ===== Cards ===== */
.card { background: #FFFFFF; border: 1px solid var(--line); border-radius: 16px; padding: 22px 26px; margin-bottom: 18px; box-shadow: var(--card-shadow); }
.card-title { font-size: 15px; font-weight: 700; color: var(--ink-900); margin: 0 0 16px 0; display: flex; align-items: center; gap: 12px; letter-spacing: -0.01em; }
.card-icon { display: inline-flex; align-items: center; justify-content: center; width: 32px; height: 32px; font-size: 16px; border-radius: 10px; background: var(--brand-50); color: var(--brand-600); }

/* Upload hint */
.upload-hint { background-color: #EFF6FF; border-radius: 12px; padding: 12px 18px; font-size: 14px; color: #1E3A8A; font-weight: 500; margin-bottom: 16px; border: 1px solid #BFDBFE; }

/* Buttons */
.stButton > button, .stDownloadButton > button {
    background: linear-gradient(135deg, var(--brand-500), var(--blue-500)) !important;
    color: white !important; border: none !important; border-radius: 12px !important;
    padding: 12px 24px !important; font-size: 14.5px !important; font-weight: 600 !important;
    box-shadow: 0 4px 14px rgba(59,130,246,0.3) !important; transition: all 0.2s ease !important;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    transform: translateY(-2px) !important; box-shadow: 0 6px 18px rgba(59,130,246,0.4) !important; filter: brightness(1.05) !important;
}
.stButton > button:active, .stDownloadButton > button:active { transform: translateY(0) !important; }

/* Streamlit components */
[data-testid="stFileUploader"] { background: #FFFFFF; border-radius: 16px; padding: 12px; border: 1px dashed var(--brand-400); box-shadow: var(--card-shadow); }
[data-testid="stFileUploaderDropzone"] { padding: 20px !important; border-radius: 12px !important; }
[data-testid="stDataFrame"] { border-radius: 16px; overflow: hidden; border: 1px solid var(--line); box-shadow: var(--card-shadow); }
[data-testid="stDateInput"] input, [data-testid="stTimeInput"] input, [data-testid="stNumberInput"] input { border-radius: 10px !important; border: 1px solid var(--line) !important; }
[data-testid="stExpander"] { border: 1px solid var(--line) !important; border-radius: 16px !important; background: #FFFFFF !important; box-shadow: var(--card-shadow) !important; }

/* Metrics */
[data-testid="stMetric"] {
    background: linear-gradient(to bottom right, #FFFFFF, #F8FAFC);
    border: 1px solid var(--line); border-radius: 16px; padding: 18px 20px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.03); border-left: 4px solid var(--brand-500);
    transition: transform 0.2s ease, box-shadow 0.2s ease;
}
[data-testid="stMetric"]:hover { transform: translateY(-2px); box-shadow: 0 8px 20px rgba(0,0,0,0.06); }
[data-testid="stMetricLabel"] { font-size: 13px !important; color: var(--ink-500) !important; font-weight: 500 !important; }
[data-testid="stMetricValue"] { font-family: 'Inter', 'Be Vietnam Pro', sans-serif !important; font-size: 26px !important; font-weight: 700 !important; color: var(--ink-900) !important; }

.stSpinner > div { border-top-color: var(--brand-500) !important; }
.stAlert { border-radius: 12px !important; border: none !important; box-shadow: 0 2px 8px rgba(0,0,0,0.03) !important; }
[data-testid="stSelectbox"] > div { border-radius: 10px !important; }
[data-testid="stMultiSelect"] > div > div { border-radius: 10px !important; }
[data-baseweb="tag"] { background-color: var(--brand-50) !important; color: var(--brand-700) !important; border-radius: 6px !important; }
.stSubheader, h3 { font-size: 16px !important; font-weight: 700 !important; color: var(--ink-900) !important; letter-spacing: -0.01em !important; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #FFFFFF 0%, #F8FAFC 70%, #DBEAFE 100%);
    border-right: 1px solid #BFDBFE;
}
</style>
""", unsafe_allow_html=True)

# ==========================================
# 2. HÀM TIỆN ÍCH
# ==========================================
def to_time_obj(val):
    try:
        is_na = pd.isna(val)
    except (TypeError, ValueError):
        is_na = False
    if is_na or val == "":
        return None
    if isinstance(val, datetime.time):
        return val
    if isinstance(val, datetime.datetime):
        return val.time()
    if isinstance(val, (int, float)):
        if 0 <= val < 1:
            total_seconds = int(val * 86400)
            return datetime.time(hour=total_seconds//3600, minute=(total_seconds%3600)//60, second=total_seconds%60)
    if isinstance(val, str):
        val = val.strip()
        if not val or val.lower() in ('nan', 'nat', 'none'):
            return None
        try:
            return pd.to_datetime(val).time()
        except:
            return None
    return None

def time_to_float(t):
    if t is None:
        return 0.0
    return t.hour + t.minute / 60.0 + t.second / 3600.0

def format_gio_lam(val):
    if val is None or (isinstance(val, float) and math.isnan(val)) or str(val).lower() in ['nan', 'nat', 'none', '']:
        return 0
    val = round(float(val), 1)
    return int(val) if val == int(val) else val

def format_gio_lam_str(val):
    return str(format_gio_lam(val))

# ==========================================
# 3. LOGIC TÍNH TOÁN
# ==========================================
def calculate_working_hours(time_in, time_out, start_chuan=8.0, end_chuan=17.0, lunch_start=12.0, lunch_end=13.0, max_hours=8.0):
    vao_trong = time_in is None or str(time_in).strip().lower() in ['', 'nan', 'none', 'nat']
    ra_trong = time_out is None or str(time_out).strip().lower() in ['', 'nan', 'none', 'nat']
    result = {'admin_hours': 0.0, 'tong_gio': 0.0, 'di_tre': 0, 've_som': 0, 'ot': 0.0, 'is_chieu': False}
    if vao_trong or ra_trong:
        return result
    t_in = to_time_obj(time_in)
    t_out = to_time_obj(time_out)
    if t_in is None or t_out is None:
        return result
    in_f = time_to_float(t_in)
    out_f = time_to_float(t_out)
    if out_f < in_f:
        result['tong_gio'] = -1
        return result
    eff_in = max(in_f, start_chuan)
    eff_out = min(out_f, end_chuan)
    if eff_out > eff_in:
        total_admin = eff_out - eff_in
        lunch_overlap = max(0.0, min(eff_out, lunch_end) - max(eff_in, lunch_start))
        admin_hours = min(max(total_admin - lunch_overlap, 0.0), max_hours)
    else:
        admin_hours = 0.0
    result['admin_hours'] = admin_hours
    
    if in_f >= lunch_start:
        result['is_chieu'] = True
        result['di_tre'] = int(max(0, (in_f - lunch_end) * 60)) if in_f > lunch_end else 0
    else:
        result['di_tre'] = int(max(0, (in_f - start_chuan) * 60)) if in_f > start_chuan else 0
        
    result['ve_som'] = int(max(0, (end_chuan - out_f) * 60)) if out_f < end_chuan else 0
    result['ot'] = max(0.0, out_f - end_chuan) + max(0.0, start_chuan - in_f)
    result['tong_gio'] = admin_hours
    return result

def clean_date(val):
    if pd.isna(val):
        return None
    try:
        if isinstance(val, datetime.datetime):
            return val
        return pd.to_datetime(str(val).strip(), dayfirst=True)
    except:
        return None

def get_fixed_holidays_for_years(years):
    dates = []
    for y in sorted(years):
        for (mo, day) in [(1,1), (4,30), (5,1), (9,2)]:
            dates.append(datetime.date(y, mo, day))
    return sorted(dates)

def is_workday_func(d_obj):
    """Trả về True nếu là ngày làm việc.
    - Thứ 7 cuối cùng của tháng = bắt buộc làm việc.
    - Các Thứ 7 khác và Chủ nhật = nghỉ.
    """
    if pd.isna(d_obj): return True
    try:
    	d = d_obj.date() if hasattr(d_obj, 'date') else d_obj
    	wd = d.weekday()
    	if wd == 6: return False  # Chủ nhật: nghỉ
    	if wd == 5:               # Thứ 7
    	    # Là Thứ 7 cuối cùng của tháng nếu cộng 7 ngày thì sang tháng khác
    	    return (d + datetime.timedelta(days=7)).month != d.month
    	return True  # Thứ 2 – Thứ 6: luôn làm việc
    except:
    	return True

def is_last_saturday_of_month(d_obj):
    """Kiểm tra ngày có phải Thứ 7 cuối cùng của tháng không."""
    try:
    	d = d_obj.date() if hasattr(d_obj, 'date') else d_obj
    	if d.weekday() != 5: return False
    	return (d + datetime.timedelta(days=7)).month != d.month
    except:
    	return False

def calculate_working_days(start_date, end_date, holidays=None, makeups=None):
    if holidays is None:
        holidays = set()
    if makeups is None:
        makeups = set()
    total = 0
    holiday_info = []
    makeup_info = []
    current = start_date
    while current <= end_date:
        is_wd = is_workday_func(current)
        if current in makeups:
            total += 1
            makeup_info.append({
                "date": current,
                "weekday_label": ["Thứ 2","Thứ 3","Thứ 4","Thứ 5","Thứ 6","Thứ 7","Chủ nhật"][current.weekday()]
            })
        elif current in holidays:
            holiday_info.append({
                "date": current,
                "weekday_label": ["Thứ 2","Thứ 3","Thứ 4","Thứ 5","Thứ 6","Thứ 7","Chủ nhật"][current.weekday()],
                "is_workday": is_wd
            })
        elif is_wd:
            total += 1
        current += datetime.timedelta(days=1)
    return total, holiday_info, makeup_info

# ==========================================
# 4. PARSER ĐỌC FILE
# ==========================================
def find_header_row(file, file_name):
    file.seek(0)
    if file_name.endswith(('.csv', '.txt', '.dat', '.tsv')):
        df_raw = pd.read_csv(file, header=None, nrows=30, sep=None, engine='python')
    else:
        df_raw = pd.read_excel(file, header=None, nrows=30)
    for i, row in df_raw.iterrows():
        row_str = [str(v).strip().lower() for v in row.values if pd.notna(v) and str(v).lower() != 'nan']
        keywords = ['mã', 'tên', 'ngày', 'vào', 'ra']
        if sum(1 for kw in keywords if any(kw in cell for cell in row_str)) >= 3:
            return i
    return 0

def parse_excel_file(uploaded_file):
    file_name = uploaded_file.name.lower()
    try:
        header_row = find_header_row(uploaded_file, file_name)
    except Exception as e:
        st.error(f"❌ Lỗi khi đọc file: {e}")
        st.stop()
    uploaded_file.seek(0)
    try:
        if file_name.endswith(('.csv', '.txt', '.dat', '.tsv')):
            df = pd.read_csv(uploaded_file, header=header_row, sep=None, engine='python')
        else:
            df = pd.read_excel(uploaded_file, header=header_row)
    except Exception as e:
        st.error(f"❌ Lỗi khi đọc bảng dữ liệu: {e}")
        st.stop()
    df = df.dropna(how='all')
    df.columns = df.columns.astype(str).str.strip()
    return df

def auto_detect_columns(df):
    mapping = {}
    for col in df.columns:
        col_lower = str(col).strip().lower()
        if 'mã' in col_lower and 'nv' not in col_lower.replace('mã',''):
            mapping['ma_nv'] = col
        elif col_lower in ['tên nhân viên', 'tên nv', 'họ tên', 'tên']:
            mapping['ten_nv'] = col
        elif 'ngày' in col_lower and 'ca' not in col_lower:
            mapping['ngay'] = col
        elif col_lower in ['vào', 'vao']:
            mapping['gio_vao'] = col
        elif col_lower == 'ra' and 'sớm' not in col_lower and 'som' not in col_lower:
            mapping['gio_ra'] = col
        elif col_lower in ['chức vụ', 'chuc vu', 'vị trí', 'title']:
            mapping['chuc_vu'] = col
        elif col_lower in ['phòng ban', 'phong ban', 'bộ phận', 'bo phan', 'department']:
            mapping['phong_ban'] = col
    return mapping

# ==========================================
# 5. EXPORTER XUẤT FILE
# ==========================================
def export_excel_tong_hop(df_filtered, mapping, start_date, end_date, total_wd):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    import pandas as pd

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi tiết chấm công"
    
    font_bold14 = Font(name='Times New Roman', size=14, bold=True)
    font_bold12 = Font(name='Times New Roman', size=12, bold=True)
    font_normal = Font(name='Times New Roman', size=12)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    border_thick = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    
    ws.merge_cells('A1:N1')
    ws['A1'] = "CHI TIẾT CHẤM CÔNG"
    ws['A1'].font = font_bold14
    ws['A1'].alignment = align_center
    
    ws.merge_cells('A2:N2')
    ws['A2'] = f"Từ ngày {start_date.strftime('%d/%m/%Y')} đến ngày {end_date.strftime('%d/%m/%Y')}"
    ws['A2'].font = font_bold12
    ws['A2'].alignment = align_center
    
    ws.merge_cells('A4:H4')
    ws['A4'] = f"Tổng số ngày làm trong tháng: {total_wd} ngày (bao gồm từ Thứ 2 đến Thứ 6 và một ngày Thứ 7)"
    ws['A4'].font = font_bold12
    ws['A4'].alignment = align_left
    
    total_hours = total_wd * 8
    ws.merge_cells('A5:H5')
    ws['A5'] = f"Tổng số giờ làm tiêu chuẩn trong tháng: {total_hours} giờ"
    ws['A5'].font = font_bold12
    ws['A5'].alignment = align_left
    
    total_off_days = (end_date - start_date).days + 1 - total_wd
    ws.merge_cells('A6:H6')
    ws['A6'] = f"Tổng số ngày nghỉ trong tháng: {total_off_days} ngày"
    ws['A6'].font = font_bold12
    ws['A6'].alignment = align_left
    
    headers = ["Mã nhân viên", "Tên nhân viên", "Phòng ban", "Chức vụ", "Ngày", "Thứ", "Vào", "Ra", "Công", "Giờ làm thực tế", "OT", "Tổng giờ làm", "Lý do tăng ca", "Ghi chú"]
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        c = ws.cell(row=8, column=col_num, value=header)
        c.font = font_bold12
        c.alignment = align_center
        c.fill = header_fill
        c.border = border
        
    row_idx = 9
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    def get_val(row, key):
        col = mapping.get(key)
        if col and col in row:
            v = row[col]
            return str(v) if pd.notna(v) else ""
        return ""
        
    for _, row in df_filtered.iterrows():
        try: d_obj = row['_parsed_date'].date()
        except: d_obj = pd.to_datetime(row['_parsed_date']).date()
        
        thu = d_obj.weekday()
        thu_str = ["Hai", "Ba", "Tư", "Năm", "Sáu", "Bảy", "CN"][thu]
        
        ws.cell(row=row_idx, column=1, value=get_val(row, 'ma_nv'))
        ws.cell(row=row_idx, column=2, value=get_val(row, 'ten_nv'))
        ws.cell(row=row_idx, column=3, value=get_val(row, 'phong_ban'))
        ws.cell(row=row_idx, column=4, value=get_val(row, 'chuc_vu'))
        ws.cell(row=row_idx, column=5, value=d_obj.strftime('%d/%m/%Y'))
        ws.cell(row=row_idx, column=6, value=thu_str)
        ws.cell(row=row_idx, column=7, value=get_val(row, 'gio_vao'))
        ws.cell(row=row_idx, column=8, value=get_val(row, 'gio_ra'))
        
        hc = float(row.get('Giờ hành chính', 0)) if pd.notna(row.get('Giờ hành chính')) else 0.0
        ws.cell(row=row_idx, column=9, value=round(hc/8, 2) if hc > 0 else 0)
        
        ws.cell(row=row_idx, column=10, value=hc if hc > 0 else "")
        
        ot = float(row.get('Giờ OT', 0)) if pd.notna(row.get('Giờ OT')) else 0.0
        ws.cell(row=row_idx, column=11, value=ot if ot > 0 else "")
        
        col_hc_letter = get_column_letter(10)
        col_ot_letter = get_column_letter(11)
        ws.cell(row=row_idx, column=12, value=f"=SUM({col_hc_letter}{row_idx}:{col_ot_letter}{row_idx})")
        
        ws.cell(row=row_idx, column=13, value=str(row.get('Lý do tăng ca', '')) if pd.notna(row.get('Lý do tăng ca')) else "")
        
        ghi_chu = str(row.get('Ghi chú', '')) if pd.notna(row.get('Ghi chú')) else ""
        loai = row.get('_loai', '') if pd.notna(row.get('_loai', '')) else ''
        
        ghi_chu_clean = ghi_chu.replace("🚫 ", "").replace("✅ ", "").replace("⚠️ ", "").replace("📝 ", "").replace("📅 ", "")
        
        is_abnormal = False
        if "nghi_khong_phep" in loai or "Thiếu giờ" in ghi_chu_clean or "Lỗi" in ghi_chu_clean or "Vắng" in ghi_chu_clean or "Làm thiếu" in ghi_chu_clean:
            is_abnormal = True

        ws.cell(row=row_idx, column=14, value=ghi_chu_clean.strip())
        
        font_red = Font(name='Times New Roman', size=12, color="FF0000")
        for col_num in range(1, 15):
            c = ws.cell(row=row_idx, column=col_num)
            c.font = font_normal
            if col_num == 14 and is_abnormal:
                c.font = font_red
            c.border = border
            if col_num in [1,2,3,4, 13, 14]:
                c.alignment = align_left
            else:
                c.alignment = align_center
            
            if thu in [5, 6] and not (thu == 5 and is_last_saturday_of_month(d_obj)):
                c.fill = gray_fill
                
        row_idx += 1
        
    for col_num, width in enumerate([10, 20, 15, 15, 12, 8, 10, 10, 8, 15, 10, 15, 20, 30], 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ==========================================
# 6. GIAO DIỆN CHÍNH
# ==========================================
if "step" not in st.session_state: st.session_state.step = 1
if "df_raw" not in st.session_state: st.session_state.df_raw = None
if "show_history" not in st.session_state: st.session_state.show_history = False

# ----- QUẢN LÝ TRẠNG THÁI TRANG (ROUTING) -----
if "app_page" not in st.session_state:
    st.session_state.app_page = "home"


def render_chatbot():
    # 7. CHATBOT BONG BÓNG LƠ LỬNG
    # ==========================================
    if 'chat_open' not in st.session_state:
        st.session_state.chat_open = False
    if 'chat_pos' not in st.session_state:
        st.session_state.chat_pos = 'right'

    def toggle_chat():
        st.session_state.chat_open = not st.session_state.chat_open

    # CSS bong bóng tròn màu xanh — dùng key container để CSS selector chính xác
    side = st.session_state.chat_pos
    st.markdown("""
    <style>
    .st-key-chat_bubble_wrap button {
        border-radius: 50% !important;
        width: 64px !important; height: 64px !important;
        background: linear-gradient(135deg, #2563EB 0%, #1D4ED8 100%) !important;
        color: white !important; border: none !important;
        box-shadow: 0 8px 24px rgba(37,99,235,0.4) !important;
        font-size: 26px !important;
        transition: transform 0.2s, box-shadow 0.2s !important;
        padding: 0 !important;
        min-height: 64px !important;
    }
    .st-key-chat_bubble_wrap button:hover {
        transform: scale(1.08) !important;
        box-shadow: 0 12px 32px rgba(37,99,235,0.5) !important;
    }
    .st-key-chat_bubble_wrap {
        position: fixed !important;
        bottom: 30px !important;
        {side}: 30px !important;
        z-index: 99999 !important;
        width: auto !important;
    }
    .st-key-chat_box_wrap {
        position: fixed !important;
        bottom: 100px !important;
        {side}: 30px !important;
        width: 360px !important;
        background: #F8FAFC !important;
        border-radius: 16px !important;
        box-shadow: 0 20px 48px rgba(0,0,0,0.2) !important;
        z-index: 99999 !important;
        padding: 0 !important;
        border: 1px solid #E2E8F0 !important;
        overflow: hidden !important;
    }
    .st-key-chat_box_wrap [data-testid="stVerticalBlock"] {
        gap: 0 !important; padding: 0 !important;
    }
    .st-key-chat_box_wrap [data-testid="stChatMessage"] {
        background: white !important; border: 1px solid #E2E8F0 !important;
        border-radius: 16px 16px 16px 4px !important; padding: 12px !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.02) !important; margin: 12px !important; width: auto !important;
    }
    .st-key-chat_box_wrap [data-testid="stChatMessage"] [data-testid="stMarkdownContainer"] p {
        font-size: 13.5px !important; color: #334155 !important; line-height: 1.5 !important;
    }
    .st-key-chat_box_wrap [data-testid="stChatInput"] {
        background: white !important; padding: 12px 16px !important; border-top: 1px solid #E2E8F0 !important;
    }
    .st-key-chat_box_wrap [data-testid="stChatInput"] textarea {
        background: #F1F5F9 !important; border-radius: 24px !important; border: 1px solid #E2E8F0 !important; font-size: 13px !important;
    }
    .st-key-close_chat_box { position: absolute !important; top: 12px !important; right: 12px !important; z-index: 100 !important; }
    .st-key-close_chat_box button { background: rgba(255,255,255,0.1) !important; color: white !important; border: none !important; border-radius: 50% !important; width: 32px !important; height: 32px !important; min-height: 32px !important; padding: 0 !important; display: flex !important; align-items: center !important; justify-content: center !important; }
    .st-key-close_chat_box button:hover { background: rgba(255,255,255,0.2) !important; color: white !important; }
    </style>
    """.replace("{side}", side), unsafe_allow_html=True)

    

    bubble = st.container(key="chat_bubble_wrap")
    with bubble:
        st.button("💬", on_click=toggle_chat, key="chat_bubble_btn", help="Trợ lý AI V-MOS")

    if st.session_state.chat_open:
        chat_box = st.container(key="chat_box_wrap")
        with chat_box:
            if st.button("✕", key="close_chat_box"):
                st.session_state.chat_open = False
                st.rerun()
            st.markdown("""
            <div style="background:#2563EB; color:white; padding:16px; display:flex; align-items:center; gap:12px;">
                <div style="width:40px; height:40px; background:linear-gradient(135deg, #F472B6 0%, #3B82F6 100%); border-radius:50%; display:flex; align-items:center; justify-content:center; font-size:24px; box-shadow:0 4px 10px rgba(0,0,0,0.1); flex-shrink:0;">🤖</div>
                <div style="display:flex; flex-direction:column;">
                    <div style="font-weight:700; font-size:15px; line-height:1.2;">V-MOS Assistant</div>
                    <div style="font-size:12px; display:flex; align-items:center; gap:6px; opacity:0.9; margin-top:4px;">
                        <span style="width:8px; height:8px; background:#10B981; border-radius:50%; display:inline-block;"></span> Đang trực tuyến
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            saved_key = load_saved_api_key()
            if saved_key:
                st.session_state['gemini_configured'] = True
            else:
                st.session_state['gemini_configured'] = False
                
            if not st.session_state['gemini_configured']:
                st.warning("⚠️ Vui lòng cấu hình Gemini API Key để Chatbot hoạt động!")
                new_key = st.text_input("Nhập Gemini API Key", type="password", placeholder="Nhập API Key tại đây...")
                if new_key:
                    import toml
                    try:
                        with open(".streamlit/secrets.toml", "r", encoding="utf-8") as f:
                            secrets = toml.load(f)
                    except:
                        secrets = {}
                    secrets["GEMINI_API_KEY"] = new_key
                    with open(".streamlit/secrets.toml", "w", encoding="utf-8") as f:
                        toml.dump(secrets, f)
                    st.session_state['gemini_configured'] = True
                    st.rerun()
            else:
                with st.expander("⚙️ Cài đặt Chatbot", expanded=False):
                    if st.button("🔄 Đổi mã API Key", use_container_width=True):
                        try:
                            import toml
                            with open(".streamlit/secrets.toml", "w", encoding="utf-8") as f:
                                toml.dump({}, f)
                        except: pass
                        st.session_state['gemini_configured'] = False
                        st.rerun()
                    
                if 'chat_messages' not in st.session_state:
                    st.session_state['chat_messages'] = [{"role": "assistant", "content": "Xin chào! Tôi có thể giúp gì cho bạn về dữ liệu chấm công?"}]

                chat_container = st.container(height=300)
                with chat_container:
                    for msg in st.session_state['chat_messages']:
                        with st.chat_message(msg["role"], avatar="🤖" if msg["role"] == "assistant" else "👤"):
                            st.markdown(msg["content"])

                prompt = st.chat_input("Hỏi AI...")
                if prompt:
                    st.session_state['chat_messages'].append({"role": "user", "content": prompt})
                    with chat_container:
                        with st.chat_message("user", avatar="👤"):
                            st.markdown(prompt)
                        with st.chat_message("assistant", avatar="🤖"):
                            with st.spinner("..."):
                                try:
                                    import requests
                                    saved_key = load_saved_api_key()
                                    if not saved_key:
                                        raise Exception("Không tìm thấy API Key. Vui lòng nhập lại API Key ở menu.")

                                    df_ctx = "Chưa có dữ liệu"
                                    if st.session_state.get('app_page') == 'mos':
                                        if 'df_mos_edited' in st.session_state and st.session_state.df_mos_edited is not None:
                                            df_ctx = "--- DỮ LIỆU TỔNG HỢP DỰ ÁN MOS ---\n" + st.session_state.df_mos_edited.to_csv(index=False)
                                        else:
                                            df_ctx = "Chưa có dữ liệu dự án MOS nào được xử lý."
                                    else:
                                        if 'df_filtered_for_chat' in st.session_state and st.session_state.df_filtered_for_chat is not None:
                                            dff = st.session_state.df_filtered_for_chat.copy()
                                            ten = st.session_state.mapping.get('ten_nv', dff.columns[1])
                                            dff['Số giờ làm thực tế'] = pd.to_numeric(dff['Số giờ làm thực tế'], errors='coerce').fillna(0)
                                            dff['Giờ OT'] = pd.to_numeric(dff['Giờ OT'], errors='coerce').fillna(0)
                                            dff['Giờ hành chính'] = pd.to_numeric(dff['Giờ hành chính'], errors='coerce').fillna(0)
    
                                            df_detail = pd.DataFrame({
                                                "Tên NV": dff[ten],
                                                "Ngày": dff["Ngày"],
                                                "Giờ làm": dff["Giờ hành chính"].round(2),
                                                "OT": dff["Giờ OT"].round(2),
                                                "Tổng": dff["Số giờ làm thực tế"].round(2),
                                                "Ghi chú": dff["Ghi chú"]
                                            })
    
                                            if len(df_detail) <= 500:
                                                df_ctx = f"--- DỮ LIỆU CHI TIẾT (Toàn bộ {len(df_detail)} dòng) ---\n" + df_detail.to_csv(index=False)
                                            else:
                                                agg = df_detail.groupby("Tên NV").agg({'Giờ làm': 'sum', 'OT': 'sum', 'Tổng': 'sum'}).reset_index().round(2)
                                                df_ctx = f"--- BẢNG TỔNG HỢP (TẤT CẢ NHÂN VIÊN) ---\n{agg.to_csv(index=False)}\n\n"
    
                                                # Nén dữ liệu cực đại để AI đọc được toàn bộ không bị mất dòng nào
                                                compressed_lines = []
                                                for name, group in df_detail.groupby("Tên NV"):
                                                    days = []
                                                    for _, row in group.iterrows():
                                                        d = str(row["Ngày"])[:5] # "15/05/2026" -> "15/05"
                                                        hc = row["Giờ làm"]
                                                        ot = row["OT"]
                                                        note = str(row["Ghi chú"]).strip()
                                                        if ot == 0 and note == "":
                                                            if hc == 8.0: days.append(d)
                                                            elif hc == 0.0: days.append(f"{d}(Nghỉ)")
                                                            else: days.append(f"{d}({hc}h)")
                                                        else:
                                                            days.append(f"{d}({hc}h|OT:{ot}|{note})")
                                                    compressed_lines.append(f"{name}: " + ", ".join(days))
    
                                                df_ctx += "--- CHI TIẾT LỊCH SỬ TỪNG NGÀY (DẠNG NÉN SIÊU TỐI ƯU) ---\n"
                                                df_ctx += "Quy ước: Nếu chỉ ghi ngày (VD: 15/05) nghĩa là làm đủ 8h không OT. Các ngày khác ghi rõ (Giờ làm|OT|Ghi chú).\n"
                                                df_ctx += "\n".join(compressed_lines)
                                        elif 'df_result' in st.session_state and st.session_state.df_result is not None:
                                            df_ctx = st.session_state.df_result.head(50).to_csv(index=False)
                                        else:
                                            df_ctx = "Chưa có dữ liệu nào được tải lên."

                                    sys_prompt = f"""Bạn là V-MOS AI, một siêu trợ lý trí tuệ nhân tạo toàn năng.
Bạn là trợ lý chính cho hệ thống "Quản lý Chấm công & Dự án Nội bộ MOS". Bạn am hiểu sâu sắc về cả việc tính toán giờ làm việc (chấm công) và tổng hợp giờ làm dự án MOS.
Bạn SẴN SÀNG TRẢ LỜI BẤT KỲ CÂU HỎI NÀO CỦA NGƯỜI DÙNG (từ phân tích số liệu bảng chấm công, báo cáo dự án MOS, cho đến viết email, làm toán, lập trình, tư vấn).
Hãy thoải mái trò chuyện và hỗ trợ người dùng như một AI thực thụ. Tuyệt đối không bao giờ từ chối trả lời vì lý do "tôi chỉ là trợ lý chấm công".

[DỮ LIỆU HIỆN TẠI MÀ NGƯỜI DÙNG ĐANG TƯƠNG TÁC]:
{df_ctx}

Luôn ưu tiên trả lời tự nhiên, thân thiện và chính xác."""

                                    try:
                                        import httpx
                                        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={saved_key}"
                                        
                                        # Chuyển đổi lịch sử chat sang định dạng Gemini
                                        gemini_history = []
                                        for msg in st.session_state['chat_messages'][-10:]:
                                            role = "user" if msg["role"] == "user" else "model"
                                            gemini_history.append({"role": role, "parts": [{"text": msg["content"]}]})
                                        
                                        # Thêm tin nhắn hiện tại
                                        gemini_history.append({"role": "user", "parts": [{"text": prompt}]})

                                        data = {
                                            "systemInstruction": {"parts": [{"text": sys_prompt}]},
                                            "contents": gemini_history,
                                            "generationConfig": {"temperature": 0.7}
                                        }
                                        
                                        res = httpx.post(url, headers={"Content-Type": "application/json"}, json=data, timeout=30.0)
                                        if res.status_code == 200:
                                            answer = res.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
                                        elif res.status_code == 400 and "API_KEY_INVALID" in res.text:
                                            answer = "⚠️ Mã Gemini API Key của bạn không hợp lệ hoặc đã hết hạn. Hệ thống đã xóa mã cũ, vui lòng bấm **Đổi API Key** bên trên để nhập lại nhé!"
                                            try:
                                                import toml
                                                with open(".streamlit/secrets.toml", "w", encoding="utf-8") as f:
                                                    toml.dump({}, f)
                                            except: pass
                                        else:
                                            answer = f"Lỗi kết nối Gemini: {res.text}" 
                                    except Exception as e:
                                        answer = f"Lỗi hệ thống: {str(e)}"

                                    st.markdown(answer)
                                    st.session_state['chat_messages'].append({"role": "assistant", "content": answer})
                                except Exception as e:
                                    st.error(f"Lỗi: {str(e)}")


# 1. GIAO DIỆN CHUNG (HOME)
logo_html_large = f'''<img src="data:image/png;base64,{LOGO_HEADER_B64}" style="height:200px;width:auto;object-fit:contain;mix-blend-mode:multiply;margin-bottom:20px; filter: drop-shadow(0 4px 10px rgba(0,0,0,0.1)); transition: transform 0.3s ease;" onmouseover="this.style.transform='scale(1.1) rotate(2deg)'" onmouseout="this.style.transform='scale(1) rotate(0)'">''' if LOGO_HEADER_B64 else '''<div style="font-size:80px; margin-bottom:10px;">📊</div>'''
if st.session_state.app_page == "home":
    # Ẩn sidebar ở trang chủ
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { display: none !important; }
    [data-testid="collapsedControl"] { display: none !important; }
    

    /* Auto-width centered buttons */

    .stButton > button {
        height: 50px;
        padding: 0 32px !important;
        border-radius: 12px !important;
        background: linear-gradient(135deg, #1E3A8A, #3B82F6) !important;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3) !important;
        transition: transform 0.2s ease, box-shadow 0.2s ease !important;
        color: white !important;
        border: none !important;
        width: auto !important; /* Force width to fit text */
    }
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 16px rgba(59, 130, 246, 0.4) !important;
    }
    .stButton > button p {
        font-size: 16px !important;
        font-weight: 600 !important;
    }
    
    div.stButton {
        text-align: center;
        display: flex;
        justify-content: center;
    }
    div.stButton > button {
        width: 85% !important; /* Đảm bảo hai nút có độ rộng cân đối bằng nhau */
        margin: 0 auto !important;
    }
    </style>
    

    """, unsafe_allow_html=True)
    
    st.markdown(f"""
    <div style="text-align:center; padding-top: 80px; padding-bottom: 100px;">
                <div style="margin-bottom:20px;">
            {logo_html_large}
        </div>
        <p style="font-size: 18px; color: #6B7280; font-weight:500;">Hệ thống Quản lý Chấm công & Dự án Nội bộ</p>
    </div>
    """, unsafe_allow_html=True)
    
    col_left, col1, col2, col_right = st.columns([1, 2, 2, 1])
    with col1:
        if st.button("Tính số giờ làm thực tế", use_container_width=True, type="primary"):
            st.session_state.app_page = "chamcong"
            st.rerun()
    with col2:
        if st.button("Tổng hợp giờ làm dự án MOS", use_container_width=True, type="primary"):
            st.session_state.app_page = "mos"
            st.rerun()
        
        
    render_chatbot()
    st.stop() # Dừng toàn bộ code phía dưới để giữ giao diện sạch


# ==========================================
# CÁC HÀM XỬ LÝ DỮ LIỆU MOS
# ==========================================
import re

import httpx
import json

def summarize_tasks_with_ai(ma_da: str, ten_da: str, tasks: list) -> str:
    """
    Gọi Claude API để tóm tắt danh sách task thành 1 câu nội dung chung.
    Nếu API lỗi, fallback về cách nối chuỗi thủ công.
    """
    if not tasks:
        return ''
    
    # Loại bỏ trùng lặp, giữ thứ tự
    tasks_unique = list(dict.fromkeys([t.strip() for t in tasks if t.strip()]))
    
    if len(tasks_unique) == 1:
        return tasks_unique[0]
    
    prompt = f"""Dưới đây là danh sách các nội dung công việc của dự án 
"{ten_da}" (mã: {ma_da}) bằng tiếng Nhật:

{chr(10).join(f'- {t}' for t in tasks_unique)}

Nhiệm vụ của bạn:
1. Dịch tên dự án "{ten_da}" sang tiếng Việt.
2. Tóm tắt danh sách công việc trên thành 1 câu súc tích bằng tiếng Nhật, sau đó dịch câu đó sang tiếng Việt.

Trả về kết quả DƯỚI DẠNG JSON với định dạng sau (không giải thích thêm):
{{
  "ten_da_song_ngu": "Tên tiếng Nhật \\n Tên tiếng Việt",
  "noi_dung_song_ngu": "Tóm tắt tiếng Nhật \\n Tóm tắt tiếng Việt"
}}
"""

    try:
        api_key = load_saved_api_key()
        
        if not api_key:
            return ' / '.join(tasks_unique)
            
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
        response = httpx.post(
            url,
            headers={"Content-Type": "application/json"},
            json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.2}
            },
            timeout=15.0
        )
        data = response.json()
        result_text = data["candidates"][0]["content"]["parts"][0]["text"].strip()
        
        import json
        if "```json" in result_text:
            result_text = result_text.split("```json")[1].split("```")[0]
        elif "```" in result_text:
            result_text = result_text.split("```")[1].split("```")[0]
            
        return json.loads(result_text)
    except Exception as e:
        return ' / '.join(tasks_unique)

def extract_ma_nv_from_filename(filename: str) -> str:
    """Lấy mã NV từ tên file: Report_VM011_ロン_2026_5.xlsx → VM011"""
    match = re.search(r'(VM\d+)', filename)
    return match.group(1) if match else filename

def extract_ten_nv_from_filename(filename: str) -> str:
    """Lấy tên Nhật từ file để hiển thị: Report_VM011_ロン_2026_5 → ロン"""
    name = re.sub(r'\.xlsx?$', '', filename)
    # Lấy phần chữ (ko phải số/gạch dưới) ngay sau cụm VMxxx
    match = re.search(r'VM\d+[\s_]+([^\d_]+)', name)
    if match:
        return match.group(1).strip()
    return filename

def parse_mos_file(file, filename: str) -> pd.DataFrame:
    """
    Đọc 1 file Report, lấy chỉ phần MOS業務.
    Trả về DataFrame với các cột đã chuẩn hóa.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    
    # Tìm header row (có chữ 案件番号)
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and any(str(v) == '案件番号' for v in row if v):
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError(f"Không tìm thấy dòng tiêu đề chuẩn (案件番号) trong file {filename}")
    
    header = rows[header_row_idx]
    
    col_map = {}
    date_cols = {}
    for ci, col in enumerate(header):
        if col is None: continue
        col_str = str(col).strip()
        if col_str == '日本担当者': col_map['tanto'] = ci
        elif col_str == '案件番号': col_map['ma_da'] = ci
        elif col_str == 'お客様': col_map['khach'] = ci
        elif col_str == '案件名': col_map['ten_da'] = ci
        elif col_str == '区分': col_map['phan_vung'] = ci
        elif col_str == 'タスク名': col_map['task'] = ci
        elif col_str == '合計': col_map['tong'] = ci
        elif '/' in col_str:
            try:
                parts = col_str.split('/')
                month, day = int(parts[0]), int(parts[1])
                date_cols[f'{month}/{day}'] = ci
            except: pass
    
    in_mos = False
    records = []
    
    for row in rows[header_row_idx + 2:]:
        if row is None: continue
        first_val = str(row[0]).strip() if row[0] else ''
        
        if first_val == 'MOS業務':
            in_mos = True
            continue
        elif first_val in ['JMOS業務', '社内業務', 'JMO S業務']:
            in_mos = False
            continue
        
        if not in_mos:
            continue
        
        ma_da = str(row[col_map.get('ma_da', 1)] or '').strip()
        if not ma_da or ma_da == 'None':
            continue
        
        tanto = str(row[col_map.get('tanto', 0)] or '').strip()
        khach = str(row[col_map.get('khach', 2)] or '').strip()
        ten_da = str(row[col_map.get('ten_da', 3)] or '').strip()
        phan_vung = str(row[col_map.get('phan_vung', 4)] or '').strip()
        task = str(row[col_map.get('task', 5)] or '').strip()
        tong = row[col_map.get('tong', 6)]
        
        try: tong = float(tong) if tong else 0.0
        except: tong = 0.0
        
        ngay_co_gio = []
        for date_str, ci in date_cols.items():
            val = row[ci]
            try:
                if val and float(val) > 0:
                    ngay_co_gio.append(date_str)
            except: pass
        
        def parse_date_str(d_str, year=2026):
            try:
                m, d = d_str.split('/')
                return datetime.date(year, int(m), int(d))
            except: return None
        
        ngay_dates = [parse_date_str(d) for d in ngay_co_gio]
        ngay_dates = [d for d in ngay_dates if d]
        
        ngay_bat_dau = min(ngay_dates).strftime('%d/%m/%Y') if ngay_dates else ''
        ngay_ket_thuc = max(ngay_dates).strftime('%d/%m/%Y') if ngay_dates else ''
        
        ql_nhat = ''
        if '/' in tanto:
            ql_nhat = tanto.split('/')[-1].replace('様', '').strip()
        elif tanto:
            ql_nhat = tanto.replace('様', '').strip()
            
        khach = khach.replace('様', '').strip()
        
        if 'シミュレーション' in phan_vung:
            phan_vung_vn = 'シミュレーション設計 \n Thiết kế mô phỏng'
        elif '電気' in phan_vung:
            phan_vung_vn = '電気設計 \n Thiết kế điện'
        elif 'メカ' in phan_vung or '機械' in phan_vung:
            phan_vung_vn = 'メカ設計 \n Thiết kế cơ khí'
        else:
            phan_vung_vn = f"{phan_vung} \n {phan_vung}" 
        
        records.append({
            'ma_nv': extract_ma_nv_from_filename(filename),
            'ten_nv': extract_ten_nv_from_filename(filename),
            'ma_da': ma_da,
            'khach': khach,
            'ten_da': ten_da,
            'phan_vung': phan_vung_vn,
            'task': task,
            'tong_gio': tong,
            'ngay_bat_dau': ngay_bat_dau,
            'ngay_ket_thuc': ngay_ket_thuc,
            'ql_nhat': ql_nhat,
        })
    if not records:
        raise ValueError(f"Không tìm thấy dữ liệu hợp lệ (phần MOS業務) trong file {filename}")
    return pd.DataFrame(records)


def tong_hop_mos(dfs: list) -> pd.DataFrame:
    """
    Gộp nhiều DataFrame từ nhiều file, tổng hợp theo mã dự án.
    """
    if not dfs:
        return pd.DataFrame()
    
    df_all = pd.concat(dfs, ignore_index=True)
    
    import streamlit as st
    result = []
    groups = list(df_all.groupby('ma_da'))
    total_groups = len(groups)
    
    if total_groups == 0:
        return pd.DataFrame()
        
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for idx, (ma_da, grp) in enumerate(groups):
        status_text.markdown(f"**🤖 Đang phân tích dự án {idx+1}/{total_groups}:** `{ma_da}`")
        ngay_bd_list = [datetime.datetime.strptime(d, '%d/%m/%Y').date() 
                        for d in grp['ngay_bat_dau'] if d]
        ngay_kt_list = [datetime.datetime.strptime(d, '%d/%m/%Y').date() 
                        for d in grp['ngay_ket_thuc'] if d]
        ngay_bd = min(ngay_bd_list).strftime('%d/%m/%Y') if ngay_bd_list else ''
        ngay_kt = max(ngay_kt_list).strftime('%d/%m/%Y') if ngay_kt_list else ''
        if ngay_bd == ngay_kt:
            ngay_kt = ''
        
        nv_co_gio = grp[grp['tong_gio'] > 0]['ten_nv'].unique()
        jp_names = [mv for mv in sorted(nv_co_gio) if mv]
        name_dict = {'ダンフン': 'Hưng', 'クアン': 'Quân', 'ハン': 'Hằng', 'グエット': 'Nguyệt', 'ロン': 'Long', 'ダオ': 'Đạo', 'フォン': 'Phương'}
        
        # Lấy nội dung công việc (task) hợp lệ đầu tiên làm đại diện duy nhất
        valid_tasks = [str(t).strip() for t in grp['task'].tolist() if str(t).strip() and str(t).lower() not in ['nan', 'none']]
        first_task = valid_tasks[0] if valid_tasks else ""
        tasks = [first_task] if first_task else []
        
        ten_da_val = grp['ten_da'].iloc[0]
        
        # Gọi AI (lấy cả tên dự án và nội dung)
        ai_res = summarize_tasks_with_ai(ma_da, ten_da_val, tasks)
        if isinstance(ai_res, dict):
            ten_da_song_ngu = ai_res.get('ten_da_song_ngu', ten_da_val)
            noi_dung = ai_res.get('noi_dung_song_ngu', first_task)
        else:
            ten_da_song_ngu = ten_da_val
            noi_dung = ai_res
            
        phan_vung = grp['phan_vung'].iloc[0]
        
        # Xử lý: Nếu có 'Phương' thì chuyển sang quản lý VN
        ql_viet_nam = None
        if 'フォン' in jp_names:
            jp_names.remove('フォン')
            ql_viet_nam = "フォン \n Phương"
            
        # Nếu nội dung liên quan đến điện thì Đạo làm quản lý
        if any(kw in str(noi_dung).lower() or kw in str(phan_vung).lower() for kw in ['điện', '電気']):
            ql_viet_nam = "ダオ \n Đạo"
            if 'ダオ' in jp_names:
                jp_names.remove('ダオ')
                
        # Nếu nội dung liên quan đến thiết kế mô phỏng thì Phương làm quản lý
        if any(kw in str(noi_dung).lower() or kw in str(phan_vung).lower() for kw in ['mô phỏng', 'シミュレーション', 'simulation']):
            ql_viet_nam = "フォン \n Phương"
            if 'フォン' in jp_names:
                jp_names.remove('フォン')
                
        vn_names = [name_dict.get(mv, mv) for mv in jp_names]
        nguoi_th = ' / '.join(jp_names) + ' \n ' + ' / '.join(vn_names) if jp_names else ''
            
        ql_nhat_goc = grp['ql_nhat'].iloc[0]
        jp_manager_dict = {'石部': 'Ishibe', '勝亦': 'Katsumata', '池谷': 'Ikeya', '田代': 'Tashiro', '井手上': 'Ideue', '山崎': 'Yamazaki', 'フォン': 'Phương', 'ロン': 'Long'}
        ql_nhat_trans = jp_manager_dict.get(ql_nhat_goc, ql_nhat_goc)
        ql_nhat_song_ngu = f"{ql_nhat_goc} \n {ql_nhat_trans}" if ql_nhat_goc else ""
        
        result.append({
            'Mã dự án': ma_da,
            'Tên dự án': ten_da_song_ngu,
            'Phân vùng': grp['phan_vung'].iloc[0],
            'Nội dung ủy thác': noi_dung,
            'Giờ làm (h)': grp['tong_gio'].sum(),
            'Ngày bắt đầu': ngay_bd,
            'Ngày kết thúc': ngay_kt,
            'Quản lý Nhật Bản': ql_nhat_song_ngu,
            'Quản lý Việt Nam': ql_viet_nam,
            'Người thực hiện': nguoi_th,
            'Trạng thái': None,
        })
        progress_bar.progress((idx + 1) / total_groups)
    
    status_text.empty()
    progress_bar.empty()
    
    df_result = pd.DataFrame(result)
    df_result.insert(0, 'STT', range(1, len(df_result)+1))
    return df_result


def render_holiday_makeup_sidebar():
    if "custom_holidays" not in st.session_state:
        st.session_state.custom_holidays = set()
    if "custom_workdays" not in st.session_state:
        st.session_state.custom_workdays = set()

    with st.sidebar.expander("📅 Ngày lễ âm lịch & nghỉ bù", expanded=False):
        st.markdown("<small style='color:green;'>✅ Các ngày lễ cố định (Tết Dương lịch, 30/4, 1/5, Quốc khánh) đã được hệ thống tự động thiết lập.</small>", unsafe_allow_html=True)

        st.markdown("<small>Chọn ngày nghỉ tùy chỉnh từ Lịch:</small>", unsafe_allow_html=True)
        selected_date = st.date_input("Chọn ngày", value=datetime.date.today(), label_visibility="collapsed", key="date_in_holiday")
        if st.button("➕ Thêm ngày", key="btn_add_custom_holiday", use_container_width=True):
            st.session_state.custom_holidays.add(selected_date)
            st.rerun()
        if st.session_state.custom_holidays:
            st.markdown(f"<small><b>Đang có {len(st.session_state.custom_holidays)} ngày nghỉ tùy chỉnh:</b></small>", unsafe_allow_html=True)
            holiday_list = [d.strftime('%d/%m/%Y') for d in sorted(st.session_state.custom_holidays)]
            selected_to_remove = st.multiselect("Xóa ngày", options=holiday_list, placeholder="Nhấp để xem/xóa ngày...", label_visibility="collapsed", key="multi_sel_del_holiday")
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("❌ Xóa đã chọn", key="btn_del_sel_holiday", use_container_width=True) and selected_to_remove:
                    for d_str in selected_to_remove:
                        try:
                            st.session_state.custom_holidays.remove(datetime.datetime.strptime(d_str, '%d/%m/%Y').date())
                        except: pass
                    st.rerun()
            with col_btn2:
                if st.button("🗑️ Xóa tất cả", key="btn_del_all_holiday", use_container_width=True):
                    st.session_state.custom_holidays = set()
                    st.rerun()

    with st.sidebar.expander("💼 Ngày làm bù", expanded=False):
        st.markdown("<small>Thêm các ngày làm bù (VD: đi làm bù Thứ 7, CN):</small>", unsafe_allow_html=True)
        selected_makeup = st.date_input("Chọn ngày làm bù", value=datetime.date.today(), label_visibility="collapsed", key="date_makeup_input")
        if st.button("➕ Thêm ngày làm bù", key="btn_add_makeup", use_container_width=True):
            st.session_state.custom_workdays.add(selected_makeup)
            st.rerun()
        if st.session_state.custom_workdays:
            st.markdown(f"<small><b>Đang có {len(st.session_state.custom_workdays)} ngày làm bù:</b></small>", unsafe_allow_html=True)
            makeup_list = [d.strftime('%d/%m/%Y') for d in sorted(st.session_state.custom_workdays)]
            selected_makeup_remove = st.multiselect("Xóa ngày làm bù", options=makeup_list, placeholder="Nhấp để xem/xóa...", label_visibility="collapsed", key="multi_sel_del_makeup")
            col_btn3, col_btn4 = st.columns(2)
            with col_btn3:
                if st.button("❌ Xóa đã chọn", key="btn_del_sel_makeup", use_container_width=True) and selected_makeup_remove:
                    for d_str in selected_makeup_remove:
                        try: st.session_state.custom_workdays.remove(datetime.datetime.strptime(d_str, '%d/%m/%Y').date())
                        except: pass
                    st.rerun()
            with col_btn4:
                if st.button("🗑️ Xóa tất cả", key="btn_del_all_makeup", use_container_width=True):
                    st.session_state.custom_workdays = set()
                    st.rerun()


def render_mos_page():
    with st.sidebar:
        render_holiday_makeup_sidebar()
        st.markdown("<br>", unsafe_allow_html=True)

    col_back, _ = st.columns([1, 5])
    with col_back:
        if st.button("⬅️ Quay lại Trang chủ", type="secondary", use_container_width=True, key="btn_mos_back_top"):
            st.session_state.app_page = "home"
            st.rerun()

    st.markdown("""
    <style>
    /* ── Main Container ── */
    .mos-main-card {
        /* Bỏ khung trắng theo yêu cầu */
        padding: 0;
    }

    /* ── MOS Page Header (Banner) ── */
    
    .mos-header {
        background: linear-gradient(135deg, #60A5FA 0%, #BFDBFE 100%);
        border-radius: 16px;
        padding: 32px 40px;
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 32px;
        box-shadow: 0 8px 25px rgba(59, 130, 246, 0.25);
        position: relative;
        overflow: hidden;
    }
    .mos-header::before {
        content: '';
        position: absolute;
        top: -50%;
        right: -10%;
        width: 300px;
        height: 300px;
        background: radial-gradient(circle, rgba(255,255,255,0.1) 0%, rgba(255,255,255,0) 70%);
        border-radius: 50%;
    }
    .mos-header-left {
        display: flex;
        gap: 20px;
        align-items: center;
        position: relative;
        z-index: 2;
    }
    .mos-header-logo {
        width: 250px;
        height: auto;
        object-fit: contain;
        filter: drop-shadow(0 4px 6px rgba(0,0,0,0.2));
        transition: transform 0.3s ease;
    }
    .mos-header-logo:hover {
        transform: scale(1.05) rotate(-2deg);
    }
    .mos-hero-title {
        font-size: 26px;
        font-weight: 800;
        color: white;
        margin: 0 0 6px 0;
        line-height: 1.2;
        letter-spacing: -0.02em;
    }
    .mos-hero-sub {
        font-size: 14px;
        color: rgba(255,255,255,0.85);
        margin: 0;
    }
    .mos-hero-badge {
        background: rgba(255,255,255,0.2);
        border: 1px solid rgba(255,255,255,0.4);
        backdrop-filter: blur(4px);
        border-radius: 20px;
        padding: 8px 18px;
        font-size: 13px;
        color: white;
        font-weight: 600;
        position: relative;
        z-index: 2;
    }

    /* ── Stepper ── */
    .mos-stepper {
        display: flex;
        align-items: center;
        justify-content: space-between;
        margin-bottom: 32px;
        padding: 0 20px;
    }
    .mos-step-item {
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 8px;
        position: relative;
        z-index: 1;
    }
    .mos-step-circle {
        width: 28px;
        height: 28px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 12px;
        font-weight: 600;
        background: white;
    }
    .step-active .mos-step-circle {
        border: 2px solid #2563EB;
        color: #2563EB;
        box-shadow: 0 0 0 4px rgba(37, 99, 235, 0.15);
        animation: pulse 2s infinite;
    }
    @keyframes pulse {
        0% { box-shadow: 0 0 0 0 rgba(37, 99, 235, 0.3); }
        70% { box-shadow: 0 0 0 10px rgba(37, 99, 235, 0); }
        100% { box-shadow: 0 0 0 0 rgba(37, 99, 235, 0); }
    }
    .step-inactive .mos-step-circle {
        border: 2px solid #E2E8F0;
        color: #94A3B8;
    }
    .step-done .mos-step-circle {
        background: #2563EB;
        border: 2px solid #2563EB;
        color: white;
    }
    .mos-step-label {
        font-size: 13px;
        font-weight: 700;
        margin-top: 4px;
    }
    .step-active .mos-step-label { color: #1E3A8A; }
    .step-inactive .mos-step-label { color: #94A3B8; }
    .step-done .mos-step-label { color: #2563EB; }
    
    .mos-stepper-line {
        flex: 1;
        height: 1px;
        background: #E2E8F0;
        margin: 0 16px;
        margin-bottom: 20px; /* offset to align with circles */
    }

    /* ── Upload zone override ── */
    [data-testid="stFileUploader"] {
        background: #F8FAFC !important;
        border: 1.5px dashed #CBD5E1 !important;
        border-radius: 12px !important;
        padding: 24px !important;
    }

    /* ── File tag ── */
    .mos-file-tag {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: white;
        border: 1px solid #E2E8F0;
        border-radius: 8px;
        padding: 8px 12px;
        font-size: 12px;
        color: #334155;
        margin: 4px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.04);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .mos-file-tag:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.08);
        border-color: #BFDBFE;
    }

    /* ── Name mapping grid ── */
    .mos-nv-label {
        background: #F1F5F9;
        color: #0F172A;
        border: 1px solid #E2E8F0;
        border-radius: 6px;
        padding: 4px 10px;
        font-size: 12px;
        font-weight: 600;
        display: inline-block;
        margin-bottom: 4px;
    }

    /* ── Result summary bar ── */
    .mos-summary {
        background: #F8FAFC;
        border: 1px solid #E2E8F0;
        border-radius: 12px;
        padding: 14px 20px;
        display: flex;
        gap: 32px;
        margin-bottom: 16px;
        flex-wrap: wrap;
    }
    .mos-summary-item .label {
        font-size: 11px;
        color: #64748B;
        text-transform: uppercase;
        letter-spacing: .05em;
    }
    .mos-summary-item .value {
        font-size: 20px;
        font-weight: 700;
        color: #0F172A;
    }

    /* Override primary button */
    .stButton > button[kind="primary"] {
        background: #3B82F6 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 500 !important;
    }
    </style>
    """, unsafe_allow_html=True)
    

    
    st.markdown(f"""
    <div class="mos-header">
        <div class="mos-header-left">
            <img src="data:image/png;base64,{LOGO_HEADER_B64}" class="mos-header-logo">
            <div>
                <p class="mos-hero-title">Tổng hợp giờ làm dự án MOS</p>
                <p class="mos-hero-sub">Upload file Report của từng thành viên · Hệ thống tự động tổng hợp theo mã dự án</p>
            </div>
        </div>
        <span class="mos-hero-badge">モス委託業務工数集計</span>
    </div>
    
    """, unsafe_allow_html=True)
    
    # Xác định bước hiện tại
    step = 1
    if 'df_mos_result' in st.session_state and st.session_state.df_mos_result is not None:
        step = 2
    if st.session_state.get('mos_saved', False):
        step = 3

    st.markdown(f"""
    <div class="mos-stepper">
        <div class="mos-step-item {'step-done' if step > 1 else 'step-active'}">
            <div class="mos-step-circle">{'✓' if step > 1 else '1'}</div>
            <div class="mos-step-label">Tải lên báo cáo</div>
        </div>
        <div class="mos-stepper-line"></div>
        <div class="mos-step-item {'step-done' if step > 2 else ('step-active' if step == 2 else 'step-inactive')}">
            <div class="mos-step-circle">{'✓' if step > 2 else '2'}</div>
            <div class="mos-step-label">Kiểm tra dữ liệu</div>
        </div>
        <div class="mos-stepper-line"></div>
        <div class="mos-step-item {'step-active' if step == 3 else 'step-inactive'}">
            <div class="mos-step-circle">3</div>
            <div class="mos-step-label">Xuất file báo cáo</div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Đã gộp vào stepper ở trên
    
    uploaded_files = st.file_uploader(
        "Upload",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        label_visibility="collapsed"
    )
    
    if not uploaded_files:
        st.info("Vui lòng tải lên các file báo cáo giờ làm của thành viên để bắt đầu.")
        return
            
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    col_btn, _ = st.columns([1, 3])
    with col_btn:
        run_btn = st.button("⚡ Tổng hợp dữ liệu", type="primary", use_container_width=True)
        
    if run_btn:
        files_key = '_'.join(sorted([f.name for f in uploaded_files]))
        if st.session_state.get('mos_files_key') == files_key and 'df_mos_result' in st.session_state:
            st.success("✅ Sử dụng kết quả đã tóm tắt (Cached). Nếu muốn tính lại, hãy chọn lại file!")
        else:
            has_key = bool(load_saved_api_key())
            
            if not has_key:
                st.warning("⚠️ Chưa cấu hình GEMINI_API_KEY. Hệ thống đang dùng chế độ nối chuỗi thủ công thay vì dùng AI!")
                
            with st.spinner("🤖 AI đang phân tích và tóm tắt nội dung ủy thác..."):
                dfs = []
                file_errors = []
                for f in uploaded_files:
                    try:
                        df_f = parse_mos_file(f, f.name)
                        if not df_f.empty:
                            dfs.append(df_f)
                    except Exception as e:
                        file_errors.append(str(e))
                
                if file_errors:
                    st.session_state['mos_file_errors'] = file_errors
                else:
                    st.session_state.pop('mos_file_errors', None)
                
                if dfs:
                    df_tong_hop = tong_hop_mos(dfs)
                    st.session_state['df_mos_raw'] = pd.concat(dfs, ignore_index=True)
                    st.session_state['df_mos_result'] = df_tong_hop
                    st.session_state['mos_files_key'] = files_key
                    st.session_state['mos_saved'] = False
                    st.success("✅ Tổng hợp và tóm tắt AI xong!")
                else:
                    st.error("Không tìm thấy dữ liệu hợp lệ (phần MOS業務) trong các file đã tải lên.")
                
    if 'df_mos_result' in st.session_state and st.session_state['df_mos_result'] is not None:
        df_result = st.session_state['df_mos_result']
        


        total_da = len(df_result)
        total_gio = df_result['Giờ làm (h)'].sum()
        total_nv = len(uploaded_files)
        
        st.markdown(f"""
        <div class="mos-summary">
            <div class="mos-summary-item">
                <div class="label">Số dự án</div>
                <div class="value">{total_da}</div>
            </div>
            <div class="mos-summary-item">
                <div class="label">Tổng giờ làm</div>
                <div class="value">{total_gio:.1f} h</div>
            </div>
            <div class="mos-summary-item">
                <div class="label">Số thành viên</div>
                <div class="value">{total_nv}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("---")
        df_raw = st.session_state.get('df_mos_raw', pd.DataFrame())
        tab_data, tab_emp, tab_dash = st.tabs(["📝 Dữ liệu Dự án (MOS)", "👥 Năng suất Cá nhân", "📊 Dashboard Thống Kê"])

        with tab_data:
            # 4. BỘ LỌC THÔNG MINH
            st.markdown("### 🔍 Bộ lọc nâng cao")
            col_f1, col_f2, col_f3 = st.columns(3)
            with col_f1:
                search_name = st.text_input("Tên dự án:", placeholder="Nhập từ khóa...")
            with col_f2:
                ma_da_list = df_result['Mã dự án'].unique().tolist()
                selected_ma_da = st.multiselect("Mã dự án:", options=ma_da_list, default=[])
            with col_f3:
                default_managers = ["フォン \n Phương", "ロン \n Long", "ダオ \n Đạo"]
                data_managers = [str(x) for x in df_result['Quản lý Việt Nam'].dropna().unique() if str(x).strip()]
                manager_list = ["Tất cả"] + sorted(list(set(default_managers + data_managers)))
                selected_manager = st.selectbox("Quản lý VN:", options=manager_list)

            df_display = df_result.copy()
            if search_name:
                df_display = df_display[df_display['Tên dự án'].str.contains(search_name, case=False, na=False)]
            if selected_ma_da:
                df_display = df_display[df_display['Mã dự án'].isin(selected_ma_da)]
            if selected_manager != "Tất cả":
                df_display = df_display[df_display['Quản lý Việt Nam'] == selected_manager]

            # 5. CẢNH BÁO BẤT THƯỜNG
            def check_anomaly(row):
                gio = row.get('Giờ làm (h)', 0)
                try: gio = float(gio)
                except: gio = 0
                if gio <= 0: return "🔴 Bất thường (0h)"
                if gio > 50: return "🟡 Chú ý (OT cao)"
                return "🟢 Bình thường"

            if 'Cảnh báo' not in df_display.columns:
                df_display.insert(1, 'Cảnh báo', df_display.apply(check_anomaly, axis=1))

            with st.form("mos_editor_form"):
                st.markdown("💡 **Mẹo:** Sửa trực tiếp trên bảng, sau đó bấm **Lưu thay đổi** ở bên dưới để cập nhật.")
                edited_display = st.data_editor(
                    df_display,
                    use_container_width=True,
                    num_rows="dynamic",
                    column_config={
                        "STT": st.column_config.NumberColumn("STT", disabled=True),
                        "Cảnh báo": st.column_config.TextColumn("Cảnh báo", disabled=True),
                        "Mã dự án": st.column_config.TextColumn("Mã dự án", disabled=False),
                        "Tên dự án": st.column_config.TextColumn("Tên dự án", disabled=False),
                        "Phân vùng": st.column_config.TextColumn("Phân vùng"),
                        "Nội dung ủy thác": st.column_config.TextColumn("Nội dung ủy thác"),
                        "Giờ làm (h)": st.column_config.NumberColumn("Giờ làm (h)", disabled=False),
                        "Ngày bắt đầu": st.column_config.TextColumn("Ngày bắt đầu", disabled=False),
                        "Ngày kết thúc": st.column_config.TextColumn("Ngày kết thúc", disabled=False),
                        "Quản lý Nhật Bản": st.column_config.TextColumn("Quản lý Nhật Bản", disabled=False),
                        "Quản lý Việt Nam": st.column_config.SelectboxColumn("Quản lý Việt Nam", options=["フォン \n Phương", "ロン \n Long", "ダオ \n Đạo"]),
                        "Người thực hiện": st.column_config.TextColumn("Người thực hiện", disabled=False),
                        "Trạng thái": st.column_config.SelectboxColumn("Trạng thái", options=["完了 \n Hoàn thành", "実行中 \n Đang tiến hành", "未着手 \n Chưa bắt đầu"]),
                    },
                    hide_index=True
                )

                submit_edits = st.form_submit_button("💾 Lưu thay đổi", type="primary")
                if submit_edits:
                    st.session_state['mos_saved'] = True
                    edited_no_warning = edited_display.drop(columns=['Cảnh báo'], errors='ignore')
                    if search_name or selected_ma_da or selected_manager != "Tất cả":
                        df_result.update(edited_no_warning)
                        st.session_state['df_mos_edited'] = df_result
                        st.session_state['df_mos_result'] = df_result
                    else:
                        st.session_state['df_mos_edited'] = edited_no_warning
                        st.session_state['df_mos_result'] = edited_no_warning
                    st.rerun()


        with tab_emp:
            st.markdown("### 👥 Thống kê Giờ làm theo Cá nhân (Chỉ tính MOS)")
            if not df_raw.empty:
                emp_stats = df_raw.groupby("ten_nv")["tong_gio"].sum().reset_index()
                emp_stats.columns = ["Tên Nhân viên", "Tổng giờ MOS"]
                emp_stats = emp_stats.sort_values("Tổng giờ MOS", ascending=False)
                st.dataframe(emp_stats, use_container_width=True)
            else:
                st.info("Chưa có dữ liệu thô để thống kê cá nhân.")

        with tab_dash:
            st.markdown("### 📊 Dashboard Thống Kê")
            import plotly.express as px
            if not df_result.empty:
                col_d1, col_d2 = st.columns(2)
                with col_d1:
                    fig_pie = px.pie(df_result, values="Giờ làm (h)", names="Tên dự án", title="Phân bổ Giờ làm theo Dự án")
                    st.plotly_chart(fig_pie, use_container_width=True)
                with col_d2:
                    mgr_stats = df_result.groupby("Quản lý Việt Nam")["Giờ làm (h)"].sum().reset_index()
                    fig_bar = px.bar(mgr_stats, x="Quản lý Việt Nam", y="Giờ làm (h)", title="Tổng giờ làm theo Quản lý VN", color="Quản lý Việt Nam")
                    st.plotly_chart(fig_bar, use_container_width=True)
            else:
                st.info("Không có dữ liệu để vẽ biểu đồ.")

        # If not submitted but it exists in session, ensure it's mapped so download works correctly
        if 'df_mos_result' not in st.session_state:
            st.session_state['df_mos_result'] = df_result
        if 'df_mos_edited' not in st.session_state:
            st.session_state['df_mos_edited'] = df_result
        

        
        def to_excel(df):
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            import re
            import datetime
            import io
            
            font_normal = Font(name='Times New Roman', size=12)
            font_bold = Font(name='Times New Roman', size=12, bold=True)
            font_title = Font(name='Times New Roman', size=16, bold=True)
            
            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
            align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            def format_bilingual(text):
                if text is None: return ""
                text_str = str(text).strip()
                if text_str.lower() in ['nan', 'nat', 'none', '']: return ""
                if '\n' not in text_str:
                    text_str = re.sub(r'([\u3040-\u30ff\u4e00-\u9faf])\s*(?:/|-)?\s*([A-Za-zÀ-ỹ])', r'\1\n\2', text_str)
                    match = re.match(r'^([A-Za-zÀ-ỹ\s/]+)\s*(?:/|-)?\s*([\u3040-\u30ff\u4e00-\u9faf\s/]+)$', text_str)
                    if match:
                        text_str = match.group(2).strip() + '\n' + match.group(1).strip()
                return text_str

            def set_cell(cell, text, bold=False, align=align_center, fill=None, font=None):
                cell.value = format_bilingual(text)
                if font:
                    cell.font = font
                else:
                    cell.font = font_bold if bold else font_normal
                cell.alignment = align
                if fill:
                    cell.fill = fill

            # --- Pre-calculate stats ---
            unique_people = set()
            if 'Người thực hiện' in df.columns:
                for val in df['Người thực hiện']:
                    v_str = str(val).strip()
                    if v_str and v_str.lower() not in ['nan', 'nat', 'none']:
                        unique_people.add(v_str)
            num_people = len(unique_people)
            
            sum_hours = 0
            for idx, row in df.iterrows():
                gio_lam = row.get('Giờ làm (h)', '')
                try: 
                    if str(gio_lam).strip() != '':
                        sum_hours += float(gio_lam)
                except: pass

            output = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Tổng hợp MOS'
            
            now = datetime.datetime.now()
            year_jp = now.year - 2000
            
            # --- Header Title ---
            ws.merge_cells('A1:J1')
            set_cell(ws['A1'], 'モス委託業務工数集計', font=font_title)
            
            ws.merge_cells('A2:J2')
            set_cell(ws['A2'], 'Bảng kê chi tiết nội dung nghiệp vụ ủy thác', font=font_title)
            
            ws.merge_cells('A4:J4')
            set_cell(ws['A4'], f'（{year_jp}年{now.month}月分）', font=font_title)
            
            ws.merge_cells('A5:J5')
            set_cell(ws['A5'], f'Phần tháng {now.month}/{now.year}', font=font_title)
            
            # --- Các bảng phụ bên trái ---
            sub_table_titles = [
                ('A10', '人数\nSố người'),
                ('A11', '一人当たり月枠稼働時間(h)\nGiờ làm việc tiêu chuẩn(h)'),
                ('A12', '月目標稼働時間(h)\nMục tiêu giờ làm(h)'),
                ('A13', '月実績稼働時間(h)\nGiờ làm thực tế(h)'),
                ('A14', '目標に対して稼働率(%)\nTỷ lệ hoàn thành(%)')
            ]
            for coord, title in sub_table_titles:
                set_cell(ws[coord], title, align=align_right)
                ws[coord].border = thin_border
                ws[f'B{coord[1:]}'].border = thin_border
                ws[f'B{coord[1:]}'].font = font_normal
                ws[f'B{coord[1:]}'].alignment = align_center

            ws['B10'] = num_people
            ws['B11'] = 168
            ws['B12'] = "=B10*B11"
            ws['B13'] = sum_hours
            ws['B14'] = "=B13/(B10*B11)"
            ws['B14'].number_format = '0.00%'
            
            # --- Cột Ngày tháng bên phải ---
            ws.merge_cells('H10:J10')
            set_cell(ws['H10'], f'作成日付： {now.year}年{now.month}月{now.day}日', align=align_left)
            ws.merge_cells('H11:J11')
            set_cell(ws['H11'], f'Ngày lập bảng kê: {now.day}/{now.month}/{now.year}', align=align_left)
            ws.merge_cells('H12:J12')
            set_cell(ws['H12'], '作成者： レータンフォン', align=align_left)
            ws.merge_cells('H13:J13')
            set_cell(ws['H13'], 'Người lập bảng kê: Lê Thanh Phương', align=align_left)
            
            # --- Table Headers (Row 16 & 17) ---
            headers = [
                ('A16:A17', '案件名\nTên dự án', 30),
                ('B16:B17', '区分\nPhân vùng', 15),
                ('C16:C17', '委託内容\nNội dung ủy thác', 50),
                ('D16:D17', '実工数(h)\nGiờ làm (h)', 15),
                ('E16:F16', '期間 Thời gian', None),
                ('G16:H16', '管理者 Người quản lý', None),
                ('I16:I17', '実施者\nNgười thực hiện', 20),
                ('J16:J17', '状態\nTrạng thái', 15)
            ]
            
            for range_str, text, width in headers:
                ws.merge_cells(range_str)
                cell = ws[range_str.split(':')[0]]
                if text in ["期間 Thời gian", "管理者 Người quản lý"]:
                    cell.value = text
                    cell.font = font_bold
                    cell.alignment = align_center
                else:
                    set_cell(cell, text, bold=True)
                if width:
                    ws.column_dimensions[cell.column_letter].width = width
            
            sub_headers = {
                'E17': '委託受領日\nNgày bắt đầu',
                'F17': '完了日\nNgày kết thúc',
                'G17': '日本\nNhật Bản',
                'H17': 'ベトナム\nViệt Nam'
            }
            for c, text in sub_headers.items():
                set_cell(ws[c], text, bold=True)
                ws.column_dimensions[ws[c].column_letter].width = 15
            
            for r in ws['A16:J17']:
                for cell in r:
                    cell.font = font_bold
                    cell.alignment = align_center
                    
            for range_str, text, _ in headers:
                cell = ws[range_str.split(':')[0]]
                if text in ["期間 Thời gian", "管理者 Người quản lý"]:
                    cell.value = text
                else:
                    cell.value = format_bilingual(text)
            for c, text in sub_headers.items():
                ws[c].value = format_bilingual(text)
            
            # --- Write Data ---
            row_idx = 18
            for idx, row in df.iterrows():
                ten_da = f"{row.get('Mã dự án', '')}_{row.get('Tên dự án', '')}"
                if ten_da == "_": ten_da = ""
                
                set_cell(ws[f'A{row_idx}'], ten_da, align=align_left)
                set_cell(ws[f'B{row_idx}'], row.get('Phân vùng', ''))
                set_cell(ws[f'C{row_idx}'], row.get('Nội dung ủy thác', ''), align=align_left)
                
                gio_lam = row.get('Giờ làm (h)', '')
                ws[f'D{row_idx}'] = gio_lam
                ws[f'D{row_idx}'].font = font_normal
                ws[f'D{row_idx}'].alignment = align_center
                
                set_cell(ws[f'E{row_idx}'], row.get('Ngày bắt đầu', ''))
                set_cell(ws[f'F{row_idx}'], row.get('Ngày kết thúc', ''))
                set_cell(ws[f'G{row_idx}'], row.get('Quản lý Nhật Bản', ''))
                set_cell(ws[f'H{row_idx}'], row.get('Quản lý Việt Nam', ''))
                set_cell(ws[f'I{row_idx}'], row.get('Người thực hiện', ''), align=align_left)
                set_cell(ws[f'J{row_idx}'], row.get('Trạng thái', ''))
                
                row_idx += 1
            
            # --- Footer ---
            ws.merge_cells(f'A{row_idx}:C{row_idx}')
            set_cell(ws[f'A{row_idx}'], '実工数合計(h)\nTổng giờ làm (h)', bold=True, align=align_right)
            
            ws[f'D{row_idx}'] = sum_hours
            ws[f'D{row_idx}'].font = font_bold
            ws[f'D{row_idx}'].alignment = align_center
            
            # Kẻ khung toàn bộ bảng (từ row 16 đến row_idx)
            for row_cells in ws.iter_rows(min_row=16, max_row=row_idx, min_col=1, max_col=10):
                for cell in row_cells:
                    cell.border = thin_border
            
            wb.save(output)
            return output.getvalue()

        excel_data = to_excel(st.session_state['df_mos_edited'])
        st.download_button(
            label="📥 Tải file Excel Báo cáo",
            data=excel_data,
            file_name=f"Tong_hop_MOS_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

# ==========================================
# GIAO DIỆN CÁC CHỨC NĂNG
# ==========================================

# Chỉ ẩn sidebar ở mos (chấm công thì hiện ra)
if st.session_state.app_page == "mos":
    st.markdown("""
    <style>
    [data-testid="stSidebar"] { display: none !important; }
    [data-testid="collapsedControl"] { display: none !important; }
    

    /* Auto-width centered buttons */

    .stButton > button {
        height: 50px;
        padding: 0 32px !important;
        border-radius: 12px !important;
        background: linear-gradient(135deg, #1E3A8A, #3B82F6) !important;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3) !important;
        transition: transform 0.2s ease, box-shadow 0.2s ease !important;
        color: white !important;
        border: none !important;
        width: auto !important; /* Force width to fit text */
    }
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 16px rgba(59, 130, 246, 0.4) !important;
    }
    .stButton > button p {
        font-size: 16px !important;
        font-weight: 600 !important;
    }
    </style>
    

    """, unsafe_allow_html=True)





    render_mos_page()
    
    render_chatbot()
    st.stop()

# ==========================================
# CHỨC NĂNG CHẤM CÔNG (Giữ nguyên như cũ)
# Nút quay lại trang chủ ở sidebar
if st.sidebar.button("⬅️ Quay lại Trang chủ", type="primary", use_container_width=True, key="btn_1018"):
    st.session_state.app_page = "home"
    st.rerun()
# ==========================================




if st.session_state.get("show_history") and app_mode == "Xử lý file chấm công":
    st.session_state.show_history = False
    st.rerun()

# ----- UPLOAD FILE -----
with st.sidebar:
    render_holiday_makeup_sidebar()
    st.markdown("---")

st.sidebar.markdown("### 📁 Bước 1: Upload file chấm công")
st.sidebar.markdown('<div class="upload-hint">✅ Hỗ trợ: <b>.xlsx, .xls, .csv, .txt, .dat, .tsv</b></div>', unsafe_allow_html=True)
uploaded_file = st.sidebar.file_uploader("Chọn file chấm công", type=["xlsx","xls","csv","txt","dat","tsv","xlsm","xlsb"], label_visibility="collapsed")

# ----- LANDING PAGE (khi chưa có file) -----
if uploaded_file is None and st.session_state.df_raw is None:
    logo_b64 = LOGO_HEADER_B64
    logo_html = _logo_img_tag(logo_b64, "height: 80px; object-fit: contain;") if logo_b64 else '<span style="color:#10B981">✦</span> VIET.MOS'
    st.markdown(f"""
<style>

/* Landing Page Wrapper */
.landing-wrapper {{
    font-family: 'Inter', 'Be Vietnam Pro', sans-serif;
    color: #0F172A;
    display: flex;
    flex-direction: column;
    gap: 48px;
    margin-top: 10px;
}}

/* Hero Section */
.hero-section {{
    background: linear-gradient(135deg, #2563EB 0%, #3B82F6 100%);
    border-radius: 24px;
    position: relative;
    overflow: hidden;
    padding: 32px 48px 64px 48px;
    color: white;
    box-shadow: 0 20px 40px rgba(37, 99, 235, 0.15);
}}
.hero-section::before {{
    content: ''; position: absolute; top: -50%; right: -10%;
    width: 600px; height: 600px;
    background: radial-gradient(circle, rgba(255,255,255,0.08) 0%, rgba(255,255,255,0) 70%);
    border-radius: 50%; pointer-events: none;
}}

/* Nav inside hero */
.hero-nav {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 60px;
}}
.hero-logo {{
    background: white;
    color: #1E3A8A;
    font-weight: 800;
    font-size: 13px;
    padding: 6px 16px;
    border-radius: 100px;
    display: inline-flex;
    align-items: center;
    gap: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.1);
}}
.hero-menu {{
    display: flex;
    gap: 32px;
    font-size: 13.5px;
    font-weight: 500;
}}
.hero-menu span {{
    cursor: pointer;
    opacity: 0.9;
    transition: opacity 0.2s;
}}
.hero-menu span:hover {{ opacity: 1; }}

/* Hero Content */
.hero-content {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 40px;
}}
.hero-text {{
    flex: 1;
    max-width: 480px;
}}
.hero-title {{
    font-size: 46px;
    font-weight: 800;
    line-height: 1.15;
    margin: 0 0 16px 0;
    letter-spacing: -0.02em;
}}
.hero-subtitle {{
    font-size: 17px;
    font-weight: 500;
    line-height: 1.5;
    margin: 0 0 12px 0;
    color: rgba(255, 255, 255, 0.95);
}}
.hero-desc {{
    font-size: 15px;
    font-weight: 400;
    margin: 0 0 32px 0;
    color: rgba(255, 255, 255, 0.75);
}}
.btn-upload-hero {{
    background: white;
    color: #1E3A8A;
    padding: 14px 24px;
    border-radius: 100px;
    font-size: 14px;
    font-weight: 600;
    border: none;
    cursor: pointer;
    box-shadow: 0 8px 24px rgba(0,0,0,0.1);
    transition: transform 0.2s;
    display: inline-flex;
    align-items: center;
    gap: 8px;
}}
.btn-upload-hero:hover {{ transform: translateY(-2px); }}

/* Animation Keyframes */
@keyframes float-slow {{
    0%, 100% {{ transform: translateY(0); }}
    50% {{ transform: translateY(-15px); }}
}}
@keyframes float-fast {{
    0%, 100% {{ transform: translateY(0); }}
    50% {{ transform: translateY(-8px); }}
}}
@keyframes float-delay {{
    0%, 100% {{ transform: translateY(0); }}
    50% {{ transform: translateY(-12px); }}
}}
@keyframes slide-arrow {{
    0%, 100% {{ transform: translateX(0) scale(1); }}
    50% {{ transform: translateX(10px) scale(1.05); }}
}}
@keyframes bar-grow {{
    0%, 100% {{ transform: scaleY(0.8); }}
    50% {{ transform: scaleY(1); }}
}}

/* Graphic Animated */
.hero-graphic {{
    flex: 1.2;
    position: relative;
    height: 380px;
    display: flex;
    align-items: center;
    justify-content: flex-end;
}}

/* Floating Icons */
.f-icon {{
    position: absolute;
    width: 48px; height: 48px;
    background: white;
    border-radius: 12px;
    box-shadow: 0 10px 25px rgba(0,0,0,0.15);
    display: flex; align-items: center; justify-content: center;
    font-size: 24px;
    z-index: 5;
    transition: transform 0.3s ease, box-shadow 0.3s ease;
}}
.f-icon:hover {{
    transform: translateY(-10px);
}}
.fi-clock {{ top: 30px; right: 420px; color: #3B82F6; }}
.fi-gear {{ top: -10px; right: 140px; color: #3B82F6; }}
.fi-chart {{ top: 40px; right: -20px; color: #10B981; }}

/* Laptop */
.laptop-container {{
    position: relative;
    width: 440px;
    z-index: 2;
    transition: transform 0.3s ease, box-shadow 0.3s ease;
}}
.laptop-container:hover {{
    transform: translateY(-10px);
}}
.laptop-screen {{
    background: #111827;
    border-radius: 16px 16px 0 0;
    padding: 12px 12px 16px 12px;
    box-shadow: 0 20px 40px rgba(0,0,0,0.3);
}}
.laptop-display {{
    background: #F1F5F9;
    border-radius: 4px;
    height: 240px;
    display: flex;
    overflow: hidden;
}}
.laptop-base {{
    height: 12px;
    background: #CBD5E1;
    border-radius: 0 0 16px 16px;
    position: relative;
    box-shadow: inset 0 -4px 8px rgba(0,0,0,0.1);
}}
.laptop-base::after {{
    content: ''; position: absolute; top: 0; left: 50%; transform: translateX(-50%);
    width: 60px; height: 4px; background: #94A3B8;
    border-radius: 0 0 4px 4px;
}}

/* Dashboard UI */
.dash-sidebar {{
    width: 25%;
    background: #1E3A8A;
    padding: 12px 8px;
    display: flex;
    flex-direction: column;
    gap: 12px;
}}
.dash-logo {{
    color: white; font-weight: 800; font-size: 10px; margin-bottom: 8px;
}}
.dash-menu-item {{
    height: 6px; background: rgba(255,255,255,0.2); border-radius: 4px; width: 80%;
}}
.dash-menu-item.active {{ background: #3B82F6; width: 100%; }}

.dash-main {{
    flex: 1; padding: 12px; display: flex; flex-direction: column; gap: 8px;
}}
.dash-header {{ height: 10px; background: #E2E8F0; width: 40px; border-radius: 4px; margin-bottom: 4px; }}
.dash-grid {{
    display: grid; grid-template-columns: 1fr 1fr; gap: 8px; flex: 1;
}}
.dash-card {{
    background: white; border-radius: 6px; padding: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    display: flex; flex-direction: column; gap: 6px;
}}
.dash-card-title {{ height: 6px; background: #E2E8F0; width: 50%; border-radius: 3px; }}
.dash-bars {{ display: flex; gap: 4px; align-items: flex-end; flex: 1; justify-content: center; }}
.dash-bar {{ width: 8px; background: #3B82F6; border-radius: 2px 2px 0 0; transform-origin: bottom; transition: transform 0.3s ease; }}
.dash-bar:hover {{ transform: scaleY(1.2); }}
.dash-bar.green {{ background: #10B981; }}
.db1 {{ height: 40%; }} .db2 {{ height: 70%; }} .db3 {{ height: 50%; }}
.db4 {{ height: 80%; }} .db5 {{ height: 40%; }} .db6 {{ height: 90%; }}

/* Excel Icon */
.excel-wrapper {{
    position: absolute;
    right: 380px;
    bottom: 60px;
    z-index: 10;
    transition: transform 0.3s ease;
}}
.excel-wrapper:hover {{
    transform: translateY(-10px);
}}
.excel-file {{
    width: 80px; height: 100px;
    background: #10B981;
    border-radius: 8px;
    box-shadow: 0 12px 24px rgba(0,0,0,0.2);
    position: relative;
    padding: 12px;
    color: white;
    display: flex;
    flex-direction: column;
    justify-content: center;
}}
.excel-file::after {{
    content: ''; position: absolute; top: 0; right: 0;
    border-width: 0 24px 24px 0;
    border-style: solid;
    border-color: rgba(255,255,255,0.8) rgba(255,255,255,0.8) #059669 #059669;
    border-radius: 0 8px 0 8px;
    box-shadow: -2px 2px 4px rgba(0,0,0,0.1);
}}
.excel-x {{ font-weight: 900; font-size: 32px; font-family: sans-serif; line-height: 1; text-align: center; margin-bottom: 8px;}}
.excel-grid {{
    display: grid; grid-template-columns: 1fr 1fr; gap: 4px;
}}
.excel-cell {{ height: 6px; background: rgba(255,255,255,0.4); border-radius: 2px;}}

.excel-check {{
    position: absolute;
    bottom: -10px; right: -10px;
    width: 40px; height: 40px;
    background: #10B981;
    border: 4px solid white;
    border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    color: white; font-weight: bold; font-size: 20px;
    box-shadow: 0 4px 8px rgba(0,0,0,0.1);
}}

/* Arrow */
.transfer-arrow {{
    position: absolute;
    right: 280px;
    bottom: 120px;
    z-index: 9;
    transform-origin: left center;
    transition: transform 0.3s ease;
}}
.transfer-arrow:hover {{
    transform: translateX(10px) scale(1.05);
}}

/* Dropdown Menu */
.menu-dropdown {{
    position: relative;
    display: inline-block;
    padding-bottom: 30px;
    margin-bottom: -30px;
}}
.dropdown-content {{
    visibility: hidden;
    opacity: 0;
    position: absolute;
    top: 100%;
    left: 50%;
    transform: translateX(-50%) translateY(10px);
    background-color: white;
    min-width: 340px;
    box-shadow: 0 20px 40px rgba(0,0,0,0.15);
    border-radius: 16px;
    padding: 16px;
    z-index: 100;
    display: flex;
    flex-direction: column;
    gap: 12px;
    transition: all 0.3s ease;
    cursor: default;
}}
.menu-dropdown:hover .dropdown-content {{
    visibility: visible;
    opacity: 1;
    transform: translateX(-50%) translateY(0);
}}
.menu-dropdown.right-align .dropdown-content {{
    left: auto;
    right: 0;
    transform: translateY(10px);
}}
.menu-dropdown.right-align:hover .dropdown-content {{
    transform: translateY(0);
}}
.feature-card {{
    background: #F8FAFC;
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 16px;
    text-align: left;
    box-shadow: 0 2px 8px rgba(0,0,0,0.02);
    display: grid;
    grid-template-columns: auto 1fr;
    column-gap: 16px;
    row-gap: 4px;
    align-items: start;
    cursor: pointer;
    transition: background 0.2s ease;
}}
.feature-card:hover {{
    background: #F1F5F9;
}}
.fc-icon {{
    font-size: 20px;
    background: white;
    width: 36px; height: 36px;
    display: inline-flex;
    align-items: center; justify-content: center;
    border-radius: 8px;
    grid-row: span 2;
    margin-bottom: 0;
}}
.feature-card h3 {{
    font-size: 14px;
    font-weight: 700;
    color: #1E293B;
    margin: 0;
}}
.feature-card p {{
    font-size: 13px;
    color: #64748B;
    margin: 0;
    line-height: 1.4;
}}
.contact-box {{
    display: flex; align-items: center; gap: 14px; background: #F8FAFC; padding: 14px; 
    border-radius: 12px; border: 1px solid #E2E8F0; transition: all 0.2s ease; cursor: default;
}}
.contact-box:hover {{
    border-color: #3B82F6; background: #EFF6FF; transform: translateY(-2px); box-shadow: 0 4px 12px rgba(59,130,246,0.1);
}}
</style>

<div class="landing-wrapper">
<div class="hero-section">
<div class="hero-nav">
<div class="hero-logo" style="background:transparent; box-shadow:none; padding:0;">
{logo_html}
</div>
<div class="hero-menu">

<div class="menu-dropdown">
    <span class="dropdown-trigger">Tính năng</span>
    <div class="dropdown-content">
        <div class="feature-card">
            <div class="fc-icon">⏱️</div>
            <h3>Tính giờ tự động</h3>
            <p>Tự tính giờ làm thực tế, trừ giờ nghỉ trưa đúng chuẩn, xoá mờ mọi lỗi thủ công</p>
        </div>
        <div class="feature-card">
            <div class="fc-icon">⚙️</div>
            <h3>Tuỳ biến linh hoạt</h3>
            <p>Tự chỉnh giờ làm chuẩn, giờ nghỉ trưa theo ca làm việc linh hoạt</p>
        </div>
        <div class="feature-card">
            <div class="fc-icon">📊</div>
            <h3>Xuất báo cáo tức thì</h3>
            <p>Xuất file Excel báo cáo chi tiết, chuyên nghiệp sẵn sàng trình sếp</p>
        </div>
    </div>
</div>
<a href="?page=huong_dan" target="_self" style="text-decoration: none; color: inherit;">Hướng dẫn</a>
<div class="menu-dropdown right-align">
    <span class="dropdown-trigger">Liên hệ</span>
    <div class="dropdown-content" style="min-width: 280px; padding: 0; overflow: hidden; border: none; box-shadow: 0 20px 40px rgba(0,0,0,0.15);">
        <div style="background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%); padding: 24px 20px; text-align: center;">
            <div style="width: 56px; height: 56px; background: white; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 26px; margin: 0 auto 12px auto; box-shadow: 0 8px 24px rgba(0,0,0,0.2);">
                👨‍💻
            </div>
            <h3 style="font-size: 16px; font-weight: 700; color: white; margin: 0;">Kỹ thuật trang web</h3>
            <p style="font-size: 13px; color: rgba(255,255,255,0.8); margin: 4px 0 0 0;">Luôn sẵn sàng hỗ trợ bạn</p>
        </div>
        <div style="padding: 16px; background: white;">
            <div class="contact-box">
                <div style="width: 42px; height: 42px; background: #DBEAFE; border-radius: 10px; display: flex; align-items: center; justify-content: center; font-size: 20px; color: #2563EB; flex-shrink: 0;">📞</div>
                <div>
                    <div style="font-size: 11px; color: #64748B; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px;">Điện thoại / Zalo</div>
                    <div style="font-size: 17px; font-weight: 800; color: #0F172A; margin-top: 2px;">0867.153.701</div>
                </div>
            </div>
        </div>
    </div>
</div>
</div>
</div>

<div class="hero-content">
<div class="hero-text">
<h1 class="hero-title">Ứng dụng<br>Bảng Chấm Công</h1>
<p class="hero-subtitle">Tính giờ làm việc thực tế từ file Excel máy chấm công</p>
<p class="hero-desc">Tự động - Chính xác - Tức thì</p>
<button class="btn-upload-hero" onclick="window.parent.document.querySelector('[data-testid=\'stFileUploader\'] input').click()">
👈 Upload file ở menu trái để bắt đầu
</button>
</div>
<div class="hero-graphic">
<div class="f-icon fi-clock">🕒</div>
<div class="f-icon fi-gear">⚙️</div>
<div class="f-icon fi-chart">📈</div>

<div class="laptop-container">
<div class="laptop-screen">
<div class="laptop-display">
<div class="dash-sidebar">
<div class="dash-logo" style="width:24px;height:8px;background:rgba(255,255,255,0.4);border-radius:4px;"></div>
<div class="dash-menu-item active"></div>
<div class="dash-menu-item"></div>
<div class="dash-menu-item"></div>
</div>
<div class="dash-main">
<div class="dash-header"></div>
<div class="dash-grid">
<div class="dash-card">
<div class="dash-card-title"></div>
<div style="flex:1; display:flex; align-items:center; justify-content:center; font-size:32px;">📄</div>
</div>
<div class="dash-card">
<div class="dash-card-title"></div>
<div class="dash-bars">
<div class="dash-bar db1"></div>
<div class="dash-bar db2 green"></div>
<div class="dash-bar db3"></div>
<div class="dash-bar db4 green"></div>
</div>
</div>
<div class="dash-card">
<div class="dash-card-title"></div>
<div class="dash-bars">
<div class="dash-bar db5 green"></div>
<div class="dash-bar db6"></div>
<div class="dash-bar db1 green"></div>
</div>
</div>
<div class="dash-card">
<div class="dash-card-title"></div>
<div style="flex:1; display:flex; align-items:center; justify-content:center;">
<svg width="100%" height="40" viewBox="0 0 100 40" preserveAspectRatio="none">
<path d="M0,35 Q25,10 50,25 T100,15" fill="none" stroke="#3B82F6" stroke-width="4"/>
<path d="M0,25 Q25,35 50,20 T100,5" fill="none" stroke="#10B981" stroke-width="4"/>
</svg>
</div>
</div>
</div>
</div>
</div>
</div>
<div class="laptop-base"></div>
</div>

<svg class="transfer-arrow" width="120" height="80" viewBox="0 0 120 80">
<path d="M 10,70 Q 50,10 100,30" fill="none" stroke="#60A5FA" stroke-width="10" stroke-linecap="round"/>
<polygon points="90,12 120,35 85,48" fill="#60A5FA"/>
</svg>

<div class="excel-wrapper">
<div class="excel-file">
<div style="display:flex; align-items:center; gap:8px; margin-bottom:8px;">
<div class="excel-x">X</div>
<div style="flex:1;">
<div class="excel-cell"></div>
<div class="excel-cell" style="margin-top:4px;"></div>
<div class="excel-cell" style="margin-top:4px;"></div>
</div>
</div>
<div class="excel-grid">
<div class="excel-cell"></div><div class="excel-cell"></div>
<div class="excel-cell"></div><div class="excel-cell"></div>
<div class="excel-cell"></div><div class="excel-cell"></div>
</div>
</div>
<div class="excel-check">✓</div>
</div>
</div>
</div>
</div>


</div>
""", unsafe_allow_html=True)

# ----- XỬ LÝ FILE -----
if uploaded_file is not None:
    if st.session_state.df_raw is None or st.session_state.get('last_uploaded') != uploaded_file.name:
        with st.spinner("Đang đọc file..."):
            st.session_state['uploaded_file_bytes'] = uploaded_file.read()
            uploaded_file.seek(0)
            df = parse_excel_file(uploaded_file)
            if df is not None and not df.empty:
                st.session_state.df_raw = df
                st.session_state.last_uploaded = uploaded_file.name
                mapping_auto = auto_detect_columns(df)
                req_keys = ['ma_nv', 'ten_nv', 'ngay', 'gio_vao', 'gio_ra']
                if all(k in mapping_auto for k in req_keys):
                    st.session_state.mapping = mapping_auto
                    st.session_state.step = 3
                    st.rerun()
                else:
                    st.session_state.step = 2
                    st.session_state.mapping_auto = mapping_auto
elif st.session_state.df_raw is not None:
    st.session_state.df_raw = None
    st.session_state.last_uploaded = None
    st.session_state.step = 1
    for _key in ("mapping", "mapping_auto"):
        if _key in st.session_state: del st.session_state[_key]
    st.rerun()

# ----- MAPPING CỘT -----
if st.session_state.step == 2 and st.session_state.df_raw is not None:
    df_raw = st.session_state.df_raw
    mapping_auto = st.session_state.mapping_auto
    st.markdown("### Dữ liệu thô (10 dòng đầu)")
    st.dataframe(df_raw.head(10), use_container_width=True, hide_index=True)
    st.sidebar.markdown("### 🔗 Bước 2: Xác nhận mapping cột")
    st.sidebar.warning("⚠️ Không nhận diện đủ 5 cột.")
    columns = ["-- Chọn cột --"] + list(df_raw.columns)
    def get_index(col_name):
        return columns.index(col_name) if col_name in columns else 0
    map_emp_id = st.sidebar.selectbox("Mã NV", columns, index=get_index(mapping_auto.get('ma_nv')))
    map_emp_name = st.sidebar.selectbox("Tên NV", columns, index=get_index(mapping_auto.get('ten_nv')))
    map_date = st.sidebar.selectbox("Ngày làm việc", columns, index=get_index(mapping_auto.get('ngay')))
    map_in = st.sidebar.selectbox("Giờ vào", columns, index=get_index(mapping_auto.get('gio_vao')))
    map_out = st.sidebar.selectbox("Giờ ra", columns, index=get_index(mapping_auto.get('gio_ra')))
    
    st.sidebar.markdown("<small style='color:#64748B'>— Tuỳ chọn (để trống nếu không có) —</small>", unsafe_allow_html=True)
    opt_cols = ["(Không có)"] + list(df_raw.columns)
    map_chuc_vu = st.sidebar.selectbox("Chức vụ (tuỳ chọn)", opt_cols, index=0)
    map_phong_ban = st.sidebar.selectbox("Phòng ban (tuỳ chọn)", opt_cols, index=0)
    
    if st.sidebar.button("Xác nhận Mapping", type="primary", use_container_width=True):
        if "-- Chọn cột --" in [map_emp_id, map_emp_name, map_date, map_in, map_out]:
            st.sidebar.error("❌ Vui lòng chọn đủ 5 cột bắt buộc!")
        else:
            new_mapping = {'ma_nv': map_emp_id, 'ten_nv': map_emp_name, 'ngay': map_date, 'gio_vao': map_in, 'gio_ra': map_out}
            if map_chuc_vu != "(Không có)": new_mapping['chuc_vu'] = map_chuc_vu
            if map_phong_ban != "(Không có)": new_mapping['phong_ban'] = map_phong_ban
            st.session_state.mapping = new_mapping
            st.session_state.step = 3
            st.rerun()

# ----- TÍNH TOÁN & KẾT QUẢ -----
if st.session_state.step >= 3 and "mapping" in st.session_state:
    m = st.session_state.mapping
    df_process = st.session_state.df_raw.copy()
    df_process = df_process.dropna(subset=[m['ma_nv']])
    df_process = df_process[df_process[m['ma_nv']].astype(str).str.strip() != '']
    df_process["_parsed_date"] = df_process[m['ngay']].apply(clean_date)
    df_process = df_process.dropna(subset=["_parsed_date"])

    if df_process.empty:
        st.warning("Không có dữ liệu hợp lệ (Lỗi đọc ngày tháng hoặc Mã NV trống).")
    else:
        min_date = df_process["_parsed_date"].min()
        max_date = df_process["_parsed_date"].max()
        default_start = datetime.date(2026, 5, 21)
        default_end = datetime.date(2026, 6, 20)
        if min_date.date() > default_end or max_date.date() < default_start:
            default_start = min_date.date()
            default_end = max_date.date()

        st.sidebar.markdown("### 📅 Bước 3: Chọn kỳ công")
        col_d1, col_d2 = st.sidebar.columns(2)
        with col_d1: start_date = st.date_input("Từ ngày", default_start)
        with col_d2: end_date = st.date_input("Đến ngày", default_end)

        years_in_range = range(start_date.year, end_date.year + 1)
        fixed_holidays = set(get_fixed_holidays_for_years(years_in_range))
        
        if "custom_holidays" not in st.session_state:
            st.session_state.custom_holidays = set()
        if "custom_workdays" not in st.session_state:
            st.session_state.custom_workdays = set()

        holiday_dates = fixed_holidays | st.session_state.custom_holidays
        makeup_dates = st.session_state.custom_workdays
        working_days, holiday_info, makeup_info = calculate_working_days(start_date, end_date, holiday_dates, makeup_dates)

        if makeup_info:
            st.markdown('<div class="card"><div class="card-title"><span class="card-icon">💼</span>Ngày làm bù trong kỳ</div></div>', unsafe_allow_html=True)
            for m in sorted(makeup_info, key=lambda x: x["date"]):
                date_str = m["date"].strftime("%d/%m/%Y")
                st.info(f"💼 **{date_str} ({m['weekday_label']})** được tính là ngày làm việc (Ngày làm bù).")

        if holiday_info:
            st.markdown('<div class="card"><div class="card-title"><span class="card-icon">📅</span>Ngày lễ trong kỳ</div></div>', unsafe_allow_html=True)
            for h in sorted(holiday_info, key=lambda x: x["date"]):
                date_str = h["date"].strftime("%d/%m/%Y")
                if h["is_workday"]:
                    st.warning(f"⚠️ **{date_str} ({h['weekday_label']})** trùng vào ngày làm việc — đã trừ 1 ngày công.")
                else:
                    st.info(f"ℹ️ {date_str} ({h['weekday_label']}) rơi vào ngày nghỉ cuối tuần.")

        with st.sidebar.expander("⚙️ Tuỳ chỉnh giờ làm chuẩn"):
            gio_vao_chuan = st.time_input("Giờ vào chuẩn", datetime.time(8, 0))
            gio_ra_chuan = st.time_input("Giờ ra chuẩn", datetime.time(17, 0))
            nghi_trua_bat_dau = st.time_input("Nghỉ trưa từ", datetime.time(12, 0))
            nghi_trua_ket_thuc = st.time_input("Nghỉ trưa đến", datetime.time(13, 0))
            so_gio_toi_da = st.number_input("Tối đa giờ/ngày", min_value=0.0, max_value=24.0, value=8.0, step=0.5)

        if time_to_float(gio_ra_chuan) <= time_to_float(gio_vao_chuan):
            st.sidebar.error("❌ Giờ ra chuẩn phải sau giờ vào chuẩn."); st.stop()
        if time_to_float(nghi_trua_ket_thuc) < time_to_float(nghi_trua_bat_dau):
            st.sidebar.error("❌ Giờ nghỉ trưa không hợp lệ."); st.stop()

        if "manual_ot" not in st.session_state:
            st.session_state.manual_ot = {}

        with st.sidebar.expander("⏱️ Nhập OT thủ công"):
            st.markdown("<small>Ghi đè số giờ tăng ca bằng tay.</small>", unsafe_allow_html=True)
            emp_list = sorted(df_process[m['ma_nv']].dropna().astype(str).unique())
            sel_emp = st.selectbox("Mã NV", [""] + emp_list, key="ot_emp")
            if sel_emp:
                emp_dates = df_process[df_process[m['ma_nv']].astype(str) == sel_emp]["_parsed_date"].dt.strftime('%d/%m/%Y').tolist()
                sel_date = st.selectbox("Ngày", emp_dates, key="ot_date")
                if sel_date:
                    current_ot = st.session_state.manual_ot.get((sel_emp, sel_date), 0.0)
                    new_ot = st.number_input("Số giờ OT", min_value=0.0, max_value=24.0, value=float(current_ot), step=0.5, key="ot_val")
                    col_btn_ot1, col_btn_ot2 = st.columns(2)
                    with col_btn_ot1:
                        if st.button("💾 Lưu OT", use_container_width=True):
                            st.session_state.manual_ot[(sel_emp, sel_date)] = float(new_ot)
                            st.rerun()
                    with col_btn_ot2:
                        if st.button("🗑️ Xoá", use_container_width=True):
                            if (sel_emp, sel_date) in st.session_state.manual_ot:
                                del st.session_state.manual_ot[(sel_emp, sel_date)]
                                st.rerun()

        mask = (df_process["_parsed_date"].dt.date >= start_date) & (df_process["_parsed_date"].dt.date <= end_date)
        df_filtered = df_process.loc[mask].copy()

        if "manual_emps" not in st.session_state:
            st.session_state.manual_emps = []

        with st.sidebar.expander("➕ Thêm nhân viên thủ công"):
            st.markdown("<small>Tạo dòng chấm công cho nhân viên mới (không có trong file gốc).</small>", unsafe_allow_html=True)
            with st.form("form_add_emp"):
                new_ma = st.text_input("Mã NV (*)")
                new_ten = st.text_input("Tên nhân viên (*)")
                new_cv = st.text_input("Chức vụ")
                new_pb = st.text_input("Phòng ban")
                submitted_emp = st.form_submit_button("Thêm nhân viên", use_container_width=True)
                if submitted_emp:
                    if new_ma and new_ten:
                        st.session_state.manual_emps.append({
                            "ma": new_ma.strip().upper(),
                            "ten": new_ten.strip(),
                            "cv": new_cv.strip(),
                            "pb": new_pb.strip()
                        })
                        st.rerun()
                    else:
                        st.error("Vui lòng nhập Mã NV và Tên NV!")
            if st.session_state.manual_emps:
                st.markdown("**Đã thêm thủ công:**")
                for i, emp in enumerate(st.session_state.manual_emps):
                    c1, c2 = st.columns([4, 1])
                    c1.markdown(f"<small>{emp['ma']} - {emp['ten']}</small>", unsafe_allow_html=True)
                    if c2.button("❌", key=f"del_emp_{i}", use_container_width=True):
                        st.session_state.manual_emps.pop(i)
                        st.rerun()

        if st.session_state.manual_emps:
            new_rows = []
            curr_d = start_date
            while curr_d <= end_date:
                for emp in st.session_state.manual_emps:
                    row_data = {
                        m['ma_nv']: emp['ma'],
                        m['ten_nv']: emp['ten'],
                        "_parsed_date": pd.to_datetime(curr_d),
                        m['ngay']: curr_d.strftime('%d/%m/%Y'),
                        m['gio_vao']: pd.NaT,
                        m['gio_ra']: pd.NaT
                    }
                    if 'chuc_vu' in m: row_data[m['chuc_vu']] = emp['cv']
                    if 'phong_ban' in m: row_data[m['phong_ban']] = emp['pb']
                    new_rows.append(row_data)
                curr_d += datetime.timedelta(days=1)
            if new_rows:
                df_new = pd.DataFrame(new_rows)
                df_filtered = pd.concat([df_filtered, df_new], ignore_index=True)

        if df_filtered.empty:
            st.warning("Không có dữ liệu trong khoảng thời gian đã chọn.")
        else:
            with st.spinner("⏳ Đang tính toán giờ làm..."):
                df_calc = df_filtered.apply(lambda row: calculate_working_hours(
                    row[m['gio_vao']], row[m['gio_ra']],
                    start_chuan=time_to_float(gio_vao_chuan), end_chuan=time_to_float(gio_ra_chuan),
                    lunch_start=time_to_float(nghi_trua_bat_dau), lunch_end=time_to_float(nghi_trua_ket_thuc),
                    max_hours=so_gio_toi_da,
                ), axis=1)
                df_filtered["Giờ hành chính"] = df_calc.apply(lambda x: x['admin_hours'] if isinstance(x, dict) else 0.0)
                df_filtered["Số giờ làm thực tế"] = df_calc.apply(lambda x: x['tong_gio'] if isinstance(x, dict) else 0.0)
                df_filtered["_is_chieu"] = df_calc.apply(lambda x: x.get('is_chieu', False) if isinstance(x, dict) else False)
                df_filtered["Phút đi trễ"] = df_calc.apply(lambda x: x['di_tre'] if isinstance(x, dict) else 0)
                df_filtered["Phút về sớm"] = df_calc.apply(lambda x: x['ve_som'] if isinstance(x, dict) else 0)
                df_filtered["Giờ OT"] = None

                df_filtered["Ngày"] = df_filtered["_parsed_date"].dt.strftime('%d/%m/%Y')
                df_filtered = df_filtered.sort_values(by=[m['ma_nv'], "_parsed_date"])
                df_filtered["_nv_label"] = df_filtered[m['ma_nv']].astype(str) + " - " + df_filtered[m['ten_nv']].astype(str)
                danh_sach_nv = sorted(df_filtered["_nv_label"].unique().tolist())

            st.sidebar.markdown("### 🔍 Bước 4: Lọc dữ liệu")
            chon_nv = st.sidebar.multiselect("Chọn nhân viên", options=danh_sach_nv, default=[], placeholder="Bỏ trống để xem tất cả", label_visibility="collapsed")

            st.markdown("## 📊 Kết quả Tổng hợp")
            if chon_nv:
                df_filtered = df_filtered[df_filtered["_nv_label"].isin(chon_nv)]


            @st.fragment
            def render_interactive_dashboard(df_base):
                df_filtered = df_base.copy()
                # Bắt sự kiện chỉnh sửa trực tiếp từ bảng (Data Editor)
                if "data_editor_ot" in st.session_state and "edited_rows" in st.session_state["data_editor_ot"]:
                    changes = st.session_state["data_editor_ot"]["edited_rows"]
                    if changes and "df_result" in st.session_state:
                        last_df = st.session_state["df_result"]
                        for row_idx_str, edits in changes.items():
                            row_idx = int(row_idx_str)
                            if row_idx < len(last_df):
                                ma_edit = str(last_df.iloc[row_idx]["Mã NV"]).strip().upper()
                                if ma_edit.endswith('.0'): ma_edit = ma_edit[:-2]
                                ngay_edit = last_df.iloc[row_idx]["Ngày"]

                                if "Giờ làm thực tế" in edits:
                                    if "manual_hc" not in st.session_state: st.session_state.manual_hc = {}
                                    val = edits["Giờ làm thực tế"]
                                    st.session_state.manual_hc[(ma_edit, ngay_edit)] = float(val) if val is not None else 0.0

                                if "OT" in edits:
                                    if "manual_ot" not in st.session_state: st.session_state.manual_ot = {}
                                    val = edits["OT"]
                                    st.session_state.manual_ot[(ma_edit, ngay_edit)] = float(val) if val is not None else 0.0

                                if "Có phép" in edits:
                                    if "manual_leave" not in st.session_state: st.session_state.manual_leave = {}
                                    if edits["Có phép"]:
                                        st.session_state.manual_leave[(ma_edit, ngay_edit)] = True
                                    else:
                                        st.session_state.manual_leave.pop((ma_edit, ngay_edit), None)

                                if "Lý do tăng ca" in edits:
                                    if "manual_ot_reason" not in st.session_state: st.session_state.manual_ot_reason = {}
                                    st.session_state.manual_ot_reason[(ma_edit, ngay_edit)] = str(edits["Lý do tăng ca"])

                                if "Ghi chú" in edits:
                                    if "manual_notes" not in st.session_state: st.session_state.manual_notes = {}
                                    if edits["Ghi chú"] is None:
                                        st.session_state.manual_notes.pop((ma_edit, ngay_edit), None)
                                    else:
                                        st.session_state.manual_notes[(ma_edit, ngay_edit)] = str(edits["Ghi chú"])

                                if "Tổng giờ" in edits:
                                    if "manual_total" not in st.session_state: st.session_state.manual_total = {}
                                    if edits["Tổng giờ"] is None:
                                        st.session_state.manual_total.pop((ma_edit, ngay_edit), None)
                                    else:
                                        st.session_state.manual_total[(ma_edit, ngay_edit)] = float(edits["Tổng giờ"])

                # Áp dụng Giờ HC thủ công
                def apply_manual_hc(row):
                    ma = str(row[m['ma_nv']]).strip().upper()
                    if ma.endswith('.0'): ma = ma[:-2]
                    ngay = row["_parsed_date"].strftime('%d/%m/%Y')
                    if "manual_hc" in st.session_state and (ma, ngay) in st.session_state.manual_hc:
                        return float(st.session_state.manual_hc[(ma, ngay)])
                    return float(row["Giờ hành chính"])
                df_filtered["Giờ hành chính"] = df_filtered.apply(apply_manual_hc, axis=1)

                # Áp dụng OT thủ công
                def apply_manual_ot(row):
                    ma = str(row[m['ma_nv']]).strip().upper()
                    if ma.endswith('.0'): ma = ma[:-2]
                    ngay = row["_parsed_date"].strftime('%d/%m/%Y')
                    if "manual_ot" in st.session_state and (ma, ngay) in st.session_state.manual_ot:
                        return float(st.session_state.manual_ot[(ma, ngay)])
                    return row["Giờ OT"]
                df_filtered["Giờ OT"] = df_filtered.apply(apply_manual_ot, axis=1)

                # Tính lại Tổng giờ = Giờ HC + Giờ OT (Chỉ áp dụng nếu không phải lỗi -1, trừ khi có sửa thủ công)
                def apply_manual_total(row):
                    ma = str(row[m['ma_nv']]).strip().upper()
                    if ma.endswith('.0'): ma = ma[:-2]
                    ngay = row["_parsed_date"].strftime('%d/%m/%Y')
                    
                    if "manual_total" in st.session_state and (ma, ngay) in st.session_state.manual_total:
                        return float(st.session_state.manual_total[(ma, ngay)])

                    has_hc_override = "manual_hc" in st.session_state and (ma, ngay) in st.session_state.manual_hc
                    has_ot_override = "manual_ot" in st.session_state and (ma, ngay) in st.session_state.manual_ot
                    
                    if row["Số giờ làm thực tế"] == -1 and not (has_hc_override or has_ot_override):
                        return -1.0
                        
                    ot = float(row["Giờ OT"]) if pd.notna(row["Giờ OT"]) and str(row["Giờ OT"]).strip() != "" else 0.0
                    hc = float(row["Giờ hành chính"]) if pd.notna(row["Giờ hành chính"]) and str(row["Giờ hành chính"]).strip() != "" else 0.0
                    return hc + ot
                df_filtered["Số giờ làm thực tế"] = df_filtered.apply(apply_manual_total, axis=1)

                # Kiểm tra lỗi check-out TRƯỚC khi replace -1 → 0
                has_checkout_error = (df_filtered["Số giờ làm thực tế"] == -1).any()
                if has_checkout_error:
                    st.warning("⚠️ Phát hiện một số dòng có giờ ra sớm hơn giờ vào (lỗi check-out), đã được tính là 0 giờ.")

                def check_anomaly(row):
                    ma = str(row[m['ma_nv']]).strip().upper()
                    if ma.endswith('.0'): ma = ma[:-2]
                    ngay = row["_parsed_date"].strftime('%d/%m/%Y')
                    
                    if "manual_notes" in st.session_state and (ma, ngay) in st.session_state.manual_notes:
                        return st.session_state.manual_notes[(ma, ngay)]

                    vao = row[m['gio_vao']]
                    ra = row[m['gio_ra']]
                    is_wd = is_workday_func(row["_parsed_date"])
                    thu = row["_parsed_date"].weekday()  # 5=T7, 6=CN
                    has_ot_override = "manual_ot" in st.session_state and (ma, ngay) in st.session_state.manual_ot
                    has_hc_override = "manual_hc" in st.session_state and (ma, ngay) in st.session_state.manual_hc
                    has_leave = st.session_state.get("manual_leave", {}).get((ma, ngay), False)

                    try:
                        vao_trong = pd.isna(vao) or str(vao).strip().lower() in ['', 'nan', 'none', 'nat']
                    except (TypeError, ValueError):
                        vao_trong = False
                    try:
                        ra_trong = pd.isna(ra) or str(ra).strip().lower() in ['', 'nan', 'none', 'nat']
                    except (TypeError, ValueError):
                        ra_trong = False

                    notes = []
                    if has_hc_override:
                        notes.append("[HC thủ công]")
                    if has_ot_override:
                        notes.append("[OT thủ công]")

                    if has_leave:
                        d_check = row["_parsed_date"].date() if hasattr(row["_parsed_date"], 'date') else row["_parsed_date"]
                        if is_wd and vao_trong and ra_trong and is_last_saturday_of_month(d_check):
                            notes.append("🔴 Vắng Thứ 7 bắt buộc")
                        return " | ".join(notes) if notes else ""

                    # Cuối tuần không có dữ liệu → không cần ghi chú, trả về lười OT thủ công nếu có
                    if not is_wd and vao_trong and ra_trong:
                        return " | ".join(notes) if notes else ""

                    if vao_trong and not ra_trong: notes.append("⚠️ Thiếu giờ vào")
                    elif ra_trong and not vao_trong: notes.append("⚠️ Thiếu giờ ra")
                    elif vao_trong and ra_trong:
                        if is_wd:
                            # Kiểm tra có phải Thứ 7 bắt buộc không
                            d_check = row["_parsed_date"].date() if hasattr(row["_parsed_date"], 'date') else row["_parsed_date"]
                            if is_last_saturday_of_month(d_check):
                                notes.append("🔴 Vắng Thứ 7 bắt buộc")
                            else:
                                notes.append("🔴 Nghỉ không phép")
                    elif row["Số giờ làm thực tế"] == -1: notes.append("🟣 Lỗi check-out")
                    elif 0 < float(row["Số giờ làm thực tế"]) < 4: notes.append("🟠 Làm thiếu giờ (< 4h)")

                    if has_leave and not (vao_trong and ra_trong):
                        notes.append("🟢 Nghỉ có phép")

                    if row.get("_is_chieu", False):
                        notes.append("🔵 Làm ca chiều")

                    return " | ".join(notes)
                df_filtered["Ghi chú"] = df_filtered.apply(check_anomaly, axis=1)
                df_filtered["Số giờ làm thực tế"] = df_filtered["Số giờ làm thực tế"].replace(-1, 0)
                df_filtered["Số giờ làm thực tế"] = df_filtered["Số giờ làm thực tế"].apply(format_gio_lam)
                df_filtered["Giờ hành chính"] = df_filtered["Giờ hành chính"].apply(format_gio_lam)
                df_filtered["Giờ OT"] = df_filtered["Giờ OT"].apply(format_gio_lam)

                def get_has_leave(row):
                    ma = str(row[m['ma_nv']]).strip().upper()
                    if ma.endswith('.0'): ma = ma[:-2]
                    ngay = row["_parsed_date"].strftime('%d/%m/%Y')
                    return st.session_state.get("manual_leave", {}).get((ma, ngay), False)

                df_filtered["Có phép"] = df_filtered.apply(get_has_leave, axis=1)

                def get_ot_reason(row):
                    ma = str(row[m['ma_nv']]).strip().upper()
                    if ma.endswith('.0'): ma = ma[:-2]
                    ngay = row["_parsed_date"].strftime('%d/%m/%Y')
                    return st.session_state.get("manual_ot_reason", {}).get((ma, ngay), "")

                df_filtered["Lý do tăng ca"] = df_filtered.apply(get_ot_reason, axis=1)

                with st.spinner("⏳ Đang tổng hợp kết quả..."):
                    total_rows = len(df_filtered)
                    total_emps = df_filtered[m['ma_nv']].nunique()
                    df_numeric = df_filtered[pd.to_numeric(df_filtered['Số giờ làm thực tế'], errors='coerce').notnull()]
                    total_hours = df_numeric['Số giờ làm thực tế'].sum() if not df_numeric.empty else 0
                    ngay_nghi = int((df_filtered["Số giờ làm thực tế"] == 0).sum())

                    st.markdown('<div class="card"><div class="card-title"><span class="card-icon">📈</span>Tổng quan kỳ công</div></div>', unsafe_allow_html=True)
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Số ngày làm việc", working_days, help="Tổng số ngày: trừ CN, Thứ 7 (trừ Thứ 7 cuối tháng), trừ ngày lễ trùng ngày làm việc.")
                    c2.metric("Số nhân viên", total_emps)
                    c3.metric("Tổng giờ làm", f"{format_gio_lam(total_hours)} giờ")
                    c4.metric("Số ngày nghỉ", ngay_nghi)

                    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

                    chuc_vu_vals = df_filtered[m['chuc_vu']].values if 'chuc_vu' in m and m['chuc_vu'] in df_filtered else [""] * len(df_filtered)
                    phong_ban_vals = df_filtered[m['phong_ban']].values if 'phong_ban' in m and m['phong_ban'] in df_filtered else [""] * len(df_filtered)

                    df_result_ui = pd.DataFrame({
                        "STT": range(1, len(df_filtered) + 1),
                        "Mã NV": df_filtered[m['ma_nv']].values,
                        "Tên nhân viên": df_filtered[m['ten_nv']].values,
                        "Chức vụ": chuc_vu_vals,
                        "Phòng ban": phong_ban_vals,
                        "Thứ": df_filtered["_parsed_date"].dt.weekday.map({0:'Hai',1:'Ba',2:'Tư',3:'Năm',4:'Sáu',5:'Bảy',6:'CN'}).values,
                        "Ngày": df_filtered["Ngày"].values,
                        "Giờ làm thực tế": df_filtered["Giờ hành chính"].values,
                        "OT": df_filtered["Giờ OT"].values,
                        "Tổng giờ": df_filtered["Số giờ làm thực tế"].values,
                        "Lý do tăng ca": df_filtered["Lý do tăng ca"].values,
                        "Có phép": df_filtered["Có phép"].values,
                        "Ghi chú": df_filtered["Ghi chú"].values
                    })
                    st.session_state['df_result'] = df_result_ui
                    st.session_state['df_filtered_for_chat'] = df_filtered.copy()

                    def get_loai_ngay(row):
                        ngay_str = row["Ngày"]
                        gio_vao_raw = row[m['gio_vao']]
                        gio_ra_raw = row[m['gio_ra']]
                        ma = str(row[m['ma_nv']]).strip().upper()
                        if ma.endswith('.0'): ma = ma[:-2]
                        has_leave = st.session_state.get("manual_leave", {}).get((ma, ngay_str), False)
                        try:
                            ngay = datetime.datetime.strptime(ngay_str, '%d/%m/%Y').date()
                        except:
                            return 'binh_thuong'
                        thu = ngay.weekday()  # 5=T7, 6=CN
                        is_wd = is_workday_func(ngay)
                        vao_trong = pd.isna(gio_vao_raw) or str(gio_vao_raw).strip().lower() in ['', 'nan', 'none', 'nat']
                        ra_trong = pd.isna(gio_ra_raw) or str(gio_ra_raw).strip().lower() in ['', 'nan', 'none', 'nat']

                        if not is_wd and vao_trong and ra_trong:
                            return 'cuoi_tuan'   # T7 / CN không có dữ liệu

                        if is_wd and vao_trong and ra_trong:
                            if has_leave:
                                return 'nghi_co_phep'
                            return 'nghi_khong_phep'

                        return 'binh_thuong'

                    df_result_ui["_loai"] = df_filtered.apply(get_loai_ngay, axis=1).values

                    def style_row(row):
                        loai = df_result_ui.loc[row.name, "_loai"]
                        ngay_str = df_result_ui.loc[row.name, "Ngày"]
                        try:
                            d_obj_row = datetime.datetime.strptime(ngay_str, '%d/%m/%Y').date()
                            thu = d_obj_row.weekday()
                            is_weekend = (thu in [5, 6]) and not (thu == 5 and is_last_saturday_of_month(d_obj_row))
                        except:
                            is_weekend = False

                        styles = [""] * len(row)
                        if is_weekend:
                            styles = ["background-color: #F1F5F9"] * len(row)
                            
                        idx_gio = list(row.index).index("Tổng giờ")
                        idx_ghi_chu = list(row.index).index("Ghi chú")
                        
                        if loai == 'cuoi_tuan':
                            styles[idx_gio] = "background-color: #F1F5F9; color: #CBD5E1"
                        elif loai == 'nghi_khong_phep':
                            styles = ["background-color: #FEE2E2; color: #991B1B"] * len(row)
                            styles[idx_gio] = "background-color: #FEE2E2; color: #991B1B; font-weight: 600"
                        elif loai == 'nghi_co_phep':
                            styles[idx_gio] = "background-color: #F1F5F9; color: #2563EB; font-weight: 600" if is_weekend else "color: #2563EB; font-weight: 600"
                        else:
                            styles[idx_gio] = "background-color: #F1F5F9; color: #2563EB; font-weight: 600" if is_weekend else "color: #2563EB; font-weight: 600"
                            
                        val_str = str(row["Ghi chú"])
                        if "Nghỉ không phép" in val_str or "Vắng Thứ 7 bắt buộc" in val_str:
                            styles[idx_ghi_chu] = 'background-color: #FEE2E2; color: #991B1B; font-weight: 600'
                        elif "Lỗi check-out" in val_str:
                            styles[idx_ghi_chu] = 'background-color: #F3E8FF; color: #6B21A8; font-weight: 500'
                        elif "Thiếu giờ vào" in val_str or "Thiếu giờ ra" in val_str or "Làm thiếu giờ" in val_str:
                            styles[idx_ghi_chu] = 'background-color: #FEF3C7; color: #92400E; font-weight: 500'
                        elif "OT thủ công" in val_str:
                            styles[idx_ghi_chu] = 'background-color: #E0F2FE; color: #0369A1; font-weight: 500'
                        elif "ca chiều" in val_str:
                            styles[idx_ghi_chu] = 'background-color: #F8FAFC; color: #475569; font-weight: 500'

                        return styles

                    st.markdown('<div class="card"><div class="card-title"><span class="card-icon">📋</span>Kết quả chi tiết</div></div>', unsafe_allow_html=True)
                    df_display = df_result_ui.drop(columns=["_loai"])

                    tab_table, tab_chart = st.tabs(["📑 Bảng Dữ liệu", "📈 Biểu đồ"])
                    with tab_table:
                        st.markdown("💡 *Mẹo: Khi bạn tích chọn, sửa OT hoặc sửa Giờ làm thực tế, bấm 'Lưu thay đổi bảng' để tính toán lại Tổng giờ.*")
                        
                        with st.form("chamcong_editor_form"):
                            st.data_editor(
                                        df_display.style.apply(style_row, axis=1),
                                        use_container_width=True, hide_index=True,
                                        height=min(600, 40 + len(df_result_ui) * 35),
                                        key="data_editor_ot",
                                        column_config={
                                            "STT": st.column_config.NumberColumn("STT", width="small", format="%d", disabled=True),
                                            "Mã NV": st.column_config.TextColumn("Mã NV", width="small", disabled=True),
                                            "Tên nhân viên": st.column_config.TextColumn("Tên nhân viên", width="medium", disabled=True),
                                            "Chức vụ": st.column_config.TextColumn("Chức vụ", width="small", disabled=True),
                                            "Phòng ban": st.column_config.TextColumn("Phòng ban", width="small", disabled=True),
                                            "Thứ": st.column_config.TextColumn("Thứ", width="small", disabled=True),
                                            "Ngày": st.column_config.TextColumn("Ngày", width="small", disabled=True),
                                            "Giờ làm thực tế": st.column_config.NumberColumn("Giờ làm thực tế", width="small", format="%g", disabled=False),
                                            "OT": st.column_config.NumberColumn("OT", width="small", format="%g", disabled=False),
                                            "Tổng giờ": st.column_config.NumberColumn("Tổng giờ", width="small", format="%g", disabled=False),
                                            "Lý do tăng ca": st.column_config.TextColumn("Lý do tăng ca", width="medium", disabled=False),
                                            "Có phép": st.column_config.CheckboxColumn("Có phép", width="small", disabled=False),
                                            "Ghi chú": st.column_config.TextColumn("Ghi chú", width="medium", disabled=False),
                                        }
                            )
                            submit_edits = st.form_submit_button("💾 Lưu thay đổi bảng", type="primary")
                            if submit_edits:
                                st.rerun()

                    with tab_chart:
                        st.markdown("### 📊 Tổng quan hiệu suất Chấm công")
                        try:
                            import plotly.express as px
                            # Ép kiểu numeric trước khi groupby để tránh lỗi sum() trên object
                            df_chart = df_filtered.copy()
                            for _col in ['Phút đi trễ', 'Phút về sớm', 'Giờ OT', 'Số giờ làm thực tế']:
                                df_chart[_col] = pd.to_numeric(df_chart[_col], errors='coerce').fillna(0)
                            df_nv = df_chart.groupby([m['ma_nv'], m['ten_nv']]).agg(
                                Tong_Tre=('Phút đi trễ', 'sum'), Tong_Som=('Phút về sớm', 'sum'),
                                Tong_OT=('Giờ OT', 'sum'), Tong_Gio=('Số giờ làm thực tế', 'sum')
                            ).reset_index()
                            df_ot = df_nv[df_nv['Tong_OT'] > 0].sort_values('Tong_OT', ascending=False).head(10)

                            if not df_ot.empty:
                                fig2 = px.bar(df_ot, x='Tong_OT', y=m['ten_nv'], orientation='h', title="🟢 Top những người tăng ca nhiều nhất", color='Tong_OT', color_continuous_scale='Greens', labels={'Tong_OT': 'Tổng giờ OT', m['ten_nv']: 'Nhân viên'})
                                fig2.update_layout(yaxis={'categoryorder':'total ascending'}, margin=dict(l=0,r=0,t=40,b=0))
                                st.plotly_chart(fig2, use_container_width=True)
                            else:
                                st.info("ℹ️ Chưa có nhân viên nào tăng ca trong kỳ này.")
                            st.markdown("<hr>", unsafe_allow_html=True)

                            df_nghi = df_result_ui[df_result_ui['_loai'] == 'nghi'].groupby(['Mã NV', 'Tên nhân viên']).size().reset_index(name='So_Ngay_Nghi')
                            df_nghi = df_nghi.sort_values('So_Ngay_Nghi', ascending=False).head(10)
                            if not df_nghi.empty:
                                fig_nghi = px.bar(df_nghi, x='So_Ngay_Nghi', y='Tên nhân viên', orientation='h', title="🔴 Top những người nghỉ nhiều nhất", color='So_Ngay_Nghi', color_continuous_scale='Reds', labels={'So_Ngay_Nghi': 'Số ngày nghỉ', 'Tên nhân viên': 'Nhân viên'})
                                fig_nghi.update_layout(yaxis={'categoryorder':'total ascending'}, margin=dict(l=0,r=0,t=40,b=0))
                                st.plotly_chart(fig_nghi, use_container_width=True)
                            else:
                                st.info("ℹ️ Không có nhân viên nào nghỉ trong kỳ này.")
                            st.markdown("<hr>", unsafe_allow_html=True)

                            st.markdown("### 📅 Xu hướng đi làm theo ngày")
                            df_ngay = df_chart.groupby("Ngày").agg(Tong_OT=('Giờ OT', 'sum')).reset_index()
                            df_ngay['Ngày_dt'] = pd.to_datetime(df_ngay['Ngày'], format='%d/%m/%Y')
                            df_ngay = df_ngay.sort_values('Ngày_dt')
                            fig3 = px.line(df_ngay, x='Ngày', y=['Tong_OT'], title="Xu hướng Tăng ca theo ngày", markers=True, labels={'value': 'Số lượng', 'variable': 'Chỉ số'})
                            fig3.for_each_trace(lambda t: t.update(name="Tổng giờ OT"))
                            fig3.update_layout(legend_title_text='')
                            st.plotly_chart(fig3, use_container_width=True)
                        except ImportError:
                            st.warning("⚠️ Đang cài đặt thư viện `plotly`... Vui lòng Refresh lại trang sau 10 giây.")

                    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
                    st.markdown("## ⬇️ Xuất file & Lưu trữ")
                    col_exp1, col_exp2, col_exp3 = st.columns([2, 1, 1])
                    with col_exp1:
                        st.markdown(f"""<div style='background:#F0FDF8;border:0.5px solid #6EE7C0;border-radius:10px;padding:12px 16px;font-size:13px;color:#0F6E56'>
    📄 File: <b>chitiet_chamcong_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx</b><br>
    <span style='color:#6B7280;font-size:12px'>Gồm {total_rows} dòng · {total_emps} nhân viên</span>
    </div>""", unsafe_allow_html=True)
                    with col_exp2:
                        if st.button("💾 Lưu vào Hệ thống", use_container_width=True):
                            with st.spinner("Đang lưu dữ liệu..."):
                                save_to_db(df_filtered, m)
                            st.success("Đã lưu thành công!")
                    with col_exp3:
                        file_name_export = f"chitiet_chamcong_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx"
                        try:
                            total_wd_tuple = calculate_working_days(start_date, end_date, st.session_state.custom_holidays, st.session_state.custom_workdays)
                            total_wd = total_wd_tuple[0]
                            excel_data = export_excel_tong_hop(df_filtered, m, start_date, end_date, total_wd)
                        except Exception as e:
                            st.error(f"❌ Lỗi xuất file: {e}")
                        if excel_data is not None:
                            st.download_button(
                                label="⬇️ Tải Excel",
                                data=excel_data,
                                file_name=file_name_export,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary", use_container_width=True,
                            )
                        else:
                            st.button("⬇️ Tải Excel", disabled=True, use_container_width=True, help="Không có dữ liệu để xuất")

    # ==========================================

            render_interactive_dashboard(df_filtered)

# Gọi chatbot ở cuối luồng (dành cho chamcong)
render_chatbot()
