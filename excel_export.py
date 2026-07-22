import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import streamlit as st
from log_config import logger
from utils import *
from translations import get_t, get_data_t

def fallback_xml_parser(file_obj):
    file_obj.seek(0)
    import xml.etree.ElementTree as ET
    try:
        tree = ET.parse(file_obj)
        root = tree.getroot()
        
        for elem in root.iter():
            if '}' in elem.tag:
                elem.tag = elem.tag.split('}', 1)[1]
                
        rows = root.findall('.//Row')
        if rows:
            data = []
            max_cols = 0
            for row in rows:
                row_data = {}
                current_idx = 0
                for cell in row.findall('.//Cell'):
                    # Check for Index attribute (handling namespaces like {urn:...}Index)
                    idx_attr = None
                    for k, v in cell.attrib.items():
                        if k.endswith('Index'):
                            idx_attr = v
                            break
                            
                    if idx_attr is not None:
                        current_idx = int(idx_attr) - 1
                        
                    data_elem = cell.find('Data')
                    if data_elem is not None:
                        row_data[current_idx] = data_elem.text
                    else:
                        row_data[current_idx] = None
                    current_idx += 1
                    
                if row_data:
                    max_cols = max(max_cols, max(row_data.keys()) + 1) if row_data.keys() else max_cols
                    data.append(row_data)
            
            # Convert dicts to list based on max_cols
            final_data = []
            for rd in data:
                row_list = [rd.get(i, None) for i in range(max_cols)]
                final_data.append(row_list)
                
            return pd.DataFrame(final_data)
        else:
            data = []
            for child in root:
                row_data = {}
                for subchild in child:
                    row_data[subchild.tag] = subchild.text
                if row_data:
                    data.append(row_data)
            return pd.DataFrame(data)
    except Exception:
        pass
        
    file_obj.seek(0)
    try:
        df = pd.read_xml(file_obj, parser='etree')
        return df
    except Exception:
        pass

    # Maybe it's HTML saved as .xls?
    file_obj.seek(0)
    try:
        dfs = pd.read_html(file_obj.read().decode('utf-8', errors='ignore'))
        if dfs:
            return dfs[0]
    except Exception:
        pass
        
    return None

def find_header_row(file, file_name):
    file.seek(0)
    if file_name.endswith(('.csv', '.txt', '.dat', '.tsv')):
        df_raw = pd.read_csv(file, header=None, nrows=30, sep=None, engine='python')
    elif file_name.endswith('.xml'):
        df_raw = fallback_xml_parser(file)
        if df_raw is not None:
            df_raw = df_raw.head(30)
        else:
            return 0
    else:
        try:
            df_raw = pd.read_excel(file, header=None, nrows=30)
        except Exception:
            return 0
            
    if df_raw is None: return 0
    for i, row in df_raw.iterrows():
        row_str = [str(v).strip().lower() for v in row.values if pd.notna(v) and str(v).lower() != 'nan']
        keywords = ['mã', 'tên', 'ngày', 'vào', 'ra', 'time', 'date', 'name', 'id']
        if sum(1 for kw in keywords if any(kw in cell for cell in row_str)) >= 3:
            return i
    return 0


def parse_excel_file(uploaded_file):
    file_name = uploaded_file.name.lower()
    try:
        header_row = find_header_row(uploaded_file, file_name)
    except Exception as e:
        st.error(f"❌ Lỗi khi đọc file: {e}")
        st.stop()
        
    uploaded_file.seek(0)
    try:
        if file_name.endswith(('.csv', '.txt', '.dat', '.tsv')):
            df = pd.read_csv(uploaded_file, header=header_row, sep=None, engine='python')
        elif file_name.endswith('.xml'):
            df = fallback_xml_parser(uploaded_file)
            if df is None:
                raise Exception("Không thể parse file XML.")
            if header_row > 0:
                df.columns = df.iloc[header_row].values
                df = df.iloc[header_row+1:].reset_index(drop=True)
            elif header_row == 0:
                df.columns = df.iloc[0].values
                df = df.iloc[1:].reset_index(drop=True)
        else:
            try:
                df = pd.read_excel(uploaded_file, header=header_row)
            except Exception as orig_e:
                # Có thể là file XML/HTML giả dạng .xls
                df = fallback_xml_parser(uploaded_file)
                if df is None:
                    raise orig_e
                if header_row > 0:
                    df.columns = df.iloc[header_row].values
                    df = df.iloc[header_row+1:].reset_index(drop=True)
                elif header_row == 0:
                    df.columns = df.iloc[0].values
                    df = df.iloc[1:].reset_index(drop=True)
    except Exception as e:
        st.error(f"❌ Lỗi khi đọc bảng dữ liệu: {e}")
        st.stop()
        
    if df is not None:
        df = df.dropna(how='all')
        df.columns = df.columns.astype(str).str.strip()
    return df


def export_excel_tong_hop(df_filtered, mapping, start_date, end_date, total_wd):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    import pandas as pd
    import streamlit as st
    
    t = st.session_state.get('cached_t', None)
    if t is None:
        from translations import get_t, get_data_t, translate_name
        t = get_t(st.session_state.lang)
        data_t = get_data_t(st.session_state.lang)
        t_name = lambda name: translate_name(name, st.session_state.lang)
    else:
        from translations import get_t, get_data_t, translate_name
        t = get_t(st.session_state.lang)
        data_t = get_data_t(st.session_state.lang)
        t_name = lambda name: translate_name(name, st.session_state.lang)

    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True
    ws = wb.active
    ws.title = "Chi tiết chấm công" if st.session_state.get('lang', 'vi') == 'vi' else "勤怠詳細"
    
    font_bold14 = Font(name='Times New Roman', size=14, bold=True)
    font_bold12 = Font(name='Times New Roman', size=12, bold=True)
    font_normal = Font(name='Times New Roman', size=12)
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    border_thick = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    
    ws.merge_cells('A1:O1')
    ws['A1'] = "CHI TIẾT CHẤM CÔNG" if st.session_state.get('lang', 'vi') == 'vi' else "勤怠詳細"
    ws['A1'].font = font_bold14
    ws['A1'].alignment = align_center
    
    ws.merge_cells('A2:N2')
    ws['A2'] = t("export_date_range", start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y'))
    ws['A2'].font = font_bold12
    ws['A2'].alignment = align_center
    
    ws.merge_cells('A4:H4')
    ws['A4'] = t("export_total_days", total_wd)
    ws['A4'].font = font_bold12
    ws['A4'].alignment = align_left
    
    total_hours = total_wd * 8
    ws.merge_cells('A5:H5')
    ws['A5'] = t("export_total_hours", total_hours)
    ws['A5'].font = font_bold12
    ws['A5'].alignment = align_left
    
    total_off_days = (end_date - start_date).days + 1 - total_wd
    ws.merge_cells('A6:H6')
    ws['A6'] = t("export_total_off", total_off_days)
    ws['A6'].font = font_bold12
    ws['A6'].alignment = align_left
    
    headers = [
        t("export_col_emp_code"), t("export_col_emp_name"), t("emp_dept"), t("emp_position"),
        t("export_col_date"), t("export_col_weekday"), t("export_col_in"), t("export_col_out"),
        t("export_col_work_day"), t("export_col_actual_hours"), t("export_col_ot"),
        t("export_col_total_hours"), t("export_col_ot_reason"), t("export_col_note")
    ]
    header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        c = ws.cell(row=8, column=col_num, value=header)
        c.font = font_bold12
        c.alignment = align_center
        c.fill = header_fill
        c.border = border
        
    row_idx = 9
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    def get_val(row, key):
        col = mapping.get(key)
        if col and col in row:
            v = row[col]
            return str(v) if pd.notna(v) else ""
        return ""
        
    for _, row in df_filtered.iterrows():
        try: d_obj = row['_parsed_date'].date()
        except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); d_obj = pd.to_datetime(row['_parsed_date']).date()
        
        thu = d_obj.weekday()
        if st.session_state.lang == 'ja':
            thu_str = ["月", "火", "水", "木", "金", "土", "日"][thu]
        else:
            thu_str = ["Hai", "Ba", "Tư", "Năm", "Sáu", "Bảy", "CN"][thu]
        
        ws.cell(row=row_idx, column=1, value=get_val(row, 'ma_nv'))
        ws.cell(row=row_idx, column=2, value=t_name(get_val(row, 'ten_nv')))
        ws.cell(row=row_idx, column=3, value=data_t(get_val(row, 'phong_ban')))
        ws.cell(row=row_idx, column=4, value=data_t(get_val(row, 'chuc_vu')))
        ws.cell(row=row_idx, column=5, value=d_obj.strftime('%d/%m/%Y'))
        ws.cell(row=row_idx, column=6, value=thu_str)
        ws.cell(row=row_idx, column=7, value=get_val(row, 'gio_vao'))
        ws.cell(row=row_idx, column=8, value=get_val(row, 'gio_ra'))
        
        m_ma = get_val(row, 'ma_nv').strip().upper()
        if m_ma.endswith('.0'): m_ma = m_ma[:-2]
        ngay_str = d_obj.strftime('%d/%m/%Y')
        
        hc = float(row.get('Giờ hành chính', 0)) if pd.notna(row.get('Giờ hành chính')) else 0.0
        if "manual_hc" in st.session_state and (m_ma, ngay_str) in st.session_state.manual_hc:
            hc = float(st.session_state.manual_hc[(m_ma, ngay_str)])
            
        ws.cell(row=row_idx, column=9, value=round(hc/8 + 1e-9, 2) if hc > 0 else "")
        ws.cell(row=row_idx, column=10, value=hc if hc > 0 else "")
        
        ot = float(row.get('Giờ OT', 0)) if pd.notna(row.get('Giờ OT')) else 0.0
        if "manual_ot" in st.session_state and (m_ma, ngay_str) in st.session_state.manual_ot:
            ot = float(st.session_state.manual_ot[(m_ma, ngay_str)])
            
        ws.cell(row=row_idx, column=11, value=ot if ot > 0 else "")
        
        total = hc + ot
        if "manual_total" in st.session_state and (m_ma, ngay_str) in st.session_state.manual_total:
            total = float(st.session_state.manual_total[(m_ma, ngay_str)])
            
        ws.cell(row=row_idx, column=12, value=total if total > 0 else "")
        
        ws.cell(row=row_idx, column=13, value=str(row.get('Lý do tăng ca', '')) if pd.notna(row.get('Lý do tăng ca')) else "")
        
        ghi_chu = str(row.get('Ghi chú', '')) if pd.notna(row.get('Ghi chú')) else ""
        loai = row.get('_loai', '') if pd.notna(row.get('_loai', '')) else ''
        
        ghi_chu_clean = ghi_chu
        for icon in ["🚫 ", "✅ ", "⚠️ ", "📝 ", "📅 ", "🔴 ", "🟣 ", "🟠 ", "🟢 ", "🔵 "]:
            ghi_chu_clean = ghi_chu_clean.replace(icon, "")
        ghi_chu_clean = ghi_chu_clean.strip()
        
        is_abnormal = False
        if "nghi_khong_phep" in loai or "Thiếu giờ" in ghi_chu_clean or "Lỗi" in ghi_chu_clean or "Vắng" in ghi_chu_clean or "Làm thiếu" in ghi_chu_clean:
            is_abnormal = True

        ws.cell(row=row_idx, column=14, value=ghi_chu_clean.strip())
        
        font_red = Font(name='Times New Roman', size=12, color="FF0000")
        for col_num in range(1, 15):
            c = ws.cell(row=row_idx, column=col_num)
            c.font = font_normal
            if col_num == 14 and is_abnormal:
                c.font = font_red
            c.border = border
            if col_num in [1,2,3,4, 13, 14]:
                c.alignment = align_left
            else:
                c.alignment = align_center
            
            if thu in [5, 6] and not (thu == 5 and is_last_saturday_of_month(d_obj)):
                c.fill = gray_fill
                
        row_idx += 1
        
    for col_num, width in enumerate([10, 20, 15, 15, 12, 8, 10, 10, 8, 15, 10, 15, 20, 30], 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


