import streamlit as st
from theme import get_theme
from log_config import logger
import pandas as pd
import re
import io
import base64
from excel_export import *
from translations import get_t, translate_name
from db import get_company_emp_options, get_company_emp_dict
from ai_chat import load_saved_api_key, smart_offline_summarize, batch_summarize_projects

def extract_ma_nv_from_filename(filename: str) -> str:
    """Lấy mã NV từ tên file: Report_VM011_ロン_2026_5.xlsx → VM011"""
    match = re.search(r'(VM\d+)', filename)
    return match.group(1) if match else filename


def extract_ten_nv_from_filename(filename: str) -> str:
    """Lấy tên Nhật từ file để hiển thị: Report_VM011_ロン_2026_5 → ロン"""
    name = re.sub(r'\.xlsx?$', '', filename)
    # Lấy phần chữ (ko phải số/gạch dưới) ngay sau cụm VMxxx
    match = re.search(r'VM\d+[\s_]+([^\d_]+)', name)
    if match:
        return match.group(1).strip()
    return filename


def robust_decode(payload_bytes, charset=None):
    if not payload_bytes: return ""
    if charset:
        try: return payload_bytes.decode(charset)
        except Exception: pass
    for enc in ['utf-8', 'shift_jis', 'iso-2022-jp', 'euc-jp', 'cp932', 'windows-1258']:
        try: return payload_bytes.decode(enc)
        except Exception: pass
    try:
        import charset_normalizer
        return str(charset_normalizer.from_bytes(payload_bytes).best())
    except Exception: pass
    return payload_bytes.decode('utf-8', errors='replace')


def parse_mos_file(file, filename: str) -> pd.DataFrame:
    """
    Đọc 1 file Report, lấy chỉ phần MOS業務.
    Trả về DataFrame với các cột đã chuẩn hóa.
    """
    import openpyxl
    try:
        if hasattr(file, 'seek'):
            file.seek(0)
    except Exception:
        pass
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        try:
            wb.close()
        except Exception:
            pass
    
    # Tìm header row (dòng chứa tiêu đề cột như 案件番号, Mã dự án, Code, Project...)
    header_row_idx = None
    for i, row in enumerate(rows):
        if not row: continue
        row_strs_clean = [str(v).strip().lower().replace('　', '').replace(' ', '').replace('\n', '').replace('\r', '') for v in row if v is not None and str(v).strip() != '']
        row_combined = " ".join(row_strs_clean)
        if any(k in row_combined for k in ['案件番号', '案件名', 'mãdựán', 'mãda', 'projectcode', 'projectname', 'projectno']) or (any(k in row_combined for k in ['code', 'mã', '番号', '案件']) and any(k in row_combined for k in ['name', 'tên', 'task', 'côngviệc', 'tanto', '담당', '담당자', 'nhật'])):
            header_row_idx = i
            break
    if header_row_idx is None:
        # Fallback: tìm dòng đầu tiên có ít nhất 4 ô có dữ liệu
        for i, row in enumerate(rows[:15]):
            if row and len([v for v in row if v is not None and str(v).strip() != '']) >= 4:
                header_row_idx = i
                break
        if header_row_idx is None:
            raise ValueError(f"Không tìm thấy dòng tiêu đề chuẩn (案件番号/Mã dự án) trong file {filename}")
    
    header = rows[header_row_idx]
    
    # Tự động xác định năm và tháng gốc của sheet/file (thay vì mặc định tháng 6/năm 2026)
    default_year = 2026
    default_month = 6 # fallback
    
    header_context = " ".join([filename] + [str(c) for r in rows[:header_row_idx+1] if r for c in r if c])
    m_yr = re.search(r'(20\d\d)', header_context)
    if m_yr:
        default_year = int(m_yr.group(1))
        
    m_mo_jp = re.search(r'(?:20\d\d[\s年._/-]*)?([1-9]|0[1-9]|1[0-2])\s*月度?', header_context)
    if m_mo_jp:
        default_month = int(m_mo_jp.group(1))
    else:
        m_mo_vn = re.search(r'(?:tháng|month|th|t|_)[\s._/-]*([1-9]|0[1-9]|1[0-2])\b', header_context, re.IGNORECASE)
        if m_mo_vn:
            default_month = int(m_mo_vn.group(1))
        else:
            m_mo_iso = re.search(r'(?:20\d\d)[-_/._\s]+(0[1-9]|1[0-2])\b|\b(0[1-9]|1[0-2])[-_/._\s]+(?:20\d\d)', filename)
            if m_mo_iso:
                default_month = int(m_mo_iso.group(1) or m_mo_iso.group(2))
    
    col_map = {}
    date_cols = {}
    for ci, col in enumerate(header):
        if col is None: continue
        if isinstance(col, (datetime.date, datetime.datetime)):
            date_cols[f'{col.month}/{col.day}'] = ci
            continue
        col_str = str(col).strip()
        col_str_low = col_str.lower().replace(' ', '').replace('　', '')
        
        # Nhận diện X月Y日 (ví dụ 6月1日)
        m_jp = re.search(r'(\d+)\s*月\s*(\d+)\s*日?', col_str)
        if m_jp:
            date_cols[f"{int(m_jp.group(1))}/{int(m_jp.group(2))}"] = ci
            continue
            
        if any(k in col_str_low for k in ['nhật', '담당', '담당자', 'tanto', 'thựchiện', '担当']): col_map['tanto'] = ci
        elif any(k in col_str_low for k in ['mã', 'number', 'code', '案件番号', 'mãdựán']): col_map['ma_da'] = ci
        elif any(k in col_str_low for k in ['khách', 'client', 'customer', 'お客様', 'kháchhàng']): col_map['khach'] = ci
        elif any(k in col_str_low for k in ['tên', 'name', 'project', '案件名', 'têndựán']): col_map['ten_da'] = ci
        elif any(k in col_str_low for k in ['phânvùng', 'phầnvùng', 'quảnlý', 'bộphận', 'khuvực', 'category', 'type', '区分', 'phânloại']): col_map['phan_vung'] = ci
        elif any(k in col_str_low for k in ['task', 'côngviệc', 'nội\ndung', 'nhiệmvụ', 'tácvụ', 'タスク']): col_map['task'] = ci
        elif any(k in col_str_low for k in ['tổng', 'total', 'sum', 'cộng', '合計', '工数', 'totalhours', 'tổnggiờ', 'tổngcộng', 'tổngsố']): col_map['tong'] = ci
        elif '/' in col_str or '-' in col_str:
            try:
                clean_d = col_str.split(' ')[0]
                sep = '/' if '/' in clean_d else '-'
                parts = clean_d.split(sep)
                if len(parts) >= 2:
                    p1, p2 = int(parts[0]), int(parts[1])
                    if p2 == default_month or (1 <= p2 <= 12 and p1 > 12):
                        month, day = p2, p1
                    elif p1 == default_month or (1 <= p1 <= 12 and p2 > 12):
                        month, day = p1, p2
                    else:
                        month, day = p2, p1 # default dd/mm
                    date_cols[f'{month}/{day}'] = ci
                    continue
            except Exception: pass
            
        # Nhận diện X日 hoặc số ngày từ 1 đến 31 (khi cột ở vị trí từ cột thứ 6 trở đi)
        m_day = re.match(r'^(\d{1,2})\s*日?$', col_str)
        if m_day and ci >= 6:
            day_num = int(m_day.group(1))
            if 1 <= day_num <= 31:
                date_cols[f'{default_month}/{day_num}'] = ci
    
    def parse_float_safe(val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(',', '.')
        s = re.sub(r'[^\d.]', '', s)
        try:
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    records = []
    in_mos_section = True
    curr_section_type = 'MOS' # Khởi tạo mặc định là MOS
    file_daily_work_hours = {} # date_str -> tổng giờ làm trong MOS, JMOS, VMOS
    
    # Biến lưu trữ thông tin dự án hiện tại để kế thừa cho các dòng merged cell (dòng trống mã dự án)
    curr_ma_da = ''
    curr_khach = ''
    curr_ten_da = ''
    curr_phan_vung = ''
    curr_tanto = ''
    
    for row in rows[header_row_idx + 1:]:
        if row is None: continue
        
        import unicodedata
        row_cells_clean = [unicodedata.normalize('NFKC', str(val)).strip().upper() for val in row if val is not None and str(val).strip() != '']
        if not row_cells_clean:
            continue
            
        ma_da_raw = str(row[col_map.get('ma_da', 1)] or '').strip()
        ma_da = unicodedata.normalize('NFKC', ma_da_raw).strip().upper() if ma_da_raw else ''
        ma_da = re.sub(r'[\u200b\ufeff\s]+', '', ma_da)
        
        row_str_all = " ".join(row_cells_clean)
        
        # Nếu dòng hoàn toàn không có dữ liệu (trừ một số cột không quan trọng), ngắt kế thừa dự án
        tanto_val_tmp = str(row[col_map.get('tanto', 0)] or '').strip()
        khach_val_tmp = str(row[col_map.get('khach', 2)] or '').strip()
        ten_da_val_tmp = str(row[col_map.get('ten_da', 3)] or '').strip()
        phan_vung_val_tmp = str(row[col_map.get('phan_vung', 4)] or '').strip()
        task_tmp = str(row[col_map.get('task', 5)] or '').strip()
        tong_val_tmp = row[col_map.get('tong', 6)] if 'tong' in col_map else None
        tong_tmp = parse_float_safe(tong_val_tmp) if 'tong' in col_map else 0.0
        
        daily_hours_tmp = 0.0
        for date_str, ci in date_cols.items():
            if ci < len(row):
                v_float = parse_float_safe(row[ci])
                if v_float > 0: daily_hours_tmp += v_float
                
        if not ma_da and not task_tmp and not ten_da_val_tmp and not phan_vung_val_tmp and tong_tmp <= 0 and daily_hours_tmp <= 0:
            curr_ma_da = ''
            continue
        
        # Xác định vùng / section hiện tại để tính tổng giờ MOS + JMOS + VMOS cho việc kiểm tra ngày 0h
        if any(k in row_str_all for k in ['JMOS']):
            curr_section_type = 'JMOS'
        elif any(k in row_str_all for k in ['VMOS']):
            curr_section_type = 'VMOS'
        elif 'MOS業務' in row_str_all or (len(row_cells_clean) == 1 and row_cells_clean[0] == 'MOS'):
            curr_section_type = 'MOS'
        elif any(k in row_str_all for k in ['社内', '自社', '応援', 'その他', '管理', '一般', '有給', '休暇', '支援']):
            curr_section_type = 'OTHER'
            
        # Tính tổng giờ làm các mục MOS, JMOS, VMOS để phục vụ kiểm tra ngày làm việc bất thường (0h)
        summary_keywords = ['合計', '小計', '総計', 'TOTAL', 'SUBTOTAL', '実工数', '工数合計', '累計', '実績', '合　計', '小　計', '総　計', '社内', '自社', '応援', '管理', '一般', '有給', '休暇', '朝礼', '清掃', '会議', '出張', '移動', '欠勤', '遅刻', '早退']
        is_summary_or_nonwork = any(k in row_str_all for k in summary_keywords)
        is_internal_code = ma_da.startswith('M') or (ma_da and re.match(r'^0+$', ma_da))
        
        if not is_summary_or_nonwork and not is_internal_code:
            if curr_section_type in ['MOS', 'JMOS', 'VMOS'] or ma_da.startswith('J') or ma_da.startswith('V') or ma_da.startswith('K') or 'JMOS' in ma_da or 'VMOS' in ma_da:
                for date_str, ci in date_cols.items():
                    if ci < len(row):
                        v_float = parse_float_safe(row[ci])
                        if v_float > 0:
                            file_daily_work_hours[date_str] = file_daily_work_hours.get(date_str, 0.0) + v_float

        # Kiểm tra nếu chuyển vùng (section header hoặc summary row)
        is_section_header = len(row_cells_clean) <= 3 and (not ma_da or ma_da in ['MOS', 'JMOS', 'VMOS', 'TOTAL', 'SUBTOTAL', 'MOS業務', 'JMOS業務', 'VMOS業務', '社内', '自社', '応援', 'その他', '管理', '一般', '有給', '休暇', '支援', '社内業務'])
        if is_section_header and any(k in row_str_all for k in ['MOS業務', 'JMOS', '社内', '自社', '応援', 'その他', '管理', '一般', '有給', '休暇', '合計', '小計', '総計', 'TOTAL', 'SUBTOTAL', '支援業務']):
            if 'JMOS' not in row_str_all and ('MOS業務' in row_str_all or 'MOS PROJECT' in row_str_all or 'DỰ ÁN MOS' in row_str_all or (len(row_cells_clean) == 1 and row_cells_clean[0] == 'MOS')):
                in_mos_section = True
                continue
            elif any(k in row_str_all for k in ['JMOS', '社内', '自社', '応援', 'その他', '管理', '一般', '有給', '休暇', '小計', '合計', '総計', 'TOTAL', 'SUBTOTAL', '支援']):
                in_mos_section = False
                curr_ma_da = ''
                continue
            
        # Theo yêu cầu: Chỉ nhận những dòng thuộc bảng MOS (hoặc mã dự án hợp lệ bắt đầu bằng K, P, V, số hợp lệ)
        is_valid_project_code = bool(ma_da and (ma_da.startswith('K') or ma_da.startswith('P') or ma_da.startswith('V') or re.match(r'^[A-Z]?\d{4,}$', ma_da)))
        if not in_mos_section:
            curr_ma_da = ''
            continue
        
        tanto_val = str(row[col_map.get('tanto', 0)] or '').strip()
        khach_val = str(row[col_map.get('khach', 2)] or '').strip()
        ten_da_val = str(row[col_map.get('ten_da', 3)] or '').strip()
        phan_vung_val = str(row[col_map.get('phan_vung', 4)] or '').strip()
        task = str(row[col_map.get('task', 5)] or '').strip()
        
        # Loại bỏ các mã dự án nội bộ/quản lý/hỗ trợ/không phải MOS (như M050001...) hoặc mã thuộc JMOS (Mã J), mã ảo toàn số 0
        is_non_mos_code = ma_da.startswith('M') or ma_da.startswith('J') or (ma_da and re.match(r'^0+$', ma_da))
        is_internal_name = not is_valid_project_code and any(k in ma_da or k in ten_da_val or k in phan_vung_val for k in ['社内', '自社', '有給', '休暇', '管理業務', '教育業務', '一般業務', 'その他業務', 'ADMIN', 'INTERNAL', '応援', '支援業務', '朝礼', '清掃', '会議', '出張', '移動', '欠勤', '遅刻', '早退', '合計', '小計', '総計', 'TOTAL', 'SUBTOTAL', '実工数', '工数合計', '累計', '実績', '合　計', '小　計', '総　計'])
        is_mos_flag = True
        if is_non_mos_code or is_internal_name or ma_da in ['社内', 'ADMIN', 'INTERNAL', 'NONE', 'NAN', 'JMOS業務', '社内業務', 'MOS業務', 'JMOS', 'JMO S業務']:
            is_mos_flag = False
            if ma_da in ['社内', 'ADMIN', 'INTERNAL', 'NONE', 'NAN', 'JMOS業務', '社内業務', 'MOS業務', 'JMOS', 'JMO S業務'] and not ma_da.startswith('M') and not ma_da.startswith('J'):
                curr_ma_da = ''
                continue
        tong_val = row[col_map.get('tong', 6)] if 'tong' in col_map else None
        
        tong = parse_float_safe(tong_val) if 'tong' in col_map else 0.0
        
        daily_hours = 0.0
        ngay_co_gio = []
        chi_tiet_ngay = {}
        for date_str, ci in date_cols.items():
            if ci < len(row):
                val = row[ci]
                v_float = parse_float_safe(val)
                if v_float > 0:
                    daily_hours += v_float
                    ngay_co_gio.append(date_str)
                    chi_tiet_ngay[date_str] = v_float
                    
        # Nếu mã dự án hợp lệ (không rỗng), cập nhật thông tin dự án hiện tại
        if ma_da and ma_da not in ['NONE', 'NAN', '']:
            curr_ma_da = ma_da
            curr_khach = khach_val
            curr_ten_da = ten_da_val
            curr_phan_vung = phan_vung_val
            curr_tanto = tanto_val
        else:
            # Nếu mã dự án bị trống (merged cell trong Excel) nhưng dòng này có số giờ làm việc > 0
            # Tuyệt đối không kế thừa nếu là dòng tổng hợp/subtotal/footer/trống task (tránh cộng nhầm dòng tổng vào giờ dự án)
            summary_keywords = ['合計', '小計', '総計', 'TOTAL', 'SUBTOTAL', '実工数', '工数合計', '累計', '実績', '合　計', '小　計', '総　計', '社内', '自社', '応援', '管理', '一般', '有給', '休暇', '朝礼', '清掃', '会議', '出張', '移動', '欠勤', '遅刻', '早退']
            is_summary_row = any(k in str(task).upper() or k in str(ten_da_val).upper() for k in summary_keywords)
            
            if curr_ma_da and not is_summary_row and (tong > 0 or daily_hours > 0):
                ma_da = curr_ma_da
                khach_val = khach_val or curr_khach
                ten_da_val = ten_da_val or curr_ten_da
                phan_vung_val = phan_vung_val or curr_phan_vung
                tanto_val = tanto_val or curr_tanto
                if ma_da.startswith('M') or ma_da.startswith('J'):
                    is_mos_flag = False
            else:
                continue
                
        # Ưu tiên lấy số giờ lớn hơn giữa ô Tổng và tổng số giờ chi tiết hàng ngày để đảm bảo tuyệt đối không thất thoát giờ
        if daily_hours > 0:
            tong = max(tong, round(daily_hours, 2))
            
        if tong <= 0 and daily_hours <= 0:
            continue
        
        def parse_date_str(d_str, year=default_year):
            try:
                m, d = d_str.split('/')
                return datetime.date(year, int(m), int(d))
            except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return None
        
        ngay_dates = [parse_date_str(d) for d in ngay_co_gio]
        ngay_dates = [d for d in ngay_dates if d]
        
        ngay_bat_dau = min(ngay_dates).strftime('%d/%m/%Y') if ngay_dates else ''
        ngay_ket_thuc = max(ngay_dates).strftime('%d/%m/%Y') if ngay_dates else ''
        if ngay_bat_dau and ngay_ket_thuc and ngay_bat_dau == ngay_ket_thuc:
            ngay_ket_thuc = ''
        
        ql_nhat = ''
        if '/' in tanto_val:
            ql_nhat = tanto_val.split('/')[-1].replace('様', '').strip()
        elif tanto_val:
            ql_nhat = tanto_val.replace('様', '').strip()
            
        khach = khach_val.replace('様', '').strip()
        
        ma_da_clean_row = str(ma_da).strip().lower()
        txt_check_row = f"{ma_da} {phan_vung_val} {ten_da_val}".lower()
        if any(k in txt_check_row for k in ['cơ khí', '機械', 'メカ', 'meco', 'machine', '機設計', 'thiết kế máy', 'cấu trúc', 'j01009', 'k230059']):
            phan_vung_vn = '機械設計 \n Thiết kế cơ khí'
        elif any(k in txt_check_row for k in ['mô phỏng', 'simulation', 'cae', 'シミュレーション', 'sim', 'phân tích', 'ansys', '解析', 'p012004', 'vmos', 'k000000', '松井', 'matsui', '石部', 'ishibe']):
            phan_vung_vn = 'シミュレーション設計 \n Thiết kế mô phỏng'
        elif any(k in txt_check_row for k in ['điện', 'điều khiển', '電気', '制御', 'elec', 'control', 'plc', 'mạch', 'pcb', 'hardware', 'phần cứng', 'cảm biến', 'sensor', 'motor', 'động cơ', 'panel', 'tủ điện', 'bảng điện', 'inverter', 'biến tần', 'circuit', 'schematic', 'ee', 'hw', 'io', 'dây', 'wiring', 'k210361', 'ocv', 'k220141', '原料', '外観', 'ngoại quan', 'bề mặt', 'pot', 'イニシャル', 'initial', 'đạo', 'ダオ']):
            phan_vung_vn = '電気設計 \n Thiết kế điện'
        else:
            phan_vung_vn = f"{phan_vung_val} \n {phan_vung_val}" if phan_vung_val else '未指定 \n Chưa xác định'
        
        records.append({
            'ma_nv': extract_ma_nv_from_filename(filename),
            'ten_nv': extract_ten_nv_from_filename(filename),
            'ma_da': ma_da,
            'khach': khach,
            'ten_da': ten_da_val,
            'phan_vung': phan_vung_vn,
            'task': task,
            'tong_gio': tong,
            'ngay_bat_dau': ngay_bat_dau,
            'ngay_ket_thuc': ngay_ket_thuc,
            'ql_nhat': ql_nhat,
            'chi_tiet_ngay': chi_tiet_ngay,
            'all_dates': list(date_cols.keys()),
            'file_year': default_year,
            'gio_mos_jmos_vmos': file_daily_work_hours,
            'is_mos': is_mos_flag
        })
    if not records:
        raise ValueError(f"Không tìm thấy dữ liệu hợp lệ (phần MOS業務) trong file {filename}")
    return pd.DataFrame(records)


def decode_mime_string(s):
    if not s or not isinstance(s, str): return ""
    import email.header
    try:
        decoded_list = email.header.decode_header(s)
        res = ""
        for part, encoding in decoded_list:
            if isinstance(part, bytes):
                res += part.decode(encoding or 'utf-8', errors='replace')
            else:
                res += str(part)
        return res.strip()
    except Exception:
        return str(s).strip()


def parse_single_email_report(file_or_text, filename="", default_year=2026, default_month=6) -> list:
    """
    Đọc và phân tích 1 file mail báo cáo từ Thunderbird/Outlook (.eml, .msg, .txt, .mbox) hoặc văn bản dán.
    Trả về danh sách báo cáo trong mail đó.
    """
    import email
    import email.utils
    import email.policy
    import streamlit as st
    from datetime import datetime
    
    content = ""
    timestamp = datetime.now()
    sender_name = ""
    sender_email = ""
    subject = ""
    
    if hasattr(file_or_text, 'read'):
        try:
            file_or_text.seek(0)
        except Exception: pass
        raw_bytes = file_or_text.read()
    else:
        raw_bytes = file_or_text

    if isinstance(raw_bytes, bytes):
        if str(filename).lower().endswith('.msg'):
            try:
                import extract_msg
                import io
                msg_obj = extract_msg.Message(io.BytesIO(raw_bytes))
                content = msg_obj.body or ""
                if not content.strip() and hasattr(msg_obj, 'htmlBody') and msg_obj.htmlBody:
                    html_raw = str(msg_obj.htmlBody)
                    clean_txt = re.sub(r'<style.*?>.*?</style>', '', html_raw, flags=re.DOTALL | re.I)
                    clean_txt = re.sub(r'<br\s*/?>', '\n', clean_txt, flags=re.I)
                    clean_txt = re.sub(r'</p>', '\n', clean_txt, flags=re.I)
                    clean_txt = re.sub(r'<tr.*?>', '\n', clean_txt, flags=re.I)
                    clean_txt = re.sub(r'<td.*?>', ' ', clean_txt, flags=re.I)
                    clean_txt = re.sub(r'<[^>]+>', ' ', clean_txt)
                    clean_txt = re.sub(r'&nbsp;', ' ', clean_txt)
                    content = clean_txt

                subject = msg_obj.subject or ""
                sender_name = msg_obj.sender or ""
                date_str = msg_obj.date
                if date_str:
                    try:
                        timestamp = datetime.strptime(date_str, "%a, %d %b %Y %H:%M:%S %z")
                    except Exception:
                        pass
                
                # Trích xuất email nếu có trong sender
                if "<" in sender_name and ">" in sender_name:
                    m_email = re.search(r'<([^>]+)>', sender_name)
                    if m_email:
                        sender_email = m_email.group(1).strip().lower()
                        sender_name = sender_name.split('<')[0].replace('"', '').replace("'", "").strip()
                
                msg_obj.close()
                msg = None # To skip eml parsing
            except Exception as e:
                import streamlit as st
                st.error(f"⚠️ Lỗi thư viện đọc .msg: {e}. Hệ thống đang đọc file dưới dạng nhị phân, có thể dẫn đến việc đọc phải rác bộ nhớ (mã cũ như K220537). Vui lòng mở Terminal và chạy lệnh: pip install extract-msg")
                # Fallback to eml parsing if extract_msg fails
                try:
                    msg = email.message_from_bytes(raw_bytes, policy=email.policy.default)
                except Exception:
                    msg = email.message_from_string(raw_bytes.decode('utf-8', errors='replace'), policy=email.policy.default)
        else:
            try:
                msg = email.message_from_bytes(raw_bytes, policy=email.policy.default)
            except Exception:
                msg = email.message_from_string(raw_bytes.decode('utf-8', errors='replace'), policy=email.policy.default)
    else:
        msg = email.message_from_string(str(raw_bytes), policy=email.policy.default)
        
    if msg:
        subject = decode_mime_string(str(msg.get('subject', '')))
        from_hdr = decode_mime_string(str(msg.get('from', '')))
        date_hdr = str(msg.get('date', ''))
        
        if date_hdr:
            try:
                dt_parsed = email.utils.parsedate_to_datetime(date_hdr)
                if dt_parsed: timestamp = dt_parsed
            except Exception: pass
            
        if from_hdr:
            m_email = re.search(r'<([^>]+)>', from_hdr)
            if m_email:
                sender_email = m_email.group(1).strip().lower()
                name_part = from_hdr.split('<')[0].replace('"', '').replace("'", "").strip()
                if name_part: sender_name = name_part
            else:
                m_mail_only = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', from_hdr)
                if m_mail_only:
                    sender_email = m_mail_only.group(0).strip().lower()
                    name_part = from_hdr.replace(sender_email, "").replace('"', '').replace("'", "").strip()
                    if name_part: sender_name = name_part
                else:
                    sender_name = from_hdr.replace('"', '').replace("'", "").replace('>', '').replace('<', '').strip()
                    
        if not sender_name and sender_email:
            user_part = sender_email.split('@')[0]
            sender_name = " ".join([w.capitalize() for w in user_part.replace('.', ' ').replace('_', ' ').split()])
        
    if msg:
        html_content = ""
        if msg.is_multipart():
            for part in msg.walk():
                ctype = part.get_content_type()
                if ctype == "text/plain":
                    content += robust_decode(part.get_payload(decode=True), part.get_content_charset()) + "\n"
                elif ctype == "text/html":
                    html_content += robust_decode(part.get_payload(decode=True), part.get_content_charset()) + "\n"
            if not content.strip() and html_content.strip():
                clean_txt = re.sub(r'<style.*?>.*?</style>', '', html_content, flags=re.DOTALL | re.I)
                clean_txt = re.sub(r'<script.*?>.*?</script>', '', clean_txt, flags=re.DOTALL | re.I)
                clean_txt = re.sub(r'<br\s*/?>', '\n', clean_txt, flags=re.I)
                clean_txt = re.sub(r'</p>', '\n', clean_txt, flags=re.I)
                clean_txt = re.sub(r'<tr.*?>', '\n', clean_txt, flags=re.I)
                clean_txt = re.sub(r'<td.*?>', ' ', clean_txt, flags=re.I)
                clean_txt = re.sub(r'<[^>]+>', ' ', clean_txt)
                clean_txt = re.sub(r'&nbsp;', ' ', clean_txt)
                content = clean_txt
        else:
            ctype = msg.get_content_type()
            raw_payload = robust_decode(msg.get_payload(decode=True), msg.get_content_charset())
            if ctype == "text/html":
                clean_txt = re.sub(r'<style.*?>.*?</style>', '', raw_payload, flags=re.DOTALL | re.I)
                clean_txt = re.sub(r'<script.*?>.*?</script>', '', clean_txt, flags=re.DOTALL | re.I)
                clean_txt = re.sub(r'<br\s*/?>', '\n', clean_txt, flags=re.I)
                clean_txt = re.sub(r'</p>', '\n', clean_txt, flags=re.I)
                clean_txt = re.sub(r'<tr.*?>', '\n', clean_txt, flags=re.I)
                clean_txt = re.sub(r'<td.*?>', ' ', clean_txt, flags=re.I)
                clean_txt = re.sub(r'<[^>]+>', ' ', clean_txt)
                clean_txt = re.sub(r'&nbsp;', ' ', clean_txt)
                content = clean_txt
            else:
                content = raw_payload
                
    import unicodedata
    if content: content = unicodedata.normalize('NFKC', content)
    if filename: filename = unicodedata.normalize('NFKC', str(filename))
    if subject: subject = unicodedata.normalize('NFKC', subject)
    if sender_name: sender_name = unicodedata.normalize('NFKC', sender_name)
                
    # Nếu là email được forward, hãy cố gắng lấy tên người gửi gốc từ nội dung email để tránh gom nhầm vào tên người forward
    m_fwd = re.search(r'(?i)(?:\n|^)\s*(?:From|Từ|差出人|送信者):\s*([^\n]+)', content)
    if m_fwd:
        fwd_from = m_fwd.group(1).strip()
        m_email = re.search(r'<([^>]+)>', fwd_from)
        if m_email:
            fwd_name = fwd_from.split('<')[0].replace('"', '').replace("'", "").strip()
            if fwd_name: sender_name = fwd_name
        else:
            m_mail_only = re.search(r'[\w\.-]+@[\w\.-]+\.\w+', fwd_from)
            if m_mail_only:
                fwd_name = fwd_from.replace(m_mail_only.group(0), "").replace('"', '').replace("'", "").strip()
                if fwd_name: sender_name = fwd_name
            else:
                fwd_name = fwd_from.replace('"', '').replace("'", "").strip()
                if fwd_name: sender_name = fwd_name
        
        
    # Xác định ma_nv, ten_nv
    ma_nv = extract_ma_nv_from_filename(filename) if filename else ""
    ten_nv = extract_ten_nv_from_filename(filename) if filename else ""
    
    if not ma_nv or ma_nv == filename:
        m_code = re.search(r'\b(VM\d+)\b', f"{filename} {sender_name} {subject} {content[:500]}", re.IGNORECASE)
        if m_code:
            ma_nv = m_code.group(1).upper()
            m_name = re.search(r'VM\d+[\s_.:-]+([^\d_.:-]+)', f"{filename} {sender_name} {subject} {content[:500]}")
            if m_name: ten_nv = m_name.group(1).strip()
            else: ten_nv = sender_name or ma_nv
        else:
            ten_nv = sender_name or "Nhân viên Mail"
            ma_nv = ten_nv if sender_name else f"Nhân viên Mail - {filename}"
            
    # Ánh xạ thông minh từ tên/email sang Mã NV & Tên NV chuẩn trong Excel / Database
    known_emp_map = {}
    try:
        from db import get_company_emp_dict
        from utils import auto_detect_columns
        db_emps = get_company_emp_dict('vi', auto_detect_columns)
        for code, name in db_emps.items():
            known_emp_map[code.lower().strip()] = (code, name)
            known_emp_map[name.lower().strip()] = (code, name)
            known_emp_map[remove_accents_mos(name).lower().strip()] = (code, name)
            known_emp_map[remove_accents_mos(name).replace(' ', '').lower().strip()] = (code, name)
            if '(' in name and ')' in name:
                sub = re.search(r'\(([^)]+)\)', name)
                if sub: known_emp_map[sub.group(1).strip().lower()] = (code, name)
    except Exception: pass
    
    try:
        for df_key in ['df_mos_raw', 'df_mos_result']:
            if df_key in st.session_state and st.session_state[df_key] is not None:
                df_ex = st.session_state[df_key]
                if 'ma_nv' in df_ex.columns and 'ten_nv' in df_ex.columns:
                    for _, r in df_ex[['ma_nv', 'ten_nv']].drop_duplicates().iterrows():
                        c = str(r['ma_nv']).strip()
                        n = str(r['ten_nv']).strip()
                        if c and n:
                            known_emp_map[c.lower()] = (c, n)
                            known_emp_map[n.lower()] = (c, n)
                            known_emp_map[remove_accents_mos(n).lower().strip()] = (c, n)
                            known_emp_map[remove_accents_mos(n).replace(' ', '').lower().strip()] = (c, n)
    except Exception: pass

    if known_emp_map:
        matched_pair = None
        if ma_nv and ma_nv.lower().strip() in known_emp_map:
            matched_pair = known_emp_map[ma_nv.lower().strip()]
        from translations import translate_name
        
        keys_to_check = [
            sender_name.lower().strip(),
            translate_name(sender_name, 'vi').lower().strip(),
            sender_email.split('@')[0].replace('.', ' ').strip().lower() if sender_email else "",
            sender_email.split('@')[0].lower() if sender_email else "",
            sender_email.split('@')[0].split('.')[-1].lower() if sender_email and '.' in sender_email else "",
            re.sub(r'\.eml|\.msg|\.txt|\.mbox', '', filename, flags=re.I).lower().strip() if filename else "",
            translate_name(re.sub(r'\.eml|\.msg|\.txt|\.mbox', '', filename, flags=re.I), 'vi').lower().strip() if filename else ""
        ]
        
        m_bracket = re.search(r'【(.*?)】', filename) if filename else None
        if m_bracket:
            kata = m_bracket.group(1).strip()
            vi_name = translate_name(kata, 'vi').lower().strip()
            if vi_name:
                keys_to_check.append(vi_name)
                keys_to_check.append(remove_accents_mos(vi_name))
        
        for k in keys_to_check:
            if not k: continue
            if k in known_emp_map:
                matched_pair = known_emp_map[k]
                break
                
            for known_k, val in known_emp_map.items():
                if len(k) >= 3 and len(known_k) >= 3 and (k == known_k or (known_k in k and len(known_k) >= 4)):
                    matched_pair = val
                    break
                    
                # Phân tích tên cuối (First Name trong tiếng Việt, ví dụ Đạo trong Hà Văn Đạo)
                c, n = val
                last_name_vi = n.strip().split()[-1].lower() if n else ""
                last_name_no_accents = remove_accents_mos(last_name_vi)
                if len(last_name_vi) >= 3:
                    if last_name_vi in k or last_name_no_accents in k:
                        matched_pair = val
                        break
                        
            if matched_pair: break
                
        if matched_pair:
            ma_nv, ten_nv = matched_pair
        else:
            ten_nv = re.sub(r'[><]', '', ten_nv).strip()
            ma_nv = re.sub(r'[><]', '', ma_nv).strip()
            
    # Xác định ngày báo cáo (date_str)
    date_str = f"{default_month}/1"
    m_date = re.search(r'(?:20\d\d|\d\d)[/\-.](\d{1,2})[/\-.](\d{1,2})|(?:ngày|date|báo cáo|report|日)[\s:_.\-/]*(\d{1,2})[/\-.](\d{1,2})|(\d{1,2})\s*月\s*(\d{1,2})\s*日|(?:\b|^)(\d{1,2})[/\-.](\d{1,2})(?:[/\-.](?:20\d\d|\d\d))?\b', f"{filename} {subject} {content[:2500]}", re.IGNORECASE)
    if m_date:
        mo = m_date.group(1) or m_date.group(3) or m_date.group(5) or m_date.group(8)
        da = m_date.group(2) or m_date.group(4) or m_date.group(6) or m_date.group(7)
        if mo and da:
            d_int, m_int = int(da), int(mo)
            if m_int == default_month or (1 <= m_int <= 12 and d_int <= 31):
                date_str = f"{m_int}/{d_int}"
            elif d_int == default_month or (1 <= d_int <= 12 and m_int <= 31):
                date_str = f"{d_int}/{m_int}"
            else:
                date_str = f"{default_month}/{d_int if d_int <= 31 else 1}"
    else:
        date_str = f"{timestamp.month}/{timestamp.day}"
        
    projects = {}
    project_sections = {} # map: mã dự án -> mục ('MOS', 'JMOS', 'SHANAI', 'UNKNOWN')
    project_managers = {} # map: mã dự án -> tên quản lý Nhật ('石部', '大瀧'...)
    project_names = {} # map: mã dự án -> tên dự án
    section_hours = {'MOS': 0.0, 'JMOS': 0.0, 'SHANAI': 0.0, 'UNKNOWN': 0.0}
    
    current_section = "UNKNOWN"
    current_manager = ""
    current_p_code = None
    task_counter = 1
    # Cắt bỏ phần lịch sử email (Forward / Reply) để tránh đọc nhầm báo cáo của người khác
    quote_patterns = [
        r'(?:\n|^)\s*_{20,}', 
        r'(?:\n|^)\s*-{5,}\s*Original Message\s*-{5,}',
        r'(?:\n|^)\s*From:\s*.*',
        r'(?:\n|^)\s*Từ:\s*.*',
        r'(?:\n|^)\s*差出人:\s*.*',
        r'(?:\n|^)\s*送信者:\s*.*',
        r'(?:\n|^)\s*On\s+.*wrote:\s*',
        r'(?:\n|^)\s*202\d[年/\-].*?:\s*\n?',
        r'(?:\n|^)\s*Vào\s+.*đã viết:\s*',
        r'(?:\n|^)\s*.*@.*wrote:\s*'
    ]
    min_idx = len(content)
    for p in quote_patterns:
        match = re.search(p, content, re.IGNORECASE)
        if match:
            min_idx = min(min_idx, match.start())
            
    if min_idx < len(content):
        top_part = content[:min_idx]
        # Chỉ cắt bỏ phần history nếu phần đầu (tin nhắn mới) CÓ chứa mã công việc hoặc số giờ
        # Nếu phần đầu trống (ví dụ admin forward lại mail), ta KHÔNG cắt để giữ lại nội dung báo cáo thật bên dưới!
        if re.search(r'\b[KJPVM]\d{4,}\b|\b\d{6}\b', top_part, re.IGNORECASE) or 'mos' in top_part.lower() or 'jmos' in top_part.lower() or re.search(r'\b\d{1,2}(?:\.\d+)?h\b', top_part, re.IGNORECASE):
            content = top_part

    # --- DEBUG START ---
    try:
        with open("debug_mail_text.txt", "a", encoding="utf-8") as f:
            f.write(f"\n\n{'='*50}\n")
            f.write(f"FILE: {filename}\n")
            f.write(f"{'='*50}\n")
            f.write(content)
    except Exception:
        pass
    # --- DEBUG END ---

    import unicodedata
    content = unicodedata.normalize('NFKC', content)
    content_flat = re.sub(r'[\r\n]+', ' ', content)
    # Xoá các mã dự án nằm trong ngoặc đơn (ví dụ: (K220537-A1), （K220537）) vì đây thường là chú thích phụ
    content_flat = re.sub(r'[\(\（][^\)\）]*(?:[KJPVM]\d{4,}|\b\d{6}\b)[^\)\）]*[\)\）]', '()', content_flat, flags=re.IGNORECASE)
    content_norm = re.sub(r'(?i)([・•◇■◆▼▲]|■\s*MOS|■\s*JMOS|■\s*社内|【\s*MOS\s*】|【\s*JMOS\s*】|【\s*社内\s*】|Dự án:|Project:|\b[KJPVM]\d{4,}\b|\b\d{6}\b|\b[KJPV]\d+\b)', r'\n\1', content_flat)
    lines = content_norm.split('\n')
    
    last_assigned_proj_key = None
    last_assigned_hours = 0.0
    
    for line in lines:
        line_clean = line.strip()
        if not line_clean or line_clean.startswith('>'): continue
        
        line_upper = line_clean.upper()
        # Nhận diện tên Quản lý Nhật trong Mail (ví dụ: 石部様, 大瀧様, 営業/石部様, 【石部様】, (石部様), Ishibe, Otaki)
        m_mgr = re.search(r'([^\s/<\[\(:：,，+]+)\s*様|\b(石部|大滝|大瀧|池谷|勝亦|田代|山崎|井手上|松井|Ishibe|Otaki|Ikeya|Katsumata|Tashiro|Yamazaki|Ideue|Matsui)\b', line_clean, re.IGNORECASE)
        if m_mgr:
            mgr_val = (m_mgr.group(1) or m_mgr.group(2) or '').replace('様', '').strip()
            if mgr_val and len(mgr_val) >= 2 and not any(k in mgr_val.upper() for k in ['MOS', 'JMOS', 'VMOS', 'TOTAL', 'SUBTOTAL', '業務', '会社', '社内', '客様', 'DỰ ÁN', 'PROJECT', 'NGÀY', 'DATE', 'TIME']):
                current_manager = mgr_val

        # Kiểm tra chuyển mục (Section Header trong Mail)
        if re.search(r'◇\s*JMOS|■\s*JMOS|【\s*JMOS\s*】|\bJMOS\s*業務|\bJMOS\s*:$|^\s*JMOS\s*$', line_upper):
            current_section = "JMOS"
            continue
        elif re.search(r'◇\s*VMOS|■\s*VMOS|【\s*VMOS\s*】|\bVMOS\s*業務|◇\s*社内|■\s*社内|【\s*社内\s*】|\b社内\s*業務|\bNỘI\s*BỘ|◇\s*INTERNAL|■\s*INTERNAL', line_upper):
            current_section = "SHANAI"
            continue
        elif re.search(r'◇\s*MOS|■\s*MOS|【\s*MOS\s*】|\bMOS\s*業務|\bMOS\s*:$|^\s*MOS\s*$', line_upper):
            current_section = "MOS"
            continue
            
        matches = list(re.finditer(r'([KJPVM]\d{4,}|\b\d{6}\b|[KJPV]\d+)|\b(MOS|JMOS|VMOS|社内|有給|休暇|管理|応援|一般|支援)\b', line_upper))
        if matches:
            real_codes = [m.group(1) for m in matches if m.group(1)]
            generic_codes = [m.group(2) for m in matches if m.group(2)]
            
            new_p_code = None
            if real_codes:
                new_p_code = real_codes[0]
            elif generic_codes:
                new_p_code = generic_codes[0]
                
            if new_p_code:
                # Không để chữ MOS/JMOS/VMOS vô tình đè mất mã dự án thật (ví dụ K12345) ở dòng trước
                if new_p_code in ['MOS', 'JMOS', 'VMOS']:
                    if not current_p_code or not re.match(r'(?:[KJPVM]\d+|\b\d{6}\b)', current_p_code):
                        current_p_code = new_p_code
                else:
                    current_p_code = new_p_code
                
                # Extract project name
                name_clean = re.sub(r'(?i)([KJPVM]\d{4,}|\b\d{6}\b|[KJPV]\d+)|\b(MOS|JMOS|VMOS|社内|有給|休暇|管理|応援|一般|支援)\b', '', line_clean)
                name_clean = re.sub(r'(?i)(\d{1,2})\s*(?:h|hr|giờ|tiếng)\s*(\d{1,2})\b', '', name_clean)
                name_clean = re.sub(r'(?i)(\d+(?:[.,]\d+)?)\s*(?:h|hr|hrs|giờ|tiếng|h/ngày|h/day)\b|\b(?:giờ|tiếng|h|hours?|thời gian|time)\s*[:=]?\s*(\d+(?:[.,]\d+)?)\b', '', name_clean)
                name_clean = re.sub(r'^[\]\)\s:,\.\-]*', '', name_clean).strip()
                if name_clean and len(name_clean) >= 3 and current_p_code:
                    if current_p_code not in project_names or len(name_clean) > len(project_names[current_p_code]):
                        project_names[current_p_code] = name_clean
            
        if any(k in line_upper for k in ['TỔNG', 'TOTAL', 'SUM', 'CỘNG', '合計']):
            current_p_code = None
            continue
            
        line_working = line_clean.lower()
        
        # Ngoại lệ: Nếu mã dự án là P012004 nhưng nội dung nhắc đến triển lãm thì chuyển sang tính là SHANAI (nội bộ/chung)
        if current_p_code and current_p_code.upper().startswith('P012004'):
            if any(k in line_working for k in ['triển lãm', 'trien lam', '展示会']):
                current_p_code = 'SHANAI'
                # Retroactively move hours if they were parsed on the previous line
                if last_assigned_proj_key and str(last_assigned_proj_key).upper().startswith('P012004'):
                    if projects.get(last_assigned_proj_key, 0) >= last_assigned_hours and last_assigned_hours > 0:
                        projects[last_assigned_proj_key] -= last_assigned_hours
                        shanai_key = f"SHANAI_NOCODE_{task_counter}"
                        task_counter += 1
                        projects[shanai_key] = projects.get(shanai_key, 0.0) + last_assigned_hours
                        
                        # Fix project names if needed
                        if last_assigned_proj_key in project_names:
                            project_names[shanai_key] = project_names[last_assigned_proj_key]
                        
                        last_assigned_proj_key = shanai_key
                
        # Tìm các chuỗi XhY (ví dụ 4h30)
        for m_hm in re.finditer(r'(?<!\d:)(\d{1,2})\s*(?:h|hr|giờ|tiếng)\s*(\d{1,2})\b', line_working):
            val_float = float(m_hm.group(1)) + float(m_hm.group(2)) / 60.0
            if val_float > 0 and val_float <= 24:
                p_code_use = current_p_code or "MOS" # Default to MOS if no code
                if p_code_use in ['MOS', 'JMOS', 'VMOS', 'SHANAI']:
                    p_code_use = f"{p_code_use}_NOCODE_{task_counter}"
                    task_counter += 1
                projects[p_code_use] = projects.get(p_code_use, 0.0) + val_float
                last_assigned_proj_key = p_code_use
                last_assigned_hours = val_float
                
                sec = current_section
                if sec == "UNKNOWN":
                    if p_code_use.startswith('J') or 'JMOS' in line_upper: sec = 'JMOS'
                    elif p_code_use.startswith('M') or p_code_use.startswith('V') or any(k in line_upper for k in ['社内', '有給', '休暇', '管理', '応援', '一般', '支援']): sec = 'SHANAI'
                    else: sec = 'MOS'
                project_sections[p_code_use] = sec
                if current_manager:
                    project_managers[p_code_use] = current_manager
                section_hours[sec] = section_hours.get(sec, 0.0) + val_float
                    
            # Xóa các chuỗi XhY đã tìm để không bị bắt trùng ở bước sau
            line_working = re.sub(r'(?<!\d:)(\d{1,2})\s*(?:h|hr|giờ|tiếng)\s*(\d{1,2})\b', '', line_working)
            
        # Bắt buộc phải có đơn vị hoặc từ khóa thời gian
        for m_hrs in re.finditer(r'(\d+(?:[.,]\d+)?)\s*(?:h|hr|hrs|giờ|tiếng|h/ngày|h/day)\b|\b(?:giờ|tiếng|h|hours?|thời gian|time)\s*[:=]?\s*(\d+(?:[.,]\d+)?)\b', line_working):
            val_str = m_hrs.group(1) or m_hrs.group(2)
            try:
                val_float = float(val_str.replace(',', '.'))
                if val_float > 0 and val_float <= 24:
                    p_code_use = current_p_code or "MOS" # Default to MOS
                    if p_code_use in ['MOS', 'JMOS', 'VMOS', 'SHANAI']:
                        p_code_use = f"{p_code_use}_NOCODE_{task_counter}"
                        task_counter += 1
                        
                    projects[p_code_use] = projects.get(p_code_use, 0.0) + val_float
                    sec = current_section
                    if sec == "UNKNOWN":
                        if p_code_use.startswith('J') or 'JMOS' in line_upper: sec = 'JMOS'
                        elif p_code_use.startswith('M') or p_code_use.startswith('V') or any(k in line_upper for k in ['社内', '有給', '休暇', '管理', '応援', '一般', '支援']): sec = 'SHANAI'
                        else: sec = 'MOS'
                    project_sections[p_code_use] = sec
                    if current_manager:
                        project_managers[p_code_use] = current_manager
                    section_hours[sec] = section_hours.get(sec, 0.0) + val_float
            except Exception: pass
                
        # Fallback: Nếu đã có mã dự án nhưng chưa bắt được giờ (do nhân viên viết thiếu chữ 'h', ví dụ K12345: 4 hoặc J12345 - 4.5)
        # Chỉ quét số cuối dòng hoặc sau dấu hai chấm.
        m_alone = re.search(r'(?:[:=\-]\s*|\s+)(\d+(?:[.,]\d+)?)\s*$', line_working)
        if m_alone and current_p_code:
            # Nếu trên dòng này đã bắt được giờ rồi thì thôi (tránh tính đúp)
            has_hours_already = False
            for p in projects:
                if p == current_p_code or p.startswith(f"{current_p_code}_NOCODE_"):
                    has_hours_already = True
            if not has_hours_already:
                try:
                    val_float = float(m_alone.group(1).replace(',', '.'))
                    if val_float > 0 and val_float <= 24:
                        p_code_use = current_p_code
                        if p_code_use in ['MOS', 'JMOS', 'VMOS', 'SHANAI']:
                            p_code_use = f"{p_code_use}_NOCODE_{task_counter}"
                            task_counter += 1
                        projects[p_code_use] = projects.get(p_code_use, 0.0) + val_float
                        last_assigned_proj_key = p_code_use
                        last_assigned_hours = val_float
                        sec = current_section
                        if sec == "UNKNOWN":
                            if p_code_use.startswith('J') or 'JMOS' in line_upper: sec = 'JMOS'
                            elif p_code_use.startswith('M') or p_code_use.startswith('V') or any(k in line_upper for k in ['社内', '有給', '休暇', '管理', '応援', '一般', '支援']): sec = 'SHANAI'
                            else: sec = 'MOS'
                        project_sections[p_code_use] = sec
                        if current_manager:
                            project_managers[p_code_use] = current_manager
                        section_hours[sec] = section_hours.get(sec, 0.0) + val_float
                except Exception: pass
                
    if not projects:
        return []
        
    return [{
        'ma_nv': ma_nv,
        'ten_nv': ten_nv,
        'date': date_str,
        'timestamp': timestamp,
        'projects': projects,
        'project_names': project_names,
        'project_sections': project_sections,
        'project_managers': project_managers,
        'section_hours': section_hours,
        'total_hours': sum(projects.values()),
        'source_file': filename or "Pasted Text",
        'sender_name': sender_name,
        'sender_email': sender_email,
        'raw_body': content
    }]


def deduplicate_email_reports(all_parsed_reports):
    """
    Deduplication:
    Nhóm theo (ma_nv/ten_nv, Ngày làm việc date). Nếu có 2 mail cho cùng 1 ngày làm việc, CHỈ lấy mail mới nhất!
    """
    staff_map = {
        'ダンフン': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng', 'VM018'),
        'フン': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng', 'VM018'),
        'hưng': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng', 'VM018'),
        'hung': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng', 'VM018'),
        'クアン': ('ファム・ホン・クアン', 'Phạm Hồng Quân', 'VM035'),
        'quân': ('ファム・ホン・クアン', 'Phạm Hồng Quân', 'VM035'),
        'quan': ('ファム・ホン・クアン', 'Phạm Hồng Quân', 'VM035'),
        'ハン': ('レ・ティ・カイン・ハン', 'Lê Thị Khánh Hằng', 'VM037'),
        'hằng': ('レ・ティ・カイン・ハン', 'Lê Thị Khánh Hằng', 'VM037'),
        'hang': ('レ・ティ・カイン・ハン', 'Lê Thị Khánh Hằng', 'VM037'),
        'グエット': ('グエン・ミン・グエット', 'Nguyễn Minh Nguyệt', 'VM038'),
        'nguyệt': ('グエン・ミン・グエット', 'Nguyễn Minh Nguyệt', 'VM038'),
        'nguyet': ('グエン・ミン・グエット', 'Nguyễn Minh Nguyệt', 'VM038'),
        'ロン': ('レ・ヴァン・ロン', 'Lê Văn Long', 'VM011'),
        'long': ('レ・ヴァン・ロン', 'Lê Văn Long', 'VM011'),
        'ダオ': ('ハ・ヴァン・ダオ', 'Hà Văn Đạo', 'VM008'),
        'đạo': ('ハ・ヴァン・ダオ', 'Hà Văn Đạo', 'VM008'),
        'dao': ('ハ・ヴァン・ダオ', 'Hà Văn Đạo', 'VM008'),
        'フォン': ('レ・タイン・フォン', 'Lê Thanh Phương', 'VM009'),
        'phương': ('レ・タイン・フォン', 'Lê Thanh Phương', 'VM009'),
        'phuong': ('レ・タイン・フォン', 'Lê Thanh Phương', 'VM009'),
        'ドゥ': ('ファム・ゴック・ドゥ', 'Phạm Ngọc Dư', 'VM033'),
        'dư': ('ファム・ゴック・ドゥ', 'Phạm Ngọc Dư', 'VM033'),
        'du': ('ファム・ゴック・ドゥ', 'Phạm Ngọc Dư', 'VM033'),
    }

    stage2_grouped = {}
    for rep in all_parsed_reports:
        raw_name = str(rep.get('ten_nv', '')).strip()
        raw_ma = str(rep.get('ma_nv', '')).strip()
        raw_name_lower = raw_name.lower()
        short_name = raw_name_lower.split()[-1] if raw_name_lower else ''
        mapped_names = staff_map.get(raw_name_lower, staff_map.get(short_name, None))
        
        if mapped_names:
            norm_key = mapped_names[2]
            rep['ten_nv'] = mapped_names[1]
            rep['ma_nv'] = mapped_names[2]
        else:
            norm_key = raw_ma or raw_name_lower
            
        work_date = rep['date']
        key2 = (norm_key, work_date)
        if key2 not in stage2_grouped:
            stage2_grouped[key2] = []
        stage2_grouped[key2].append(rep)
        
    dict_latest = {}
    logs = []
    import streamlit as st
    is_vi = st.session_state.get('lang', 'vi') == 'vi'
    for (norm_key, w_date), reps in stage2_grouped.items():
        reps_sorted = sorted(reps, key=lambda x: x['timestamp'])
        latest = reps_sorted[-1]
        
        raw_name = str(latest.get('ten_nv', '')).strip()
        raw_name_lower = raw_name.lower()
        short_name = raw_name_lower.split()[-1] if raw_name_lower else ''
        mapped_names = staff_map.get(raw_name_lower, staff_map.get(short_name, None))
        
        final_name = (mapped_names[1] if is_vi else mapped_names[0]) if mapped_names else raw_name
        dict_latest[(norm_key, w_date)] = latest
        
        if len(reps_sorted) > 1:
            if is_vi:
                old_times = ", ".join([r['timestamp'].strftime('%d/%m lúc %H:%M') for r in reps_sorted[:-1]])
                new_time = latest['timestamp'].strftime('%d/%m lúc %H:%M')
                logs.append(f"🔄 **{final_name}** - Ngày làm việc {w_date}: Có {len(reps_sorted)} mail báo cáo cho cùng ngày làm việc. Đã thay thế mail cũ [{old_times}] bằng mail mới nhất [{new_time}]!")
            else:
                old_times = ", ".join([r['timestamp'].strftime('%d/%m %H:%M') for r in reps_sorted[:-1]])
                new_time = latest['timestamp'].strftime('%d/%m %H:%M')
                logs.append(f"🔄 **{final_name}** - 稼働日 {w_date}: 同じ日に {len(reps_sorted)} 件の報告メールがあります。古いメール [{old_times}] を最新のメール [{new_time}] に置き換えました！")
            
    return dict_latest, logs


def remove_accents_mos(input_str):
    import unicodedata
    s1 = unicodedata.normalize('NFD', str(input_str))
    return ''.join(c for c in s1 if unicodedata.category(c) != 'Mn').replace('đ', 'd').replace('Đ', 'D')


def is_same_employee_mos(ex_ma, ex_ten, mail_key, mail_rep=None):
    ex_ma = str(ex_ma).strip().lower()
    ex_ten = str(ex_ten).strip().lower()
    mail_key_str = str(mail_key).strip().lower()
    mail_ten = str(mail_rep.get('ten_nv', '') if mail_rep else '').strip().lower()
    mail_ma = str(mail_rep.get('ma_nv', '') if mail_rep else '').strip().lower()
    mail_file = str(mail_rep.get('source_file', '') if mail_rep else '').strip().lower()
    mail_sender_name = str(mail_rep.get('sender_name', '') if mail_rep else '').strip().lower()
    mail_sender_email = str(mail_rep.get('sender_email', '') if mail_rep else '').strip().lower()
    
    # 1. So sánh chính xác Mã NV (nếu có mã chuẩn dạng VMxxx)
    if ex_ma and re.match(r'^vm\d+$', ex_ma, re.I):
        for m_val in [mail_key_str, mail_ten, mail_ma, mail_file, mail_sender_name]:
            if not m_val: continue
            if ex_ma == m_val or ex_ma in m_val:
                return True
                
    # 2. So sánh chính xác Tên NV hoặc Email
    for m_val in [mail_key_str, mail_ten, mail_ma, mail_sender_name, mail_sender_email]:
        if not m_val: continue
        if ex_ten and ex_ten == m_val: return True
        if ex_ma and ex_ma == m_val: return True
        
        # Nhận diện theo từng từ (Word-based subset matching)
        # Đặc biệt hữu ích với tên ngắn như "Dư", "An", "Hà"
        ex_words = set(remove_accents_mos(ex_ten).split())
        m_val_clean = re.sub(r'[^\w\s]', ' ', m_val)
        m_words = set(remove_accents_mos(m_val_clean).split())
        if ex_words and ex_words.issubset(m_words):
            return True
        
        ex_clean = remove_accents_mos(ex_ten).replace(' ', '').strip()
        m_clean = remove_accents_mos(m_val).replace(' ', '').strip()
        if ex_clean and m_clean and len(ex_clean) >= 3 and len(m_clean) >= 3:
            if ex_clean == m_clean:
                return True
            # Chỉ so sánh chứa (in) nếu cả 2 chuỗi đều dài từ 4 ký tự trở lên
            if len(ex_clean) >= 4 and len(m_clean) >= 4:
                if ex_clean in m_clean or m_clean in ex_clean:
                    return True
    return False


def is_dummy_project_code(code_str):
    """Kiểm tra mã dự án tạm chưa quyết định (như K000000, 000000, K00000, 未定...)"""
    c = str(code_str).strip().upper()
    c = re.sub(r'\s*\(\d+\)$', '', c) # Ignore suffix like " (2)"
    return c == 'K000000' or c == '000000' or c == 'K00000' or c == '00000' or c == '未定' or 'CHƯA QUYẾT ĐỊNH' in c or (c.startswith('K0000') and len(c) >= 6)


def resolve_dummy_code_with_mail(ma_nv, ql_nhat_excel, dict_latest_reports, used_real_codes=None, ten_nv='', target_hours=None, target_date=None):
    """
    Khi trong Excel có mã dự án tạm (K000000...), tìm trong Mail báo cáo của nhân viên xem mã thực tế là gì.
    Nếu có >= 2 mã tạm K000000 hoặc nhiều mã trong Mail, đối chiếu theo Tên quản lý Nhật Bản (ql_nhat/tanto) hoặc số giờ làm (target_hours) để phân biệt chính xác.
    """
    if not dict_latest_reports:
        return None
        
    import streamlit as st
    known_code_to_mgr = {}
    for df_source in [st.session_state.get('df_mos_raw'), st.session_state.get('df_mos_result')]:
        if df_source is not None and not df_source.empty:
            for _, r_s in df_source.iterrows():
                c_s = str(r_s.get('ma_da' if 'ma_da' in r_s else 'Mã dự án', '')).strip().upper()
                m_s = str(r_s.get('ql_nhat' if 'ql_nhat' in r_s else (r_s.get('khach', '') or r_s.get('tanto', '') or r_s.get('Quản lý Nhật Bản', '')), '')).replace('様', '').strip()
                if c_s and not is_dummy_project_code(c_s) and m_s:
                    m_clean = m_s.split('/')[-1].strip() if '/' in m_s else m_s
                    m_clean = m_s.split('/')[-1].strip() if '/' in m_s else m_s
                    known_code_to_mgr[c_s] = m_clean

    # Đã bỏ logic invalid_candidates vì nó chặn mất việc thay thế mã dự án nếu trùng ngày (thường xuyên xảy ra khi cập nhật mã)
    invalid_candidates = set()

    candidate_projects = {}
    mail_projs_dict = {}
    for (nv_key, d_str), mail_rep in dict_latest_reports.items():
        if target_date and d_str != target_date:
            continue
        if is_same_employee_mos(ma_nv, ten_nv, nv_key, mail_rep):
            mail_projs = mail_rep.get('projects', {})
            project_sections = mail_rep.get('project_sections', {})
            project_managers = mail_rep.get('project_managers', {})
            
            for p_code, p_hrs in mail_projs.items():
                p_upper = str(p_code).strip().upper()
                sec = project_sections.get(p_code, 'UNKNOWN')
                if not is_dummy_project_code(p_upper):
                    if p_upper not in candidate_projects:
                        candidate_projects[p_upper] = set()
                        mail_projs_dict[p_upper] = 0.0
                    mail_projs_dict[p_upper] += p_hrs
                    
                    mgr_mail = project_managers.get(p_code) or project_managers.get(p_upper) or known_code_to_mgr.get(p_upper) or ''
                    if mgr_mail:
                        m_mail_clean = str(mgr_mail).split('/')[-1].replace('様', '').strip()
                        candidate_projects[p_upper].add(m_mail_clean)
                    elif p_upper in known_code_to_mgr:
                        candidate_projects[p_upper].add(known_code_to_mgr[p_upper])

    if used_real_codes:
        if isinstance(used_real_codes, dict):
            for u_code, ex_hrs in used_real_codes.items():
                m_hrs = mail_projs_dict.get(u_code, 0)
                rem_hrs = m_hrs - ex_hrs
                if rem_hrs <= 0.01:
                    candidate_projects.pop(u_code, None)
                else:
                    mail_projs_dict[u_code] = rem_hrs
        # KHÔNG tự ý xóa (pop) nếu used_real_codes chỉ là list/set vì sẽ gây mất mã đúng!
            
    if not candidate_projects:
        return None
        
    if len(candidate_projects) == 1:
        return list(candidate_projects.keys())[0]
        
    ql_clean = str(ql_nhat_excel or '').split('/')[-1].replace('様', '').strip().lower()
    
    # Strategy 1: Project Name Matching across emails
    # Tìm trong mail xem nhân viên đã thay đổi mã K000000 thành mã thật nào cho cùng một tên dự án
    dummy_names = set()
    real_names = {}
    ignore_names = ['chuaxacdinh', 'chuaquyetdinh', 'unknown', 'none', 'null', 'mitei', '未定']
    
    for (nv_key, d_str), mail_rep in dict_latest_reports.items():
        if is_same_employee_mos(ma_nv, ten_nv, nv_key, mail_rep):
            p_names = mail_rep.get('project_names', {})
            mgrs = mail_rep.get('project_managers', {})
            for p_code, p_name in p_names.items():
                p_upper = str(p_code).strip().upper()
                p_name_clean = re.sub(r'[^a-zA-Z0-9]', '', str(p_name).lower())
                if p_name_clean and not any(ign in p_name_clean for ign in ignore_names):
                    if is_dummy_project_code(p_upper):
                        m_mail_str = str(mgrs.get(p_code, '')).strip().lower()
                        if not ql_clean or ql_clean in m_mail_str or m_mail_str in ql_clean:
                            dummy_names.add(p_name_clean)
                    else:
                        real_names[p_name_clean] = p_upper

    for d_name in dummy_names:
        for r_name, r_code in real_names.items():
            if len(d_name) >= 4 and len(r_name) >= 4:
                if d_name in r_name or r_name in d_name:
                    return r_code
                    
    if ql_clean:
        matched_by_mgr = []
        projects_with_no_mgr = []
        for p_code, mgr_set in candidate_projects.items():
            if not mgr_set or all(not str(m).strip() for m in mgr_set):
                projects_with_no_mgr.append(p_code)
                continue
                
            for m_mail in mgr_set:
                m_mail_clean = str(m_mail).strip().lower()
                if ql_clean in m_mail_clean or m_mail_clean in ql_clean:
                    matched_by_mgr.append(p_code)
                    break
                    
        if len(matched_by_mgr) == 1:
            return matched_by_mgr[0]
        elif len(matched_by_mgr) > 1:
            # Nếu có nhiều mã cùng trùng tên quản lý, chỉ giữ lại các mã này để phân xử tiếp bằng số giờ
            candidate_projects = {p: candidate_projects[p] for p in matched_by_mgr}
        elif projects_with_no_mgr:
            # Nếu không có mã nào trùng tên quản lý, chỉ cho phép lấy các dự án KHÔNG CÓ quản lý trong mail
            candidate_projects = {p: candidate_projects[p] for p in projects_with_no_mgr}
        else:
            # Tuyệt đối không cho phép cướp dự án đã có quản lý khác!
            return None
                    
    if target_hours is not None and target_hours > 0:
        matched_by_hours = [p for p in candidate_projects.keys() if abs(mail_projs_dict.get(p, 0) - target_hours) < 0.01]
        if len(matched_by_hours) == 1:
            return matched_by_hours[0]
            
        # Tìm dự án có số giờ gần giống nhất
        closest_p = None
        min_diff = 999999
        for p in candidate_projects.keys():
            diff = abs(mail_projs_dict.get(p, 0) - target_hours)
            if diff < min_diff:
                min_diff = diff
                closest_p = p
                
        if closest_p:
            return closest_p
            
    # Fallback cuối cùng: Chọn dự án có số lượng giờ lớn nhất để thay thế K000000
    # (Nếu K000000 thực chất bao gồm 2 dự án thật, dự án lớn nhất sẽ đè lên K000000, dự án còn lại sẽ tự động tạo dòng mới)
    if candidate_projects:
        best_p = max(candidate_projects.keys(), key=lambda p: mail_projs_dict.get(p, 0))
        return best_p
        
    return None


def reconcile_mail_vs_excel(dict_latest_reports, df_mos_raw):
    """
    Đối chiếu dữ liệu giữa Mail mới nhất và bảng Excel (df_mos_raw).
    Trả về DataFrame bảng đối chiếu chênh lệch và số lượng bất thường.
    """
    comparison_rows = []
    discrepancy_count = 0
    
    excel_map = {}
    valid_excel_dates = set()
    ex_employees = {(str(row.get('ma_nv', '')).strip(), str(row.get('ten_nv', '')).strip()) for _, row in df_mos_raw.iterrows() if str(row.get('ma_nv', '')).strip() or str(row.get('ten_nv', '')).strip()}
    
    # Gom nhóm dữ liệu Excel theo nhân viên
    excel_by_emp = {}
    
    for _, row in df_mos_raw.iterrows():
        ma_nv = str(row.get('ma_nv', '')).strip()
        ten_nv = str(row.get('ten_nv', '')).strip()
        nv_key = ma_nv.lower()
        if nv_key:
            if nv_key not in excel_by_emp: excel_by_emp[nv_key] = []
            excel_by_emp[nv_key].append(row)

        ma_da = str(row.get('ma_da', '')).strip()
        all_d = row.get('all_dates', [])
        if isinstance(all_d, list):
            valid_excel_dates.update(str(d) for d in all_d)
        chi_tiet = row.get('chi_tiet_ngay', {})
        if isinstance(chi_tiet, dict):
            valid_excel_dates.update(str(d) for d in chi_tiet.keys())
            for d_str, hrs in chi_tiet.items():
                hrs_float = float(hrs or 0.0)
                if hrs_float > 0:
                    key = (ma_nv, d_str)
                    if key not in excel_map:
                        excel_map[key] = {'projects': {}, 'total_hours': 0.0, 'managers': {}}
                    
                    mgr_row = str(row.get('ql_nhat') or row.get('tanto') or row.get('khach') or '').strip()
                    ma_da_key = ma_da
                    if is_dummy_project_code(ma_da):
                        if ma_da_key in excel_map[key]['projects'] and excel_map[key]['managers'].get(ma_da_key) != mgr_row:
                            suffix = 2
                            while f"{ma_da} ({suffix})" in excel_map[key]['projects']:
                                suffix += 1
                            ma_da_key = f"{ma_da} ({suffix})"
                            
                    excel_map[key]['projects'][ma_da_key] = excel_map[key]['projects'].get(ma_da_key, 0.0) + hrs_float
                    excel_map[key]['total_hours'] += hrs_float
                    if mgr_row:
                        excel_map[key]['managers'][ma_da_key] = mgr_row
                    
    for (mail_nv_key, d_str), mail_rep in dict_latest_reports.items():
        if valid_excel_dates and d_str not in valid_excel_dates:
            continue
        matched_ex_ma = None
        matched_ex_ten = None
        for ex_ma, ex_ten in ex_employees:
            if is_same_employee_mos(ex_ma, ex_ten, mail_nv_key, mail_rep):
                matched_ex_ma = ex_ma
                matched_ex_ten = ex_ten
                break
                
                
        if not matched_ex_ma:
            matched_ex_ma = None # Force None to trigger warning below
            matched_ex_ten = ten_nv

        ten_nv = mail_rep['ten_nv']
        mail_projs = mail_rep['projects']
        project_sections = mail_rep.get('project_sections', {})
        section_hours = mail_rep.get('section_hours', {'MOS': 0.0, 'JMOS': 0.0, 'SHANAI': 0.0, 'UNKNOWN': 0.0})
        mail_tot = mail_rep['total_hours']
        
        # Lọc CHỈ lấy các dự án thuộc mục MOS trong Mail để đối chiếu chuẩn với bảng Excel MOS
        mail_mos_projs = {}
        for k, v in mail_projs.items():
            sec = project_sections.get(k, 'UNKNOWN')
            k_upper = str(k).upper()
            if sec == 'MOS':
                mail_mos_projs[k] = v
            elif sec == 'UNKNOWN':
                is_internal = (k_upper.startswith('J') and 'JMOS' not in k_upper) or (k_upper.startswith('M') and 'MOS' not in k_upper) or (k_upper.startswith('V') and 'VMOS' not in k_upper)
                if not is_internal:
                    mail_mos_projs[k] = v
        
        mail_mos_tot = sum(mail_mos_projs.values())
        if mail_mos_tot == 0 and mail_tot > 0 and not any(s in ['MOS', 'JMOS', 'SHANAI'] for s in project_sections.values()):
            mail_mos_projs = {k: v for k, v in mail_projs.items() if not (str(k).upper().startswith('J') and 'JMOS' not in str(k).upper()) and not (str(k).upper().startswith('M') and 'MOS' not in str(k).upper()) and not (str(k).upper().startswith('V') and 'VMOS' not in str(k).upper())}
            mail_mos_tot = sum(mail_mos_projs.values())
            
        ex_data = excel_map.get((matched_ex_ma, d_str), {'projects': {}, 'total_hours': 0.0, 'managers': {}}) if matched_ex_ma else {'projects': {}, 'total_hours': 0.0, 'managers': {}}
        ex_projs = ex_data['projects']
        ex_mgrs = ex_data.get('managers', {})
        ex_tot = ex_data['total_hours']
        
        status = "✅ Khớp số liệu"
        notes = []
        is_diff = False
        
        if not matched_ex_ma:
            status = "⚠️ Có Mail nhưng không có trong Excel"
            notes.append("Nhân viên này có gửi báo cáo Mail nhưng không có tên trong file Excel hiện tại!")
            is_diff = True

        
        # Giải quyết mã tạm K000000 trong Excel sang mã thực tế từ Mail (đối chiếu theo Quản lý Nhật Bản)
        ex_projs_resolved = {}
        used_real = set()
        real_codes_ex = {k: v for k, v in ex_projs.items() if not is_dummy_project_code(k)}
        for ex_code, ex_h in ex_projs.items():
            if is_dummy_project_code(ex_code):
                mgr_ex = ex_mgrs.get(ex_code, '')
                real_code = resolve_dummy_code_with_mail(matched_ex_ma, mgr_ex, dict_latest_reports, real_codes_ex, ten_nv=matched_ex_ten, target_hours=ex_h, target_date=None)
                if real_code:
                    ex_projs_resolved[real_code] = ex_projs_resolved.get(real_code, 0.0) + ex_h
                    used_real.add(real_code)
                    notes.append(f"💡 Đã cập nhật mã tạm {ex_code} ➔ {real_code} (QL Nhật: {mgr_ex or 'Nhật'})")
                else:
                    ex_projs_resolved[ex_code] = ex_projs_resolved.get(ex_code, 0.0) + ex_h
            else:
                ex_projs_resolved[ex_code] = ex_projs_resolved.get(ex_code, 0.0) + ex_h
        
        # Giải quyết mã tạm K000000 trong Mail sang mã thực tế (đối chiếu theo Quản lý Nhật Bản trong Mail hoặc Excel)
        project_managers = mail_rep.get('project_managers', {})
        mail_mos_projs_resolved = {}
        mail_projs_resolved = {}
        real_codes_mail = {k: v for k, v in mail_mos_projs.items() if not is_dummy_project_code(k)}
        for m_code, m_h in mail_mos_projs.items():
            if is_dummy_project_code(m_code):
                mgr_mail = project_managers.get(m_code) or ex_mgrs.get(m_code, '')
                
                # Ưu tiên 1: Lấy mã thật từ Excel của chính ngày hôm nay (nếu khớp số giờ hoặc khớp tên quản lý)
                matched_ex_real = None
                for ex_code, ex_h in ex_projs_resolved.items():
                    if not is_dummy_project_code(ex_code) and ex_code not in mail_mos_projs_resolved and ex_code not in real_codes_mail:
                        mgr_ex = ex_mgrs.get(ex_code, '')
                        if m_h == ex_h or (mgr_mail and mgr_ex and (mgr_mail in mgr_ex or mgr_ex in mgr_mail)):
                            matched_ex_real = ex_code
                            break
                            
                if matched_ex_real:
                    real_code = matched_ex_real
                else:
                    # Ưu tiên 2: Nếu Excel cũng ko có, đành phải lục lại lịch sử email cũ để tìm mã thật
                    real_code = resolve_dummy_code_with_mail(matched_ex_ma, mgr_mail, dict_latest_reports, real_codes_mail, ten_nv=matched_ex_ten, target_hours=m_h, target_date=None)
                
                if real_code:
                    mail_mos_projs_resolved[real_code] = mail_mos_projs_resolved.get(real_code, 0.0) + m_h
                    if m_code in project_sections: project_sections[real_code] = project_sections[m_code]
                else:
                    mail_mos_projs_resolved[m_code] = mail_mos_projs_resolved.get(m_code, 0.0) + m_h
            else:
                mail_mos_projs_resolved[m_code] = mail_mos_projs_resolved.get(m_code, 0.0) + m_h
                
        for m_code, m_h in mail_projs.items():
            if is_dummy_project_code(m_code):
                mgr_mail = project_managers.get(m_code) or ex_mgrs.get(m_code, '')
                real_code = resolve_dummy_code_with_mail(matched_ex_ma, mgr_mail, dict_latest_reports, None, ten_nv=matched_ex_ten, target_hours=m_h, target_date=None)
                if real_code:
                    mail_projs_resolved[real_code] = mail_projs_resolved.get(real_code, 0.0) + m_h
                    if m_code in project_sections: project_sections[real_code] = project_sections[m_code]
                else:
                    mail_projs_resolved[m_code] = mail_projs_resolved.get(m_code, 0.0) + m_h
            else:
                mail_projs_resolved[m_code] = mail_projs_resolved.get(m_code, 0.0) + m_h
        
        # 0. Auto-map mã con và mã cha (ví dụ: J01009 trong Mail với J01009-0032 trong Excel)
        keys_to_update = {}
        for m_code in list(mail_projs_resolved.keys()):
            for ex_code in ex_projs_resolved.keys():
                if m_code != ex_code:
                    # Nếu một mã chứa mã kia kèm dấu gạch ngang (VD: J01009 và J01009-0032)
                    if ex_code.startswith(m_code + "-") or m_code.startswith(ex_code + "-"):
                        keys_to_update[m_code] = ex_code
                        break
        for m_code, ex_code in keys_to_update.items():
            mail_projs_resolved[ex_code] = mail_projs_resolved.get(ex_code, 0.0) + mail_projs_resolved.pop(m_code)
            if m_code in project_sections:
                project_sections[ex_code] = project_sections[m_code]
        
        # 0.5 Tính tổng giờ MOS thực sự trong Mail
        mail_mos_tot = sum(v for k, v in mail_projs_resolved.items() if (project_sections.get(k) == 'MOS' or project_sections.get(k) == 'UNKNOWN') and not str(k).upper().startswith('J') and not str(k).upper().startswith('M') and not str(k).upper().startswith('V'))

        
        # 1. Đối chiếu từng dự án trong Excel với Mail
        for p_code, ex_h in ex_projs_resolved.items():
            m_h = mail_projs_resolved.get(p_code, 0.0)
            if m_h == 0:
                if not is_dummy_project_code(p_code):
                    status = "⚠️ Thừa dự án trong Excel"
                    notes.append(f"Excel có dự án {p_code} ({ex_h}h) nhưng trong Mail không có!")
                    is_diff = True
            elif abs(ex_h - m_h) > 0.1:
                status = "⚠️ Lệch giờ dự án"
                notes.append(f"Dự án {p_code}: Mail ({round(m_h, 2)}h) ≠ Excel ({round(ex_h, 2)}h)")
                is_diff = True
                
        # 2. Đối chiếu các dự án MOS trong Mail xem có bị thiếu trong Excel không
        for p_code, m_h in mail_projs_resolved.items():
            is_mos = (project_sections.get(p_code) in ['MOS', 'UNKNOWN']) and not str(p_code).upper().startswith('J') and not str(p_code).upper().startswith('M') and not str(p_code).upper().startswith('V')
            if is_mos and p_code not in ex_projs_resolved:
                if not is_dummy_project_code(p_code):
                    status = "⚠️ Thiếu dự án trong Excel"
                    notes.append(f"Excel thiếu dự án MOS {p_code} ({m_h}h)")
                    is_diff = True

        # 3. Kiểm tra sâu chuyên đề Mã J (JMOS) ghi sai mục trong Mail
        for p_code, p_hrs in mail_projs_resolved.items():
            sec = project_sections.get(p_code, 'UNKNOWN')
            if str(p_code).upper().startswith('J') and sec == 'MOS':
                status = "🚫 Lỗi sai mục trong Mail (Mã J)"
                notes.append(f"🚫 Lỗi Mail: Dự án [{p_code}] có đầu mã J (thuộc JMOS) nhưng trong Mail lại ghi sai vào mục ◇MOS業務!")
                is_diff = True
                        

            
        if is_diff:
            discrepancy_count += 1
            
        # Hiển thị rõ các mục trong cột Giờ Mail và Dự án Mail
        mail_hrs_display = f"{mail_mos_tot}h (MOS)"
        if section_hours.get('JMOS', 0) > 0: mail_hrs_display += f" | {section_hours['JMOS']}h (JMOS)"
        if section_hours.get('SHANAI', 0) > 0: mail_hrs_display += f" | {section_hours['SHANAI']}h (Nội bộ)"
        
        projs_display = []
        for k, v in mail_projs_resolved.items():
            sec_tag = project_sections.get(k, '')
            tag_label = "Nội bộ" if sec_tag == 'SHANAI' else sec_tag
            tag_str = f" [{tag_label}]" if sec_tag in ['MOS', 'JMOS', 'SHANAI'] else ""
            projs_display.append(f"{k} ({v}h){tag_str}")
            
        import streamlit as st
        is_vi = st.session_state.get('lang', 'vi') == 'vi'
        staff_map = {
            'ダンフン': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng'), 'フン': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng'), 'hưng': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng'),
            'クアン': ('ファム・ホン・クアン', 'Phạm Hồng Quân'), 'quân': ('ファム・ホン・クアン', 'Phạm Hồng Quân'),
            'ハン': ('レ・ティ・カイン・ハン', 'Lê Thị Khánh Hằng'), 'hằng': ('レ・ティ・カイン・ハン', 'Lê Thị Khánh Hằng'),
            'グエット': ('グエン・ミン・グエット', 'Nguyễn Minh Nguyệt'), 'nguyệt': ('グエン・ミン・グエット', 'Nguyễn Minh Nguyệt'),
            'ロン': ('レ・ヴァン・ロン', 'Lê Văn Long'), 'long': ('レ・ヴァン・ロン', 'Lê Văn Long'),
            'ダオ': ('ハ・ヴァン・ダオ', 'Hà Văn Đạo'), 'đạo': ('ハ・ヴァン・ダオ', 'Hà Văn Đạo'),
            'フォン': ('レ・タイン・フォン', 'Lê Thanh Phương'), 'phương': ('レ・タイン・フォン', 'Lê Thanh Phương'),
            'ドゥ': ('ファム・ゴック・ドゥ', 'Phạm Ngọc Dư'), 'dư': ('ファム・ゴック・ドゥ', 'Phạm Ngọc Dư'),
        }
        raw_name = matched_ex_ten or ten_nv
        raw_name_lower = str(raw_name).strip().lower()
        mapped_names = staff_map.get(raw_name, staff_map.get(raw_name_lower, None))
        final_name = (mapped_names[1] if is_vi else mapped_names[0]) if mapped_names else raw_name
        
        comparison_rows.append({
            'Mã NV': matched_ex_ma,
            'Tên NV': final_name,
            'Ngày': d_str,
            'Trạng thái': status,
            'Giờ Mail': mail_hrs_display,
            'Giờ Excel': f"{ex_tot}h",
            'Dự án Mail': ", ".join(projs_display),
            'Dự án Excel': ", ".join([f"{k} ({v}h)" for k, v in ex_projs_resolved.items()]),
            'Chi tiết chênh lệch / Cảnh báo': " | ".join(notes) if notes else "Khớp hoàn toàn"
        })
        

    if comparison_rows:
        try:
            comparison_rows.sort(key=lambda x: (x['Tên NV'], x['Mã NV'], tuple(map(int, x['Ngày'].split('/')))))
        except: pass
        
    return pd.DataFrame(comparison_rows), discrepancy_count

    # Sắp xếp lại bảng theo Tên -> Mã -> Ngày để dễ nhìn



def apply_reconciliation_to_excel(dict_latest_reports, df_mos_raw):
    """
    Chuẩn hóa df_mos_raw theo dữ liệu trong dict_latest_reports.
    Trả về df_mos_raw đã chuẩn hóa.
    """
    df_clean = df_mos_raw.copy()
    # 0. Giải quyết các mã tạm K000000 trong Excel sang mã thực tế từ Mail (đối chiếu theo Tên quản lý Nhật Bản)
    used_real_by_emp = {}
    # Pre-collect all valid real codes per employee and their total hours in Excel
    for idx, row in df_clean.iterrows():
        ma_nv = str(row.get('ma_nv', '')).strip()
        ma_da_row = str(row.get('ma_da', '')).strip().upper()
        if not is_dummy_project_code(ma_da_row):
            if ma_nv not in used_real_by_emp: used_real_by_emp[ma_nv] = {}
            used_real_by_emp[ma_nv][ma_da_row] = used_real_by_emp[ma_nv].get(ma_da_row, 0.0) + float(row.get('tong_gio') or 0.0)
    for idx, row in df_clean.iterrows():
        ma_da_row = str(row.get('ma_da', '')).strip().upper()
        if is_dummy_project_code(ma_da_row):
            ma_nv = str(row.get('ma_nv', '')).strip()
            ten_nv = str(row.get('ten_nv', '')).strip()
            mgr_ex = str(row.get('ql_nhat') or row.get('tanto') or row.get('khach') or '').strip()
            real_code = resolve_dummy_code_with_mail(ma_nv, mgr_ex, dict_latest_reports, used_real_by_emp.get(ma_nv, {}), ten_nv=ten_nv, target_hours=row.get('tong_gio'))
            if real_code:
                df_clean.at[idx, 'ma_da_original'] = ma_da_row
                df_clean.at[idx, 'ma_da'] = real_code
                if ma_nv not in used_real_by_emp: used_real_by_emp[ma_nv] = {}
                used_real_by_emp[ma_nv][real_code] = used_real_by_emp[ma_nv].get(real_code, 0.0) + float(row.get('tong_gio') or 0.0)
    valid_excel_dates = set()
    for _, row in df_clean.iterrows():
        all_d = row.get('all_dates', [])
        if isinstance(all_d, list):
            valid_excel_dates.update(str(d) for d in all_d)
        chi_t = row.get('chi_tiet_ngay', {})
        if isinstance(chi_t, dict):
            valid_excel_dates.update(str(d) for d in chi_t.keys())
    for (mail_nv_key, d_str), mail_rep in dict_latest_reports.items():
        if valid_excel_dates and d_str not in valid_excel_dates:
            continue
        mail_projs = mail_rep['projects']
        project_sections = mail_rep.get('project_sections', {})
        project_managers = mail_rep.get('project_managers', {})
        # CHỈ lấy các dự án thuộc mục MOS trong Mail để chuẩn hóa vào bảng Excel MOS
        mail_mos_projs = {k: v for k, v in mail_projs.items() if not str(k).upper().startswith('J') and not str(k).upper().startswith('M') and not str(k).upper().startswith('V') and (project_sections.get(k) == 'MOS' or project_sections.get(k) == 'UNKNOWN')}
        if sum(mail_mos_projs.values()) == 0 and sum(mail_projs.values()) > 0 and not any(s in ['MOS', 'JMOS', 'SHANAI'] for s in project_sections.values()):
            mail_mos_projs = {k: v for k, v in mail_projs.items() if not str(k).upper().startswith('J') and not str(k).upper().startswith('M') and not str(k).upper().startswith('V')}
        # Giải quyết mã tạm K000000 trong Mail sang mã thực tế
        mail_mos_projs_resolved = {}
        real_codes_mail = {k for k in mail_mos_projs.keys() if not is_dummy_project_code(k)}
        for m_code, m_h in mail_mos_projs.items():
            if is_dummy_project_code(m_code):
                mgr_mail = project_managers.get(m_code, '')
                if not mgr_mail:
                    for _, r_ex in df_clean.iterrows():
                        if is_same_employee_mos(r_ex.get('ma_nv', ''), r_ex.get('ten_nv', ''), mail_nv_key, mail_rep):
                            mgr_mail = str(r_ex.get('ql_nhat') or r_ex.get('tanto') or r_ex.get('khach') or '').strip()
                            if mgr_mail: break
                real_code = resolve_dummy_code_with_mail(mail_nv_key, mgr_mail, dict_latest_reports, real_codes_mail, ten_nv=mail_rep.get('ten_nv', ''), target_hours=m_h)
                if real_code:
                    mail_mos_projs_resolved[real_code] = mail_mos_projs_resolved.get(real_code, 0.0) + m_h
                else:
                    mail_mos_projs_resolved[m_code] = mail_mos_projs_resolved.get(m_code, 0.0) + m_h
            else:
                mail_mos_projs_resolved[m_code] = mail_mos_projs_resolved.get(m_code, 0.0) + m_h
        mail_mos_projs = mail_mos_projs_resolved
        # 1. Cập nhật các dòng hiện có trong df_clean cho ma_nv
        allocated_projects = set()
        for idx, row in df_clean.iterrows():
            if is_same_employee_mos(row.get('ma_nv', ''), row.get('ten_nv', ''), mail_nv_key, mail_rep):
                ma_da = str(row.get('ma_da', '')).strip()
                chi_tiet = dict(row.get('chi_tiet_ngay', {}))
                if ma_da in mail_mos_projs:
                    if ma_da not in allocated_projects:
                        chi_tiet[d_str] = mail_mos_projs[ma_da]
                        allocated_projects.add(ma_da)
                    else:
                        if d_str in chi_tiet: chi_tiet[d_str] = 0.0
                else:
                    if d_str in chi_tiet and float(chi_tiet[d_str] or 0) > 0:
                        chi_tiet[d_str] = 0.0
                df_clean.at[idx, 'chi_tiet_ngay'] = chi_tiet
                df_clean.at[idx, 'tong_gio'] = sum([float(v or 0) for v in chi_tiet.values()])
                
                # Cập nhật tên dự án
                old_name = str(row.get('ten_da', '')).strip()
                is_placeholder = any(x in old_name.lower() for x in ['未定', 'chưa xác định', 'unknown', 'none', 'nan'])
                if not old_name or is_placeholder or is_dummy_project_code(str(row.get('ma_da_original', ma_da))):
                    p_names = mail_rep.get('project_names', {})
                    if ma_da in p_names and p_names[ma_da]:
                        df_clean.at[idx, 'ten_da'] = p_names[ma_da]
                    elif is_dummy_project_code(ma_da):
                        for k, v in p_names.items():
                            if is_dummy_project_code(k) and v:
                                df_clean.at[idx, 'ten_da'] = v
                                break
        # 2. Kiểm tra nếu Mail có dự án MOS mới mà trong Excel của ma_nv chưa có dòng nào cho dự án đó
        matched_ex_ma = None
        for _, r_emp in df_clean.iterrows():
            if is_same_employee_mos(r_emp.get('ma_nv', ''), r_emp.get('ten_nv', ''), mail_nv_key, mail_rep):
                matched_ex_ma = str(r_emp.get('ma_nv', '')).strip()
                break
        if not matched_ex_ma:
            continue
        existing_projs = set(df_clean[df_clean['ma_nv'] == matched_ex_ma]['ma_da'].astype(str).str.strip())
        for p_code, p_hrs in mail_mos_projs.items():
            if p_code not in existing_projs and p_hrs > 0:
                sample_row = df_clean[df_clean['ma_nv'] == matched_ex_ma].iloc[0] if not df_clean[df_clean['ma_nv'] == matched_ex_ma].empty else None
                
                # Lấy tên dự án từ Mail nếu có
                mail_p_name = mail_rep.get('project_names', {}).get(p_code, '')
                if not mail_p_name and is_dummy_project_code(p_code):
                    for k, v in mail_rep.get('project_names', {}).items():
                        if is_dummy_project_code(k) and v:
                            mail_p_name = v
                            break
                new_ten_da = mail_p_name if mail_p_name else f"Dự án từ Mail ({p_code})"
                
                new_row = {
                    'ma_nv': matched_ex_ma,
                    'ten_nv': sample_row['ten_nv'] if sample_row is not None else mail_rep['ten_nv'],
                    'ma_da': p_code,
                    'khach': sample_row['khach'] if sample_row is not None else '',
                    'ten_da': new_ten_da,
                    'phan_vung': sample_row['phan_vung'] if sample_row is not None else '機械設計 \n Thiết kế cơ khí',
                    'task': 'Báo cáo từ Mail Thunderbird',
                    'tong_gio': p_hrs,
                    'ngay_bat_dau': sample_row['ngay_bat_dau'] if sample_row is not None else '',
                    'ngay_ket_thuc': sample_row['ngay_ket_thuc'] if sample_row is not None else '',
                    'ql_nhat': sample_row['ql_nhat'] if sample_row is not None else '',
                    'chi_tiet_ngay': {d_str: p_hrs},
                    'all_dates': sample_row['all_dates'] if sample_row is not None else [d_str],
                    'file_year': sample_row['file_year'] if sample_row is not None else 2026,
                    'gio_mos_jmos_vmos': {d_str: p_hrs}
                }
                df_clean = pd.concat([df_clean, pd.DataFrame([new_row])], ignore_index=True)
    return df_clean
def _clean_jp_to_vn(text: str) -> str:
    import re
    if not text: return ""
    s_trim = str(text).strip()
    exact_map = {
        "機械設計・モデリング及び作図支援業務": "Nghiệp vụ hỗ trợ thiết kế cơ khí, mô hình hóa 3D & lập bản vẽ",
        "CAEシミュレーション解析・評価支援業務": "Nghiệp vụ hỗ trợ phân tích mô phỏng CAE & đánh giá kỹ thuật",
        "電気制御設計・回路図作成及び評価業務": "Nghiệp vụ thiết kế điện điều khiển, lập sơ đồ mạch & đánh giá",
        "ソフトウェア開発・プログラミング支援業務": "Nghiệp vụ hỗ trợ phát triển phần mềm & lập trình hệ thống",
        "技術ドキュメント作成・翻訳支援業務": "Nghiệp vụ hỗ trợ soạn thảo tài liệu kỹ thuật & dịch thuật",
        "総合技術サポート及びプロジェクト推進業務": "Nghiệp vụ hỗ trợ kỹ thuật tổng hợp & triển khai dự án",
    }
    if s_trim in exact_map:
        return exact_map[s_trim]
    trans = text
    
    # Sử dụng Regex để hoán đổi cấu trúc ngữ pháp Tiếng Nhật (Tân ngữ + Động từ) sang Tiếng Việt (Động từ + Tân ngữ)
    # Ví dụ: レイアウト 作成 -> Tạo レイアウト (Thay vì "レイアウト Tạo")
    trans = re.sub(r'([^\s,・]+)\s*(?:を作成する|を作成|の作成|作成する|作成)', r'Tạo \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を実施する|を実施|の実施|を実施し|実施する|実施)', r'Thực hiện \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を行う|を行い|の実行|を実行する|実行)', r'Thực hiện \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を修正する|を修正|の修正|修正する|修正)', r'Chỉnh sửa \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を変更する|の変更|を変更|変更する|変更)', r'Thay đổi \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を確認する|を確認|の確認|確認する|確認)', r'Kiểm tra \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:をチェックする|をチェック|のチェック|チェックする|チェック)', r'Kiểm tra \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を検討する|を検討|の検討|検討する|検討)', r'Nghiên cứu \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を評価する|を評価|の評価|評価する|評価)', r'Đánh giá \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:をテストする|のテスト|テストする|テスト)', r'Kiểm thử \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:をサポートする|をサポート|のサポート|を支援する|の支援|サポートする|サポート|支援する|支援)', r'Hỗ trợ \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:に参加する|に参加|への参加|参加する|参加)', r'Tham gia \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を準備する|の准备|準備する|準備)', r'Chuẩn bị \1', trans)
    trans = re.sub(r'([^\s,・]+)\s*(?:を対応|の対応|対応する|対応)', r'Xử lý \1', trans)
    
    terms = [
        # Cụm động từ/nghiệp vụ đầy đủ thường gặp
        ("シミュレーションを作成する", "Xây dựng mô phỏng"), ("シミュレーションを作成", "Xây dựng mô phỏng"),
        ("シミュレーションの作成", "Xây dựng mô phỏng"), ("シミュレーション作成", "Xây dựng mô phỏng"),
        ("シミュレーションを実施する", "Thực hiện mô phỏng"), ("シミュレーションを実施", "Thực hiện mô phỏng"),
        ("シミュレーションの実施", "Thực hiện mô phỏng"), ("シミュレーション実施", "Thực hiện mô phỏng"),
        ("3Dモデリングを作成する", "Xây dựng mô hình 3D"), ("3Dモデリングを作成", "Xây dựng mô hình 3D"),
        ("3Dモデルを作成する", "Xây dựng mô hình 3D"), ("3Dモデルを作成", "Xây dựng mô hình 3D"),
        ("3Dモデルの作成", "Xây dựng mô hình 3D"), ("モデル作成", "Xây dựng mô hình"),
        ("2D図面を作成する", "Lập bản vẽ 2D"), ("2D図面を作成", "Lập bản vẽ 2D"),
        ("図面を作成する", "Lập bản vẽ"), ("図面を作成", "Lập bản vẽ"), ("図面の作成", "Lập bản vẽ"),
        ("図面を修正する", "Chỉnh sửa bản vẽ"), ("図面を修正", "Chỉnh sửa bản vẽ"), ("図面の修正", "Chỉnh sửa bản vẽ"),
        ("設計を実施する", "Thực hiện thiết kế"), ("設計を行う", "Thực hiện thiết kế"),
        ("検討を実施する", "Thực hiện nghiên cứu"), ("検討を行う", "Thực hiện nghiên cứu"),
        ("解析を実施する", "Thực hiện phân tích"), ("解析を行う", "Thực hiện phân tích"),
        ("評価を実施する", "Thực hiện đánh giá"), ("評価を行う", "Thực hiện đánh giá"),
        ("検査を実施する", "Thực hiện kiểm tra"), ("検査を行う", "Thực hiện kiểm tra"),
        ("打合せを行う", "Thực hiện thảo luận"), ("打ち合わせを行う", "Thực hiện thảo luận"),
        ("会議に参加する", "Tham gia họp"),
        # Thuật ngữ nghiệp vụ
        ("技術開発業務委託費", "Phí ủy thác nghiệp vụ phát triển kỹ thuật"),
        ("業務委託費", "Phí ủy thác nghiệp vụ"),
        ("技術開発", "Phát triển kỹ thuật"),
        ("業務委託", "Ủy thác nghiệp vụ"),
        ("委託費", "Phí ủy thác"),
        ("3D モデリング", "Mô hình hóa 3D"), ("モデリング", "Mô hình hóa 3D"),
        ("実装", "Lắp ráp"), ("ネジ締め機", "Máy siết ốc"),
        ("5 号機", "máy số 5"), ("1 号機", "máy số 1"), ("2 号機", "máy số 2"),
        ("3 号機", "máy số 3"), ("4 号機", "máy số 4"), ("号機", "máy số"),
        ("対応改造", "Cải tiến tương thích"), ("改造", "Cải tiến/Cải tạo"), ("改善", "Cải tiến"),
        ("外観検査機", "Máy kiểm tra ngoại quan"), ("自動検査機", "Máy kiểm tra tự động"),
        ("検査装置", "Thiết bị kiểm tra"), ("検査機", "Máy kiểm tra"), ("検査", "Kiểm tra"),
        ("原料ポット", "Hũ nguyên liệu"), ("原料", "Nguyên liệu"), ("外観", "Ngoại quan"),
        ("水側", "Đường nước"), ("リーク", "Rò rỉ"), ("検討", "Nghiên cứu/Thẩm định"),
        ("条件", "Điều kiện"), ("ハブ", "Hub"), ("接着", "Dán keo"),
        ("自動機", "Máy tự động"), ("自動", "Tự động"), ("装置", "Thiết bị"),
        ("2D 図面化", "Lên bản vẽ 2D"), ("図面化", "Lên bản vẽ"),
        ("図面修正", "Sửa bản vẽ"), ("図面", "Bản vẽ"), ("製図", "Lập bản vẽ"),
        ("構想設計", "Thiết kế ý tưởng"), ("詳細設計", "Thiết kế chi tiết"), ("変更設計", "Thiết kế thay đổi"),
        ("回路設計", "Thiết kế mạch"), ("電気設計", "Thiết kế điện"), ("制御設計", "Thiết kế điều khiển"),
        ("機械設計", "Thiết kế cơ khí"), ("シミュレーションセッケイ", "Thiết kế mô phỏng"), ("シミュレーション設計", "Thiết kế mô phỏng"),
        ("設計", "Thiết kế"),
        ("シミュレーション", "Mô phỏng"), ("CAE解析", "Phân tích CAE"), ("解析", "Phân tích"), ("制作", "Chế tạo"), ("製作", "Chế tạo"),
        ("治具", "Đồ gá (Jig)"), ("組立", "Lắp ráp"), ("部品", "Linh kiện"),
        ("評価", "Đánh giá"), ("試験", "Thử nghiệm"), ("テスト", "Kiểm thử"),
        ("マニュアル", "Tài liệu"), ("仕様書", "Đặc tả kỹ thuật"),
        ("研究", "Nghiên cứu"), ("調査", "Khảo sát"), ("分析", "Phân tích"),
        ("測定", "Đo lường"), ("調整", "Điều chỉnh"), ("保守", "Bảo trì"), ("メンテナンス", "Bảo trì"),
        ("技術", "Kỹ thuật/Công nghệ"), ("開発", "Phát triển"), ("業務", "Nghiệp vụ"),
        ("委託", "Ủy thác"), ("費用", "Chi phí"), ("費", "Chi phí"),
        ("管理", "Quản lý"), ("支援", "Hỗ trợ"), ("対応", "Xử lý/Tương thích"),
        ("翻訳", "Dịch thuật"), ("通訳", "Thông dịch"),
        ("ドキュメント作成", "Soạn thảo tài liệu"), ("資料作成", "Soạn thảo tài liệu"),
        ("修正", "Chỉnh sửa"), ("サポート", "Hỗ trợ"), ("その他", "Khác"), ("未定", "Chưa xác định"),
        ("打ち合わせ", "Họp"), ("打合せ", "Họp"), ("会議", "Họp"),
        ("レイアウト", "Layout"), ("アニメーション", "Animation"), ("データ", "Dữ liệu"), ("デザイン", "Thiết kế / Design"),
        ("プログラム", "Chương trình"), ("システム", "Hệ thống"),
        # Các từ Katakana/Kanji kỹ thuật thường gặp (lỗi gõ nhanh bằng Katakana hoặc tiếng Anh hóa)
        ("サイクルタイム", "Cycle time"), ("タイムチャート", "Biểu đồ thời gian"),
        ("セッケイ", "Thiết kế"), ("セイズ", "Lập bản vẽ"), ("カイセキ", "Phân tích"),
        ("ヒョウカ", "Đánh giá"), ("ケンサ", "Kiểm tra"), ("ケントウ", "Nghiên cứu / Thẩm định"),
        ("シュウセイ", "Chỉnh sửa"), ("サクセイ", "Xây dựng / Lập"), ("ジッシ", "Thực hiện"),
        ("カイハツ", "Phát triển"), ("ギジュツ", "Kỹ thuật"), ("ジグ", "Đồ gá (Jig)"),
        ("クミタテ", "Lắp ráp"), ("ズメン", "Bản vẽ"), ("ブヒン", "Linh kiện"),
        ("シケン", "Thử nghiệm"), ("チョウセイ", "Điều chỉnh"), ("ホシュ", "Bảo trì"),
        ("ギョウム", "Nghiệp vụ"), ("シエン", "Hỗ trợ"), ("タイオウ", "Xử lý / Tương thích"),
        ("ソノタ", "Khác"), ("ミテイ", "Chưa xác định"), ("ウチアワセ", "Họp / Thảo luận"),
        ("カイギ", "Họp"),
        ("モデリング", "Mô hình hóa 3D"), ("ズメンカ", "Lên bản vẽ"),
        ("キカイセッケイ", "Thiết kế cơ khí"), ("セイギョセッケイ", "Thiết kế điều khiển"),
        ("デンキセッケイ", "Thiết kế điện"), ("カイロセッケイ", "Thiết kế mạch"),
        ("コウソウセッケイ", "Thiết kế ý tưởng"), ("ショウサイセッケイ", "Thiết kế chi tiết"),
        ("ヘンコウセッケイ", "Thiết kế thay đổi"),
        ("ポンプ", "Bơm"), ("バルブ", "Van"), ("センサ", "Cảm biến"), ("センサー", "Cảm biến"),
        ("モータ", "Động cơ"), ("モーター", "Động cơ"), ("アセンブリ", "Lắp ráp (Assembly)"),
        ("ユニット", "Cụm (Unit)"), ("フレーム", "Khung"), ("カバー", "Vỏ bảo vệ"),
        ("ブラケット", "Gá đỡ"), ("プレート", "Tấm"), ("ケース", "Hộp/Vỏ"),
        ("シャフト", "Trục"), ("ギヤ", "Bánh răng"), ("ギア", "Bánh răng"),
        ("ベアリング", "Vòng bi"), ("ケーブル", "Cáp"), ("コネクタ", "Đầu nối"),
        ("ボルト", "Bu lông"), ("ナット", "Đai ốc"), ("ネジ", "Ốc vít"),
        # Các đuôi động từ / ngữ pháp tiếng Nhật (đặt ở cuối danh sách replacement)
        ("を作成する", " Tạo"), ("を作成", " Tạo"), ("の作成", " Tạo"), ("作成する", " Tạo"), ("作成", " Tạo"),
        ("を実施する", " - Thực hiện"), ("を実施", " - Thực hiện"), ("の実施", " - Thực hiện"), ("を実施し", " - Thực hiện"), ("を実施", " - Thực hiện"),
        ("を行う", " - Thực hiện"), ("を行い", " - Thực hiện"), ("の実行", " - Thực hiện"), ("を実行する", " - Thực hiện"), ("実行", " - Thực hiện"),
        ("を修正する", " - Chỉnh sửa"), ("を修正", " - Chỉnh sửa"), ("の修正", " - Chỉnh sửa"), ("修正する", " - Chỉnh sửa"), ("修正", " - Chỉnh sửa"),
        ("を変更する", " - Thay đổi"), ("の変更", " - Thay đổi"), ("を変更", " - Thay đổi"), ("変更", " - Thay đổi"),
        ("を確認する", " - Kiểm tra/Xác nhận"), ("を確認", " - Kiểm tra/Xác nhận"), ("の確認", " - Kiểm tra/Xác nhận"), ("確認する", " - Kiểm tra/Xác nhận"),
        ("をチェックする", " - Kiểm tra"), ("をチェック", " - Kiểm tra"), ("のチェック", " - Kiểm tra"),
        ("を検討する", " - Nghiên cứu/Khảo sát"), ("を検討", " - Nghiên cứu/Khảo sát"), ("の検討", " - Nghiên cứu/Khảo sát"), ("検討する", " - Nghiên cứu/Khảo sát"),
        ("を評価する", " - Đánh giá"), ("を評価", " - Đánh giá"), ("の評価", " - Đánh giá"), ("評価する", " - Đánh giá"),
        ("をテストする", " - Kiểm thử"), ("のテスト", " - Kiểm thử"), ("テストする", " - Kiểm thử"),
        ("をサポートする", " - Hỗ trợ"), ("をサポート", " - Hỗ trợ"), ("のサポート", " - Hỗ trợ"), ("を支援する", " - Hỗ trợ"), ("の支援", " - Hỗ trợ"),
        ("に参加する", " - Tham gia"), ("に参加", " - Tham gia"), ("への参加", " - Tham gia"), ("参加する", " - Tham gia"),
        ("を準備する", " - Chuẩn bị"), ("の准备", " - Chuẩn bị"), ("準備する", " - Chuẩn bị"),
        ("について", " (về)"), ("に関する", " (liên quan)"), ("における", " (tại)"), ("等の", " (các loại)"), ("など", " (v.v.)")
    ]
    for jp, vn in terms:
        if jp in trans:
            trans = trans.replace(jp, vn)
    trans = re.sub(r'\s+-\s+-\s+', ' - ', trans)
    trans = re.sub(r'^\s*-\s*', '', trans)
    trans = re.sub(r'\s*-\s*$', '', trans)
    trans = re.sub(r'\s+', ' ', trans).strip()
    return trans
def translate_project_name_bilingual(jp_name: str) -> str:
    if not jp_name or pd.isna(jp_name): return ""
    s = str(jp_name).strip()
    if '\n' in s:
        parts = [p.strip() for p in s.split('\n')]
        if len(parts) >= 2 and parts[1] and parts[1] != parts[0]:
            return s
        jp_part = parts[0]
        vn_part = _clean_jp_to_vn(jp_part)
        if vn_part and vn_part != jp_part:
            return f"{jp_part} \n {vn_part}"
        return s
    if '/' in s: return s
    exact_dict = {
        "機械設計・モデリング及び作図支援業務": "機械設計・モデリング及び作図支援業務 \n Nghiệp vụ hỗ trợ thiết kế cơ khí, mô hình hóa 3D & lập bản vẽ",
        "CAEシミュレーション解析・評価支援業務": "CAEシミュレーション解析・評価支援業務 \n Nghiệp vụ hỗ trợ phân tích mô phỏng CAE & đánh giá kỹ thuật",
        "電気制御設計・回路図作成及び評価業務": "電気制御設計・回路図作成及び評価業務 \n Nghiệp vụ thiết kế điện điều khiển, lập sơ đồ mạch & đánh giá",
        "ソフトウェア開発・プログラミング支援業務": "ソフトウェア開発・プログラミング支援業務 \n Nghiệp vụ hỗ trợ phát triển phần mềm & lập trình hệ thống",
        "技術ドキュメント作成・翻訳支援業務": "技術ドキュメント作成・翻訳支援業務 \n Nghiệp vụ hỗ trợ soạn thảo tài liệu kỹ thuật & dịch thuật",
        "総合技術サポート及びプロジェクト推進業務": "総合技術サポート及びプロジェクト推進業務 \n Nghiệp vụ hỗ trợ kỹ thuật tổng hợp & triển khai dự án",
        "VMOS技術開発業務委託費": "VMOS技術開発業務委託費 \n Phí ủy thác nghiệp vụ phát triển kỹ thuật VMOS",
        "3D モデリング": "3D モデリング \n Mô hình hóa 3D",
        "FM 実装ネジ締め機 5 号機 Ver.3 対応改造": "FM 実装ネジ締め機 5 号機 Ver.3 対応改造 \n Cải tiến tương thích Ver.3 máy siết ốc lắp ráp FM số 5",
        "OCV3検査装置": "OCV3検査装置 \n Thiết bị kiểm tra OCV3",
        "原料ポット外観検査機": "原料ポット外観検査機 \n Máy kiểm tra ngoại quan hũ nguyên liệu",
        "未定": "未定 \n Chưa xác định",
        "ミテイ": "ミテイ \n Chưa xác định",
        "水側リーク検査1号機": "水側リーク検査1号機 \n Máy kiểm tra rò rỉ đường nước số 1",
        "ProberUPH検討(条件41-57)": "ProberUPH検討(条件41-57) \n Nghiên cứu Prober UPH (Điều kiện 41-57)",
        "ハブ接着自動機": "ハブ接着自動機 \n Máy tự động dán Hub",
        "ARP自動検査機": "ARP自動検査機 \n Máy kiểm tra tự động ARP",
        "2D 図面化": "2D 図面化 \n Lên bản vẽ 2D",
        "図面化": "図面化 \n Lên bản vẽ",
        "ズメンカ": "ズメンカ \n Lên bản vẽ",
        "設計検討": "設計検討 \n Nghiên cứu thiết kế",
        "シミュレーション": "シミュレーション \n Mô phỏng",
        "シミュレーションセッケイ": "シミュレーションセッケイ \n Thiết kế mô phỏng",
        "シミュレーション設計": "シミュレーション設計 \n Thiết kế mô phỏng",
        "モデリング": "モデリング \n Mô hình hóa 3D",
        "構想設計": "構想設計 \n Thiết kế ý tưởng",
        "詳細設計": "詳細設計 \n Thiết kế chi tiết",
        "変更設計": "変更設計 \n Thiết kế thay đổi",
        "回路設計": "回路設計 \n Thiết kế mạch",
        "電気設計": "電気設計 \n Thiết kế điện",
        "制御設計": "制御設計 \n Thiết kế điều khiển",
        "機械設計": "機械設計 \n Thiết kế cơ khí",
        "キカイセッケイ": "キカイセッケイ \n Thiết kế cơ khí",
        "その他": "その他 \n Khác",
        "ソノタ": "ソノタ \n Khác"
    }
    if s in exact_dict:
        return exact_dict[s]
    trans = _clean_jp_to_vn(s)
    if trans and trans != s:
        return f"{s} \n {trans}"
    return s
def translate_task_bilingual(jp_task: str) -> str:
    if not jp_task or pd.isna(jp_task): return ""
    s = str(jp_task).strip()
    if '\n' in s:
        parts = [p.strip() for p in s.split('\n')]
        if len(parts) >= 2 and parts[1] and parts[1] != parts[0]:
            return s
        jp_part = parts[0]
        vn_part = _clean_jp_to_vn(jp_part)
        if vn_part and vn_part != jp_part:
            return f"{jp_part} \n {vn_part}"
        return s
    if '/' in s: return s
    exact_dict = {
        "機械設計・モデリング及び作図支援業務": "機械設計・モデリング及び作図支援業務 \n Nghiệp vụ hỗ trợ thiết kế cơ khí, mô hình hóa 3D & lập bản vẽ",
        "CAEシミュレーション解析・評価支援業務": "CAEシミュレーション解析・評価支援業務 \n Nghiệp vụ hỗ trợ phân tích mô phỏng CAE & đánh giá kỹ thuật",
        "電気制御設計・回路図作成及び評価業務": "電気制御設計・回路図作成及び評価業務 \n Nghiệp vụ thiết kế điện điều khiển, lập sơ đồ mạch & đánh giá",
        "ソフトウェア開発・プログラミング支援業務": "ソフトウェア開発・プログラミング支援業務 \n Nghiệp vụ hỗ trợ phát triển phần mềm & lập trình hệ thống",
        "技術ドキュメント作成・翻訳支援業務": "技術ドキュメント作成・翻訳支援業務 \n Nghiệp vụ hỗ trợ soạn thảo tài liệu kỹ thuật & dịch thuật",
        "総合技術サポート及びプロジェクト推進業務": "総合技術サポート及びプロジェクト推進業務 \n Nghiệp vụ hỗ trợ kỹ thuật tổng hợp & triển khai dự án",
        "製図": "製図 \n Lập bản vẽ",
        "セイズ": "セイズ \n Lập bản vẽ",
        "設計": "設計 \n Thiết kế",
        "セッケイ": "セッケイ \n Thiết kế",
        "機械設計": "機械設計 \n Thiết kế cơ khí",
        "キカイセッケイ": "キカイセッケイ \n Thiết kế cơ khí",
        "制御設計": "制御設計 \n Thiết kế điều khiển",
        "セイギョセッケイ": "セイギョセッケイ \n Thiết kế điều khiển",
        "電気設計": "電気設計 \n Thiết kế điện",
        "デンキセッケイ": "デンキセッケイ \n Thiết kế điện",
        "回路設計": "回路設計 \n Thiết kế mạch",
        "カイロセッケイ": "カイロセッケイ \n Thiết kế mạch",
        "詳細設計": "詳細設計 \n Thiết kế chi tiết",
        "ショウサイセッケイ": "ショウサイセッケイ \n Thiết kế chi tiết",
        "構想設計": "構想設計 \n Thiết kế ý tưởng",
        "コウソウセッケイ": "コウソウセッケイ \n Thiết kế ý tưởng",
        "変更設計": "変更設計 \n Thiết kế thay đổi",
        "ヘンコウセッケイ": "ヘンコウセッケイ \n Thiết kế thay đổi",
        "3Dモデリング": "3Dモデリング \n Mô hình hóa 3D",
        "3D モデリング": "3D モデリング \n Mô hình hóa 3D",
        "モデリング": "モデリング \n Mô hình hóa 3D",
        "2D図面化": "2D図面化 \n Lên bản vẽ 2D",
        "2D 図面化": "2D 図面化 \n Lên bản vẽ 2D",
        "図面化": "図面化 \n Lên bản vẽ",
        "ズメンカ": "ズメンカ \n Lên bản vẽ",
        "図面修正": "図面修正 \n Sửa bản vẽ",
        "修正": "修正 \n Chỉnh sửa",
        "シュウセイ": "シュウセイ \n Chỉnh sửa",
        "シミュレーション": "シミュレーション \n Mô phỏng",
        "シミュレーションセッケイ": "シミュレーションセッケイ \n Thiết kế mô phỏng",
        "シミュレーション設計": "シミュレーション設計 \n Thiết kế mô phỏng",
        "解析": "解析 \n Phân tích",
        "カイセキ": "カイセキ \n Phân tích",
        "CAE解析": "CAE解析 \n Phân tích CAE",
        "評価": "評価 \n Đánh giá",
        "ヒョウカ": "ヒョウカ \n Đánh giá",
        "検査": "検査 \n Kiểm tra",
        "ケンサ": "ケンサ \n Kiểm tra",
        "検討": "検討 \n Nghiên cứu / Thẩm định",
        "ケントウ": "ケントウ \n Nghiên cứu / Thẩm định",
        "翻訳": "翻訳 \n Dịch thuật",
        "通訳": "通訳 \n Thông dịch",
        "ドキュメント作成": "ドキュメント作成 \n Soạn thảo tài liệu",
        "資料作成": "資料作成 \n Soạn thảo tài liệu",
        "サポート": "サポート \n Hỗ trợ",
        "その他": "その他 \n Khác",
        "ソノタ": "ソノタ \n Khác",
        "打合せ": "打合せ \n Họp / Thảo luận",
        "打ち合わせ": "打ち合わせ \n Họp / Thảo luận",
        "ウチアワセ": "ウチアワセ \n Họp / Thảo luận",
        "会議": "会議 \n Họp",
        "カイギ": "カイギ \n Họp",
        "シミュレーションを作成する": "シミュレーションを作成する \n Xây dựng mô phỏng",
        "シミュレーションを作成": "シミュレーションを作成 \n Xây dựng mô phỏng",
        "シミュレーションの作成": "シミュレーションの作成 \n Xây dựng mô phỏng"
    }
    if s in exact_dict:
        return exact_dict[s]
        
    # Bỏ qua dịch tự động cho các chuỗi quá dài hoặc chứa dấu phẩy (nhiều tác vụ) để tránh lỗi mix ngôn ngữ lộn xộn
    if len(s) > 40 or ',' in s or '、' in s:
        return s
        
    trans = _clean_jp_to_vn(s)
    if trans and trans != s:
        return f"{s} \n {trans}"
    return s


def convert_emails_to_df(dict_latest_emails):
    import datetime
    records = []
    for (mail_nv_key, d_str), mail_rep in dict_latest_emails.items():
        ma_nv = mail_rep.get('ma_nv') or mail_nv_key
        ten_nv = mail_rep.get('ten_nv') or mail_nv_key
        
        try:
            m, d = map(int, d_str.split('/'))
            file_year = mail_rep['timestamp'].year
            dt_obj = datetime.date(file_year, m, d)
            date_fmt = dt_obj.strftime('%d/%m/%Y')
        except:
            file_year = 2026
            date_fmt = d_str
            
        projs = mail_rep.get('projects', {})
        mgrs = mail_rep.get('project_managers', {})
        sects = mail_rep.get('project_sections', {})
        project_names = mail_rep.get('project_names', {})
        
        for p_code, p_hrs in projs.items():
            if p_hrs <= 0: continue
            
            sec = sects.get(p_code, 'UNKNOWN')
            is_internal_m = p_code.startswith('M') and 'MOS' not in p_code
            is_internal_j = p_code.startswith('J') and 'JMOS' not in p_code
            
            if sec == 'MOS':
                is_mos = True
            elif sec == 'UNKNOWN':
                is_mos = not (is_internal_m or is_internal_j)
            else:
                is_mos = False
                
            if 'NOCODE' in p_code:
                is_mos = True
            
            if p_code in ['MOS', 'JMOS', 'VMOS', 'SHANAI']:
                p_code = 'K000000' # Dummy code
                
            records.append({
                'ma_nv': ma_nv,
                'ten_nv': ten_nv,
                'ma_da': p_code,
                'khach': '',
                'ten_da': project_names.get(p_code, '(Từ Mail)'),
                'phan_vung': 'Email',
                'task': 'Báo cáo từ Mail Thunderbird',
                'tong_gio': p_hrs,
                'ngay_bat_dau': date_fmt,
                'ngay_ket_thuc': date_fmt,
                'ql_nhat': mgrs.get(p_code, ''),
                'chi_tiet_ngay': {d_str: p_hrs},
                'all_dates': [d_str],
                'file_year': file_year,
                'gio_mos_jmos_vmos': {d_str: mail_rep.get('total_hours', p_hrs)},
                'is_mos': is_mos
            })
    return pd.DataFrame(records)

def tong_hop_mos(dfs: list) -> pd.DataFrame:
    """
    Gộp nhiều DataFrame từ nhiều file, tổng hợp theo mã dự án.
    """
    if not dfs:
        return pd.DataFrame()
    df_all = pd.concat(dfs, ignore_index=True)
    if 'is_mos' in df_all.columns:
        df_all = df_all[df_all['is_mos'] == True].copy()
    import streamlit as st
    # Tự động tìm và thay thế mã tạm K000000 sang mã thực tế từ Mail (đối chiếu theo Tên quản lý Nhật Bản)
    # [Tạm vô hiệu hóa] Theo yêu cầu: giữ nguyên mã K000000 vì đây là 2 dự án khác nhau
    # dict_latest_emails = st.session_state.get('mos_latest_email_reports', {})
    # if dict_latest_emails and not df_all.empty and 'ma_da' in df_all.columns:
    #     used_real_by_emp = {}
    #     for idx, row in df_all.iterrows():
    #         ma_da_row = str(row.get('ma_da', '')).strip().upper()
    #         if is_dummy_project_code(ma_da_row):
    #             ma_nv = str(row.get('ma_nv', '')).strip()
    #             mgr_ex = str(row.get('ql_nhat') or row.get('tanto') or row.get('khach') or '').strip()
    #             if ma_nv not in used_real_by_emp:
    #                 used_real_by_emp[ma_nv] = set()
    #             real_code = resolve_dummy_code_with_mail(ma_nv, mgr_ex, dict_latest_emails, used_real_by_emp[ma_nv], ten_nv=str(row.get('ten_nv', '')).strip(), target_hours=row.get('tong_gio'))
    #             if real_code:
    #                 df_all.at[idx, 'ma_da'] = real_code
    #                 used_real_by_emp[ma_nv].add(real_code)
    result = []
    # 1. Tính tổng số giờ làm theo từng ngày của từng nhân viên để phát hiện ngày làm bất thường (> 8h/ngày hoặc 0h vào ngày T2-T6)
    emp_daily_hours = {} # (ten_nv, date_str) -> total_hours (MOS)
    emp_daily_projects = {} # (ten_nv, date_str) -> set of ma_da
    emp_all_dates = {} # ten_nv -> set of date_str
    emp_years = {} # ten_nv -> year
    emp_all_work_hours = {} # (ten_nv, date_str) -> tổng giờ làm MOS + JMOS + VMOS
    for _, r_all in df_all.iterrows():
        t_nv = str(r_all.get('ten_nv', '')).strip()
        m_da = str(r_all.get('ma_da', '')).strip().upper()
        c_ngay = r_all.get('chi_tiet_ngay', {})
        f_dates = r_all.get('all_dates', [])
        f_yr = r_all.get('file_year', 2026)
        g_all = r_all.get('gio_mos_jmos_vmos', {})
        if t_nv not in emp_all_dates:
            emp_all_dates[t_nv] = set()
            emp_years[t_nv] = f_yr
        for d_s in f_dates:
            emp_all_dates[t_nv].add(d_s)
        if isinstance(g_all, dict):
            for d_s, h_val in g_all.items():
                key_w = (t_nv, d_s)
                emp_all_work_hours[key_w] = max(emp_all_work_hours.get(key_w, 0.0), float(h_val))
                emp_all_dates[t_nv].add(d_s)
        if isinstance(c_ngay, dict):
            for d_str, hrs in c_ngay.items():
                if hrs > 0:
                    key = (t_nv, d_str)
                    emp_daily_hours[key] = emp_daily_hours.get(key, 0.0) + float(hrs)
                    if key not in emp_daily_projects:
                        emp_daily_projects[key] = set()
                    emp_daily_projects[key].add(m_da)
                emp_all_dates[t_nv].add(d_str)
    def get_dt_obj(d_str, year):
        try:
            parts = str(d_str).strip().split('/')
            if len(parts) == 2:
                return datetime.date(year, int(parts[0]), int(parts[1]))
        except Exception: pass
        return None
    def get_vn_weekday_name(dt_obj):
        if not dt_obj: return ""
        w_map = {0: "Thứ 2", 1: "Thứ 3", 2: "Thứ 4", 3: "Thứ 5", 4: "Thứ 6", 5: "Thứ 7", 6: "Chủ nhật"}
        return w_map.get(dt_obj.weekday(), "")
    # Không còn kiểm tra ngày bất thường theo yêu cầu người dùng
    st.session_state['mos_abnormal_days'] = []


    def get_group_key(row):
        ma_da_row = str(row.get('ma_da', '')).strip().upper()
        ma_da_row = re.sub(r'_(未定|CHƯA XÁC ĐỊNH|CHUA XAC DINH)$', '', ma_da_row, flags=re.I)
        ma_da_out = ma_da_row
        if is_dummy_project_code(ma_da_row) and 'NOCODE' not in ma_da_row:
            ma_da_out = 'K000000'
        mgr = str(row.get('ql_nhat') or row.get('khach') or row.get('tanto') or '').strip()
        return f"{ma_da_out}___{mgr}"
        
    df_all['_group_key'] = df_all.apply(get_group_key, axis=1)
    groups = list(df_all.groupby('_group_key'))
    total_groups = len(groups)
    
    if total_groups == 0:
        return pd.DataFrame()
        
    progress_bar = st.progress(0)
    status_text = st.empty()
    is_vi = st.session_state.get('lang', 'vi') == 'vi'
    if is_vi:
        status_text.markdown(f"**🤖 AI đang dịch và tóm tắt {total_groups} dự án MOS...**")
    else:
        status_text.markdown(f"**🤖 AIが {total_groups} 件のMOSプロジェクトを翻訳・要約しています...**")
    
    def filter_and_split_tasks(t_list, ma_da_val):
        raw_tasks = []
        for t in t_list:
            t_str = str(t).strip()
            if not t_str or t_str.lower() in ['nan', 'none'] or 'タスクは指定されていません' in t_str:
                continue
            
            parts = re.split(r'[,/]', t_str)
            for p in parts:
                p = p.strip()
                if not p: continue
                if str(ma_da_val).strip().upper() == 'P012004':
                    # Loại bỏ cả các task học tập simulation/CAD/modeling liên quan đến triển lãm
                    bad_keywords = ['triển lãm', 'trien lam', '展示', 'シミュレーション', 'モデリング', 'realvirtual', 'cadデータ']
                    if any(k in p.lower() for k in bad_keywords):
                        continue
                raw_tasks.append(p)
        return list(dict.fromkeys(raw_tasks))

    projects_batch = []
    for g_key, grp in groups:
        ma_da_real = str(grp['ma_da'].iloc[0]).strip().upper()
        ma_da_real = re.sub(r'_(未定|CHƯA XÁC ĐỊNH|CHUA XAC DINH)$', '', ma_da_real, flags=re.I)
        ten_da_val = grp['ten_da'].iloc[0]
        
        if ma_da_real in ['K000000', '000000']:
            continue
            
        known_names = {
            'K230093': 'ARP自動検査機',
            'K230062': 'ハブ接着自動機',
            'K220141': '原料ポット外観検査機',
            'K210361': 'OCV3検査装置',
            'K230059': 'FM 実装ネジ締め機 5 号機 Ver.3 対応改造'
        }
        if ma_da_real in known_names and (pd.isna(ten_da_val) or str(ten_da_val).strip() in ['未定', 'Chưa xác định', '', '(Từ Mail)', 'ミテイ']):
            ten_da_val = known_names[ma_da_real]
            
        valid_tasks_u = filter_and_split_tasks(grp['task'].tolist(), ma_da_real)
        nv_u = grp[grp['tong_gio'] > 0]['ten_nv'].unique()
        tasks_to_send = valid_tasks_u
        projects_batch.append({'ma_da': g_key, 'ten_da': ten_da_val, 'tasks': tasks_to_send})
        
    ai_batch_res = batch_summarize_projects(projects_batch)
    
    for idx, (g_key, grp) in enumerate(groups):
        ma_da = str(grp['ma_da'].iloc[0]).strip().upper()
        ma_da = re.sub(r'_(未定|CHƯA XÁC ĐỊNH|CHUA XAC DINH)$', '', ma_da, flags=re.I)
        if ma_da in ['K000000', '000000']:
            continue
            
        if is_vi:
            status_text.markdown(f"**🤖 Đang tổng hợp dự án {idx+1}/{total_groups}:** `{ma_da}`")
        else:
            status_text.markdown(f"**🤖 プロジェクトを集計中 {idx+1}/{total_groups}:** `{ma_da}`")
        ngay_bd_list = [datetime.datetime.strptime(d, '%d/%m/%Y').date() 
                        for d in grp['ngay_bat_dau'] if d]
        ngay_kt_list = [datetime.datetime.strptime(d, '%d/%m/%Y').date() 
                        for d in grp['ngay_ket_thuc'] if d]
        ngay_bd = min(ngay_bd_list).strftime('%d/%m/%Y') if ngay_bd_list else ''
        ngay_kt = max(ngay_kt_list).strftime('%d/%m/%Y') if ngay_kt_list else ''
        if not ngay_kt and ngay_bd:
            ngay_kt = ngay_bd
            
        # Theo yêu cầu người dùng: Đảm bảo dự án K000000 của Matsui hiển thị đúng ngày kết thúc là 19/6
        if str(ma_da).strip().upper() == 'K000000' and any(k in str(grp.get('ql_nhat', '')).lower() + str(grp.get('ten_da', '')).lower() for k in ['松井', 'matsui', 'mô phỏng', 'simulation']):
            if not ngay_kt or '19/06' not in ngay_kt:
                year_str = ngay_bd.split('/')[-1] if ngay_bd and '/' in ngay_bd else '2024'
                ngay_kt = f"19/06/{year_str}"
                if not ngay_bd:
                    ngay_bd = ngay_kt
        
        nv_co_gio = grp[grp['tong_gio'] > 0]['ten_nv'].unique()
        staff_map = {
            'ダンフン': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng'), 'フン': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng'), 'hưng': ('グエン・ダン・フン', 'Nguyễn Đăng Hưng'),
            'クアン': ('ファム・ホン・クアン', 'Phạm Hồng Quân'), 'quân': ('ファム・ホン・クアン', 'Phạm Hồng Quân'),
            'ハン': ('レ・ティ・カイン・ハン', 'Lê Thị Khánh Hằng'), 'hằng': ('レ・ティ・カイン・ハン', 'Lê Thị Khánh Hằng'),
            'グエット': ('グエン・ミン・グエット', 'Nguyễn Minh Nguyệt'), 'nguyệt': ('グエン・ミン・グエット', 'Nguyễn Minh Nguyệt'),
            'ロン': ('レ・ヴァン・ロン', 'Lê Văn Long'), 'long': ('レ・ヴァン・ロン', 'Lê Văn Long'),
            'ダオ': ('ハ・ヴァン・ダオ', 'Hà Văn Đạo'), 'đạo': ('ハ・ヴァン・ダオ', 'Hà Văn Đạo'),
            'フォン': ('レ・タイン・フォン', 'Lê Thanh Phương'), 'phương': ('レ・タイン・フォン', 'Lê Thanh Phương'),
            'ドゥ': ('ファム・ゴック・ドゥ', 'Phạm Ngọc Dư'), 'dư': ('ファム・ゴック・ドゥ', 'Phạm Ngọc Dư'),
        }
        
        valid_tasks_unique = filter_and_split_tasks(grp['task'].tolist(), ma_da)
        first_task = valid_tasks_unique[0] if valid_tasks_unique else ""
        
        ten_da_val = grp['ten_da'].iloc[0]
        if ma_da in known_names and (pd.isna(ten_da_val) or str(ten_da_val).strip() in ['未定', 'Chưa xác định', '', '(Từ Mail)', 'ミテイ']):
            ten_da_val = known_names[ma_da]
        
        ai_res = ai_batch_res.get(g_key, {})
        if isinstance(ai_res, dict) and ('ten_da_song_ngu' in ai_res or 'noi_dung_song_ngu' in ai_res):
            ten_da_song_ngu = ai_res.get('ten_da_song_ngu', ten_da_val)
            noi_dung = ai_res.get('noi_dung_song_ngu', first_task) if ai_res.get('noi_dung_song_ngu') else first_task
        elif isinstance(ai_res, str) and ai_res:
            ten_da_song_ngu = ten_da_val
            noi_dung = ai_res
        else:
            ten_da_song_ngu = ten_da_val
            noi_dung = first_task
            
        if len(valid_tasks_unique) > 1 and (' / ' in str(noi_dung) or len(str(noi_dung).split(' / ')) > 1):
            noi_dung = smart_offline_summarize(ma_da, ten_da_val, valid_tasks_unique, grp['phan_vung'].iloc[0])
            
        ten_da_song_ngu = translate_project_name_bilingual(ten_da_song_ngu)
            
        # Đảm bảo dịch song ngữ cho tất cả nội dung ủy thác (kể cả khi chỉ có 1 task)
        noi_dung = translate_task_bilingual(noi_dung)
        
        # Override tóm tắt song ngữ cho P012004 theo đúng ý người dùng (ngắn gọn, chuẩn xác)
        if str(ma_da).strip().upper() == 'P012004':
            noi_dung = "Unity版スマートデバッグの開発と確認 \n Phát triển và xác nhận Smart Debug bản Unity"
            ten_da_song_ngu = "VMOS技術開発業務委託費 \n Chi phí ủy thác phát triển công nghệ VMOS"
        
        if not valid_tasks_unique:
            noi_dung = ""
            
        phan_vung = grp['phan_vung'].iloc[0]
        
        raw_staff = []
        for s in nv_co_gio:
            if s and pd.notna(s):
                for part in re.split(r'[/,&]', str(s)):
                    p = part.strip()
                    if p and p not in raw_staff:
                        raw_staff.append(p)
                        
        jp_list = []
        vn_list = []
        for p in sorted(raw_staff):
            p_low = p.lower()
            if p in staff_map:
                j, v = staff_map[p]
            elif p_low in staff_map:
                j, v = staff_map[p_low]
            else:
                j, v = p, p
            j_short = j
            if isinstance(j, str):
                if '・' in j:
                    j_short = j.split('・')[-1]
                elif ' ' in j:
                    j_short = j.split()[-1]
            if j_short not in jp_list: jp_list.append(j_short)
            
            v_short = v.split()[-1] if isinstance(v, str) and v.strip() else v
            if v_short not in vn_list: vn_list.append(v_short)

        ql_viet_nam = None
        ma_da_clean = str(ma_da).strip().lower()
        pv_goc = str(phan_vung).strip()
        pv_lower = pv_goc.lower()
        txt_check_grp = f"{ma_da} {pv_goc} {ten_da_val}".lower()

        # Phân vùng theo từ khóa chuẩn xác (mã dự án, phân vùng, tên dự án, nội dung)
        if any(k in txt_check_grp for k in ['cơ khí', '機械', 'メカ', 'meco', 'machine', '機設計', 'thiết kế máy', 'cấu trúc', 'j01009', 'k230059']):
            phan_vung = "機械設計 \n Thiết kế cơ khí"
            ql_viet_nam = "ロン \n Long"
        elif any(k in txt_check_grp for k in ['mô phỏng', 'simulation', 'cae', 'シミュレーション', 'sim', 'phân tích', 'ansys', '解析', 'p012004', 'vmos', 'k000000', '松井', 'matsui', '石部', 'ishibe']):
            phan_vung = "シミュレーション設計 \n Thiết kế mô phỏng"
            ql_viet_nam = "フォン \n Phương"
        elif any(k in txt_check_grp for k in ['điện', 'điều khiển', '電気', '制御', 'elec', 'control', 'plc', 'mạch', 'pcb', 'hardware', 'phần cứng', 'cảm biến', 'sensor', 'motor', 'động cơ', 'panel', 'tủ điện', 'bảng điện', 'inverter', 'biến tần', 'circuit', 'schematic', 'ee', 'hw', 'io', 'dây', 'wiring', 'k210361', 'ocv', 'k220141', '原料', '外観', 'ngoại quan', 'bề mặt', 'pot', 'イニシャル', 'initial', 'đạo', 'ダオ']):
            phan_vung = "制御設計 \n Thiết kế điện điều khiển"
            ql_viet_nam = "ダオ \n Đạo"
        else:
            # Nếu không khớp từ khóa đặc biệt nào, phân loại dựa theo từ khóa trong phân vùng gốc từ file
            if 'cơ khí' in pv_lower or '機械' in pv_lower:
                phan_vung = "機械設計 \n Thiết kế cơ khí"
                ql_viet_nam = "ロン \n Long"
            elif 'điện' in pv_lower or '制御' in pv_lower or '電気' in pv_lower:
                phan_vung = "制御設計 \n Thiết kế điện điều khiển"
                ql_viet_nam = "ダオ \n Đạo"
            elif 'mô phỏng' in pv_lower or 'シミュレーション' in pv_lower:
                phan_vung = "シミュレーション設計 \n Thiết kế mô phỏng"
                ql_viet_nam = "フォン \n Phương"
            else:
                phan_vung = f"{pv_goc} \n {pv_goc}" if pv_goc else "未指定 \n Chưa xác định"
                ql_viet_nam = ""

        # Theo yêu cầu: Quản lý cơ khí (Long) và điện (Đạo) kiêm luôn người thực hiện
        # Riêng quản lý mô phỏng (Phương) thì không trực tiếp làm, chỉ quản lý
        if "機械設計" in phan_vung:
            j_long, v_long = staff_map.get('long', ('レ・ヴァン・ロン', 'Lê Văn Long', 'VM011'))[:2]
            j_long_short = j_long.split('・')[-1] if isinstance(j_long, str) and '・' in j_long else j_long
            if j_long_short not in jp_list: jp_list.append(j_long_short)
            v_long_short = v_long.split()[-1] if isinstance(v_long, str) and v_long.strip() else v_long
            if v_long_short not in vn_list: vn_list.append(v_long_short)
        elif "制御設計" in phan_vung:
            j_dao, v_dao = staff_map.get('đạo', ('ハ・ヴァン・ダオ', 'Hà Văn Đạo', 'VM008'))[:2]
            j_dao_short = j_dao.split('・')[-1] if isinstance(j_dao, str) and '・' in j_dao else j_dao
            if j_dao_short not in jp_list: jp_list.append(j_dao_short)
            v_dao_short = v_dao.split()[-1] if isinstance(v_dao, str) and v_dao.strip() else v_dao
            if v_dao_short not in vn_list: vn_list.append(v_dao_short)
        elif "シミュレーション設計" in phan_vung:
            # Phương chỉ quản lý, KHÔNG trực tiếp thực hiện dự án mô phỏng
            jp_list = [p for p in jp_list if 'フォン' not in p]
            vn_list = [v for v in vn_list if 'phương' not in v.lower() and 'phuong' not in v.lower()]

        if jp_list and jp_list != vn_list:
            nguoi_th = ' / '.join(jp_list) + ' \n ' + ' / '.join(vn_list)
        elif jp_list:
            nguoi_th = ' / '.join(jp_list)
        elif vn_list:
            nguoi_th = ' / '.join(vn_list)
        else:
            nguoi_th = ''
            
        ql_nhat_goc = str(grp['ql_nhat'].iloc[0]).strip() if not grp['ql_nhat'].empty and pd.notna(grp['ql_nhat'].iloc[0]) else ''
        jp_manager_dict = {
            '大滝': 'Otaki', 'オオタキ': 'Otaki',
            '松井': 'Matsui', 'マツイ': 'Matsui',
            '室伏': 'Murofushi', 'ムロフシ': 'Murofushi',
            '和田': 'Wada', 'ワダ': 'Wada',
            '田代': 'Tashiro', 'タシロ': 'Tashiro',
            '石部': 'Ishibe', 'イシベ': 'Ishibe',
            '勝亦': 'Katsumata', 'カツマタ': 'Katsumata',
            '池谷': 'Ikeya', 'イケヤ': 'Ikeya',
            '井手上': 'Ideue', 'イデウエ': 'Ideue',
            '山崎': 'Yamazaki', 'ヤマザキ': 'Yamazaki',
            '伏見': 'Fushimi', 'フシミ': 'Fushimi',
            'フォン': 'Phương', 'ロン': 'Long', 'ダオ': 'Đạo'
        }
        ql_nhat_trans = jp_manager_dict.get(ql_nhat_goc, '')
        if not ql_nhat_trans:
            for k, v in jp_manager_dict.items():
                if k in ql_nhat_goc:
                    ql_nhat_trans = v
                    break
        if ql_nhat_trans and ql_nhat_trans != ql_nhat_goc:
            ql_nhat_song_ngu = f"{ql_nhat_goc} \n {ql_nhat_trans}"
        else:
            ql_nhat_song_ngu = ql_nhat_goc
        
        tong_gio_goc = round(grp['tong_gio'].sum(), 2)
        tong_gio_val = tong_gio_goc
        
        # Theo yêu cầu của người dùng: Ghi đè cứng số giờ cho các dự án cụ thể (do khác biệt với dữ liệu thô)
        ma_da_upper_check = str(ma_da).strip().upper()
        if ma_da_upper_check == 'P012004':
            tong_gio_val = 149.5
            tong_gio_goc = 149.5
        elif ma_da_upper_check == 'K220596' or ma_da_upper_check == '220596':
            tong_gio_val = 67.5
            tong_gio_goc = 67.5
        elif ma_da_upper_check in ['K230062', '230062', 'K230063', '230063'] or any(k in str(ten_da_song_ngu).lower() for k in ['háp', 'hub', 'dán hub', 'ハブ接着']):
            tong_gio_val = 177.0
            tong_gio_goc = 177.0
        elif ma_da_upper_check == 'K220141' or ma_da_upper_check == '220141':
            tong_gio_val = 91.0
            tong_gio_goc = 91.0
        elif ma_da_upper_check == 'K230093' or ma_da_upper_check == '230093':
            tong_gio_val = 22.0
            tong_gio_goc = 22.0
                
        ma_da_upper = str(ma_da).strip().upper()
        is_j_code = ma_da_upper.startswith('J')
        
        warnings_list = []
        if is_j_code:
            warnings_list.append(f"⚠️ Mã J (JMOS) - Không thuộc bảng MOS (gốc: {tong_gio_goc}h)")
            tong_gio_val = 0.0 # Không ghi nhận giờ làm cho mã J trên bảng UI
            

            
        canh_bao_str = "🔴" if warnings_list else ""
                
        result.append({
            'Cảnh báo': canh_bao_str,
            'Mã dự án': ma_da,
            'Tên dự án': ten_da_song_ngu,
            'Phân vùng': phan_vung,
            'Nội dung ủy thác': noi_dung,
            'Giờ làm (h)': tong_gio_val,
            'Giờ làm gốc (h)': tong_gio_goc,
            'Ngày bắt đầu': ngay_bd,
            'Ngày kết thúc': ngay_kt,
            'Quản lý Nhật Bản': ql_nhat_song_ngu,
            'Quản lý Việt Nam': ql_viet_nam,
            'Người thực hiện': nguoi_th,
            'Trạng thái': None,
        })
        progress_bar.progress((idx + 1) / total_groups)
    
    status_text.empty()
    progress_bar.empty()
    
    df_result = pd.DataFrame(result)
    if not df_result.empty and 'Phân vùng' in df_result.columns:
        def get_pv_order(pv_str):
            pv = str(pv_str).lower()
            if 'cơ khí' in pv or '機械' in pv or 'メカ' in pv: return 1
            elif 'điện' in pv or 'điều khiển' in pv or '制御' in pv or '電気' in pv: return 2
            elif 'mô phỏng' in pv or 'simulation' in pv or 'cae' in pv or 'シミュレーション' in pv: return 3
            else: return 4
        df_result['_pv_order'] = df_result['Phân vùng'].apply(get_pv_order)
        df_result = df_result.sort_values(by=['_pv_order', 'Mã dự án'], ascending=[True, True]).drop(columns=['_pv_order']).reset_index(drop=True)
        
    df_result.insert(0, 'STT', range(1, len(df_result)+1))
    return df_result


def render_mos_page():
    t = get_t(st.session_state.get('lang', 'vi'))
    is_vi = st.session_state.get('lang', 'vi') == 'vi'
    lbl_ph = """Ví dụ:
Báo cáo ngày 05/06 - VM038 Nguyễn Minh Nguyệt
1. K230059: 4h
2. J01009: 4h
...""" if is_vi else """例
日報 05/06 - VM038 グエン・ミン・グエット
1. K230059: 4h
2. J01009: 4h
..."""
    # 8. GIAO DIỆN XỬ LÝ MOS
    theme_mode = st.session_state.get('theme_mode', 'light')
    gT = get_theme(theme_mode)
    is_sepia = (theme_mode == 'sepia')

    # ── Bảng theme trung tâm – dùng chung toàn page_mos ───────────────────
    T = {
        "bg_header":     f"{gT['bg_card']}E6",
        "bg_card":       f"{gT['bg_card']}F2",
        "bg_section":    gT['bg_app'],
        "bg_input":      f"{gT['bg_content']}E6",
        "bg_tag":        f"{gT['primary']}1A",
        "bg_result":     gT['bg_app'],
        "bg_upload":     gT['bg_app'],
        "bg_step_active":f"{gT['primary']}1A",
        "border_header": gT['border'],
        "border_card":   gT['border'],
        "border_input":  gT['border'],
        "border_upload": gT['primary'],
        "border_result": gT['border'],
        "text_h1":       gT['text_primary'],
        "text_h2":       gT['text_primary'],
        "text_body":     gT['text_secondary'],
        "text_muted":    gT['text_tertiary'],
        "text_accent":   gT['primary'],
        "text_label":    gT['text_secondary'],
        "shadow_header": gT['shadow'],
        "shadow_card":   gT['shadow'],
        "accent_btn":    gT['primary'],
        "accent_btn_hover": gT['primary_hover'],
        "accent_border_focus": gT['primary'],
        "pulse_color":   gT['primary'],
    }
    # ──────────────────────────────────────────────────────────────────────

    st.markdown(f"""
    <style>
    /* ── Main Container ── */
    .mos-main-card {{
        /* Bỏ khung trắng theo yêu cầu */
        padding: 0;
    }}

    /* ── MOS Page Header (Banner) ── */
    
    .mos-header {{
        background: {gT['bg_card']}E6;
        backdrop-filter: blur(24px);
        -webkit-backdrop-filter: blur(24px);
        border: 1.5px solid {gT['border']};
        border-radius: 20px;
        padding: 32px 40px;
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-top: -90px !important;
        margin-bottom: 32px;
        box-shadow: 0 16px 40px {gT['shadow']};
        position: relative;
        overflow: hidden;
    }}
    
    .mos-header::before {{
        content: '';
        position: absolute;
        top: -50%;
        right: -10%;
        width: 300px;
        height: 300px;
        background: radial-gradient(circle, {'rgba(217, 119, 6, 0.1)' if is_sepia else 'rgba(14, 165, 233, 0.15)'} 0%, rgba(255,255,255,0) 70%);
        border-radius: 50%;
    }}
    .mos-header-left {{
        display: flex;
        gap: 20px;
        align-items: center;
        position: relative;
        z-index: 2;
    }}
    .mos-hero-title {{
        font-size: 28px;
        font-weight: 800;
        color: {gT['text_primary']};
        margin: 0 0 8px 0;
        line-height: 1.2;
        letter-spacing: -0.02em;
        {"" if is_sepia else "background: linear-gradient(135deg, #0EA5E9 0%, #0284C7 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent;"}
    }}
    .mos-hero-sub {{
        font-size: 15px;
        color: {gT['text_secondary']};
        margin: 0;
        font-weight: 500;
    }}
    .mos-hero-badge {{
        background: {'rgba(217, 119, 6, 0.1)' if is_sepia else 'linear-gradient(135deg, #0EA5E9 0%, #0284C7 100%)'};
        box-shadow: 0 4px 12px {'rgba(217, 119, 6, 0.15)' if is_sepia else 'rgba(14, 165, 233, 0.25)'};
        padding: 8px 18px;
        border-radius: 20px;
        color: {gT['text_primary']};
        font-size: 13px;
        font-weight: 700;
        letter-spacing: 0.05em;
        border: none;
        position: relative;
        z-index: 2;
    }}

    /* ── Stepper (Có nền trắng kính mờ để nổi bật trên nền Sakura) ── */
    .mos-stepper {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        margin-bottom: 24px;
        padding: 18px 28px;
        background: {gT['bg_card']}F2;
        backdrop-filter: blur(16px);
        border-radius: 18px;
        border: 1.5px solid {'rgba(217, 119, 6, 0.25)' if is_sepia else 'rgba(14, 165, 233, 0.25)'};
        box-shadow: 0 4px 18px rgba(0, 0, 0, 0.06);
    }}
    .mos-step-item {{
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 8px;
        position: relative;
        z-index: 1;
    }}
    .mos-step-circle {{
        width: 30px;
        height: 30px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 13px;
        font-weight: 700;
        background: {gT['bg_content']};
    }}
    .step-active .mos-step-circle {{
        border: 2.5px solid {gT['primary']};
        color: {gT['primary']};
        box-shadow: 0 0 0 4px {gT['primary']}2E;
        animation: pulse 2s infinite;
    }}
    @keyframes pulse {{
        0% {{ box-shadow: 0 0 0 0 {gT['primary']}59; }}
        70% {{ box-shadow: 0 0 0 10px {gT['primary']}00; }}
        100% {{ box-shadow: 0 0 0 0 {gT['primary']}00; }}
    }}
    .step-inactive .mos-step-circle {{
        border: 2px solid {gT['border']};
        color: {gT['text_secondary']};
    }}
    .step-done .mos-step-circle {{
        background: {gT['primary']};
        border: 2px solid {gT['primary']};
        color: white;
    }}
    .mos-step-label {{
        font-size: 13.5px;
        font-weight: 800;
        margin-top: 4px;
    }}
    .step-active .mos-step-label {{ color: {gT['primary']}; }}
    .step-inactive .mos-step-label {{ color: {gT['text_secondary']}; }}
    .step-done .mos-step-label {{ color: {gT['primary']}; }}
    
    .mos-stepper-line {{
        flex: 1;
        height: 2px;
        background: {gT['border']};
        margin: 0 16px;
        margin-bottom: 24px; /* offset to align with circles */
    }}

    /* ── Upload zone override (Glassmorphic & Responsive) ── */
    [data-testid="stFileUploader"] section {{
        background: {gT['bg_card']} !important;
        border: 2px dashed {gT['border']} !important;
        border-radius: clamp(16px, 2vw, 22px) !important;
        padding: clamp(24px, 3vw, 42px) clamp(16px, 2vw, 28px) !important;
        text-align: center !important;
        box-shadow: {gT['shadow']} !important;
        transition: all 0.3s ease !important;
    }}
    [data-testid="stFileUploader"] section:hover {{
        border-color: {gT['primary']} !important;
        background: {gT['bg_content']} !important;
        box-shadow: {gT['shadow_glow']} !important;
        transform: translateY(-2px) !important;
    }}
    [data-testid="stFileUploader"] section > div > div > p {{
        color: {gT['text_primary']} !important;
        font-size: clamp(15px, 1.3vw, 17px) !important;
        font-weight: 700 !important;
        font-family: 'Plus Jakarta Sans', 'Inter', sans-serif !important;
        margin-bottom: 4px !important;
    }}
    [data-testid="stFileUploaderDropzone"] svg {{
        color: {gT['primary']} !important;
        width: clamp(34px, 3.5vw, 48px) !important;
        height: clamp(34px, 3.5vw, 48px) !important;
        margin-bottom: clamp(6px, 1vw, 10px) !important;
    }}

    .mos-upload-header-box {{
        background: {gT['bg_card']}E6;
        backdrop-filter: blur(16px);
        border: 1px solid {gT['border']};
        border-radius: 16px;
        padding: 16px 20px;
        margin-bottom: 16px;
        box-shadow: {gT['shadow']};
    }}
    /* Nâng cấp hộp container chung chứa toàn bộ Dashboard tải file MOS */
    div[data-testid="stVerticalBlockBorderWrapper"] {{
        background: {gT['bg_card']}FA !important;
        backdrop-filter: blur(24px) !important;
        border: 1.5px solid {gT['border']} !important;
        border-radius: 24px !important;
        padding: 24px 28px !important;
        box-shadow: 0 20px 50px rgba(0, 0, 0, 0.12) !important;
        margin-bottom: 24px !important;
        position: relative !important;
        z-index: 10 !important;
    }}
    div[data-testid="stVerticalBlockBorderWrapper"] > div[data-testid="stVerticalBlock"] {{
        background: transparent !important;
        padding: 0 !important;
        position: relative !important;
        z-index: 1 !important;
    }}
    /* Khi đặt bên trong container chung, loại bỏ viền lặp lại để liền mạch trọn vẹn với uploader bên dưới */
    .mos-upload-header-box-clean {{
        background: transparent !important;
        border: none !important;
        padding: 0 0 16px 0 !important;
        margin-bottom: 16px !important;
        border-bottom: 1px dashed {gT['border']} !important;
        box-shadow: none !important;
        position: relative !important;
        z-index: 2 !important;
    }}
    .mos-upload-box-title {{
        font-size: clamp(16px, 1.5vw, 19px);
        font-weight: 800;
        color: {gT['text_primary']};
        margin-bottom: 6px;
        display: flex;
        align-items: center;
        gap: 8px;
    }}
    .mos-upload-box-sub {{
        font-size: clamp(12.5px, 1.1vw, 14px);
        color: {gT['text_secondary']};
        line-height: 1.45;
        margin: 0;
    }}


    /* Tự động ẩn hoàn toàn các file upload bị báo lỗi đỏ hoặc file tạm trong danh sách của Streamlit */
    [data-testid="stFileUploaderFile"]:has([data-testid*="Error"]),
    [data-testid="stFileUploaderFile"]:has([data-testid*="Alert"]),
    [data-testid="stFileUploaderFile"]:has(svg[fill*="ff4b"]),
    [data-testid="stFileUploaderFile"]:has(svg[stroke*="ff4b"]),
    [data-testid="stFileUploaderFile"]:has(svg[fill*="ff2b"]),
    [data-testid="stFileUploaderFile"]:has(svg[stroke*="ff2b"]),
    [data-testid="stFileUploaderFile"]:has([class*="error"]),
    [data-testid="stFileUploaderFile"]:has([class*="Error"]),
    [data-testid="stFileUploaderFile"]:has([class*="alert"]),
    [data-testid="stUploadedFile"]:has([data-testid*="Error"]),
    [data-testid="stUploadedFile"]:has([class*="error"]) {{
        display: none !important;
        opacity: 0 !important;
        height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
        overflow: hidden !important;
    }}
    /* Ẩn box thông báo lỗi đỏ dưới uploader nếu có */
    [data-testid="stFileUploader"] [data-testid="stAlert"],
    [data-testid="stFileUploader"] [class*="stAlert"] {{
        display: none !important;
    }}

    /* ── File tag ── */
    .mos-file-tag {{
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: {gT['bg_content']};
        border: 1px solid {gT['border']};
        border-radius: 8px;
        padding: 8px 12px;
        font-size: 12px;
        color: {gT['text_secondary']};
        margin: 4px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.04);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }}
    .mos-file-tag:hover {{
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.08);
        border-color: {gT['primary']};
    }}
    function enableFolderUploaders() {{
        const uploaders = document.querySelectorAll('[data-testid="stFileUploader"]');
        uploaders.forEach(uploader => {{
            const text = uploader.innerText;
            if (text.includes("Folder")) {{
                const input = uploader.querySelector('input[type="file"]');
                if (input && !input.hasAttribute('webkitdirectory')) {{
                    input.setAttribute('webkitdirectory', '');
                    input.setAttribute('directory', '');
                    input.setAttribute('multiple', '');
                    input.setAttribute('mozdirectory', '');
                }}
            }} else if (text.includes("File") || text.includes("ZIP") || text.includes("eml")) {{
                const input = uploader.querySelector('input[type="file"]');
                if (input && input.hasAttribute('webkitdirectory')) {{
                    input.removeAttribute('webkitdirectory');
                    input.removeAttribute('directory');
                    input.removeAttribute('mozdirectory');
                }}
            }}
        }});
    }}
    cleanUploaderErrors();
    enableFolderUploaders();

    /* ── Name mapping grid ── */
    .mos-nv-label {{
        background: {gT['bg_app']};
        color: {gT['text_primary']};
        border: 1px solid {gT['border']};
        border-radius: 6px;
        padding: 4px 10px;
        font-size: 12px;
        font-weight: 600;
        display: inline-block;
        margin-bottom: 4px;
    }}

    /* ── Result summary bar ── */
    .mos-summary {{
        background: {gT['bg_app']};
        border: 1px solid {gT['border']};
        border-radius: 12px;
        padding: 14px 20px;
        display: flex;
        gap: 32px;
        margin-bottom: 16px;
        flex-wrap: wrap;
    }}
    .mos-summary-item .label {{
        font-size: 11px;
        color: {gT['text_secondary']};
        text-transform: uppercase;
        letter-spacing: .05em;
    }}
    .mos-summary-item .value {{
        font-size: 20px;
        font-weight: 700;
        color: {gT['text_primary']};
    }}

    /* Override primary button */
    .stButton > button[kind="primary"] {{
        background: {gT['primary']} !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 500 !important;
        transition: all 0.3s ease;
    }}
    .stButton > button[kind="primary"]:hover {{
        background: {gT['primary_hover']} !important;
        box-shadow: 0 4px 12px {gT['shadow']} !important;
    }}
    /* Kéo toàn bộ cụm Tabs và nội dung bên dưới lên cao để lấp khoảng trắng */
    div[data-testid="stTabs"] {{
        margin-top: -130px !important;
    }}
    </style>
    """, unsafe_allow_html=True)
    

    

    # Xác định bước hiện tại
    step = 1
    if 'df_mos_result' in st.session_state and st.session_state.df_mos_result is not None and len(st.session_state.df_mos_result) > 0:
        step = 2
    if st.session_state.get('mos_saved', False):
        step = 3


    if "internal_mos_tab" not in st.session_state:
        st.session_state.internal_mos_tab = 0

    tab_names = [
        "📝 Dữ liệu Dự án (MOS)" if is_vi else "📝 プロジェクトデータ (MOS)",
        "📊 Năng suất & Thống kê" if is_vi else "📊 生産性・統計",
        "📈 Báo cáo tổng hợp" if is_vi else "📈 総合レポート"
    ]

    _tab0, _tab1, _tab2 = st.tabs(tab_names)
    with _tab0:
        import streamlit.components.v1 as components
        


        if 'mos_show_email_compare' not in st.session_state:
            st.session_state['mos_show_email_compare'] = False

        total_stored_check = len(st.session_state.get('mos_accumulated_raw_emails', []))
        is_email_open = st.session_state['mos_show_email_compare'] or (total_stored_check > 0)

        st.markdown("<div style='height: 32px'></div>", unsafe_allow_html=True)

        # Khu vực Dropzone Excel thu gọn vào chính giữa
        _space_l, col_upload, _space_r = st.columns([1, 2, 1])
        with col_upload:
            uploaded_files = st.file_uploader(
                "Upload Excel Report",
                type=["xlsx", "xls"],
                accept_multiple_files=True,
                label_visibility="collapsed",
                key="mos_multi_file_uploader"
            )
        # Container 2 riêng biệt cho Đối chiếu Mail Báo Cáo, xếp chồng bên dưới, rộng 100%, không bao giờ bị bóp hẹp hay mất nền
        if not is_email_open:
            st.markdown("<div style='height: 4px'></div>", unsafe_allow_html=True)
            col_space, col_btn = st.columns([5, 1.5])
            with col_btn:
                btn_open_lbl = "📧 Mở Đối chiếu Mail" if is_vi else "📧 照合機能を開く"
                if st.button(btn_open_lbl, key="btn_open_mos_email_compare", use_container_width=True, type="secondary"):
                    st.session_state['mos_show_email_compare'] = True
                    st.rerun()
        if is_email_open:
            st.markdown("<div style='height: 50px;'></div>", unsafe_allow_html=True)
            with st.container(border=False):
                # Di chuyển tiêu đề ra giữa
                _, col_eml_center, col_eml_right = st.columns([1, 4, 1], gap="medium")
                with col_eml_center:
                    st.markdown(f"""
                    <div class="mos-upload-header-box-clean" style="border-left: 4px solid #0EA5E9; padding-left: 12px; text-align: center; border-left: none;">
                        <div class="mos-upload-box-title" style="text-align: center;">📧 {'2. Đối chiếu Mail Báo Cáo (Thunderbird)' if is_vi else '2. メール報告の自動照合 (Thunderbird)'}</div>
                        <p class="mos-upload-box-sub" style="text-align: center;">{'Tải lên thư mục (Folder) hoặc file nén (.ZIP/.EML) từ Thunderbird để kiểm tra lệch giờ tự động.' if is_vi else 'Thunderbird等のメールフォルダやZIPをアップロードし、Excel工数と自動で照合・確認します。'}</p>
                    </div>
                    """, unsafe_allow_html=True)
                with col_eml_right:
                    btn_close_lbl2 = "✕ Đóng lại" if is_vi else "✕ 閉じる"
                    if st.button(btn_close_lbl2, key="btn_hide_mos_email_compare_2", use_container_width=True):
                        st.session_state['mos_show_email_compare'] = False
                        st.rerun()
                
                if 'mos_accumulated_raw_emails' not in st.session_state:
                    st.session_state['mos_accumulated_raw_emails'] = []
                    
                if 'mos_email_uploader_key' not in st.session_state:
                    st.session_state['mos_email_uploader_key'] = 0
                up_key = st.session_state['mos_email_uploader_key']
                
                # Di chuyển khu vực upload ra chính giữa
                _, col_upload, _ = st.columns([1, 2, 1])
                with col_upload:
                    upload_mode = st.radio(
                        "Chọn chế độ tải lên:" if is_vi else "アップロードモード:", 
                        ["📁 Thư mục (Folder)", "📄 File lẻ (.eml, .zip)"], 
                        horizontal=True, 
                        key=f"mos_upload_mode_{up_key}"
                    )
                    
                    if "Folder" in upload_mode:
                        lbl_upload = "📁 Chọn Thư mục (Folder) Mail" if is_vi else "📁 メールフォルダを選ぶ"
                        accepted_types = None
                    else:
                        lbl_upload = "📄 Tải File Mail lẻ hoặc File nén (.ZIP)" if is_vi else "📄 個別メールまたはZIPファイルをアップロード"
                        accepted_types = ["eml", "msg", "txt", "mbox", "zip"]
                        
                    uploaded_email_folders = st.file_uploader(lbl_upload, type=accepted_types, accept_multiple_files=True, key=f"mos_email_folder_uploader_{up_key}", label_visibility="collapsed")
                
                if 'mos_processed_email_files' not in st.session_state:
                    st.session_state['mos_processed_email_files'] = {}
                    
                newly_parsed = []
                all_inputs = []
                if uploaded_email_folders: all_inputs.extend(uploaded_email_folders)
                
                if all_inputs:
                    import zipfile
                    for ef in all_inputs:
                        file_id = f"{ef.name}_{ef.size}"
                        if file_id in st.session_state['mos_processed_email_files']:
                            continue
                        
                        st.session_state['mos_processed_email_files'][file_id] = True
                        
                        if ef.name.lower().endswith('.zip'):
                            try:
                                with zipfile.ZipFile(ef, 'r') as z:
                                    for zname in z.namelist():
                                        if not zname.startswith('__MACOSX') and not zname.endswith('/'):
                                            ext = zname.lower().split('.')[-1] if '.' in zname else ''
                                            if ext in ['eml', 'msg', 'txt', 'mbox', 'html', 'htm'] or not ext:
                                                with z.open(zname) as zfile:
                                                    reps = parse_single_email_report(zfile, zname.split('/')[-1])
                                                    newly_parsed.extend(reps)
                            except Exception as e:
                                logger.error(f"Lỗi giải nén file zip {ef.name}: {e}")
                        else:
                            reps = parse_single_email_report(ef, ef.name)
                            newly_parsed.extend(reps)
                    
                if newly_parsed:
                    existing_keys = {(r['source_file'], str(r['timestamp']), r['raw_body'][:100]) for r in st.session_state['mos_accumulated_raw_emails']}
                    for r in newly_parsed:
                        k = (r['source_file'], str(r['timestamp']), r['raw_body'][:100])
                        if k not in existing_keys:
                            st.session_state['mos_accumulated_raw_emails'].append(r)
                            existing_keys.add(k)
                            
                total_stored = len(st.session_state['mos_accumulated_raw_emails'])
                if total_stored > 0:
                    dict_latest, dedup_logs = deduplicate_email_reports(st.session_state['mos_accumulated_raw_emails'])
                    st.session_state['mos_latest_email_reports'] = dict_latest
                    st.session_state['mos_email_dedup_logs'] = dedup_logs
                    
                    col_st1, col_st2 = st.columns([3, 1])
                    with col_st1:
                        if is_vi:
                            msg = f"📧 Đã tích lũy thành công tổng cộng **{total_stored}** email từ các folder/file. Đã lọc chọn **{len(dict_latest)}** mail mới nhất từng ngày!"
                        else:
                            msg = f"📧 **{total_stored}** 件のメールをフォルダ/ファイルから正常に蓄積しました。日ごとの最新メール **{len(dict_latest)}** 件を抽出しました！"
                        st.success(msg)
                    with col_st2:
                        btn_lbl = "🗑️ Xóa làm lại từ đầu" if is_vi else "🗑️ 全てクリアしてやり直す"
                        if st.button(btn_lbl, key="btn_clear_accumulated_emails", use_container_width=True):
                            st.session_state['mos_accumulated_raw_emails'] = []
                            st.session_state['mos_processed_email_files'] = {}
                            st.session_state.pop('mos_latest_email_reports', None)
                            st.session_state.pop('mos_email_dedup_logs', None)
                            st.session_state['mos_email_uploader_key'] = st.session_state.get('mos_email_uploader_key', 0) + 1
                            st.rerun()
                else:
                    st.session_state.pop('mos_latest_email_reports', None)
                    st.session_state.pop('mos_email_dedup_logs', None)
        else:
            uploaded_email_folders = None

        components.html("""
        <script>
        const doc = window.parent.document;
        function cleanUploaderErrors() {
            const uploaders = doc.querySelectorAll('[data-testid="stFileUploader"]');
            uploaders.forEach(uploader => {
                const fileItems = uploader.querySelectorAll('[data-testid="stFileUploaderFile"], [data-testid="stUploadedFile"], [data-testid="fileUploaderFile"], li, .uploadedFile');
                fileItems.forEach(item => {
                    const text = item.textContent || "";
                    const hasError = item.querySelector('[data-testid*="Error"], [data-testid*="Alert"], [data-testid*="error"], svg[fill*="ff4b"], svg[stroke*="ff4b"], svg[fill*="ff2b"], svg[stroke*="ff2b"], [class*="error"], [class*="Error"]');
                    const isTempSize = text.includes("165.0B") || text.includes("165B") || text.includes("0.0B") || text.includes("0B") || text.includes("~$") || text.includes("._");
                    if (hasError || isTempSize) {
                        item.style.setProperty('display', 'none', 'important');
                        item.style.setProperty('opacity', '0', 'important');
                        item.style.setProperty('height', '0', 'important');
                        item.style.setProperty('margin', '0', 'important');
                        item.style.setProperty('padding', '0', 'important');
                        item.style.setProperty('overflow', 'hidden', 'important');
                    }
                });
                const errorBoxes = uploader.querySelectorAll('[data-testid="stAlert"], [data-testid="stNotification"], [class*="stAlert"]');
                errorBoxes.forEach(box => {
                    const text = box.textContent || "";
                    if (text.includes("165") || text.includes("~$") || text.includes("error") || text.includes("lỗi") || text.includes("Error") || text.includes("0.0B") || text.includes("0B")) {
                        box.style.setProperty('display', 'none', 'important');
                    }
                });
            });
        }
        function enableFolderUploaders() {
            const uploaders = doc.querySelectorAll('[data-testid="stFileUploader"]');
            uploaders.forEach(uploader => {
                const text = uploader.textContent || "";
                if (text.includes("Folder") || text.includes("Thư mục") || text.includes("folder") || text.includes("thư mục")) {
                    const input = uploader.querySelector('input[type="file"]');
                    if (input && !input.hasAttribute('webkitdirectory')) {
                        input.setAttribute('webkitdirectory', '');
                        input.setAttribute('directory', '');
                        input.setAttribute('multiple', '');
                        input.setAttribute('mozdirectory', '');
                        console.log("Đã kích hoạt chế độ chọn/kéo thả Folder cho input");
                    }
        cleanUploaderErrors();
        enableFolderUploaders();
        if (!doc.window_clean_uploader_observer) {
            let timeout = null;
            const observer = new MutationObserver(() => {
                if (timeout) clearTimeout(timeout);
                timeout = setTimeout(() => {
                    cleanUploaderErrors();
                    enableFolderUploaders();
                }, 200);
            });
            observer.observe(doc.body, { childList: true, subtree: true });
            doc.window_clean_uploader_observer = observer;
        }
        </script>
        """, height=0, width=0)

        has_mos_result = ('df_mos_result' in st.session_state and st.session_state['df_mos_result'] is not None and len(st.session_state['df_mos_result']) > 0)
        has_emails = 'mos_latest_email_reports' in st.session_state and len(st.session_state['mos_latest_email_reports']) > 0
                
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        use_mail_only = False
        if has_emails:
            use_mail_only = st.checkbox("Chỉ tổng hợp từ Mail (Tạm bỏ qua Excel)" if is_vi else "メールのみ集計", key="mos_use_mail_only", value=not bool(uploaded_files))
            
        col_btn, _ = st.columns([1.5, 4.5])
        with col_btn:
            run_btn = st.button("🔄 Tổng hợp dữ liệu" if is_vi else "🔄 データを集計", type="primary", use_container_width=True)
            
        if run_btn:
            if not uploaded_files and not use_mail_only:
                st.error("⚠️ Vui lòng tải lên file Excel hoặc tích chọn 'Chỉ tổng hợp từ Mail'.")
                st.stop()
                
            files_key = '_'.join(sorted([f.name for f in uploaded_files])) if not use_mail_only else f"MAIL_ONLY_{len(st.session_state['mos_latest_email_reports'])}"
            if st.session_state.get('mos_files_key') == files_key and 'df_mos_result' in st.session_state:
                st.success("✅ Sử dụng kết quả đã tóm tắt (Cached). Nếu muốn tính lại, hãy chọn lại file!")
            else:
                has_key = bool(load_saved_api_key())
                
                if not has_key:
                    st.warning("⚠️ Chưa cấu hình GEMINI_API_KEY. Hệ thống đang dùng chế độ nối chuỗi thủ công thay vì dùng AI!")
                    
                with st.spinner("🤖 AI đang phân tích và tóm tắt nội dung ủy thác..."):
                    dfs = []
                    file_errors = []
                    
                    if use_mail_only:
                        df_mail = convert_emails_to_df(st.session_state['mos_latest_email_reports'])
                        if not df_mail.empty:
                            dfs.append(df_mail)
                    else:
                        for f in uploaded_files:
                            if f.name.startswith('~$') or f.name.startswith('.') or getattr(f, 'size', 1000) < 500:
                                continue
                            if not (f.name.lower().endswith('.xlsx') or f.name.lower().endswith('.xls')):
                                continue
                            try:
                                df_f = parse_mos_file(f, f.name)
                                if not df_f.empty:
                                    dfs.append(df_f)
                            except Exception as e:
                                file_errors.append(f"{f.name}: {str(e)}")
                    
                    if file_errors:
                        st.session_state['mos_file_errors'] = file_errors
                        for err in file_errors:
                            st.warning(f"⚠️ Bỏ qua file do lỗi định dạng: {err}")
                    else:
                        st.session_state.pop('mos_file_errors', None)
                    
                    if dfs:
                        df_tong_hop = tong_hop_mos(dfs)
                        st.session_state['df_mos_raw'] = pd.concat(dfs, ignore_index=True)
                        st.session_state['df_mos_result'] = df_tong_hop
                        st.session_state['mos_result_version'] = st.session_state.get('mos_result_version', 0) + 1
                        st.session_state['mos_files_key'] = files_key
                        raw_concat = st.session_state['df_mos_raw']
                        st.session_state['mos_num_people'] = raw_concat['ma_nv'].nunique() if 'ma_nv' in raw_concat.columns and not raw_concat.empty else len(dfs)
                        
                        # Xóa các override cũ nếu có để KPI table nhận số liệu mới
                        for k in ["override_mos_nv", "override_mos_std", "override_mos_target", "override_mos_actual"]:
                            st.session_state.pop(k, None)
                        
                        # Automatically detect month and year from the uploaded data
                        if len(df_tong_hop) > 0:
                            def get_kpi_m(val):
                                if pd.isna(val) or str(val).strip() == '': return None
                                try:
                                    if isinstance(val, (datetime.date, datetime.datetime, pd.Timestamp)):
                                        return val.month
                                    s = str(val).strip()
                                    dt = pd.to_datetime(s, format='%d/%m/%Y', errors='coerce')
                                    if pd.isna(dt):
                                        dt = pd.to_datetime(s, errors='coerce')
                                    return dt.month if pd.notna(dt) else None
                                except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return None

                            kpi_ms = pd.Series(dtype=float)
                            if 'Ngày kết thúc' in df_tong_hop.columns:
                                kpi_ms = df_tong_hop['Ngày kết thúc'].apply(get_kpi_m).dropna()
                            if len(kpi_ms) == 0 and 'Ngày bắt đầu' in df_tong_hop.columns:
                                kpi_ms = df_tong_hop['Ngày bắt đầu'].apply(get_kpi_m).dropna()
                                
                            if len(kpi_ms) > 0:
                                kpi_m = int(kpi_ms.mode()[0])
                                st.session_state['kpi_month'] = kpi_m
                                st.session_state['mos_kpi_month'] = kpi_m
                                
                                # Estimate year from the most common date
                                col_to_use = 'Ngày kết thúc' if 'Ngày kết thúc' in df_tong_hop.columns and len(df_tong_hop['Ngày kết thúc'].dropna()) > 0 else 'Ngày bắt đầu'
                                if col_to_use in df_tong_hop.columns:
                                    dates_for_year = pd.to_datetime(df_tong_hop[col_to_use], format='%d/%m/%Y', errors='coerce').dropna()
                                    if len(dates_for_year) > 0:
                                        st.session_state['kpi_year'] = dates_for_year.dt.year.mode()[0]
                                        st.session_state['mos_kpi_year'] = dates_for_year.dt.year.mode()[0]

                        st.success("✅ Tổng hợp và tóm tắt AI xong!" if is_vi else "✅ AIによる集計と要約が完了しました！")
                    else:
                        st.error("Không tìm thấy dữ liệu hợp lệ (phần MOS業務) trong các file đã tải lên." if is_vi else "アップロードされたファイルに有効なデータ（MOS業務部分）が見つかりませんでした。")

    if 'df_mos_result' in st.session_state and st.session_state['df_mos_result'] is not None and len(st.session_state['df_mos_result']) > 0:
        # Làm tròn một lần duy nhất vào session state để giữ data ổn định, tránh mất focus khi enter
        if 'Giờ làm (h)' in st.session_state['df_mos_result'].columns:
            def safe_round_float(val):
                if pd.isna(val) or str(val).strip() == '': return 0.0
                try: return round(float(val), 1)
                except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return 0.0
                
            st.session_state['df_mos_result']['Giờ làm (h)'] = st.session_state['df_mos_result']['Giờ làm (h)'].apply(safe_round_float).astype(float)

        df_result = st.session_state['df_mos_result']
        # Theo yêu cầu của người dùng: Sửa lại trong bảng dữ liệu mã máy dán Hub (K230062 / ハブ接着) từ 179 giờ thành 175 giờ
        if 'Mã dự án' in df_result.columns and 'Giờ làm (h)' in df_result.columns:
            for idx_r in df_result.index:
                m_code = str(df_result.at[idx_r, 'Mã dự án']).strip().upper()
                t_name = str(df_result.at[idx_r, 'Tên dự án']).lower() if 'Tên dự án' in df_result.columns else ''
                if m_code in ['K230062', '230062', 'K230063', '230063'] or any(k in t_name for k in ['háp', 'hub', 'dán hub', 'ハブ接着']):
                    if df_result.at[idx_r, 'Giờ làm (h)'] == 179.0 or df_result.at[idx_r, 'Giờ làm (h)'] != 177.0 or round(df_result.at[idx_r, 'Giờ làm (h)']) == 179:
                        df_result.at[idx_r, 'Giờ làm (h)'] = 177.0
                        
        # Theo yêu cầu của người dùng: Khôi phục ngày kết thúc (19/6) cho mã dự án K000000 của Matsui và các dự án 1 ngày
        if 'Ngày kết thúc' in df_result.columns and 'Ngày bắt đầu' in df_result.columns:
            for idx_r in df_result.index:
                curr_kt = str(df_result.at[idx_r, 'Ngày kết thúc'] or '').strip()
                curr_bd = str(df_result.at[idx_r, 'Ngày bắt đầu'] or '').strip()
                if (not curr_kt or curr_kt.lower() in ['nan', 'none', '']) and curr_bd and curr_bd.lower() not in ['nan', 'none', '']:
                    df_result.at[idx_r, 'Ngày kết thúc'] = curr_bd
                
                m_code = str(df_result.at[idx_r, 'Mã dự án']).strip().upper() if 'Mã dự án' in df_result.columns else ''
                mgr_val = str(df_result.at[idx_r, 'Quản lý Nhật Bản']).lower() if 'Quản lý Nhật Bản' in df_result.columns else ''
                if m_code == 'K000000' and any(k in mgr_val for k in ['松井', 'matsui']):
                    if not curr_kt or curr_kt.lower() in ['nan', 'none', ''] or '19/06' not in curr_kt:
                        year_str = curr_bd.split('/')[-1] if curr_bd and '/' in curr_bd else (str(st.session_state.get('mos_kpi_year', '2024')))
                        df_result.at[idx_r, 'Ngày kết thúc'] = f"19/06/{year_str}"
                        if not curr_bd or curr_bd.lower() in ['nan', 'none', '']:
                            df_result.at[idx_r, 'Ngày bắt đầu'] = f"19/06/{year_str}"
                            
        # Theo yêu cầu của người dùng: Tóm tắt thành 1 nội dung ủy thác duy nhất nếu bị ghép nối dài dòng
        if 'Nội dung ủy thác' in df_result.columns and 'Mã dự án' in df_result.columns:
            for idx_r in df_result.index:
                curr_nd = str(df_result.at[idx_r, 'Nội dung ủy thác'] or '').strip()
                if ' / ' in curr_nd or len(curr_nd.split(' / ')) > 1:
                    m_code = str(df_result.at[idx_r, 'Mã dự án']).strip()
                    t_name = str(df_result.at[idx_r, 'Tên dự án']).strip() if 'Tên dự án' in df_result.columns else ''
                    pv_name = str(df_result.at[idx_r, 'Phân vùng']).strip() if 'Phân vùng' in df_result.columns else ''
                    ts_list = [t.strip() for t in curr_nd.split(' / ') if t.strip()]
                    if len(ts_list) > 1:
                        df_result.at[idx_r, 'Nội dung ủy thác'] = smart_offline_summarize(m_code, t_name, ts_list, pv_name)
                
                # Đảm bảo dịch song ngữ cho tất cả nội dung ủy thác đang hiển thị
                curr_nd_after = str(df_result.at[idx_r, 'Nội dung ủy thác'] or '').strip()
                df_result.at[idx_r, 'Nội dung ủy thác'] = translate_task_bilingual(curr_nd_after)
                        
        # Theo yêu cầu của người dùng: Cảnh báo bất thường (Mã dự án bắt đầu bằng J)
        if 'Mã dự án' in df_result.columns:
            if 'Cảnh báo' not in df_result.columns:
                df_result.insert(1 if 'STT' in df_result.columns else 0, 'Cảnh báo', '')
            if 'Giờ làm gốc (h)' not in df_result.columns and 'Giờ làm (h)' in df_result.columns:
                df_result['Giờ làm gốc (h)'] = df_result['Giờ làm (h)']
            if 'Đơn giá' not in df_result.columns:
                idx = df_result.columns.get_loc('Giờ làm (h)') + 1
                df_result.insert(idx, 'Đơn giá', 2500.0)
                gio_h = pd.to_numeric(df_result['Giờ làm (h)'], errors='coerce').fillna(0.0)
                if 'Tổng tiền' not in df_result.columns:
                    df_result.insert(idx + 1, 'Tổng tiền', gio_h * 2500.0)
                else:
                    df_result['Tổng tiền'] = gio_h * 2500.0
            else:
                df_result['Đơn giá'] = pd.to_numeric(df_result['Đơn giá'], errors='coerce').replace([0.0, 0], 2500.0).fillna(2500.0)
                gio_h = pd.to_numeric(df_result.get('Giờ làm (h)', 0.0), errors='coerce').fillna(0.0)
                if 'Tổng tiền' not in df_result.columns:
                    idx = df_result.columns.get_loc('Đơn giá') + 1
                    df_result.insert(idx, 'Tổng tiền', gio_h * df_result['Đơn giá'])
                else:
                    df_result['Tổng tiền'] = gio_h * df_result['Đơn giá']

            
            for idx_r in df_result.index:
                m_code = str(df_result.at[idx_r, 'Mã dự án']).strip().upper()
                w_list = []
                if m_code.startswith('J'):
                    w_list.append(f"⚠️ Mã J (JMOS) - Không thuộc bảng MOS")
                    if 'Giờ làm (h)' in df_result.columns and df_result.at[idx_r, 'Giờ làm (h)'] != 0.0:
                        df_result.at[idx_r, 'Giờ làm (h)'] = 0.0
                
                old_w = str(df_result.at[idx_r, 'Cảnh báo']).strip()
                df_result.at[idx_r, 'Cảnh báo'] = "🔴" if (w_list or old_w) else ""
                        
        df_raw = st.session_state.get('df_mos_raw', pd.DataFrame())

        with _tab0:
            st.markdown("---")
            # 3.5. BẢNG ĐỐI CHIẾU MAIL BÁO CÁO (THUNDERBIRD/OUTLOOK) VS EXCEL
            dict_latest_emails = st.session_state.get('mos_latest_email_reports', {})
            dedup_logs = st.session_state.get('mos_email_dedup_logs', [])
            
            if dict_latest_emails and not df_raw.empty:
                is_mail_only = st.session_state.get('mos_use_mail_only', False)
                if is_mail_only:
                    st.markdown("### 📧 Chi tiết dữ liệu tổng hợp từ Mail" if is_vi else "### 📧 メールからの集計データの詳細")
                else:
                    st.markdown("### 📧 Đối chiếu Mail báo cáo (Thunderbird) & File Excel" if is_vi else "### 📧 メール報告とExcelの照合")
                
                if dedup_logs:
                    with st.expander("🔄 Nhật ký tự động lọc Mail mới nhất (Khi 1 ngày có >1 mail)" if is_vi else "🔄 最新メール自動フィルタリングログ（1日に複数メールがある場合）", expanded=True):
                        for log_item in dedup_logs:
                            st.markdown(f"- {log_item}")
                            
                df_compare, diff_cnt = reconcile_mail_vs_excel(dict_latest_emails, df_raw)
                
                df_display = df_compare.copy()
                if not df_display.empty and not is_vi:
                    # Translate status and notes
                    df_display = df_display.replace({
                        "✅ Khớp số liệu": "✅ データ一致",
                        "⚠️ Khớp mã, lệch số liệu": "⚠️ コード一致、データ相違",
                        "❌ Lệch/Thiếu giờ Excel": "❌ Excel時間相違/不足",
                        "❌ Lệch giờ Mail": "❌ メール時間相違",
                        "Khớp hoàn toàn": "完全一致",
                        r"\(Nội bộ\)": "(社内)",
                        r"🚫 Lỗi sai mục trong Mail \(Mã J\)": "🚫 メールの記載項目エラー (Jコード)",
                        r"🚫 Lỗi Mail: Dự án \[(.*?)\] có đầu mã J \(thuộc JMOS\) nhưng trong Mail lại ghi sai vào mục ◇MOS業務!": r"🚫 メールエラー: プロジェクト [\1] はJコード (JMOS) ですが、メール内で誤って◇MOS業務に記載されています！",
                        r"⚠️ Thiếu dự án trong Excel": "⚠️ Excelにプロジェクトがありません",
                        r"Excel thiếu dự án MOS (.*)": r"ExcelにMOSプロジェクト \1 がありません",
                        r"Dự án (.*): Mail \((.*)h\) ≠ Excel \((.*)h\)": r"プロジェクト \1: メール (\2h) ≠ Excel (\3h)"
                    }, regex=True)
                    df_display.rename(columns={
                        'Trạng thái': 'ステータス',
                        'Giờ Mail': 'メール時間',
                        'Giờ Excel': 'Excel時間',
                        'Dự án Mail': 'メールPJ',
                        'Dự án Excel': 'ExcelPJ',
                        'Chi tiết chênh lệch / Cảnh báo': '差異詳細 / 警告',
                        'Mã NV': '社員コード',
                        'Tên NV': '氏名',
                        'Ngày': '日付'
                    }, inplace=True)

                if is_mail_only:
                    cols_to_drop = ['Trạng thái', 'Giờ Excel', 'Dự án Excel', 'Chi tiết chênh lệch / Cảnh báo', 'ステータス', 'Excel時間', 'ExcelPJ', '差異詳細 / 警告']
                    df_display = df_display.drop(columns=[c for c in cols_to_drop if c in df_display.columns], errors='ignore')
                    if df_display.empty:
                        st.warning("⚠️ Không có dữ liệu trích xuất từ Mail hợp lệ." if is_vi else "⚠️ 有効なメール抽出データがありません。")
                    else:
                        st.dataframe(df_display, use_container_width=True, hide_index=True)
                else:
                    if df_display.empty:
                        st.warning("⚠️ Không tìm thấy dữ liệu trùng khớp để đối chiếu! Vui lòng kiểm tra lại sự trùng khớp về Tên nhân viên và Ngày làm việc giữa Mail và File Excel." if is_vi else "⚠️ 照合するデータが見つかりません！メールとExcelファイルの間で担当者名と日付が一致しているか確認してください。")
                    else:
                        if diff_cnt > 0:
                            st.warning(f"⚠️ Phát hiện **{diff_cnt} điểm chênh lệch/bất thường** giữa Mail báo cáo mới nhất và dữ liệu Excel!" if is_vi else f"⚠️ 最新の報告メールとExcelデータの間に **{diff_cnt} 箇所の相違/異常** が見つかりました！")
                        else:
                            st.success("✅ Tuyệt vời! Tất cả dữ liệu từ Mail báo cáo mới nhất khớp hoàn toàn 100% với bảng Excel." if is_vi else "✅ 素晴らしい！最新の報告メールのすべてのデータがExcelと100%一致しています。")
                            
                        theme_mode = st.session_state.get('theme_mode', 'light')
                        def highlight_abnormal(row):
                            val = row.get('Chi tiết chênh lệch / Cảnh báo', row.get('差異詳細 / 警告', 'Khớp hoàn toàn'))
                            if val not in ["Khớp hoàn toàn", "完全一致"]:
                                bg_color = '#FADBD8' if theme_mode == 'sepia' else '#FEF2F2'
                                return [f'background-color: {bg_color}; color: #991B1B'] * len(row)
                            return [''] * len(row)
                        
                        styled_df = df_display.style.apply(highlight_abnormal, axis=1)
                        
                        # Ẩn cột cảnh báo vì đã tô đỏ dòng
                        col_cfg = {
                            'Chi tiết chênh lệch / Cảnh báo': None,
                            '差異詳細 / 警告': None
                        }
                        st.dataframe(styled_df, use_container_width=True, hide_index=True, column_config=col_cfg)
                        
                        col_sync1, col_sync2 = st.columns([2.5, 3])
                        with col_sync1:
                            btn_lbl = "✨ Đồng bộ & Chuẩn hóa Excel theo Mail mới nhất" if is_vi else "✨ 最新メールに合わせてExcelを同期・正規化"
                            if st.button(btn_lbl, type="primary", use_container_width=True, key="btn_sync_mail_excel"):
                                df_raw_clean = apply_reconciliation_to_excel(dict_latest_emails, df_raw)
                                st.session_state['df_mos_raw'] = df_raw_clean
                                df_res_clean = tong_hop_mos([df_raw_clean])
                                st.session_state['df_mos_result'] = df_res_clean
                                st.session_state['mos_result_version'] = st.session_state.get('mos_result_version', 0) + 1
                                st.success("✅ Đã chuẩn hóa thành công 100% số liệu Excel theo Mail mới nhất!" if is_vi else "✅ 最新の報告メールに合わせてExcelデータを100%正常に正規化しました！")
                                st.rerun()
                        with col_sync2:
                            st.markdown(f"<div style='font-size:13px; color:#64748B; padding-top:8px;'>💡 {'Bấm nút trên để tự động lấy dữ liệu từ Mail mới nhất ghi đè vào các chỗ sai/thiếu trong Excel.' if is_vi else '上のボタンをクリックすると、最新の報告メールから自動的にデータを取得し、Excelの誤りや不足部分を上書きします。'}</div>", unsafe_allow_html=True)
                st.markdown("---")

            if not st.session_state.get('mos_use_mail_only', False):
                # 4. BỘ LỌC THÔNG MINH
                st.markdown("### 🔍 Bộ lọc nâng cao" if is_vi else "### 🔍 高度なフィルター")
                col_f1, col_f2, col_f3 = st.columns(3)
                with col_f1:
                    search_name = st.text_input("Tên dự án:" if is_vi else "プロジェクト名:", placeholder="Nhập từ khóa..." if is_vi else "キーワードを入力...")
                with col_f2:
                    ma_da_list = df_result['Mã dự án'].unique().tolist()
                    selected_ma_da = st.multiselect("Mã dự án" if is_vi else "プロジェクトコード", options=ma_da_list, default=[], placeholder="Chọn mã dự án..." if is_vi else "プロジェクトコードを選択...")
                with col_f3:
                    default_managers = ["フォン \n Phương", "ロン \n Long", "ダオ \n Đạo"]
                    data_managers = [str(x) for x in df_result['Quản lý Việt Nam'].dropna().unique() if str(x).strip()]
                    manager_list = ["Tất cả" if is_vi else "すべて"] + sorted(list(set(default_managers + data_managers)))
                    selected_manager = st.selectbox("Quản lý VN:" if is_vi else "VN側管理者:", options=manager_list)

                df_display = df_result.copy()
                if search_name:
                    df_display = df_display[df_display['Tên dự án'].str.contains(search_name, case=False, na=False, regex=False)]
                if selected_ma_da:
                    df_display = df_display[df_display['Mã dự án'].isin(selected_ma_da)]
                all_label = "Tất cả" if is_vi else "すべて"
                if selected_manager != all_label:
                    df_display = df_display[df_display['Quản lý Việt Nam'] == selected_manager]

                # HIỂN THỊ BANNER CẢNH BÁO BẤT THƯỜNG (Mã J)
                j_codes_list = [r['Mã dự án'] for _, r in df_result.iterrows() if str(r.get('Mã dự án', '')).strip().upper().startswith('J')]
            
                if j_codes_list:
                    with st.container():
                        st.markdown(f"""
                        <div style="background: linear-gradient(135deg, rgba(254, 242, 242, 0.95) 0%, rgba(255, 241, 242, 0.95) 100%); backdrop-filter: blur(12px); border-left: 5px solid #EF4444; border-radius: 12px; padding: 16px 20px; margin-bottom: 18px; box-shadow: 0 4px 15px rgba(239, 68, 68, 0.1);">
                            <div style="display: flex; align-items: center; gap: 10px; margin-bottom: 10px;">
                                <span style="font-size: 22px;">🚨</span>
                                <span style="font-size: 16px; font-weight: 700; color: #991B1B;">
                                    {"PHÁT HIỆN CẢNH BÁO BẤT THƯỜNG TRONG DỮ LIỆU MOS" if is_vi else "MOSデータ内に異常警告が検出されました"}
                                </span>
                            </div>
                        """, unsafe_allow_html=True)
                    
                        j_str = ", ".join([f"<b>{c}</b>" for c in sorted(list(set(j_codes_list)))])
                        msg_j = f"• <b>Mã dự án thuộc bảng JMOS (Mã J):</b> Phát hiện các mã {j_str} trong bảng MOS. Theo quy định, các mã này thuộc JMOS, không thuộc bảng MOS và <b>không được ghi nhận giờ làm trong báo cáo MOS (0h)</b>." if is_vi else f"• <b>JMOSプロジェクトコード (Jで始まる):</b> {j_str} が検出されました。規定によりMOS対象外となり、<b>労働時間は計上されません (0h)</b>。"
                        st.markdown(f"<div style='color: #B91C1C; font-size: 14px; margin-left: 32px; margin-bottom: 8px;'>{msg_j}</div>", unsafe_allow_html=True)
                                
                        st.markdown("</div>", unsafe_allow_html=True)

                # Ẩn cột STT và Giờ làm gốc khỏi hiển thị trên bảng UI
                cols_to_hide = [c for c in ['STT', 'Giờ làm gốc (h)'] if c in df_display.columns]
                df_display_show = df_display.drop(columns=cols_to_hide, errors='ignore')

                st.markdown("💡 **Mẹo:** Sửa trực tiếp trên bảng, sau đó bấm **Lưu thay đổi** ở bên dưới để cập nhật." if is_vi else "💡 **ヒント:** 表上で直接編集し、下の**変更を保存**をクリックして更新してください。")
            
                @st.fragment
                def render_mos_editor():
                    edit_mode = st.toggle("Bật chế độ chỉnh sửa" if st.session_state.get('lang', 'vi') == 'vi' else "編集モードを有効にする", key="mos_main_edit_toggle", value=False)
                
                    def fmt_num(x):
                        try:
                            if pd.isna(x) or str(x).strip() == '': return ''
                            v = round(float(x), 1)
                            return f"{int(v)}" if v.is_integer() else f"{v}"
                        except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return x

                    df_editor_input = df_display_show.copy()
                    if 'Giờ làm (h)' in df_editor_input.columns:
                        df_editor_input['Giờ làm (h)'] = df_editor_input['Giờ làm (h)'].apply(fmt_num)

                    mos_editor_col_cfg = {
                        "STT": st.column_config.Column("STT" if is_vi else "No."),
                        "Cảnh báo": st.column_config.Column("Cảnh báo" if is_vi else "警告"),
                        "Mã dự án": st.column_config.TextColumn("Mã dự án" if is_vi else "プロジェクトコード", disabled=False),
                        "Tên dự án": st.column_config.TextColumn("Tên dự án" if is_vi else "プロジェクト名", disabled=False),
                        "Phân vùng": st.column_config.TextColumn("Phân vùng" if is_vi else "分野"),
                        "Nội dung ủy thác": st.column_config.TextColumn("Nội dung ủy thác" if is_vi else "委託内容"),
                        "Giờ làm (h)": "Giờ làm (h)" if is_vi else "労働時間 (h)",
                        "Đơn giá": st.column_config.NumberColumn(
                            "Đơn giá (Yên)" if is_vi else "単価 (¥)",
                            format="%,.0f ¥",
                            default=2500.0,
                            step=100.0
                        ),
                        "Tổng tiền": st.column_config.NumberColumn(
                            "Tổng tiền (Yên)" if is_vi else "合計 (¥)",
                            format="%,.0f ¥",
                            disabled=True
                        ),
                        "Ngày bắt đầu": st.column_config.TextColumn("Ngày bắt đầu" if is_vi else "開始日", disabled=False),
                        "Ngày kết thúc": st.column_config.TextColumn("Ngày kết thúc" if is_vi else "終了日", disabled=False),
                        "Quản lý Nhật Bản": st.column_config.TextColumn("Quản lý Nhật Bản" if is_vi else "JP側管理者", disabled=False),
                        "Quản lý Việt Nam": st.column_config.SelectboxColumn("Quản lý Việt Nam" if is_vi else "VN側管理者", options=["フォン \n Phương", "ロン \n Long", "ダオ \n Đạo"]),
                        "Người thực hiện": st.column_config.TextColumn("Người thực hiện" if is_vi else "担当者", disabled=False),
                        "Trạng thái": st.column_config.SelectboxColumn("Trạng thái" if is_vi else "ステータス", options=["完了 \n Hoàn thành", "実行中 \n Đang tiến hành", "未着手 \n Chưa bắt đầu"]),
                    }
                    if not edit_mode:
                        st.dataframe(df_editor_input, use_container_width=True, hide_index=True, column_config=mos_editor_col_cfg)
                    else:
                        edited_display = st.data_editor(
                            df_editor_input,
                            use_container_width=True,
                            num_rows="dynamic",
                            column_config=mos_editor_col_cfg,
                            hide_index=True,
                            key="mos_data_editor"
                        )
    
                        if st.button("💾 Lưu thay đổi bảng" if is_vi else "💾 テーブルを保存", type="primary", key="btn_mos_save_edits"):
                            st.session_state['mos_saved'] = True
                            edited_no_warning = edited_display.copy()
                        
                            if 'Giờ làm (h)' in edited_no_warning.columns:
                                def parse_and_round(val):
                                    try:
                                        return round(float(str(val).replace(',', '.')), 1)
                                    except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return 0.0
                                edited_no_warning['Giờ làm (h)'] = edited_no_warning['Giờ làm (h)'].apply(parse_and_round).astype(float)
                                
                            if 'Đơn giá' in edited_no_warning.columns:
                                def parse_float(val):
                                    try: return float(str(val).replace(',', '').replace('¥', '').replace('Yên', '').strip())
                                    except: return 2500.0
                                edited_no_warning['Đơn giá'] = edited_no_warning['Đơn giá'].apply(parse_float).astype(float)
                                edited_no_warning['Tổng tiền'] = edited_no_warning['Giờ làm (h)'] * edited_no_warning['Đơn giá']
                            
                            if search_name or selected_ma_da or selected_manager != all_label:
                                df_result.update(edited_no_warning)
                                st.session_state['df_mos_edited'] = df_result
                                st.session_state['df_mos_result'] = df_result
                            else:
                                st.session_state['df_mos_edited'] = edited_no_warning
                                st.session_state['df_mos_result'] = edited_no_warning
                            st.session_state['mos_result_version'] = st.session_state.get('mos_result_version', 0) + 1
                            st.rerun()

                render_mos_editor()
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            total_da = len(df_result)
            total_gio = df_result['Giờ làm (h)'].sum()
            total_nv = st.session_state.get('mos_num_people', 0)
            
            lang = st.session_state.get('lang', 'vi')
            lbl_title = "Thống kê & KPI" if lang == 'vi' else "統計・KPI"
            lbl_month = "Tháng" if lang == 'vi' else "月"
            lbl_year = "Năm" if lang == 'vi' else "年"
            lbl_holiday = "Số ngày lễ (Âm lịch)" if lang == 'vi' else "祝日数"
            
            st.markdown(f"### 📊 {lbl_title}")
            col_kpi1, col_kpi2, col_kpi3 = st.columns(3)
            with col_kpi1:
                current_month = datetime.date.today().month
                selected_month = st.selectbox(lbl_month, range(1, 13), index=current_month-1, key="kpi_month")
            with col_kpi2:
                current_year = datetime.date.today().year
                selected_year = st.selectbox(lbl_year, range(current_year-2, current_year+3), index=2, key="kpi_year")
            with col_kpi3:
                manual_holidays = st.number_input(lbl_holiday, min_value=0, max_value=20, value=0, key="kpi_holidays")
                
            import calendar
            num_days_in_month = calendar.monthrange(selected_year, selected_month)[1]
            month_days = [datetime.date(selected_year, selected_month, day) for day in range(1, num_days_in_month + 1)]
            
            saturdays = sum(1 for d in month_days if d.weekday() == 5)
            sundays = sum(1 for d in month_days if d.weekday() == 6)
            
            # Căn cứ chính xác theo số ngày làm việc trong tháng, trừ đi T7 CN, cộng thêm 1 ngày làm việc cuối tháng
            standard_days = len(month_days) - saturdays - sundays + 1 - manual_holidays
            standard_days = max(0, standard_days)
            std_hours_per_person = standard_days * 8
            target_hours = total_nv * std_hours_per_person
            completion_rate = (total_gio / target_hours * 100) if target_hours > 0 else 0
            
            if st.session_state.get('last_kpi_month') != selected_month or st.session_state.get('last_kpi_year') != selected_year:
                st.session_state['last_kpi_month'] = selected_month
                st.session_state['last_kpi_year'] = selected_year
                for k in ["override_mos_nv", "override_mos_std", "override_mos_target", "override_mos_actual", "mos_kpi_interactive_editor"]:
                    st.session_state.pop(k, None)
            
            st.session_state['mos_kpi_month'] = selected_month
            st.session_state['mos_kpi_year'] = selected_year
            st.session_state['mos_kpi_std_hours'] = std_hours_per_person
            
            # ----- Xử lý chỉnh sửa thủ công KPI -----
            editor_key_kpi = "mos_kpi_interactive_editor"
            if editor_key_kpi in st.session_state and st.session_state[editor_key_kpi].get("edited_rows"):
                edits = st.session_state[editor_key_kpi]["edited_rows"]
                for r_idx, changes in edits.items():
                    if "Giá trị" in changes:
                        val = float(changes["Giá trị"])
                        if int(r_idx) == 0: st.session_state["override_mos_nv"] = val
                        elif int(r_idx) == 1: st.session_state["override_mos_std"] = val
                        elif int(r_idx) == 2: st.session_state["override_mos_target"] = val
                        elif int(r_idx) == 3: st.session_state["override_mos_actual"] = val
                        elif int(r_idx) == 5: st.session_state["override_mos_price"] = val
                        elif int(r_idx) == 6: st.session_state["override_mos_total_money"] = val

            cur_nv = st.session_state.get("override_mos_nv", float(total_nv))
            cur_std = st.session_state.get("override_mos_std", float(std_hours_per_person))
            
            # Nếu người dùng tự nhập target thì dùng, không thì lấy nv * std
            if "override_mos_target" in st.session_state:
                cur_target = st.session_state["override_mos_target"]
            else:
                cur_target = cur_nv * cur_std

            cur_actual = st.session_state.get("override_mos_actual", float(total_gio))
            cur_rate = (cur_actual / cur_target * 100) if cur_target > 0 else 0.0

            cur_price = st.session_state.get("override_mos_price", 2500.0)
            if "override_mos_total_money" in st.session_state:
                cur_total_money = st.session_state["override_mos_total_money"]
            else:
                cur_total_money = cur_actual * cur_price

            df_kpi_ui = pd.DataFrame({
                "Chỉ tiêu": [
                    "人数 \n Số người",
                    "一人当たり月稼働時間 (h) \n Giờ làm việc tiêu chuẩn(h)",
                    "月目標稼働時間 (h) \n Mục tiêu giờ làm(h)",
                    "月実稼働時間 (h) \n Giờ làm thực tế(h)",
                    "目標に対して稼働率 (%) \n Tỷ lệ hoàn thành(%)",
                    "単価 (¥) \n Đơn giá (JPY)",
                    "実稼働に対して総金額 \n Tổng tiền theo giờ thực tế (JPY)"
                ],
                "Giá trị": [
                    round(cur_nv, 1), 
                    round(cur_std, 1), 
                    round(cur_target, 1), 
                    round(cur_actual, 1), 
                    round(cur_rate, 1),
                    round(cur_price, 0),
                    round(cur_total_money, 0)
                ]
            })

            col_tip, col_rst = st.columns([3.5, 1.5])
            with col_tip:
                st.markdown("💡 *Mẹo: Bạn có thể click đúp vào cột **Giá trị** bên dưới để sửa đổi thủ công.*" if lang == 'vi' else "💡 *ヒント: 下記の「数値」列をダブルクリックして手動で変更できます。*")
            with col_rst:
                if st.button("🔄 Lấy lại dữ liệu gốc" if lang == 'vi' else "🔄 リセット", key="reset_mos_kpi_btn", use_container_width=True):
                    for k in ["override_mos_nv", "override_mos_std", "override_mos_target", "override_mos_actual", "override_mos_price", "override_mos_total_money", editor_key_kpi]:
                        st.session_state.pop(k, None)
                    st.rerun()

            st.data_editor(
                df_kpi_ui,
                use_container_width=True,
                hide_index=True,
                key=editor_key_kpi,
                column_config={
                    "Chỉ tiêu": st.column_config.TextColumn("Hạng mục / Chỉ tiêu KPI" if is_vi else "KPI項目 / 指標", disabled=True, width="large"),
                    "Giá trị": st.column_config.NumberColumn("Giá trị" if is_vi else "数値", disabled=False, format="%g", step=0.1)
                }
            )
            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        with _tab1:
            st.markdown("### 📊 Thống kê Năng suất & Tổng quan" if is_vi else "### 📊 生産性・概要統計")
            if 'df_raw' in locals() and not df_raw.empty:
                emp_stats = df_raw.groupby("ten_nv")["tong_gio"].sum().reset_index()
                emp_stats.columns = ["Tên nhân viên" if is_vi else "氏名", "Tổng giờ MOS" if is_vi else "MOS総労働時間"]
                emp_stats["Tổng giờ MOS" if is_vi else "MOS総労働時間"] = emp_stats["Tổng giờ MOS" if is_vi else "MOS総労働時間"].apply(
                    lambda x: round(float(x), 1) if pd.notna(x) else x
                )
                emp_stats = emp_stats.sort_values("Tổng giờ MOS" if is_vi else "MOS総労働時間", ascending=False)
                
                # Bọc dataframe vào một cột nhỏ để bảng không bị kéo dài toàn màn hình
                col_tbl, _ = st.columns([2.5, 3.5])
                with col_tbl:
                    st.dataframe(emp_stats, use_container_width=True, hide_index=True)
            else:
                st.info("Chưa có dữ liệu thô để thống kê cá nhân." if is_vi else "個人統計用の生データがありません。")
            
            st.markdown("---")
            
            try:
                import plotly.express as px
                if 'df_result' in locals() and not df_result.empty:
                    col_d1, col_d2 = st.columns(2)
                    with col_d1:
                        fig_pie = px.pie(df_result, values="Giờ làm (h)", names="Tên dự án", title="Phân bổ Giờ làm theo Dự án" if is_vi else "プロジェクト別労働時間分布")
                        st.plotly_chart(fig_pie, use_container_width=True)
                    with col_d2:
                        mgr_stats = df_result.groupby("Quản lý Việt Nam")["Giờ làm (h)"].sum().reset_index()
                        fig_bar = px.bar(mgr_stats, x="Quản lý Việt Nam", y="Giờ làm (h)", title="Tổng giờ làm theo Quản lý VN" if is_vi else "VN側管理者別総労働時間", color="Quản lý Việt Nam")
                        st.plotly_chart(fig_bar, use_container_width=True)
                else:
                    st.info("Không có dữ liệu để vẽ biểu đồ." if is_vi else "グラフを描画するデータがありません。")
            except ImportError:
                st.info("Không có dữ liệu để vẽ biểu đồ." if is_vi else "グラフを描画するデータがありません。")

        with _tab2:
            st.markdown("### 📊 Báo cáo tổng hợp (12 Tháng)" if st.session_state.get('lang', 'vi') == 'vi' else "### 📊 総合レポート (12ヶ月)")
            st.markdown("💡 *Chỉnh sửa bất kỳ ô nào (Cơ khí, Điều khiển, Mô phỏng, Khác, Số người, Giờ sở hữu...). Các ô % và Tiền sẽ tự động cập nhật.*" if st.session_state.get('lang', 'vi') == 'vi' else "💡 *任意のセル（機械、制御、シミュレーション、その他、人数、保有工数...）を編集できます。%と金額のセルは自動的に更新されます。*")
            
            @st.fragment
            def render_mos_report_tab():

                import json
                import os
                import datetime
                history_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mos_history.json')
                
                def format_val_init(x):
                    try:
                        if pd.isna(x) or str(x).strip() == '': return ''
                        v = float(x)
                        if v.is_integer():
                            return f"{int(v):,}".replace(',', '.')
                        else:
                            return f"{v:,.2f}".rstrip('0').rstrip('.').replace(',', 'X').replace('.', ',').replace('X', '.')
                    except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return x
                
                def format_pct_init(x):
                    try:
                        if pd.isna(x) or str(x).strip() == '': return ''
                        return str(int(round(float(x))))
                    except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return x
                
                def format_hour_init(x):
                    try:
                        if pd.isna(x) or str(x).strip() == '': return ''
                        v = float(x)
                        if v.is_integer():
                            return str(int(v))
                        else:
                            s = f"{v:.2f}".rstrip('0').rstrip('.')
                            return s.replace('.', ',') if st.session_state.get('lang', 'vi') == 'vi' else s
                    except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return x
                
                if "mos_report_df_v3" not in st.session_state or st.session_state.get("mos_report_version", -1) != st.session_state.get("mos_result_version", 0):
                    st.session_state["mos_report_version"] = st.session_state.get("mos_result_version", 0)
                    history_data = {}
                    if os.path.exists(history_file):
                        try:
                            with open(history_file, 'r', encoding='utf-8') as f:
                                history_data = json.load(f)
                        except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); pass
                        
                    tong_gio = 0
                    cat_coki, cat_dieukhien, cat_mophong, cat_khac = 0, 0, 0, 0
                    for idx, row in df_result.iterrows():
                        m_da_row = str(row.get('Mã dự án', '')).strip().upper()
                        if m_da_row.startswith('J') and 'Giờ làm gốc (h)' in row and pd.notna(row['Giờ làm gốc (h)']):
                            gio_lam = row['Giờ làm gốc (h)']
                        else:
                            gio_lam = row.get('Giờ làm (h)', 0)
                        try: gio_lam = float(str(gio_lam).strip().replace(',', '.'))
                        except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); gio_lam = 0
                        if gio_lam > 0:
                            tong_gio += gio_lam
                            phan_vung = str(row.get('Phân vùng', ''))
                            import unicodedata
                            pv_norm = unicodedata.normalize('NFD', phan_vung).encode('ascii', 'ignore').decode('utf-8').lower().replace(" ", "").replace("-", "").replace("_", "")
                            
                            if 'coki' in pv_norm or 'cokhi' in pv_norm or '機械' in phan_vung or 'cơ khí' in phan_vung.lower(): 
                                cat_coki += gio_lam
                            elif 'dieukhien' in pv_norm or 'dien' in pv_norm or 'mach' in pv_norm or '制御' in phan_vung or '電気' in phan_vung or '回路' in phan_vung or 'điều khiển' in phan_vung.lower() or 'điện' in phan_vung.lower() or 'mạch' in phan_vung.lower(): 
                                cat_dieukhien += gio_lam
                            elif 'mophong' in pv_norm or 'cae' in pv_norm or 'simulation' in pv_norm or 'シミュレーション' in phan_vung or 'mô phỏng' in phan_vung.lower(): 
                                cat_mophong += gio_lam
                            else: 
                                cat_khac += gio_lam
                            
                    kpi_m = str(st.session_state.get('mos_kpi_month', datetime.datetime.now().month))
                    
                    # Lấy dữ liệu đã được chỉnh sửa (override) từ bảng KPI ở Bước 2
                    default_nv = st.session_state.get('mos_num_people', 8)
                    default_std = st.session_state.get('mos_kpi_std_hours', 168)
                    num_ppl = float(st.session_state.get('override_mos_nv', default_nv))
                    std_hours = float(st.session_state.get('override_mos_std', default_std))
                    tong_gio = float(st.session_state.get('override_mos_actual', tong_gio))
                    
                    history_data[kpi_m] = {
                        'MOS注文人数': num_ppl,
                        'MOS保有工数': num_ppl * std_hours,
                        '総合': tong_gio,
                        '機械設計': cat_coki,
                        '制御設計': cat_dieukhien,
                        'シミュレーション': cat_mophong,
                        'その他': cat_khac
                    }
                    
                    rows = []
                    # Chỉ hiển thị 1 dòng cho tháng hiện tại
                    m_str = kpi_m
                    if m_str in history_data:
                        d = history_data[m_str]
                        
                        def s_f_h(val):
                            if pd.isna(val) or str(val).strip() == "": return 0.0
                            try: 
                                s = str(val).strip().replace(' ', '')
                                if '.' in s and ',' in s:
                                    s = s.replace('.', '').replace(',', '.')
                                elif ',' in s:
                                    s = s.replace(',', '.')
                                elif '.' in s:
                                    if s.count('.') > 1 or len(s.split('.')[-1]) == 3:
                                        s = s.replace('.', '')
                                return float(s)
                            except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return 0.0
                            
                        tong = s_f_h(d.get('総合', 0))
                        tien_thuc = tong * 2500
                        so_gio_sh = s_f_h(d.get('MOS保有工数', 0))
                        hieu_suat = (tong / so_gio_sh * 100) if so_gio_sh > 0 else 0
                        
                        rows.append({
                            "MonthNum": m_str,
                            "Tháng": f"{m_str}",
                            "Số người": s_f_h(d.get('MOS注文人数', 0)),
                            "Giờ sở hữu": so_gio_sh,
                            "Tổng giờ": tong,
                            "Cơ khí (h)": s_f_h(d.get('機械設計', 0)),
                            "Điều khiển (h)": s_f_h(d.get('制御設計', 0)),
                            "Mô phỏng (h)": s_f_h(d.get('シミュレーション', 0)),
                            "Khác (h)": s_f_h(d.get('その他', 0)),
                            "Đơn giá (JPY)": 2500.0,
                            "Tiền thực tế (JPY)": float(tien_thuc),
                            "Tiền ủy thác (JPY)": 2000000.0,
                            "Hiệu suất (%)": float(hieu_suat),
                            "Cơ khí (%)": (s_f_h(d.get('機械設計', 0))/tong*100) if tong > 0 else 0,
                            "Điều khiển (%)": (s_f_h(d.get('制御設計', 0))/tong*100) if tong > 0 else 0,
                            "Mô phỏng (%)": (s_f_h(d.get('シミュレーション', 0))/tong*100) if tong > 0 else 0,
                            "Khác (%)": (s_f_h(d.get('その他', 0))/tong*100) if tong > 0 else 0,
                        })
                    st.session_state.mos_report_df_v3 = pd.DataFrame(rows)
                    
                    is_vi_format = st.session_state.get('lang', 'vi') == 'vi'
                    
                    money_cols = ["Đơn giá (JPY)", "Tiền thực tế (JPY)", "Tiền ủy thác (JPY)"]
                    hour_cols = ["Số người", "Giờ sở hữu", "Tổng giờ", "Cơ khí (h)", "Điều khiển (h)", "Mô phỏng (h)", "Khác (h)"]
                    pct_cols = ["Hiệu suất (%)", "Cơ khí (%)", "Điều khiển (%)", "Mô phỏng (%)", "Khác (%)"]
                    for c in money_cols:
                        if c in st.session_state.mos_report_df_v3.columns:
                            st.session_state.mos_report_df_v3[c] = st.session_state.mos_report_df_v3[c].apply(format_val_init).astype(str)
                    for c in hour_cols:
                        if c in st.session_state.mos_report_df_v3.columns:
                            st.session_state.mos_report_df_v3[c] = st.session_state.mos_report_df_v3[c].apply(format_hour_init).astype(str)
                    for c in pct_cols:
                        if c in st.session_state.mos_report_df_v3.columns:
                            st.session_state.mos_report_df_v3[c] = st.session_state.mos_report_df_v3[c].apply(format_pct_init).astype(str)
                            
                    edit_cols = ["Số người", "Cơ khí (h)", "Điều khiển (h)", "Mô phỏng (h)", "Khác (h)"]
                    for c in edit_cols:
                        if c in st.session_state.mos_report_df_v3.columns:
                            st.session_state.mos_report_df_v3[c] = st.session_state.mos_report_df_v3[c].apply(format_hour_init).astype(str)

                is_vi = st.session_state.get('lang', 'vi') == 'vi'
                st.markdown("---")
                st.markdown("#### 📝 " + ("Chỉnh sửa nhanh thông tin Báo cáo tổng hợp" if is_vi else "総合レポート情報のクイック編集"))
                edit_mode = st.toggle("Bật chế độ chỉnh sửa bảng dữ liệu" if is_vi else "テーブル編集モードを有効にする", value=False)
                
                mos_col_cfg={
                    "MonthNum": None,
                    "Tháng": st.column_config.TextColumn("Tháng" if is_vi else "月", disabled=True),
                    "Số người": st.column_config.TextColumn("Số người" if is_vi else "人数"),
                    "Giờ sở hữu": st.column_config.TextColumn("Giờ sở hữu" if is_vi else "所定時間", disabled=True),
                    "Tổng giờ": st.column_config.TextColumn("Tổng giờ" if is_vi else "総時間", disabled=True),
                    "Cơ khí (h)": st.column_config.TextColumn("Cơ khí (h)" if is_vi else "機械(h)"),
                    "Điều khiển (h)": st.column_config.TextColumn("Điều khiển (h)" if is_vi else "制御(h)"),
                    "Mô phỏng (h)": st.column_config.TextColumn("Mô phỏng (h)" if is_vi else "シミュ(h)"),
                    "Khác (h)": st.column_config.TextColumn("Khác (h)" if is_vi else "その他(h)"),
                    "Đơn giá (JPY)": st.column_config.TextColumn("Đơn giá (JPY)" if is_vi else "単価(JPY)"),
                    "Tiền thực tế (JPY)": st.column_config.TextColumn("Tiền thực tế (JPY)" if is_vi else "実績金額(JPY)", disabled=True),
                    "Tiền ủy thác (JPY)": st.column_config.TextColumn("Tiền ủy thác (JPY)" if is_vi else "委託金額(JPY)", disabled=True),
                    "Hiệu suất": st.column_config.TextColumn("Hiệu suất" if is_vi else "効率(%)", disabled=True),
                    "Cơ khí %": st.column_config.TextColumn("Cơ khí %" if is_vi else "機械%", disabled=True),
                    "Điều khiển %": st.column_config.TextColumn("Điều khiển %" if is_vi else "制御%", disabled=True),
                    "Mô phỏng %": st.column_config.TextColumn("Mô phỏng %" if is_vi else "シミュ%", disabled=True),
                    "Khác %": st.column_config.TextColumn("Khác %" if is_vi else "その他%", disabled=True),
                }
                
                if not edit_mode:
                    st.dataframe(
                        st.session_state.mos_report_df_v3,
                        use_container_width=True,
                        hide_index=True,
                        column_config=mos_col_cfg
                    )
                else:
                    widget_key = f"mos_report_editor_widget_{st.session_state.get('lang', 'vi')}"
                    
                    edited_df = st.data_editor(
                        st.session_state.mos_report_df_v3,
                        key=widget_key,
                        use_container_width=True,
                        hide_index=True,
                        num_rows="fixed",
                        column_config=mos_col_cfg
                    )
                    
                    if st.button("💾 Lưu thay đổi" if is_vi else "💾 変更を保存", type="primary"):
                        if not edited_df.equals(st.session_state.mos_report_df_v3):
                            df = edited_df.copy()
                            def format_val_local(x):
                                try:
                                    if pd.isna(x) or str(x).strip() == '': return ''
                                    v = round(float(x), 1)
                                    return f"{int(v)}" if v.is_integer() else f"{v}"
                                except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return x
        
                            for r_idx in df.index:
                                # Re-calculate derived columns based on edited values
                                def s_f(v):
                                    if pd.isna(v) or str(v).strip() == "": return 0.0
                                    try: 
                                        s = str(v).strip().replace(' ', '')
                                        if '.' in s and ',' in s:
                                            s = s.replace('.', '').replace(',', '.')
                                        elif ',' in s:
                                            s = s.replace(',', '.')
                                        elif '.' in s:
                                            if s.count('.') > 1 or len(s.split('.')[-1]) == 3:
                                                s = s.replace('.', '')
                                        return float(s)
                                    except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return 0.0
                                    
                                ck = s_f(df.at[r_idx, "Cơ khí (h)"])
                                dk = s_f(df.at[r_idx, "Điều khiển (h)"])
                                mp = s_f(df.at[r_idx, "Mô phỏng (h)"])
                                kh = s_f(df.at[r_idx, "Khác (h)"])
                                sn = s_f(df.at[r_idx, "Số người"])
                                
                                tong = ck + dk + mp + kh
                                
                                df.at[r_idx, "Cơ khí (h)"] = format_hour_init(ck)
                                df.at[r_idx, "Điều khiển (h)"] = format_hour_init(dk)
                                df.at[r_idx, "Mô phỏng (h)"] = format_hour_init(mp)
                                df.at[r_idx, "Khác (h)"] = format_hour_init(kh)
                                df.at[r_idx, "Số người"] = format_hour_init(sn)
                                df.at[r_idx, "Tổng giờ"] = format_hour_init(tong)
                                
                                don_gia = s_f(df.at[r_idx, "Đơn giá (JPY)"])
                                df.at[r_idx, "Tiền thực tế (JPY)"] = format_val_init(tong * don_gia)
                                
                                so_gio_sh = s_f(df.at[r_idx, "Giờ sở hữu"])
                                df.at[r_idx, "Giờ sở hữu"] = format_hour_init(so_gio_sh)
                                df.at[r_idx, "Hiệu suất (%)"] = format_pct_init((tong / so_gio_sh * 100) if so_gio_sh > 0 else 0)
                                
                                df.at[r_idx, "Cơ khí (%)"] = format_pct_init((ck/tong*100) if tong > 0 else 0)
                                df.at[r_idx, "Điều khiển (%)"] = format_pct_init((dk/tong*100) if tong > 0 else 0)
                                df.at[r_idx, "Mô phỏng (%)"] = format_pct_init((mp/tong*100) if tong > 0 else 0)
                                df.at[r_idx, "Khác (%)"] = format_pct_init((kh/tong*100) if tong > 0 else 0)
                                
                            st.session_state.mos_report_df_v3 = df
                            st.rerun()
                

                if st.button("💾 Lưu Lịch Sử Lên Hệ Thống" if st.session_state.get('lang', 'vi') == 'vi' else "💾 システムに履歴を保存", type="primary"):
                    if os.path.exists(history_file):
                        try:
                            with open(history_file, 'r', encoding='utf-8') as f:
                                h_data = json.load(f)
                        except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); h_data = {}
                    else:
                        h_data = {}
                        
                    for idx, r in st.session_state.mos_report_df_v3.iterrows():
                        def safe_f(v):
                            if pd.isna(v) or str(v).strip() == "": return 0.0
                            try: 
                                s = str(v).strip().replace(' ', '')
                                if '.' in s and ',' in s:
                                    s = s.replace('.', '').replace(',', '.')
                                elif ',' in s:
                                    s = s.replace(',', '.')
                                elif '.' in s:
                                    if s.count('.') > 1 or len(s.split('.')[-1]) == 3:
                                        s = s.replace('.', '')
                                return float(s)
                            except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return 0.0
                            
                        tong = safe_f(r.get("Tổng giờ"))
                        if tong > 0:
                            m_str = str(r.get("MonthNum"))
                            h_data[m_str] = {
                                'MOS注文人数': safe_f(r.get("Số người")),
                                'MOS保有工数': safe_f(r.get("Giờ sở hữu")),
                                '総合': tong,
                                '機械設計': safe_f(r.get("Cơ khí (h)")),
                                '制御設計': safe_f(r.get("Điều khiển (h)")),
                                'シミュレーション': safe_f(r.get("Mô phỏng (h)")),
                                'その他': safe_f(r.get("Khác (h)"))
                            }
                    try:
                        with open(history_file, 'w', encoding='utf-8') as f:
                            json.dump(h_data, f, ensure_ascii=False, indent=2)
                        st.success("Lưu dữ liệu lịch sử MOS thành công!" if st.session_state.get('lang', 'vi') == 'vi' else "MOSの履歴データを正常に保存しました！")
                    except Exception as e:
                        st.error(f"Lỗi khi lưu lịch sử: {e}" if st.session_state.get('lang', 'vi') == 'vi' else f"履歴保存エラー: {e}")
                        
                with st.expander("👀 Xem lại dữ liệu lịch sử đã lưu (Dạng Bảng)" if st.session_state.get('lang', 'vi') == 'vi' else "👀 保存された履歴データを確認 (表形式)"):
                    if os.path.exists(history_file):
                        try:
                            with open(history_file, 'r', encoding='utf-8') as f:
                                h_data = json.load(f)
                            if h_data:
                                hist_rows = []
                                for m, data in h_data.items():
                                    row = {"Tháng": f"{m}"}
                                    row.update(data)
                                    hist_rows.append(row)
                                df_hist = pd.DataFrame(hist_rows)
                                df_hist = df_hist.rename(columns={
                                    'MOS注文人数': 'Số người',
                                    'MOS保有工数': 'Giờ sở hữu',
                                    '総合': 'Tổng giờ',
                                    '機械設計': 'Cơ khí (h)',
                                    '制御設計': 'Điều khiển (h)',
                                    'シミュレーション': 'Mô phỏng (h)',
                                    'その他': 'Khác (h)'
                                })
                                
                                # Format numbers for history table display
                                for c in df_hist.columns:
                                    if c != "Tháng":
                                        df_hist[c] = df_hist[c].apply(format_hour_init).astype(str)
                                
                                
                                st.dataframe(df_hist, use_container_width=True, hide_index=True, column_config=mos_col_cfg)
                                
                                st.markdown("<br>", unsafe_allow_html=True)
                                is_vn_lang = st.session_state.get('lang', 'vi') == 'vi'
                                col_del1, col_del2 = st.columns([1, 1])
                                with col_del1:
                                    with st.container(border=True):
                                        st.markdown("**Xóa từng tháng:**" if is_vn_lang else "**月別に削除:**")
                                        del_m = st.selectbox("Chọn tháng:" if is_vn_lang else "月を選択:", list(h_data.keys()), key="del_mos_hist_m")
                                        if st.button("🗑️ Xóa tháng này" if is_vn_lang else "🗑️ 選択した月を削除", use_container_width=True):
                                            st.session_state.confirm_del_mos_m = True
                                        if st.session_state.get('confirm_del_mos_m', False):
                                            st.warning("Xác nhận xóa tháng này?" if is_vn_lang else "削除を確認?")
                                            col_y, col_n = st.columns(2)
                                            if col_y.button("✔️ Có", key="del_mos_m_y"):
                                                try:
                                                    if del_m in h_data:
                                                        del h_data[del_m]
                                                        with open(history_file, 'w', encoding='utf-8') as f:
                                                            json.dump(h_data, f, ensure_ascii=False, indent=2)
                                                        st.session_state.confirm_del_mos_m = False
                                                        st.rerun()
                                                except Exception as e:
                                                    st.error(f"Lỗi khi xóa: {e}" if is_vn_lang else f"削除エラー: {e}")
                                            if col_n.button("❌ Hủy", key="del_mos_m_n"):
                                                st.session_state.confirm_del_mos_m = False
                                                st.rerun()
                                
                                with col_del2:
                                    with st.container(border=True):
                                        st.markdown("**Xóa tất cả:**" if is_vn_lang else "**すべて削除:**")
                                        st.markdown("<br>", unsafe_allow_html=True)
                                        if st.button("🚨 Xóa toàn bộ lịch sử" if is_vn_lang else "🚨 履歴をすべて削除", type="primary", use_container_width=True):
                                            st.session_state.confirm_del_mos_all = True
                                        if st.session_state.get('confirm_del_mos_all', False):
                                            st.warning("Bạn chắc chắn xóa toàn bộ?" if is_vn_lang else "全削除を確認?")
                                            col_y2, col_n2 = st.columns(2)
                                            if col_y2.button("✔️ Có", key="del_mos_all_y"):
                                                try:
                                                    if os.path.exists(history_file):
                                                        os.remove(history_file)
                                                    st.session_state.confirm_del_mos_all = False
                                                    st.rerun()
                                                except Exception as e:
                                                    st.error(f"Lỗi khi xóa: {e}" if is_vn_lang else f"削除エラー: {e}")
                                            if col_n2.button("❌ Hủy", key="del_mos_all_n"):
                                                st.session_state.confirm_del_mos_all = False
                                                st.rerun()
                            else:
                                st.info("Dữ liệu lịch sử hiện đang trống." if st.session_state.get('lang', 'vi') == 'vi' else "履歴データは現在空です。")
                        except Exception as e:
                            st.error(f"Lỗi đọc dữ liệu: {e}" if st.session_state.get('lang', 'vi') == 'vi' else f"データ読み込みエラー: {e}")
                    else:
                        st.info("Chưa có dữ liệu lịch sử nào được lưu trong hệ thống." if st.session_state.get('lang', 'vi') == 'vi' else "システムに保存された履歴データはありません。")

            render_mos_report_tab()


            # If not submitted but it exists in session, ensure it's mapped so download works correctly
            if 'df_mos_result' not in st.session_state:
                st.session_state['df_mos_result'] = df_result
            if 'df_mos_edited' not in st.session_state:
                st.session_state['df_mos_edited'] = df_result
            

            
            _cached_formula_dict = {}

            def set_formula(cell, formula, cached_val=None):
                cell.value = formula
                if cached_val is not None:
                    _cached_formula_dict[id(cell)] = cached_val

            import openpyxl.worksheet._writer as _writer
            if not getattr(_writer, '_has_cached_val_patch', False):
                _orig_write_cell = _writer.write_cell
                _g = _writer.write_cell.__globals__
                _set_attributes = _g['_set_attributes']
                _Element = _g['Element']
                _SubElement = _g['SubElement']
                _safe_string = _g['safe_string']
                _ArrayFormula = _g['ArrayFormula']
                _DataTableFormula = _g['DataTableFormula']
                _whitespace = _g['whitespace']
                _CellRichText = _g['CellRichText']

                def _patched_write_cell(xf, worksheet, cell, styled=None):
                    value, attributes = _set_attributes(cell, styled)
                    if value is None or value == '':
                        el = _Element('c', attributes)
                        xf.write(el)
                        return

                    if cell.data_type == 'f':
                        el = _Element('c', attributes)
                        attrib = {}
                        if isinstance(value, _ArrayFormula):
                            attrib = dict(value)
                            value = value.text
                        elif isinstance(value, _DataTableFormula):
                            attrib = dict(value)
                            value = None

                        formula = _SubElement(el, 'f', attrib)
                        if value is not None and not attrib.get('t') == 'dataTable':
                            formula.text = value[1:]
                            value = _cached_formula_dict.get(id(cell), None)

                        if value is not None:
                            cell_content = _SubElement(el, 'v')
                            cell_content.text = _safe_string(value)
                        xf.write(el)
                        return

                    _orig_write_cell(xf, worksheet, cell, styled)

                _writer.write_cell = _patched_write_cell
                _writer._has_cached_val_patch = True

            def to_excel(df, df_report=None):
                import openpyxl
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                import re
                import datetime
                import io
                
                font_normal = Font(name='Times New Roman', size=12)
                font_bold = Font(name='Times New Roman', size=12, bold=True)
                font_title = Font(name='Times New Roman', size=16, bold=True)
                
                align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
                align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
                
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                def format_bilingual(text):
                    if text is None: return ""
                    text_str = str(text).strip()
                    if text_str.lower() in ['nan', 'nat', 'none', '']: return ""
                    if '\n' not in text_str:
                        text_str = re.sub(r'([\u3040-\u30ff\u4e00-\u9faf])\s*(?:/|-)?\s*([A-Za-zÀ-ỹ])', r'\1\n\2', text_str)
                        match = re.match(r'^([A-Za-zÀ-ỹ\s/]+)\s*(?:/|-)?\s*([\u3040-\u30ff\u4e00-\u9faf\s/]+)$', text_str)
                        if match:
                            text_str = match.group(2).strip() + '\n' + match.group(1).strip()
                    return text_str

                def set_cell(cell, text, bold=False, align=align_center, fill=None, font=None):
                    cell.value = format_bilingual(text)
                    if font:
                        cell.font = font
                    else:
                        cell.font = font_bold if bold else font_normal
                    cell.alignment = align
                    if fill:
                        cell.fill = fill

                # --- Pre-calculate stats ---
                unique_people = set()
                if 'Người thực hiện' in df.columns:
                    for val in df['Người thực hiện']:
                        v_str = str(val).strip()
                        if v_str and v_str.lower() not in ['nan', 'nat', 'none']:
                            unique_people.add(v_str)
                num_people = len(unique_people)
                
                sum_hours = 0
                for idx, row in df.iterrows():
                    m_da_row = str(row.get('Mã dự án', '')).strip().upper()
                    if m_da_row.startswith('J') and 'Giờ làm gốc (h)' in row and pd.notna(row['Giờ làm gốc (h)']):
                        gio_lam = row['Giờ làm gốc (h)']
                    else:
                        gio_lam = row.get('Giờ làm (h)', '')
                    try: 
                        if str(gio_lam).strip() != '':
                            sum_hours += float(gio_lam)
                    except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); pass

                output = io.BytesIO()
                wb = openpyxl.Workbook()
                wb.calculation.fullCalcOnLoad = True
                ws = wb.active
                is_vi = st.session_state.get('lang', 'vi') == 'vi'
                ws.title = 'Tổng hợp MOS' if is_vi else 'MOS工数集計'
                
                now = datetime.datetime.now()
                
                kpi_month = st.session_state.get('mos_kpi_month', now.month)
                kpi_year = st.session_state.get('mos_kpi_year', now.year)
                kpi_std_hours = st.session_state.get('override_mos_std', st.session_state.get('mos_kpi_std_hours', 168))
                num_people_kpi = st.session_state.get('override_mos_nv', st.session_state.get('mos_num_people', num_people))
                kpi_target = st.session_state.get("override_mos_target", None)
                kpi_actual = st.session_state.get("override_mos_actual", sum_hours)
                year_jp = kpi_year - 2000 if kpi_year >= 2000 else kpi_year
                
                # --- Header Title ---
                ws.merge_cells('A1:J1')
                set_cell(ws['A1'], 'モス委託業務工数集計', font=font_title)
                
                ws.merge_cells('A2:J2')
                set_cell(ws['A2'], 'Bảng kê chi tiết nội dung nghiệp vụ ủy thác', font=font_title)
                
                ws.merge_cells('A4:J4')
                set_cell(ws['A4'], f'（{year_jp}年{kpi_month}月分）', font=font_title)
                
                ws.merge_cells('A5:J5')
                set_cell(ws['A5'], f'Phần tháng {kpi_month}/{kpi_year}', font=font_title)
                
                # --- Các bảng phụ bên trái ---
                sub_table_titles = [
                    ('A10', '人数\nSố người'),
                    ('A11', '一人当たり月稼働時間 (h)\nGiờ làm việc tiêu chuẩn(h)'),
                    ('A12', '月目標稼働時間 (h)\nMục tiêu giờ làm(h)'),
                    ('A13', '月実稼働時間 (h)\nGiờ làm thực tế(h)'),
                    ('A14', '目標に対して稼働率 (%)\nTỷ lệ hoàn thành(%)'),
                    ('A15', '実稼働に対して総金額\nTổng tiền theo giờ thực tế')
                ]
                for coord, title in sub_table_titles:
                    set_cell(ws[coord], title, align=align_right)
                    ws[coord].border = thin_border
                    ws[f'B{coord[1:]}'].border = thin_border
                    ws[f'B{coord[1:]}'].font = font_normal
                    ws[f'B{coord[1:]}'].alignment = align_center

                val_b10 = float(num_people_kpi)
                val_b11 = float(kpi_std_hours)
                val_b13 = float(kpi_actual)

                ws['B10'] = val_b10
                ws['B11'] = val_b11

                if kpi_target is not None:
                    val_b12 = float(kpi_target)
                    ws['B12'] = val_b12
                    target_ref = "B12"
                else:
                    val_b12 = val_b10 * val_b11
                    set_formula(ws['B12'], "=B10*B11", val_b12)
                    target_ref = "B12"

                ws['B13'] = val_b13

                val_b14 = (val_b13 / val_b12) if val_b12 > 0 else 0.0
                set_formula(ws['B14'], f"=IF({target_ref}>0, B13/{target_ref}, 0)", val_b14)
                ws['B14'].number_format = '0%'
                
                last_data_row = 20 + len(df)
                if len(df) > 0:
                    ws['B15'] = f"=SUM(F21:F{last_data_row})"
                else:
                    ws['B15'] = 0
                ws['B15'].number_format = '"¥"#,##0'
                ws['B15'].font = font_bold
                
                # --- Cột Ngày tháng bên phải ---
                ws.merge_cells('H10:J10')
                set_cell(ws['H10'], f'作成日付： {now.year}年{now.month}月{now.day}日', align=align_left)
                ws.merge_cells('H11:J11')
                set_cell(ws['H11'], f'Ngày lập bảng kê: {now.day}/{now.month}/{now.year}', align=align_left)
                ws.merge_cells('H12:J12')
                set_cell(ws['H12'], '作成者： レータンフォン', align=align_left)
                ws.merge_cells('H13:J13')
                set_cell(ws['H13'], 'Người lập bảng kê: Lê Thanh Phương', align=align_left)
                
                # --- Table Headers (Row 19 & 20) ---
                headers = [
                    ('A19:A20', '案件名\nTên dự án', 30),
                    ('B19:B20', '区分\nPhân vùng', 15),
                    ('C19:C20', '委託内容\nNội dung ủy thác', 50),
                    ('D19:D20', '実工数(h)\nGiờ làm (h)', 15),
                    ('E19:E20', '単価 (¥)\nĐơn giá (Yên)', 15),
                    ('F19:F20', '合計金額 (¥)\nTổng tiền (Yên)', 15),
                    ('G19:H19', '期間 Thời gian', None),
                    ('I19:J19', '管理 Người quản lý', None),
                    ('K19:K20', '実施者\nNgười thực hiện', 20),
                    ('L19:L20', '状態\nTrạng thái', 15)
                ]
                
                for range_str, text, width in headers:
                    ws.merge_cells(range_str)
                    cell = ws[range_str.split(':')[0]]
                    if text in ["期間 Thời gian", "管理者 Người quản lý"]:
                        cell.value = text
                        cell.font = font_bold
                        cell.alignment = align_center
                    else:
                        set_cell(cell, text, bold=True)
                    if width:
                        ws.column_dimensions[cell.column_letter].width = width
                
                sub_headers = {
                    'G20': '委託受領日\nNgày bắt đầu',
                    'H20': '完了日\nNgày kết thúc',
                    'I20': '日本\nNhật Bản',
                    'J20': 'ベトナム\nViệt Nam'
                }
                for c, text in sub_headers.items():
                    set_cell(ws[c], text, bold=True)
                    ws.column_dimensions[ws[c].column_letter].width = 15
                
                for r in ws['A19:L20']:
                    for cell in r:
                        cell.font = font_bold
                        cell.alignment = align_center
                        
                for range_str, text, _ in headers:
                    cell = ws[range_str.split(':')[0]]
                    if text in ["期間 Thời gian", "管理者 Người quản lý"]:
                        cell.value = text
                    else:
                        cell.value = format_bilingual(text)
                for c, text in sub_headers.items():
                    ws[c].value = format_bilingual(text)
                
                # --- Write Data ---
                row_idx = 21
                total_tien = 0.0
                total_gio_data = 0.0
                for idx, row in df.iterrows():
                    ma_da = str(row.get('Mã dự án', '')).strip()
                    raw_ten = str(row.get('Tên dự án', '')).strip()
                    formatted_ten = format_bilingual(raw_ten)
                    if ma_da:
                        lines = formatted_ten.split('\n')
                        ten_da = '\n'.join([f"{ma_da} {line}" for line in lines])
                    else:
                        ten_da = formatted_ten
                    if not ten_da.strip(): ten_da = ""
                    
                    set_cell(ws[f'A{row_idx}'], ten_da, align=align_left)
                    set_cell(ws[f'B{row_idx}'], row.get('Phân vùng', ''))
                    set_cell(ws[f'C{row_idx}'], row.get('Nội dung ủy thác', ''), align=align_left)
                    
                    m_da_row = str(row.get('Mã dự án', '')).strip().upper()
                    if m_da_row.startswith('J') and 'Giờ làm gốc (h)' in row and pd.notna(row['Giờ làm gốc (h)']):
                        gio_lam = row['Giờ làm gốc (h)']
                    else:
                        gio_lam = row.get('Giờ làm (h)', '')
                    
                    try:
                        val_gio = float(gio_lam) if pd.notna(gio_lam) and str(gio_lam).strip() != '' else 0.0
                    except:
                        val_gio = 0.0
                    total_gio_data += val_gio

                    ws[f'D{row_idx}'] = val_gio if val_gio > 0 else ""
                    ws[f'D{row_idx}'].font = font_normal
                    ws[f'D{row_idx}'].alignment = align_center
                    
                    val_dg = row.get('Đơn giá', 2500.0)
                    if pd.isna(val_dg) or val_dg == 0 or str(val_dg).strip() == '': val_dg = 2500.0
                    val_dg = float(val_dg)
                    ws[f'E{row_idx}'] = val_dg
                    ws[f'E{row_idx}'].font = font_normal
                    ws[f'E{row_idx}'].alignment = align_center
                    ws[f'E{row_idx}'].number_format = '#,##0 "¥"'
                    
                    tong_tien_row = val_gio * val_dg
                    total_tien += tong_tien_row
                    
                    # Cột tổng tiền = Giờ làm * Đơn giá (Dùng công thức Excel + Giá trị tính sẵn)
                    set_formula(ws[f'F{row_idx}'], f"=D{row_idx}*E{row_idx}", tong_tien_row)
                    ws[f'F{row_idx}'].font = font_normal
                    ws[f'F{row_idx}'].alignment = align_center
                    ws[f'F{row_idx}'].number_format = '#,##0 "¥"'
                    
                    set_cell(ws[f'G{row_idx}'], row.get('Ngày bắt đầu', ''))
                    set_cell(ws[f'H{row_idx}'], row.get('Ngày kết thúc', ''))
                    set_cell(ws[f'I{row_idx}'], row.get('Quản lý Nhật Bản', ''))
                    set_cell(ws[f'J{row_idx}'], row.get('Quản lý Việt Nam', ''))
                    set_cell(ws[f'K{row_idx}'], row.get('Người thực hiện', ''), align=align_left)
                    set_cell(ws[f'L{row_idx}'], row.get('Trạng thái', ''))
                    
                    row_idx += 1

                # --- Cập nhật B15 với công thức + giá trị tính sẵn ---
                last_data_row = row_idx - 1
                if len(df) > 0:
                    set_formula(ws['B15'], f"=SUM(F21:F{last_data_row})", total_tien)
                else:
                    ws['B15'] = 0
                ws['B15'].number_format = '"¥"#,##0'
                ws['B15'].font = font_bold

                # --- Footer ---
                ws.merge_cells(f'A{row_idx}:C{row_idx}')
                set_cell(ws[f'A{row_idx}'], '実工数合計(h)\nTổng giờ làm (h)', bold=True, align=align_right)
                
                if len(df) > 0:
                    set_formula(ws[f'D{row_idx}'], f"=SUM(D21:D{row_idx-1})", total_gio_data)
                    set_formula(ws[f'F{row_idx}'], f"=SUM(F21:F{row_idx-1})", total_tien)
                else:
                    ws[f'D{row_idx}'] = 0
                    ws[f'F{row_idx}'] = 0

                ws[f'D{row_idx}'].font = font_bold
                ws[f'D{row_idx}'].alignment = align_center
                
                ws[f'F{row_idx}'].font = font_bold
                ws[f'F{row_idx}'].alignment = align_center
                ws[f'F{row_idx}'].number_format = '#,##0 "¥"'
                
                # Kẻ khung toàn bộ bảng (từ row 18 đến row_idx)
                for row_cells in ws.iter_rows(min_row=18, max_row=row_idx, min_col=1, max_col=12):
                    for cell in row_cells:
                        cell.border = thin_border
                
                if df_report is not None:
                    export_excel_mos_aggregated(df_report, wb=wb)
                    
                if wb.views:
                    wb.views[0].windowWidth = 24000
                    wb.views[0].windowHeight = 12000
                    wb.views[0].xWindow = 240
                    wb.views[0].yWindow = 240
                wb.save(output)
                return output.getvalue()

            def export_excel_mos_aggregated(df_report, wb=None):
                import openpyxl
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                import io
                
                is_vi = st.session_state.get('lang', 'vi') == 'vi'
                sheet_title = 'Báo cáo 12 tháng' if is_vi else '12ヶ月集計'
                is_new_wb = False
                if wb is None:
                    wb = openpyxl.Workbook()
                    wb.calculation.fullCalcOnLoad = True
                    ws = wb.active
                    ws.title = sheet_title
                    is_new_wb = True
                else:
                    ws = wb.create_sheet(sheet_title)
                
                font_bold = Font(name='Times New Roman', size=12, bold=True)
                font_normal = Font(name='Times New Roman', size=12)
                align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                align_right = Alignment(horizontal='right', vertical='center')
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                ws.merge_cells('A1:A2')
                ws.merge_cells('B1:B2')
                ws['B1'] = 'MOS注文\n人数'
                ws.merge_cells('C1:C2')
                ws['C1'] = 'MOS保有\n工数(h)'
                
                ws.merge_cells('D1:H1')
                ws['D1'] = '実稼働工数(h)'
                ws['D2'] = '総合'
                ws['E2'] = '機械設計'
                ws['F2'] = '制御設計'
                ws['G2'] = 'シミュレーション'
                ws['H2'] = 'その他'
                
                ws.merge_cells('I1:I2')
                ws['I1'] = '単価\n(JPY)'
                ws.merge_cells('J1:J2')
                ws['J1'] = '実稼働に対して\n金額'
                ws.merge_cells('K1:K2')
                ws['K1'] = 'MOS月委託\n金額'
                ws.merge_cells('L1:L2')
                ws['L1'] = '保有に対して\n稼働率 (%)'
                
                ws.merge_cells('M1:P1')
                ws['M1'] = '実工数に対して各区分委託率 (%)'
                ws['M2'] = '機械設計'
                ws['N2'] = '制御設計'
                ws['O2'] = 'シミュレーション'
                ws['P2'] = 'その他'
                
                for row in ws['A1:P2']:
                    for cell in row:
                        cell.font = font_bold
                        cell.alignment = align_center
                        cell.border = thin_border
                
                for idx, r in df_report.iterrows():
                    row_idx = idx + 3
                    
                    ws[f'A{row_idx}'] = r.get("Tháng", "")
                    
                    def safe_num(val):
                        if pd.isna(val) or str(val).strip() == '': return 0.0
                        try:
                            if isinstance(val, (int, float)):
                                return float(val)
                            s = str(val).strip().replace(' ', '')
                            if '.' in s and ',' in s:
                                s = s.replace('.', '').replace(',', '.')
                            elif ',' in s:
                                s = s.replace(',', '.')
                            elif '.' in s:
                                if s.count('.') > 1 or len(s.split('.')[-1]) == 3:
                                    s = s.replace('.', '')
                            return float(s)
                        except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return 0.0
                        
                    ws[f'B{row_idx}'] = safe_num(r.get("Số người"))
                    ws[f'C{row_idx}'] = safe_num(r.get("Giờ sở hữu"))
                    ws[f'E{row_idx}'] = safe_num(r.get("Cơ khí (h)"))
                    ws[f'F{row_idx}'] = safe_num(r.get("Điều khiển (h)"))
                    ws[f'G{row_idx}'] = safe_num(r.get("Mô phỏng (h)"))
                    ws[f'H{row_idx}'] = safe_num(r.get("Khác (h)"))
                    ws[f'I{row_idx}'] = safe_num(r.get("Đơn giá (JPY)"))
                    
                    val_c = safe_num(r.get("Giờ sở hữu"))
                    val_e = safe_num(r.get("Cơ khí (h)"))
                    val_f = safe_num(r.get("Điều khiển (h)"))
                    val_g = safe_num(r.get("Mô phỏng (h)"))
                    val_h = safe_num(r.get("Khác (h)"))
                    val_i = safe_num(r.get("Đơn giá (JPY)"))
                    
                    val_d = val_e + val_f + val_g + val_h
                    val_j = val_d * val_i
                    val_k = safe_num(r.get("Tiền ủy thác (JPY)"))
                    val_l = (val_d / val_c) if val_c > 0 else 0.0
                    val_m = (val_e / val_d) if val_d > 0 else 0.0
                    val_n = (val_f / val_d) if val_d > 0 else 0.0
                    val_o = (val_g / val_d) if val_d > 0 else 0.0
                    val_p = (val_h / val_d) if val_d > 0 else 0.0

                    # Formulas with cached values
                    set_formula(ws[f'D{row_idx}'], f"=SUM(E{row_idx}:H{row_idx})", val_d)
                    set_formula(ws[f'J{row_idx}'], f"=D{row_idx}*I{row_idx}", val_j)
                    ws[f'K{row_idx}'] = val_k
                    set_formula(ws[f'L{row_idx}'], f'=IF(C{row_idx}>0, D{row_idx}/C{row_idx}, 0)', val_l)
                    set_formula(ws[f'M{row_idx}'], f'=IF(D{row_idx}>0, E{row_idx}/D{row_idx}, 0)', val_m)
                    set_formula(ws[f'N{row_idx}'], f'=IF(D{row_idx}>0, F{row_idx}/D{row_idx}, 0)', val_n)
                    set_formula(ws[f'O{row_idx}'], f'=IF(D{row_idx}>0, G{row_idx}/D{row_idx}, 0)', val_o)
                    set_formula(ws[f'P{row_idx}'], f'=IF(D{row_idx}>0, H{row_idx}/D{row_idx}, 0)', val_p)
                    
                    # Formatting
                    ws[f'I{row_idx}'].number_format = '#,##0'
                    ws[f'J{row_idx}'].number_format = '#,##0'
                    ws[f'K{row_idx}'].number_format = '#,##0'
                    ws[f'L{row_idx}'].number_format = '0%'
                    ws[f'M{row_idx}'].number_format = '0%'
                    ws[f'N{row_idx}'].number_format = '0%'
                    ws[f'O{row_idx}'].number_format = '0%'
                    ws[f'P{row_idx}'].number_format = '0%'
                    
                    for col in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']:
                        cell = ws[f'{col}{row_idx}']
                        cell.font = font_normal
                        cell.alignment = align_center
                        cell.border = thin_border
                
                ws.merge_cells('A15:I15')
                ws['A15'] = '合計'
                ws['A15'].alignment = align_right
                ws['A15'].font = font_bold
                
                tot_j = sum((safe_num(r.get("Cơ khí (h)")) + safe_num(r.get("Điều khiển (h)")) + safe_num(r.get("Mô phỏng (h)")) + safe_num(r.get("Khác (h)"))) * safe_num(r.get("Đơn giá (JPY)")) for _, r in df_report.iterrows()) if df_report is not None else 0.0
                tot_k = sum(safe_num(r.get("Tiền ủy thác (JPY)")) for _, r in df_report.iterrows()) if df_report is not None else 0.0

                set_formula(ws['J15'], f"=SUM(J3:J14)", tot_j)
                ws['J15'].number_format = '#,##0'
                set_formula(ws['K15'], f"=SUM(K3:K14)", tot_k)
                ws['K15'].number_format = '#,##0'
                
                for col in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P']:
                    cell = ws[f'{col}15']
                    cell.border = thin_border
                    cell.font = font_bold
                    if col in ['J', 'K']:
                        cell.alignment = align_center
                        
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
                for c in ['E','F','G','H']: ws.column_dimensions[c].width = 14
                ws.column_dimensions['I'].width = 12
                ws.column_dimensions['J'].width = 18
                ws.column_dimensions['K'].width = 18
                ws.column_dimensions['L'].width = 15
                for c in ['M','N','O','P']: ws.column_dimensions[c].width = 14
                if is_new_wb:
                    output = io.BytesIO()
                    if wb.views:
                        wb.views[0].windowWidth = 24000
                        wb.views[0].windowHeight = 12000
                        wb.views[0].xWindow = 240
                        wb.views[0].yWindow = 240
                    wb.save(output)
                    return output.getvalue()
                return None

            kpi_m = st.session_state.get('mos_kpi_month', datetime.datetime.now().month)
            kpi_y = st.session_state.get('mos_kpi_year', datetime.datetime.now().year)
            std_hours = st.session_state.get('mos_kpi_std_hours', 168)
            num_ppl = st.session_state.get('mos_num_people', 8)
            
            # Build 12-month dataframe for export by merging history and current UI state
            def build_12m_export_df():
                import json
                import os
                history_file = os.path.join(APP_DIR, 'mos_history.json') if 'APP_DIR' in globals() else os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mos_history.json')
                history_data = {}
                if os.path.exists(history_file):
                    try:
                        with open(history_file, 'r', encoding='utf-8') as f:
                            history_data = json.load(f)
                    except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); pass
                    
                if "mos_report_df_v3" in st.session_state:
                    for idx, r in st.session_state.mos_report_df_v3.iterrows():
                        if pd.notna(r.get("Tổng giờ")):
                            m_str = str(r.get("MonthNum"))
                            history_data[m_str] = {
                                'MOS注文人数': r.get("Số người"),
                                'MOS保有工数': r.get("Giờ sở hữu"),
                                '総合': r.get("Tổng giờ"),
                                '機械設計': r.get("Cơ khí (h)"),
                                '制御設計': r.get("Điều khiển (h)"),
                                'シミュレーション': r.get("Mô phỏng (h)"),
                                'その他': r.get("Khác (h)")
                            }
                            
                rows = []
                for m in range(1, 13):
                    m_str = str(m)
                    if m_str in history_data:
                        d = history_data[m_str]
                        
                        def s_f_h(val):
                            if pd.isna(val) or str(val).strip() == "": return 0.0
                            try:
                                if isinstance(val, (int, float)):
                                    return float(val)
                                s = str(val).strip().replace(' ', '')
                                if '.' in s and ',' in s:
                                    s = s.replace('.', '').replace(',', '.')
                                elif ',' in s:
                                    s = s.replace(',', '.')
                                elif '.' in s:
                                    if s.count('.') > 1 or len(s.split('.')[-1]) == 3:
                                        s = s.replace('.', '')
                                return float(s)
                            except Exception as e: logger.warning(f"Lỗi: {e}", exc_info=True); return 0.0
                            
                        tong = s_f_h(d.get('総合', 0))
                        tien_thuc = tong * 2500
                        so_gio_sh = s_f_h(d.get('MOS保有工数', 0))
                        hieu_suat = (tong / so_gio_sh * 100) if so_gio_sh > 0 else 0
                        
                        rows.append({
                            "Tháng": f"{m}月",
                            "Số người": s_f_h(d.get('MOS注文人数', 0)),
                            "Giờ sở hữu": so_gio_sh,
                            "Tổng giờ": tong,
                            "Cơ khí (h)": s_f_h(d.get('機械設計', 0)),
                            "Điều khiển (h)": s_f_h(d.get('制御設計', 0)),
                            "Mô phỏng (h)": s_f_h(d.get('シミュレーション', 0)),
                            "Khác (h)": s_f_h(d.get('その他', 0)),
                            "Đơn giá (JPY)": 2500.0,
                            "Tiền thực tế (JPY)": float(tien_thuc),
                            "Tiền ủy thác (JPY)": 2000000.0,
                            "Hiệu suất (%)": round(float(hieu_suat)),
                            "Cơ khí (%)": round((s_f_h(d.get('機械設計', 0))/tong*100) if tong > 0 else 0),
                            "Điều khiển (%)": round((s_f_h(d.get('制御設計', 0))/tong*100) if tong > 0 else 0),
                            "Mô phỏng (%)": round((s_f_h(d.get('シミュレーション', 0))/tong*100) if tong > 0 else 0),
                            "Khác (%)": round((s_f_h(d.get('その他', 0))/tong*100) if tong > 0 else 0),
                        })
                    else:
                        rows.append({
                            "Tháng": f"{m}月",
                            "Số người": None, "Giờ sở hữu": None, "Tổng giờ": None,
                            "Cơ khí (h)": None, "Điều khiển (h)": None, "Mô phỏng (h)": None, "Khác (h)": None,
                            "Đơn giá (JPY)": None, "Tiền thực tế (JPY)": None, "Tiền ủy thác (JPY)": None,
                            "Hiệu suất (%)": None, "Cơ khí (%)": None, "Điều khiển (%)": None, "Mô phỏng (%)": None, "Khác (%)": None,
                        })
                return pd.DataFrame(rows)
                
            with st.spinner("⏳ Đang kết xuất dữ liệu ra file Excel..."):
                excel_data = to_excel(st.session_state['df_mos_edited'], df_report=build_12m_export_df())
            
            for t_idx, tab_obj in enumerate([_tab0, _tab1]):
                with tab_obj:
                    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
                    st.markdown("---")
                    st.download_button(
                        label="📥 Tải dữ liệu Excel" if st.session_state.get('lang', 'vi') == 'vi' else "📥 Excelデータをダウンロード",
                        data=excel_data,
                        file_name=f"{kpi_m}月委託業務工数集計.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key=f"dl_mos_multi_sheet_tab_{t_idx}"
                    )



if __name__ == '__main__':
    render_mos_page()
